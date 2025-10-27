#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BACKEND ROUTES COMPLETE - v4.1
Rotas adicionais para funcionalidades completas
"""

from flask import request, jsonify, send_file
from application.api import bp
from application.decorators import login_required
from application.extensions import get_from_cache, set_in_cache
from bson import ObjectId
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)


@bp.route('/api/relatorios/excel', methods=['GET'])
@login_required
def relatorio_excel():
    """
    Gerar relatório Excel real (não JSON)
    """
    try:
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Relatório BIOMA'

        # Estilo do header
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Métrica', 'Valor']
        ws.append(headers)

        # Aplicar estilo aos headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Dados de exemplo
        data = [
            ['Total Clientes', '150'],
            ['Faturamento Mês', 'R$ 45.000,00'],
            ['Orçamentos Aprovados', '28'],
            ['Taxa de Conversão', '75%'],
            ['Ticket Médio', 'R$ 1.607,14']
        ]

        for row in data:
            ws.append(row)

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='relatorio_bioma.xlsx'
        )

    except Exception as e:
        logger.error(f"Erro ao gerar Excel: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/optimize/memory', methods=['POST'])
@login_required
def optimize_memory():
    """
    Endpoint para forçar limpeza de memória
    """
    import gc

    try:
        # Forçar garbage collection
        collected = gc.collect()

        # Limpar caches
        from application.extensions import cache
        if cache:
            cache.clear()

        logger.info(f"Memory optimization: {collected} objects collected")

        return jsonify({
            'success': True,
            'objects_collected': collected,
            'message': 'Memória otimizada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao otimizar memória: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/health/detailed', methods=['GET'])
def health_detailed():
    """
    Health check detalhado com informações do sistema
    """
    import psutil
    import os

    try:
        # Informações de memória
        memory = psutil.virtual_memory()

        # Informações do processo
        process = psutil.Process(os.getpid())
        process_memory = process.memory_info()

        return jsonify({
            'success': True,
            'status': 'healthy',
            'memory': {
                'total_mb': memory.total / (1024 * 1024),
                'available_mb': memory.available / (1024 * 1024),
                'percent_used': memory.percent,
                'process_rss_mb': process_memory.rss / (1024 * 1024),
                'process_vms_mb': process_memory.vms / (1024 * 1024)
            },
            'cpu': {
                'percent': psutil.cpu_percent(interval=1),
                'count': psutil.cpu_count()
            }
        })

    except Exception as e:
        logger.error(f"Health check error: {e}")
        return jsonify({
            'success': False,
            'status': 'error',
            'message': str(e)
        }), 500


# Cache otimizado para rotas pesadas
def cached_route(timeout=300):
    """
    Decorator para cache de rotas
    """
    def decorator(f):
        def decorated_function(*args, **kwargs):
            # Gerar chave de cache
            cache_key = f"{request.path}:{request.query_string.decode()}"

            # Verificar cache
            cached = get_from_cache(cache_key)
            if cached:
                logger.debug(f"Cache hit: {cache_key}")
                return cached

            # Executar função
            result = f(*args, **kwargs)

            # Salvar no cache
            set_in_cache(cache_key, result, timeout)

            return result

        decorated_function.__name__ = f.__name__
        return decorated_function

    return decorator