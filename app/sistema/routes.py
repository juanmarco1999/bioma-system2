#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BIOMA v3.7 - Sistema Routes
Auto-gerado pelo script de migração
"""

from flask import request, jsonify, session, current_app, send_file, render_template
from bson import ObjectId
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import logging
import io
import csv
import json

from app.sistema import bp
from app.decorators import login_required, permission_required, get_user_permissions
from app.utils import convert_objectid, allowed_file, registrar_auditoria
from app.extensions import db as get_db, get_from_cache, set_in_cache

logger = logging.getLogger(__name__)

@bp.route('/')
def index():
    return render_template('index.html')

@bp.route('/health')


@bp.route('/api/users', methods=['GET'])
@login_required
@permission_required('Admin')
def list_users():
    """Listar todos os usuários do sistema (Admin only)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        users = list(db.users.find({}, {'password': 0}).sort('name', ASCENDING))
        return jsonify({'success': True, 'users': convert_objectid(users)})

    except Exception as e:
        logger.error(f"Erro ao listar usuários: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/users/<id>/tipo-acesso', methods=['PUT'])


@bp.route('/api/users/<id>', methods=['GET'])
@login_required
def get_user(id):
    """Obter detalhes de um usuário específico"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        user = db.users.find_one({'_id': ObjectId(id)}, {'password': 0})
        if not user:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404

        return jsonify({'success': True, 'user': convert_objectid(user)})

    except Exception as e:
        logger.error(f"Erro ao buscar usuário: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/system/status')


@bp.route('/api/importar', methods=['POST'])
@login_required
def importar():
    if db is None:
        return jsonify({'success': False}), 500
    if 'file' not in request.files:
        return jsonify({'success': False}), 400
    file = request.files['file']
    tipo = request.form.get('tipo', 'produtos')
    if not file.filename:
        return jsonify({'success': False}), 400
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if ext not in ['csv', 'xlsx', 'xls']:
        return jsonify({'success': False}), 400
    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        count_success = 0
        count_error = 0
        rows = []
        if ext == 'csv':
            encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
            for encoding in encodings:
                try:
                    with open(filepath, 'r', encoding=encoding) as csvfile:
                        rows = list(csv.DictReader(csvfile))
                        break
                except:
                    continue
        else:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i in range(len(headers)):
                    if i < len(row) and row[i] is not None:
                        val = row[i]
                        if isinstance(val, str):
                            val = val.strip()
                        row_dict[headers[i]] = val
                if row_dict:
                    rows.append(row_dict)
            wb.close()
        if tipo == 'produtos':
            for idx, row in enumerate(rows, 1):
                try:
                    r = {k.lower().strip(): v for k, v in row.items() if k and v is not None}
                    if not r or all(not v for v in r.values()):
                        continue
                    nome = None
                    for col in ['nome', 'produto', 'name']:
                        if col in r and r[col]:
                            nome = str(r[col]).strip()
                            break
                    if not nome or len(nome) < 2:
                        count_error += 1
                        continue
                    marca = ''
                    for col in ['marca', 'brand']:
                        if col in r and r[col]:
                            marca = str(r[col]).strip()
                            break
                    sku = f"PROD-{count_success+1}"
                    for col in ['sku', 'codigo']:
                        if col in r and r[col]:
                            sku = str(r[col]).strip()
                            break
                    preco = 0.0
                    for col in ['preco', 'preço', 'price', 'valor']:
                        if col in r and r[col]:
                            try:
                                val = str(r[col]).replace('R$', '').strip()
                                if ',' in val:
                                    val = val.replace(',', '.')
                                preco = float(val)
                                break
                            except:
                                continue
                    if preco <= 0:
                        count_error += 1
                        continue
                    custo = 0.0
                    for col in ['custo', 'cost']:
                        if col in r and r[col]:
                            try:
                                val = str(r[col]).replace('R$', '').strip()
                                if ',' in val:
                                    val = val.replace(',', '.')
                                custo = float(val)
                                break
                            except:
                                continue
                    estoque = 0
                    for col in ['estoque', 'quantidade', 'qtd']:
                        if col in r and r[col]:
                            try:
                                estoque = int(float(r[col]))
                                break
                            except:
                                continue
                    categoria = 'Produto'
                    for col in ['categoria', 'category']:
                        if col in r and r[col]:
                            categoria = str(r[col]).strip().title()
                            break
                    db.produtos.insert_one({'nome': nome, 'marca': marca, 'sku': sku, 'preco': preco, 'custo': custo, 'estoque': estoque, 'estoque_minimo': 5, 'categoria': categoria, 'ativo': True, 'created_at': datetime.now()})
                    count_success += 1
                except:
                    count_error += 1
        elif tipo == 'servicos':
            # Importação de serviços
            for idx, row in enumerate(rows, 1):
                try:
                    r = {k.lower().strip(): v for k, v in row.items() if k and v is not None}
                    if not r or all(not v for v in r.values()):
                        continue
                    
                    # Nome do serviço
                    nome = None
                    for col in ['nome', 'servico', 'name']:
                        if col in r and r[col]:
                            nome = str(r[col]).strip()
                            break
                    if not nome or len(nome) < 2:
                        count_error += 1
                        continue
                    
                    # Categoria
                    categoria = 'Serviço'
                    for col in ['categoria', 'category']:
                        if col in r and r[col]:
                            categoria = str(r[col]).strip().title()
                            break
                    
                    # Preços por tamanho
                    tamanhos_map = {
                        'kids': ['kids', 'crianca', 'infantil'],
                        'masculino': ['masculino', 'male', 'homem'],
                        'curto': ['curto', 'short'],
                        'medio': ['medio', 'médio', 'medium'],
                        'longo': ['longo', 'long'],
                        'extra_longo': ['extra_longo', 'extra longo', 'extralongo', 'extralong']
                    }
                    
                    tamanhos_precos = {}
                    for tamanho_key, col_aliases in tamanhos_map.items():
                        preco = 0.0
                        for col_alias in col_aliases:
                            if col_alias in r and r[col_alias]:
                                try:
                                    val = str(r[col_alias]).replace('R$', '').strip()
                                    if ',' in val:
                                        val = val.replace(',', '.')
                                    preco = float(val)
                                    break
                                except:
                                    continue
                        if preco > 0:
                            tamanhos_precos[tamanho_key] = preco
                    
                    # Se não há nenhum preço válido, erro
                    if not tamanhos_precos:
                        count_error += 1
                        continue
                    
                    # Criar um serviço para cada tamanho com preço definido
                    tamanhos_labels = {
                        'kids': 'Kids',
                        'masculino': 'Masculino',
                        'curto': 'Curto',
                        'medio': 'Médio',
                        'longo': 'Longo',
                        'extra_longo': 'Extra Longo'
                    }
                    
                    for tamanho_key, preco in tamanhos_precos.items():
                        tamanho_label = tamanhos_labels.get(tamanho_key, tamanho_key.title())
                        sku = f"{nome.upper().replace(' ', '-')}-{tamanho_label.upper().replace(' ', '-')}"
                        
                        db.servicos.insert_one({
                            'nome': nome,
                            'sku': sku,
                            'tamanho': tamanho_label,
                            'preco': preco,
                            'categoria': categoria,
                            'duracao': 60,
                            'ativo': True,
                            'created_at': datetime.now()
                        })
                    
                    count_success += 1
                except Exception as e:
                    logger.error(f"Erro ao importar serviço: {e}")
                    count_error += 1
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': True, 'message': f'{count_success} importados!', 'count_success': count_success, 'count_error': count_error})
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': False}), 500


@bp.route('/api/estoque/importar', methods=['POST'])


@bp.route('/api/template/download/<tipo>')
@login_required
def download_template(tipo):
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    if tipo == 'produtos':
        ws.title = 'Produtos BIOMA'
        headers = ['nome','marca','sku','preco','custo','estoque','estoque_minimo','categoria']
    elif tipo == 'clientes':
        ws.title = 'Clientes BIOMA'
        headers = ['cpf','nome','email','telefone','genero','data_nascimento','endereco']
    elif tipo == 'profissionais':
        ws.title = 'Profissionais BIOMA'
        headers = ['cpf','nome','especialidade','email','telefone','comissao_perc','assistente_id','comissao_assistente_perc']
    else:
        # serviços (padrão)
        ws.title = 'Servicos BIOMA'
        headers = ['nome','tamanho','sku','preco','categoria','duracao_min']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    # largura de colunas básica (A..I)
    for i, col in enumerate(['A','B','C','D','E','F','G','H','I'], start=1):
        try:
            ws.column_dimensions[col].width = 18
        except Exception:
            pass
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'template_{tipo}_bioma.xlsx')



@bp.route('/api/clientes/formularios', methods=['GET'])


@bp.route('/api/config', methods=['GET', 'POST'])
@login_required
def config():
    if db is None:
        return jsonify({'success': False}), 500
    if request.method == 'GET':
        try:
            cfg = db.config.find_one({'key': 'unidade'}) or {}
            return jsonify({'success': True, 'config': convert_objectid(cfg)})
        except:
            return jsonify({'success': False}), 500
    data = request.json
    try:
        # Atualizar configurações incluindo logo_url
        config_data = {
            'key': 'unidade',
            'nome_empresa': data.get('nome_empresa', 'BIOMA UBERABA'),
            'logo_url': data.get('logo_url', ''),
            'logo_login_url': data.get('logo_login_url', ''),
            'endereco': data.get('endereco', ''),
            'telefone': data.get('telefone', ''),
            'email': data.get('email', ''),
            'cnpj': data.get('cnpj', ''),
            'cor_primaria': data.get('cor_primaria', '#7C3AED'),
            'cor_secundaria': data.get('cor_secundaria', '#EC4899'),
            'updated_at': datetime.now()
        }
        
        db.config.update_one({'key': 'unidade'}, {'$set': config_data}, upsert=True)
        logger.info("✅ Configurações atualizadas")
        return jsonify({'success': True, 'message': 'Configurações salvas com sucesso!'})
    except Exception as e:
        logger.error(f"Erro ao salvar configurações: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
        return jsonify({'success': False}), 500

@bp.route('/api/relatorios/completo', methods=['GET'])


@bp.route('/api/auditoria', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def consultar_auditoria():
    """Consultar logs de auditoria do sistema (Admin/Gestão only)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        # Paginação
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        skip = (page - 1) * per_page
        
        # Filtros
        filtro = {}
        
        # Filtro por usuário
        username = request.args.get('username')
        if username:
            filtro['username'] = {'$regex': username, '$options': 'i'}
        
        # Filtro por ação
        acao = request.args.get('acao')
        if acao:
            filtro['acao'] = acao
        
        # Filtro por entidade
        entidade = request.args.get('entidade')
        if entidade:
            filtro['entidade'] = entidade
        
        # Filtro por data
        data_inicio = request.args.get('data_inicio')
        if data_inicio:
            try:
                filtro['timestamp'] = {'$gte': datetime.fromisoformat(data_inicio)}
            except:
                pass
        
        data_fim = request.args.get('data_fim')
        if data_fim:
            try:
                if 'timestamp' in filtro:
                    filtro['timestamp']['$lte'] = datetime.fromisoformat(data_fim)
                else:
                    filtro['timestamp'] = {'$lte': datetime.fromisoformat(data_fim)}
            except:
                pass
        
        # Total de registros
        total = db.auditoria.count_documents(filtro)
        
        # Buscar registros
        registros = list(db.auditoria.find(filtro)
                        .sort('timestamp', DESCENDING)
                        .skip(skip)
                        .limit(per_page))
        
        # Estatísticas rápidas
        stats = {
            'total_acoes': total,
            'acoes_por_tipo': list(db.auditoria.aggregate([
                {'$match': filtro},
                {'$group': {'_id': '$acao', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}}
            ])),
            'usuarios_ativos': list(db.auditoria.aggregate([
                {'$match': filtro},
                {'$group': {'_id': '$username', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}},
                {'$limit': 10}
            ]))
        }
        
        return jsonify({
            'success': True,
            'registros': convert_objectid(registros),
            'stats': stats,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': (total + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao consultar auditoria: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500



@bp.route('/api/upload/logo', methods=['POST'])
@login_required
def upload_logo():
    """Upload de logo da empresa (armazenado como base64, SEM arquivos externos)"""
    try:
        if 'logo' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['logo']
        tipo = request.form.get('tipo', 'principal')  # principal ou login

        if file.filename == '':
            return jsonify({'success': False, 'message': 'Arquivo vazio'}), 400

        if file and allowed_file(file.filename):
            import base64

            # Ler arquivo em memória (SEM salvar em /tmp)
            file_content = file.read()
            foto_base64 = base64.b64encode(file_content).decode('utf-8')

            # Determinar tipo MIME
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
            mime_types = {
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'webp': 'image/webp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            # Montar data URI (armazenado diretamente no MongoDB)
            data_uri = f"data:{mime_type};base64,{foto_base64}"

            # Salvar referência no banco COM data URI
            db.uploads.insert_one({
                'tipo': f'logo_{tipo}',
                'filename': secure_filename(file.filename),
                'data_uri': data_uri,  # Armazena base64 diretamente
                'mime_type': mime_type,
                'data_upload': datetime.now()
            })

            logger.info(f"✅ Logo {tipo} salvo como base64 no MongoDB (SEM arquivos externos)")

            return jsonify({
                'success': True,
                'message': 'Logo enviado com sucesso',
                'url': data_uri  # Retorna Data URI diretamente
            })

        return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido'}), 400
    except Exception as e:
        logger.error(f"Erro no upload de logo: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 1.1. Upload de Imagem (Alias para compatibilidade)
@bp.route('/api/upload/imagem', methods=['POST'])


@bp.route('/api/config/logo', methods=['GET'])
def get_logo():
    """Obter Data URI do logo configurado (sem arquivos externos)"""
    try:
        tipo = request.args.get('tipo', 'principal')
        logo = db.uploads.find_one({'tipo': f'logo_{tipo}'}, sort=[('data_upload', DESCENDING)])

        if logo:
            # Retorna data_uri em vez de url de arquivo
            return jsonify({'success': True, 'url': logo.get('data_uri')})
        return jsonify({'success': True, 'url': None})
    except Exception as e:
        logger.error(f"Erro ao obter logo: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 3. Servir Arquivos Uploaded (DEPRECATED - Sistema usa Data URI agora)
@bp.route('/uploads/<filename>')


@bp.route('/api/upload/foto-profissional', methods=['POST'])
@login_required
def upload_foto_profissional_form():
    """Upload de foto de perfil de profissional via form data (armazenado como base64, SEM arquivos externos)"""
    try:
        if 'foto' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['foto']
        profissional_id = request.form.get('profissional_id')

        if not profissional_id:
            return jsonify({'success': False, 'message': 'ID do profissional não fornecido'}), 400

        if file.filename == '':
            return jsonify({'success': False, 'message': 'Arquivo vazio'}), 400

        if file and allowed_file(file.filename):
            import base64

            # Ler arquivo em memória (SEM salvar em /tmp)
            file_content = file.read()
            foto_base64 = base64.b64encode(file_content).decode('utf-8')

            # Determinar tipo MIME
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
            mime_types = {
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'webp': 'image/webp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            # Montar data URI (armazenado diretamente no MongoDB)
            data_uri = f"data:{mime_type};base64,{foto_base64}"

            # Atualizar profissional com a foto (Data URI em vez de URL de arquivo)
            db.profissionais.update_one(
                {'_id': ObjectId(profissional_id)},
                {'$set': {
                    'foto': data_uri,
                    'foto_url': data_uri,
                    'foto_atualizada_em': datetime.now()
                }}
            )

            # Salvar referência no banco COM data URI
            db.uploads.insert_one({
                'tipo': 'foto_profissional',
                'profissional_id': ObjectId(profissional_id),
                'filename': secure_filename(file.filename),
                'data_uri': data_uri,  # Armazena base64 diretamente
                'mime_type': mime_type,
                'data_upload': datetime.now()
            })

            logger.info(f"✅ Foto de profissional {profissional_id} salva como base64 no MongoDB (SEM arquivos externos)")

            return jsonify({
                'success': True,
                'message': 'Foto enviada com sucesso',
                'url': data_uri  # Retorna Data URI diretamente
            })

        return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido'}), 400
    except Exception as e:
        logger.error(f"Erro no upload de foto: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 5. Calcular Comissões Multiníveis
@bp.route('/api/comissoes/calcular-orcamento', methods=['POST'])


@bp.route('/api/admin/database-stats', methods=['GET'])
@login_required
def admin_database_stats():
    """Estatísticas do banco de dados (apenas Admin)"""
    try:
        # Verificar se o usuário é Admin
        user_id = session.get('user_id')
        user = db.users.find_one({'_id': ObjectId(user_id)})

        if not user or user.get('tipo_acesso') != 'Admin':
            return jsonify({'success': False, 'message': 'Apenas administradores'}), 403

        stats = {
            'clientes': db.clientes.count_documents({}),
            'profissionais': db.profissionais.count_documents({}),
            'servicos': db.servicos.count_documents({}),
            'produtos': db.produtos.count_documents({}),
            'agendamentos': db.agendamentos.count_documents({}),
            'fila': db.fila.count_documents({}),
            'orcamentos': db.orcamentos.count_documents({}),
            'contratos': db.contratos.count_documents({}),
            'anamneses': db.anamneses.count_documents({}),
            'prontuarios': db.prontuarios.count_documents({}),
            'usuarios': db.users.count_documents({}),
            'auditoria': db.auditoria.count_documents({}),
        }

        return jsonify({'success': True, 'stats': stats, 'total': sum(stats.values())})

    except Exception as e:
        logger.error(f"❌ Erro ao obter stats: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== FIM ROTAS ADMINISTRATIVAS ====================

# ==================== ROTAS ADICIONAIS ====================

@bp.route('/api/stream')
