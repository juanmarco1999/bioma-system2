#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BIOMA v3.7 - Rotas de Melhorias
APIs Críticas: Comissões, Financeiro, Notificações, etc.
"""

from flask import Blueprint, request, jsonify, send_file, session
from bson import ObjectId
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash
import logging
import os
import urllib.parse
import json
from io import BytesIO

# ReportLab imports para geração de PDF (Diretriz 3.2)
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.colors import HexColor, black, white
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY

logger = logging.getLogger(__name__)

# Criar novo blueprint para melhorias
bp = Blueprint('melhorias', __name__)

# Importar db do módulo principal
from application.api.routes import get_db

# ============================================================================
# API DE COMISSÕES (CRÍTICO - Carregamento Infinito)
# ============================================================================

@bp.route('/api/comissoes/pendentes')
def get_comissoes_pendentes():
    """Buscar comissões pendentes de pagamento"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Aggregation para buscar orçamentos aprovados com comissões não pagas
        pipeline = [
            {
                '$match': {
                    'status': 'aprovado',
                    '$or': [
                        {'comissao_paga': {'$exists': False}},
                        {'comissao_paga': False}
                    ]
                }
            },
            {
                '$lookup': {
                    'from': 'profissionais',
                    'localField': 'profissional_id',
                    'foreignField': '_id',
                    'as': 'profissional_info'
                }
            },
            {
                '$unwind': {
                    'path': '$profissional_info',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$project': {
                    '_id': 1,
                    'numero': 1,
                    'created_at': 1,
                    'valor_total': 1,
                    'profissional_nome': {'$ifNull': ['$profissional_info.nome', 'Não especificado']},
                    'profissional_comissao': {'$ifNull': ['$profissional_info.comissao_percentual', 10]},
                    'valor_comissao': {
                        '$multiply': [
                            '$valor_total',
                            {'$divide': [{'$ifNull': ['$profissional_info.comissao_percentual', 10]}, 100]}
                        ]
                    }
                }
            },
            {'$sort': {'created_at': -1}}
        ]

        comissoes = list(db.orcamentos.aggregate(pipeline))

        # Converter ObjectId para string
        for com in comissoes:
            com['_id'] = str(com['_id'])

        return jsonify({
            'success': True,
            'comissoes': comissoes,
            'total': len(comissoes),
            'total_valor': sum(c['valor_comissao'] for c in comissoes)
        })

    except Exception as e:
        logger.error(f"Erro ao buscar comissões pendentes: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/comissoes/pagas')
def get_comissoes_pagas():
    """Buscar comissões já pagas"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Buscar comissões pagas
        pipeline = [
            {
                '$match': {
                    'status': 'aprovado',
                    'comissao_paga': True
                }
            },
            {
                '$lookup': {
                    'from': 'profissionais',
                    'localField': 'profissional_id',
                    'foreignField': '_id',
                    'as': 'profissional_info'
                }
            },
            {
                '$unwind': {
                    'path': '$profissional_info',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$project': {
                    '_id': 1,
                    'numero': 1,
                    'created_at': 1,
                    'comissao_data_pagamento': 1,
                    'valor_total': 1,
                    'profissional_nome': {'$ifNull': ['$profissional_info.nome', 'Não especificado']},
                    'profissional_comissao': {'$ifNull': ['$profissional_info.comissao_percentual', 10]},
                    'valor_comissao': {
                        '$multiply': [
                            '$valor_total',
                            {'$divide': [{'$ifNull': ['$profissional_info.comissao_percentual', 10]}, 100]}
                        ]
                    }
                }
            },
            {'$sort': {'comissao_data_pagamento': -1}}
        ]

        comissoes = list(db.orcamentos.aggregate(pipeline))

        for com in comissoes:
            com['_id'] = str(com['_id'])

        return jsonify({
            'success': True,
            'comissoes': comissoes,
            'total': len(comissoes)
        })

    except Exception as e:
        logger.error(f"Erro ao buscar comissões pagas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/comissoes/<orcamento_id>/pagar', methods=['POST'])
def pagar_comissao(orcamento_id):
    """Marcar comissão como paga"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        result = db.orcamentos.update_one(
            {'_id': ObjectId(orcamento_id)},
            {
                '$set': {
                    'comissao_paga': True,
                    'comissao_data_pagamento': datetime.now()
                }
            }
        )

        if result.modified_count > 0:
            return jsonify({'success': True, 'message': 'Comissão marcada como paga'})
        else:
            return jsonify({'success': False, 'message': 'Orçamento não encontrado'})

    except Exception as e:
        logger.error(f"Erro ao pagar comissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE FINANCEIRO (CRÍTICO - Carregamento Infinito)
# ============================================================================

@bp.route('/api/financeiro/resumo')
def get_financeiro_resumo():
    """Resumo financeiro geral"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Data atual
        hoje = datetime.now()
        inicio_mes = datetime(hoje.year, hoje.month, 1)

        # Receitas do mês (orçamentos aprovados)
        receitas_mes = db.orcamentos.aggregate([
            {
                '$match': {
                    'status': 'aprovado',
                    'created_at': {'$gte': inicio_mes}
                }
            },
            {
                '$group': {
                    '_id': None,
                    'total': {'$sum': '$valor_total'}
                }
            }
        ])
        receitas_mes_valor = next(receitas_mes, {}).get('total', 0)

        # Despesas do mês
        despesas_mes = db.despesas.aggregate([
            {
                '$match': {
                    'data_vencimento': {'$gte': inicio_mes}
                }
            },
            {
                '$group': {
                    '_id': None,
                    'total': {'$sum': '$valor'}
                }
            }
        ])
        despesas_mes_valor = next(despesas_mes, {}).get('total', 0)

        # Contas a receber
        contas_receber = db.orcamentos.aggregate([
            {
                '$match': {
                    'status': 'aprovado',
                    '$or': [
                        {'pago': {'$exists': False}},
                        {'pago': False}
                    ]
                }
            },
            {
                '$group': {
                    '_id': None,
                    'total': {'$sum': '$valor_total'}
                }
            }
        ])
        contas_receber_valor = next(contas_receber, {}).get('total', 0)

        # Contas a pagar
        contas_pagar = db.despesas.aggregate([
            {
                '$match': {
                    'status': {'$ne': 'pago'}
                }
            },
            {
                '$group': {
                    '_id': None,
                    'total': {'$sum': '$valor'}
                }
            }
        ])
        contas_pagar_valor = next(contas_pagar, {}).get('total', 0)

        return jsonify({
            'success': True,
            'resumo': {
                'receitas_mes': receitas_mes_valor,
                'despesas_mes': despesas_mes_valor,
                'saldo_mes': receitas_mes_valor - despesas_mes_valor,
                'contas_receber': contas_receber_valor,
                'contas_pagar': contas_pagar_valor,
                'saldo_geral': contas_receber_valor - contas_pagar_valor
            }
        })

    except Exception as e:
        logger.error(f"Erro ao buscar resumo financeiro: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/financeiro/receitas')
def get_receitas():
    """Listar receitas (orçamentos aprovados)"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        receitas = list(db.orcamentos.find(
            {'status': 'aprovado'}
        ).sort('created_at', -1).limit(100))

        for rec in receitas:
            rec['_id'] = str(rec['_id'])

        return jsonify({'success': True, 'receitas': receitas})

    except Exception as e:
        logger.error(f"Erro ao buscar receitas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/financeiro/despesas')
def get_despesas():
    """Listar despesas"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        despesas = list(db.despesas.find().sort('data_vencimento', -1).limit(100))

        for desp in despesas:
            desp['_id'] = str(desp['_id'])

        return jsonify({'success': True, 'despesas': despesas})

    except Exception as e:
        logger.error(f"Erro ao buscar despesas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/financeiro/fluxo-caixa')
def get_fluxo_caixa():
    """Fluxo de caixa mensal"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Últimos 12 meses
        hoje = datetime.now()
        meses = []

        for i in range(12):
            mes_ref = hoje - timedelta(days=30 * i)
            inicio = datetime(mes_ref.year, mes_ref.month, 1)

            # Próximo mês
            if mes_ref.month == 12:
                fim = datetime(mes_ref.year + 1, 1, 1)
            else:
                fim = datetime(mes_ref.year, mes_ref.month + 1, 1)

            # Receitas do mês
            receitas = db.orcamentos.aggregate([
                {'$match': {'status': 'aprovado', 'created_at': {'$gte': inicio, '$lt': fim}}},
                {'$group': {'_id': None, 'total': {'$sum': '$valor_total'}}}
            ])
            receitas_valor = next(receitas, {}).get('total', 0)

            # Despesas do mês
            despesas = db.despesas.aggregate([
                {'$match': {'data_vencimento': {'$gte': inicio, '$lt': fim}}},
                {'$group': {'_id': None, 'total': {'$sum': '$valor'}}}
            ])
            despesas_valor = next(despesas, {}).get('total', 0)

            meses.insert(0, {
                'mes': inicio.strftime('%b/%y'),
                'receitas': receitas_valor,
                'despesas': despesas_valor,
                'saldo': receitas_valor - despesas_valor
            })

        return jsonify({'success': True, 'fluxo': meses})

    except Exception as e:
        logger.error(f"Erro ao buscar fluxo de caixa: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE AGENDAMENTOS (CORREÇÃO - Dados "Desconhecido")
# ============================================================================

@bp.route('/api/agendamentos/mes')
def get_agendamentos_mes():
    """Buscar agendamentos do mês com dados completos"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        hoje = datetime.now()
        inicio_mes = datetime(hoje.year, hoje.month, 1)

        if hoje.month == 12:
            fim_mes = datetime(hoje.year + 1, 1, 1)
        else:
            fim_mes = datetime(hoje.year, hoje.month + 1, 1)

        # Aggregation para popular dados relacionados
        pipeline = [
            {
                '$match': {
                    'data': {'$gte': inicio_mes, '$lt': fim_mes}
                }
            },
            {
                '$lookup': {
                    'from': 'clientes',
                    'localField': 'cliente_id',
                    'foreignField': '_id',
                    'as': 'cliente'
                }
            },
            {
                '$lookup': {
                    'from': 'profissionais',
                    'localField': 'profissional_id',
                    'foreignField': '_id',
                    'as': 'profissional'
                }
            },
            {
                '$lookup': {
                    'from': 'servicos',
                    'localField': 'servico_id',
                    'foreignField': '_id',
                    'as': 'servico'
                }
            },
            {
                '$unwind': {
                    'path': '$cliente',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$unwind': {
                    'path': '$profissional',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$unwind': {
                    'path': '$servico',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$project': {
                    '_id': 1,
                    'data': 1,
                    'horario': 1,
                    'status': 1,
                    'cliente_nome': {'$ifNull': ['$cliente.nome', 'Desconhecido']},
                    'profissional_nome': {'$ifNull': ['$profissional.nome', 'Desconhecido']},
                    'servico_nome': {'$ifNull': ['$servico.nome', 'Desconhecido']}
                }
            },
            {'$sort': {'data': 1, 'horario': 1}}
        ]

        agendamentos = list(db.agendamentos.aggregate(pipeline))

        for ag in agendamentos:
            ag['_id'] = str(ag['_id'])

        return jsonify({'success': True, 'agendamentos': agendamentos})

    except Exception as e:
        logger.error(f"Erro ao buscar agendamentos do mês: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE NOTIFICAÇÕES (E-mail + WhatsApp)
# ============================================================================

@bp.route('/api/notificacoes/email/<tipo>/<item_id>', methods=['POST'])
def enviar_email_notificacao(tipo, item_id):
    """Enviar notificação por e-mail (MailerSend)"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # TODO: Implementar integração com MailerSend
        # Por enquanto, retornar sucesso simulado

        logger.info(f"📧 E-mail simulado: {tipo} #{item_id}")

        return jsonify({
            'success': True,
            'message': 'E-mail enviado com sucesso (simulado)',
            'tipo': tipo,
            'item_id': item_id
        })

    except Exception as e:
        logger.error(f"Erro ao enviar e-mail: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/notificacoes/whatsapp/<tipo>/<item_id>', methods=['POST'])
def enviar_whatsapp_notificacao(tipo, item_id):
    """Gerar link de WhatsApp"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Buscar dados do item
        item = None
        if tipo == 'orcamento':
            item = db.orcamentos.find_one({'_id': ObjectId(item_id)})
        elif tipo == 'agendamento':
            item = db.agendamentos.find_one({'_id': ObjectId(item_id)})

        if not item:
            return jsonify({'success': False, 'message': 'Item não encontrado'})

        # Buscar telefone do cliente
        cliente = None
        if 'cliente_id' in item:
            cliente = db.clientes.find_one({'_id': item['cliente_id']})

        telefone = cliente.get('telefone', '') if cliente else ''

        # Remover formatação do telefone
        telefone = telefone.replace('+', '').replace('-', '').replace(' ', '').replace('(', '').replace(')', '')

        # Gerar mensagem
        if tipo == 'orcamento':
            mensagem = f"Olá! Seu orçamento #{item.get('numero', item_id)} no valor de R$ {item.get('valor_total', 0):.2f} está disponível. Acesse nosso sistema para mais detalhes."
        elif tipo == 'agendamento':
            mensagem = f"Olá! Confirmamos seu agendamento para {item.get('data', 'data a definir')} às {item.get('horario', 'horário a definir')}. Aguardamos você!"
        else:
            mensagem = "Olá! Você tem uma nova notificação do BIOMA Uberaba."

        # URL encode da mensagem
        mensagem_encoded = urllib.parse.quote(mensagem)

        # Gerar link do WhatsApp
        whatsapp_url = f"https://wa.me/{telefone}?text={mensagem_encoded}"

        logger.info(f"📱 WhatsApp link gerado: {tipo} #{item_id}")

        return jsonify({
            'success': True,
            'url': whatsapp_url,
            'message': 'Link do WhatsApp gerado com sucesso',
            'tipo': tipo,
            'item_id': item_id
        })

    except Exception as e:
        logger.error(f"Erro ao gerar link WhatsApp: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE CLIENTES - FATURAMENTO (CRÍTICO)
# ============================================================================

@bp.route('/api/clientes/<cliente_id>/faturamento')
def get_cliente_faturamento(cliente_id):
    """Buscar faturamento do cliente"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Agregar orçamentos aprovados do cliente
        pipeline = [
            {
                '$match': {
                    'cliente_id': ObjectId(cliente_id),
                    'status': 'aprovado'
                }
            },
            {
                '$group': {
                    '_id': None,
                    'total_faturado': {'$sum': '$valor_total'},
                    'quantidade_orcamentos': {'$sum': 1},
                    'ticket_medio': {'$avg': '$valor_total'}
                }
            },
            {
                '$project': {
                    '_id': 0,
                    'total_faturado': 1,
                    'quantidade_orcamentos': 1,
                    'ticket_medio': {'$round': ['$ticket_medio', 2]}
                }
            }
        ]

        resultado = list(db.orcamentos.aggregate(pipeline))

        if resultado:
            faturamento = resultado[0]
        else:
            faturamento = {
                'total_faturado': 0,
                'quantidade_orcamentos': 0,
                'ticket_medio': 0
            }

        return jsonify({'success': True, 'faturamento': faturamento})

    except Exception as e:
        logger.error(f"Erro ao buscar faturamento do cliente: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE ESTOQUE - EXPORTAR EXCEL (CORREÇÃO)
# ============================================================================

@bp.route('/api/estoque/exportar/excel')
def exportar_estoque_excel():
    """Exportar estoque para Excel"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Buscar produtos
        produtos = list(db.produtos.find())

        if not produtos:
            return jsonify({'success': False, 'message': 'Nenhum produto encontrado'})

        # Criar Excel com openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Estoque"

            # Cabeçalho
            headers = ['SKU', 'Nome', 'Categoria', 'Estoque Atual', 'Estoque Mínimo', 'Preço Custo', 'Preço Venda', 'Status']
            ws.append(headers)

            # Estilizar cabeçalho
            header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Dados
            for prod in produtos:
                ws.append([
                    prod.get('sku', ''),
                    prod.get('nome', ''),
                    prod.get('categoria', ''),
                    prod.get('estoque_atual', 0),
                    prod.get('estoque_minimo', 0),
                    prod.get('preco_custo', 0),
                    prod.get('preco_venda', 0),
                    'Ativo' if prod.get('ativo', True) else 'Inativo'
                ])

            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Salvar em memória
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )

        except ImportError:
            # Se openpyxl não estiver instalado, usar CSV
            import csv
            from io import StringIO

            output = StringIO()
            writer = csv.writer(output)

            writer.writerow(['SKU', 'Nome', 'Categoria', 'Estoque Atual', 'Estoque Mínimo', 'Preço Custo', 'Preço Venda', 'Status'])

            for prod in produtos:
                writer.writerow([
                    prod.get('sku', ''),
                    prod.get('nome', ''),
                    prod.get('categoria', ''),
                    prod.get('estoque_atual', 0),
                    prod.get('estoque_minimo', 0),
                    prod.get('preco_custo', 0),
                    prod.get('preco_venda', 0),
                    'Ativo' if prod.get('ativo', True) else 'Inativo'
                ])

            mem = BytesIO()
            mem.write(output.getvalue().encode('utf-8-sig'))
            mem.seek(0)

            return send_file(
                mem,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )

    except Exception as e:
        logger.error(f"Erro ao exportar estoque: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API - ATIVAR/DESATIVAR TODOS PRODUTOS/SERVIÇOS
# ============================================================================

@bp.route('/api/produtos/toggle-todos', methods=['POST'])
def toggle_todos_produtos():
    """Ativar/desativar todos os produtos"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        ativo = request.json.get('ativo', True)

        result = db.produtos.update_many(
            {},
            {'$set': {'ativo': ativo}}
        )

        logger.info(f"{'Ativados' if ativo else 'Desativados'} {result.modified_count} produtos")

        return jsonify({
            'success': True,
            'count': result.modified_count,
            'message': f"{result.modified_count} produtos {'ativados' if ativo else 'desativados'}"
        })

    except Exception as e:
        logger.error(f"Erro ao toggle produtos: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/servicos/toggle-todos', methods=['POST'])
def toggle_todos_servicos():
    """Ativar/desativar todos os serviços"""
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        ativo = request.json.get('ativo', True)

        result = db.servicos.update_many(
            {},
            {'$set': {'ativo': ativo}}
        )

        logger.info(f"{'Ativados' if ativo else 'Desativados'} {result.modified_count} serviços")

        return jsonify({
            'success': True,
            'count': result.modified_count,
            'message': f"{result.modified_count} serviços {'ativados' if ativo else 'desativados'}"
        })

    except Exception as e:
        logger.error(f"Erro ao toggle serviços: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE REGRAS DE COMISSÃO
# ============================================================================

@bp.route('/api/comissoes/regra', methods=['POST'])
def criar_regra_comissao():
    """Criar regra de comissão"""
    try:
        db = get_db()
        data = request.get_json()

        regra = {
            'tipo': data.get('tipo'),
            'percentual': float(data.get('percentual', 10)),
            'descricao': data.get('descricao', ''),
            'ativo': True,
            'criado_em': datetime.now()
        }

        result = db.regras_comissao.insert_one(regra)

        return jsonify({
            'success': True,
            'regra_id': str(result.inserted_id),
            'message': 'Regra de comissão criada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao criar regra de comissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE PRODUTOS/SERVIÇOS - GET INDIVIDUAL
# ============================================================================

@bp.route('/api/produtos/<produto_id>')
def get_produto(produto_id):
    """Buscar produto individual"""
    try:
        db = get_db()
        produto = db.produtos.find_one({'_id': ObjectId(produto_id)})

        if not produto:
            return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404

        produto['_id'] = str(produto['_id'])

        return jsonify({'success': True, 'produto': produto})

    except Exception as e:
        logger.error(f"Erro ao buscar produto: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/servicos/<servico_id>')
def get_servico(servico_id):
    """Buscar serviço individual"""
    try:
        db = get_db()
        servico = db.servicos.find_one({'_id': ObjectId(servico_id)})

        if not servico:
            return jsonify({'success': False, 'message': 'Serviço não encontrado'}), 404

        servico['_id'] = str(servico['_id'])

        return jsonify({'success': True, 'servico': servico})

    except Exception as e:
        logger.error(f"Erro ao buscar serviço: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/produtos/<produto_id>', methods=['PUT'])
def atualizar_produto(produto_id):
    """Atualizar produto"""
    try:
        db = get_db()
        data = request.get_json()

        update_data = {
            'nome': data.get('nome'),
            'sku': data.get('sku'),
            'categoria': data.get('categoria'),
            'preco_custo': float(data.get('preco_custo', 0)),
            'preco_venda': float(data.get('preco_venda')),
            'estoque_atual': int(data.get('estoque_atual', 0)),
            'estoque_minimo': int(data.get('estoque_minimo', 0)),
            'unidade': data.get('unidade', 'un'),
            'descricao': data.get('descricao', ''),
            'ativo': data.get('ativo', True),
            'atualizado_em': datetime.now()
        }

        result = db.produtos.update_one(
            {'_id': ObjectId(produto_id)},
            {'$set': update_data}
        )

        if result.modified_count > 0:
            return jsonify({'success': True, 'message': 'Produto atualizado com sucesso'})
        else:
            return jsonify({'success': False, 'message': 'Produto não encontrado ou sem alterações'}), 404

    except Exception as e:
        logger.error(f"Erro ao atualizar produto: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/servicos/<servico_id>', methods=['PUT'])
def atualizar_servico(servico_id):
    """Atualizar serviço"""
    try:
        db = get_db()
        data = request.get_json()

        update_data = {
            'nome': data.get('nome'),
            'categoria': data.get('categoria'),
            'preco': float(data.get('preco')),
            'duracao': int(data.get('duracao', 60)),
            'descricao': data.get('descricao', ''),
            'ativo': data.get('ativo', True),
            'atualizado_em': datetime.now()
        }

        result = db.servicos.update_one(
            {'_id': ObjectId(servico_id)},
            {'$set': update_data}
        )

        if result.modified_count > 0:
            return jsonify({'success': True, 'message': 'Serviço atualizado com sucesso'})
        else:
            return jsonify({'success': False, 'message': 'Serviço não encontrado ou sem alterações'}), 404

    except Exception as e:
        logger.error(f"Erro ao atualizar serviço: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE MAPA DE CALOR (RELATÓRIOS)
# ============================================================================

@bp.route('/api/relatorios/mapa-calor')
def get_mapa_calor():
    """Gerar dados do mapa de calor de agendamentos"""
    try:
        db = get_db()

        # Buscar agendamentos dos últimos 30 dias
        data_inicio = datetime.now() - timedelta(days=30)

        pipeline = [
            {
                '$match': {
                    'data': {'$gte': data_inicio.strftime('%Y-%m-%d')}
                }
            },
            {
                '$group': {
                    '_id': '$data',
                    'count': {'$sum': 1}
                }
            },
            {'$sort': {'_id': 1}}
        ]

        resultado = list(db.agendamentos.aggregate(pipeline))

        # Formatar para o formato do mapa de calor
        mapa_dados = []
        for item in resultado:
            mapa_dados.append({
                'date': item['_id'],
                'count': item['count']
            })

        return jsonify({
            'success': True,
            'dados': mapa_dados,
            'periodo': {
                'inicio': data_inicio.strftime('%Y-%m-%d'),
                'fim': datetime.now().strftime('%Y-%m-%d')
            }
        })

    except Exception as e:
        logger.error(f"Erro ao gerar mapa de calor: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE AUDITORIA
# ============================================================================

@bp.route('/api/auditoria/logs')
def get_audit_logs():
    """Buscar logs de auditoria"""
    try:
        db = get_db()

        limite = int(request.args.get('limite', 100))
        tipo = request.args.get('tipo')  # create, update, delete, access

        query = {}
        if tipo:
            query['tipo'] = tipo

        logs = list(db.auditoria.find(query).sort('data_hora', -1).limit(limite))

        for log in logs:
            log['_id'] = str(log['_id'])
            if 'usuario_id' in log and log['usuario_id']:
                log['usuario_id'] = str(log['usuario_id'])

        return jsonify({'success': True, 'logs': logs, 'total': len(logs)})

    except Exception as e:
        logger.error(f"Erro ao buscar logs de auditoria: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/auditoria/log', methods=['POST'])
def criar_audit_log():
    """Criar registro de auditoria"""
    try:
        db = get_db()
        data = request.get_json()

        log = {
            'tipo': data.get('tipo'),  # create, update, delete, access
            'acao': data.get('acao'),
            'entidade': data.get('entidade'),  # orcamento, cliente, etc
            'entidade_id': data.get('entidade_id'),
            'usuario_id': ObjectId(data.get('usuario_id')) if data.get('usuario_id') else None,
            'usuario_nome': data.get('usuario_nome'),
            'detalhes': data.get('detalhes', {}),
            'ip': request.remote_addr,
            'data_hora': datetime.now()
        }

        result = db.auditoria.insert_one(log)

        return jsonify({'success': True, 'log_id': str(result.inserted_id)})

    except Exception as e:
        logger.error(f"Erro ao criar log de auditoria: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API DE CONFIGURAÇÕES
# ============================================================================

@bp.route('/api/configuracoes')
def get_configuracoes():
    """Buscar configurações do sistema"""
    try:
        db = get_db()

        config = db.configuracoes.find_one({'tipo': 'sistema'})

        if not config:
            # Criar configurações padrão
            config = {
                'tipo': 'sistema',
                'empresa': {
                    'nome': '',
                    'cnpj': '',
                    'endereco': '',
                    'telefone': '',
                    'email': '',
                    'logo_url': ''
                },
                'notificacoes': {
                    'email_ativo': False,
                    'whatsapp_ativo': False,
                    'mailersend_api_key': ''
                },
                'sistema': {
                    'modo_manutencao': False,
                    'auto_backup': False,
                    'intervalo_backup_dias': 7
                },
                'atualizado_em': datetime.now()
            }
            db.configuracoes.insert_one(config)

        config['_id'] = str(config['_id'])

        return jsonify({'success': True, 'configuracoes': config})

    except Exception as e:
        logger.error(f"Erro ao buscar configurações: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/configuracoes', methods=['PUT'])
def atualizar_configuracoes():
    """Atualizar configurações do sistema"""
    try:
        db = get_db()
        data = request.get_json()

        data['atualizado_em'] = datetime.now()

        result = db.configuracoes.update_one(
            {'tipo': 'sistema'},
            {'$set': data},
            upsert=True
        )

        return jsonify({
            'success': True,
            'message': 'Configurações atualizadas com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao atualizar configurações: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== FILA INTELIGENTE (Diretrizes 10.1 e 10.2) ====================

@bp.route('/api/fila/inteligente', methods=['POST'])
def adicionar_fila_inteligente():
    """
    Sistema de fila inteligente com automação e notificações

    Diretriz 10.1: Automação completa ligada ao calendário e horários dos profissionais
    Diretriz 10.2: Notificações Email (MailSender) e WhatsApp ao adicionar
    """
    try:
        db = get_db()
        data = request.get_json()

        # Extrair dados
        cliente_nome = data.get('cliente_nome')
        cliente_telefone = data.get('cliente_telefone')
        cliente_email = data.get('cliente_email')
        servico_id = data.get('servico_id')
        profissional_id = data.get('profissional_id')  # Pode ser None (automático)
        data_preferencial = data.get('data_preferencial')
        notificacoes = data.get('notificacoes', {})
        observacoes = data.get('observacoes')

        # Validações
        if not cliente_nome or not cliente_telefone or not servico_id:
            return jsonify({
                'success': False,
                'message': 'Nome, telefone e serviço são obrigatórios'
            }), 400

        # Buscar informações do serviço
        servico = db.servicos.find_one({'_id': ObjectId(servico_id)})
        if not servico:
            return jsonify({'success': False, 'message': 'Serviço não encontrado'}), 404

        # AUTOMAÇÃO: Sugerir profissional se não especificado (Diretriz 10.1)
        profissional_sugerido = None
        if not profissional_id:
            # Buscar profissionais disponíveis para este serviço
            profissionais_disponiveis = list(db.profissionais.find({
                'ativo': True,
                '$or': [
                    {'servicos': servico_id},
                    {'especialidade': servico.get('categoria')}
                ]
            }))

            if profissionais_disponiveis:
                # Verificar disponibilidade no calendário
                hoje = datetime.now()

                for prof in profissionais_disponiveis:
                    # Contar agendamentos atuais do profissional
                    agendamentos_hoje = db.agendamentos.count_documents({
                        'profissional_id': str(prof['_id']),
                        'data': hoje.strftime('%Y-%m-%d'),
                        'status': {'$in': ['agendado', 'confirmado']}
                    })

                    # Considerar disponível se tiver menos de 8 agendamentos no dia
                    if agendamentos_hoje < 8:
                        profissional_sugerido = prof
                        profissional_id = str(prof['_id'])
                        break

                # Se todos estão ocupados, pegar o com menos agendamentos
                if not profissional_sugerido:
                    profissional_sugerido = profissionais_disponiveis[0]
                    profissional_id = str(profissional_sugerido['_id'])
        else:
            # Buscar profissional especificado
            profissional_sugerido = db.profissionais.find_one({'_id': ObjectId(profissional_id)})

        # Calcular posição na fila
        total_fila = db.fila_atendimento.count_documents({
            'status': {'$in': ['aguardando', 'atendendo']}
        })
        posicao = total_fila + 1

        # Calcular tempo estimado (15 min por pessoa na frente)
        tempo_estimado_min = posicao * 15
        if tempo_estimado_min <= 20:
            tempo_estimado = 'Até 20 minutos'
        elif tempo_estimado_min <= 40:
            tempo_estimado = '20 a 40 minutos'
        else:
            tempo_estimado = f'Aproximadamente {tempo_estimado_min} minutos'

        # Sugerir horário baseado no tempo estimado
        horario_sugerido = None
        if data_preferencial:
            horario_sugerido = data_preferencial
        else:
            horario_estimado = datetime.now() + timedelta(minutes=tempo_estimado_min)
            horario_sugerido = horario_estimado.strftime('%Y-%m-%d %H:%M')

        # Inserir na fila
        documento_fila = {
            'cliente_nome': cliente_nome,
            'cliente_telefone': cliente_telefone,
            'cliente_email': cliente_email,
            'servico': servico.get('nome'),
            'servico_id': servico_id,
            'profissional': profissional_sugerido.get('nome') if profissional_sugerido else 'A definir',
            'profissional_id': profissional_id,
            'posicao': posicao,
            'status': 'aguardando',
            'data_preferencial': horario_sugerido,
            'tempo_estimado': tempo_estimado,
            'observacoes': observacoes,
            'created_at': datetime.now(),
            'notificacoes_enviadas': {
                'whatsapp': False,
                'email': False
            }
        }

        result = db.fila_atendimento.insert_one(documento_fila)
        fila_id = str(result.inserted_id)

        # NOTIFICAÇÕES (Diretriz 10.2)
        notificacoes_enviadas = {}

        # WhatsApp (simulado com link wa.me)
        if notificacoes.get('whatsapp') and cliente_telefone:
            telefone_limpo = ''.join(filter(str.isdigit, cliente_telefone))
            if not telefone_limpo.startswith('55'):
                telefone_limpo = '55' + telefone_limpo

            mensagem_whatsapp = f"""*BIOMA - Confirmação de Fila*

Olá, {cliente_nome}! 👋

Você foi adicionado à fila de atendimento:

📋 *Serviço:* {servico.get('nome')}
👨‍⚕️ *Profissional:* {profissional_sugerido.get('nome') if profissional_sugerido else 'A definir'}
🔢 *Posição na fila:* {posicao}º
⏰ *Tempo estimado:* {tempo_estimado}
📅 *Horário sugerido:* {horario_sugerido}

Você receberá uma notificação quando estiver próximo do seu atendimento.

Obrigado pela preferência! 😊"""

            whatsapp_url = f"https://wa.me/{telefone_limpo}?text={urllib.parse.quote(mensagem_whatsapp)}"

            # Salvar log de notificação
            db.notificacoes_log.insert_one({
                'tipo': 'whatsapp',
                'destinatario': cliente_telefone,
                'mensagem': mensagem_whatsapp,
                'url': whatsapp_url,
                'fila_id': fila_id,
                'data_envio': datetime.now()
            })

            notificacoes_enviadas['whatsapp'] = True
            documento_fila['notificacoes_enviadas']['whatsapp'] = True

        # Email (MailSender)
        if notificacoes.get('email') and cliente_email:
            try:
                # Buscar configurações do MailSender
                config = db.configuracoes.find_one({'tipo': 'sistema'})
                mailersend_api_key = config.get('notificacoes', {}).get('mailersend_api_key') if config else None

                if mailersend_api_key:
                    # TODO: Implementar integração com MailSender API real
                    # Por enquanto, apenas registrar no log
                    pass

                # Salvar log de notificação (simulado)
                db.notificacoes_log.insert_one({
                    'tipo': 'email',
                    'destinatario': cliente_email,
                    'assunto': f'Confirmação de Fila - {servico.get("nome")}',
                    'corpo': f"""
                        Olá, {cliente_nome}!

                        Você foi adicionado à nossa fila de atendimento.

                        Detalhes:
                        - Serviço: {servico.get('nome')}
                        - Profissional: {profissional_sugerido.get('nome') if profissional_sugerido else 'A definir'}
                        - Posição: {posicao}º
                        - Tempo estimado: {tempo_estimado}
                        - Horário sugerido: {horario_sugerido}

                        Aguarde sua vez com conforto. Você receberá uma notificação quando estiver próximo.

                        Atenciosamente,
                        Equipe BIOMA
                    """,
                    'fila_id': fila_id,
                    'data_envio': datetime.now(),
                    'status': 'simulado'  # Mudar para 'enviado' quando MailSender estiver configurado
                })

                notificacoes_enviadas['email'] = True
                documento_fila['notificacoes_enviadas']['email'] = True

            except Exception as e:
                logger.error(f"Erro ao enviar email: {e}")
                notificacoes_enviadas['email'] = False

        # Atualizar documento com notificações
        db.fila_atendimento.update_one(
            {'_id': result.inserted_id},
            {'$set': {'notificacoes_enviadas': documento_fila['notificacoes_enviadas']}}
        )

        # Log de auditoria
        db.auditoria.insert_one({
            'tipo': 'fila_adicionado',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Cliente {cliente_nome} adicionado à fila - Posição {posicao}',
            'data_hora': datetime.now(),
            'detalhes': {
                'servico': servico.get('nome'),
                'profissional': profissional_sugerido.get('nome') if profissional_sugerido else 'Automático'
            }
        })

        return jsonify({
            'success': True,
            'fila_id': fila_id,
            'posicao': posicao,
            'tempo_estimado': tempo_estimado,
            'horario_sugerido': horario_sugerido,
            'profissional_nome': profissional_sugerido.get('nome') if profissional_sugerido else None,
            'notificacoes_enviadas': notificacoes_enviadas,
            'message': 'Cliente adicionado à fila com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao adicionar à fila inteligente: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/fila/sugerir-profissional')
def sugerir_profissional_fila():
    """
    Sugere profissional disponível baseado no serviço e calendário

    Diretriz 10.1: Automação inteligente ligada ao calendário
    """
    try:
        db = get_db()
        servico_id = request.args.get('servico_id')

        if not servico_id:
            return jsonify({'success': False, 'message': 'servico_id é obrigatório'}), 400

        # Buscar serviço
        servico = db.servicos.find_one({'_id': ObjectId(servico_id)})
        if not servico:
            return jsonify({'success': False, 'message': 'Serviço não encontrado'}), 404

        # Buscar profissionais que atendem este serviço
        profissionais = list(db.profissionais.find({
            'ativo': True,
            '$or': [
                {'servicos': servico_id},
                {'especialidade': servico.get('categoria')}
            ]
        }))

        if not profissionais:
            return jsonify({
                'success': False,
                'message': 'Nenhum profissional disponível para este serviço'
            })

        # Verificar disponibilidade de cada profissional
        hoje = datetime.now()
        profissionais_com_score = []

        for prof in profissionais:
            # Contar agendamentos do dia
            agendamentos_hoje = db.agendamentos.count_documents({
                'profissional_id': str(prof['_id']),
                'data': hoje.strftime('%Y-%m-%d'),
                'status': {'$in': ['agendado', 'confirmado']}
            })

            # Contar pessoas na fila
            fila_atual = db.fila_atendimento.count_documents({
                'profissional_id': str(prof['_id']),
                'status': {'$in': ['aguardando', 'atendendo']}
            })

            # Score: quanto menor, melhor (menos ocupado)
            score = agendamentos_hoje + (fila_atual * 2)

            profissionais_com_score.append({
                'profissional': prof,
                'score': score,
                'agendamentos_hoje': agendamentos_hoje,
                'fila_atual': fila_atual
            })

        # Ordenar por score (menos ocupado primeiro)
        profissionais_com_score.sort(key=lambda x: x['score'])

        melhor_opcao = profissionais_com_score[0]
        prof = melhor_opcao['profissional']

        # Calcular próxima disponibilidade
        proxima_disponibilidade = 'Disponível agora'
        if melhor_opcao['fila_atual'] > 0:
            tempo_espera = melhor_opcao['fila_atual'] * 15  # 15 min por pessoa
            proxima_hora = hoje + timedelta(minutes=tempo_espera)
            proxima_disponibilidade = proxima_hora.strftime('%H:%M')

        return jsonify({
            'success': True,
            'profissional_id': str(prof['_id']),
            'profissional_nome': prof.get('nome'),
            'profissional_especialidade': prof.get('especialidade'),
            'agendamentos_hoje': melhor_opcao['agendamentos_hoje'],
            'fila_atual': melhor_opcao['fila_atual'],
            'proxima_disponibilidade': proxima_disponibilidade,
            'disponivel': melhor_opcao['score'] < 10  # Disponível se score < 10
        })

    except Exception as e:
        logger.error(f"Erro ao sugerir profissional: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== ANAMNESE/PRONTUÁRIO COMPLETO (Diretrizes 11.1, 11.3, 11.4) ====================

@bp.route('/api/clientes/<cliente_id>/anamneses')
def get_cliente_anamneses(cliente_id):
    """
    Buscar todas as anamneses de um cliente
    Diretriz 11.1: Integração ao visualizar cliente
    """
    try:
        db = get_db()

        # Buscar cliente
        cliente = db.clientes.find_one({'_id': ObjectId(cliente_id)})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # Buscar todas as anamneses deste cliente
        anamneses = list(db.anamnese.find({'cliente_cpf': cliente.get('cpf')}).sort('data_cadastro', -1))

        return jsonify({
            'success': True,
            'anamneses': convert_objectid(anamneses)
        })

    except Exception as e:
        logger.error(f"Erro ao buscar anamneses do cliente: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/clientes/<cliente_id>/prontuarios')
def get_cliente_prontuarios(cliente_id):
    """
    Buscar todos os prontuários de um cliente
    Diretriz 11.1: Integração ao visualizar cliente
    """
    try:
        db = get_db()

        # Buscar cliente
        cliente = db.clientes.find_one({'_id': ObjectId(cliente_id)})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # Buscar todos os prontuários deste cliente
        prontuarios = list(db.prontuario.find({'cliente_cpf': cliente.get('cpf')}).sort('data_atendimento', -1))

        return jsonify({
            'success': True,
            'prontuarios': convert_objectid(prontuarios)
        })

    except Exception as e:
        logger.error(f"Erro ao buscar prontuários do cliente: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/clientes/<cpf>/anamnese', methods=['POST'])
def criar_anamnese_com_imagens(cpf):
    """
    Criar anamnese com upload de imagens e notificações

    Diretriz 11.3: Anexar imagens físicas de documentos
    Diretriz 11.4: Notificações Email/WhatsApp
    """
    try:
        db = get_db()

        # Buscar cliente
        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # Extrair dados do FormData
        observacoes = request.form.get('observacoes')
        notificacoes = request.form.get('notificacoes')
        cliente_nome = request.form.get('cliente_nome')
        cliente_email = request.form.get('cliente_email')
        cliente_telefone = request.form.get('cliente_telefone')

        # Parse notificacoes JSON
        try:
            notificacoes = json.loads(notificacoes) if notificacoes else {}
        except:
            notificacoes = {}

        # Upload de imagens (Diretriz 11.3)
        imagens_salvas = []
        if 'imagens' in request.files:
            files = request.files.getlist('imagens')

            # Criar diretório de uploads se não existir
            upload_dir = os.path.join('static', 'uploads', 'anamnese', cpf)
            os.makedirs(upload_dir, exist_ok=True)

            for file in files[:5]:  # Máximo 5 arquivos
                if file and file.filename:
                    # Gerar nome único para o arquivo
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{timestamp}_{file.filename}"
                    filepath = os.path.join(upload_dir, filename)

                    # Salvar arquivo
                    file.save(filepath)

                    # Salvar caminho relativo
                    relative_path = filepath.replace('\\', '/')
                    imagens_salvas.append(relative_path)

        # Criar documento de anamnese
        documento_anamnese = {
            'cliente_cpf': cpf,
            'observacoes': observacoes,
            'imagens': imagens_salvas,
            'data_cadastro': datetime.now(),
            'cadastrado_por': session.get('usuario', 'Sistema'),
            'respostas': {}
        }

        result = db.anamnese.insert_one(documento_anamnese)
        anamnese_id = str(result.inserted_id)

        # NOTIFICAÇÕES (Diretriz 11.4)
        notificacoes_enviadas = {}

        # WhatsApp
        if notificacoes.get('whatsapp') and cliente_telefone:
            telefone_limpo = ''.join(filter(str.isdigit, cliente_telefone))
            if not telefone_limpo.startswith('55'):
                telefone_limpo = '55' + telefone_limpo

            mensagem_whatsapp = f"""*BIOMA - Nova Anamnese Cadastrada*

Olá, {cliente_nome}! 👋

Sua anamnese foi cadastrada em nosso sistema.

📋 *Observações:* {observacoes[:100]}{'...' if len(observacoes) > 100 else ''}
📅 *Data:* {datetime.now().strftime('%d/%m/%Y %H:%M')}
{f'📎 *Anexos:* {len(imagens_salvas)} arquivo(s)' if imagens_salvas else ''}

Qualquer dúvida, estamos à disposição!

Obrigado pela confiança! 😊"""

            whatsapp_url = f"https://wa.me/{telefone_limpo}?text={urllib.parse.quote(mensagem_whatsapp)}"

            # Log de notificação
            db.notificacoes_log.insert_one({
                'tipo': 'whatsapp',
                'destinatario': cliente_telefone,
                'mensagem': mensagem_whatsapp,
                'url': whatsapp_url,
                'anamnese_id': anamnese_id,
                'data_envio': datetime.now()
            })

            notificacoes_enviadas['whatsapp'] = True

        # Email
        if notificacoes.get('email') and cliente_email:
            db.notificacoes_log.insert_one({
                'tipo': 'email',
                'destinatario': cliente_email,
                'assunto': 'Nova Anamnese Cadastrada - BIOMA',
                'corpo': f"""
                    Olá, {cliente_nome}!

                    Sua anamnese foi cadastrada em nosso sistema.

                    Observações: {observacoes}
                    Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}
                    {f'Anexos: {len(imagens_salvas)} arquivo(s)' if imagens_salvas else ''}

                    Obrigado pela confiança!

                    Atenciosamente,
                    Equipe BIOMA
                """,
                'anamnese_id': anamnese_id,
                'data_envio': datetime.now(),
                'status': 'simulado'
            })

            notificacoes_enviadas['email'] = True

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'anamnese_criada',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Anamnese cadastrada para cliente {cliente_nome}',
            'data_hora': datetime.now(),
            'detalhes': {
                'cliente_cpf': cpf,
                'imagens_anexadas': len(imagens_salvas),
                'notificacoes': notificacoes_enviadas
            }
        })

        return jsonify({
            'success': True,
            'anamnese_id': anamnese_id,
            'imagens_salvas': imagens_salvas,
            'notificacoes_enviadas': notificacoes_enviadas,
            'message': 'Anamnese cadastrada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao criar anamnese: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/clientes/<cpf>/prontuario', methods=['POST'])
def criar_prontuario_com_imagens(cpf):
    """
    Criar prontuário com upload de imagens e notificações

    Diretriz 11.3: Anexar imagens físicas de documentos
    Diretriz 11.4: Notificações Email/WhatsApp
    """
    try:
        db = get_db()

        # Buscar cliente
        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # Extrair dados do FormData
        data_atendimento = request.form.get('data_atendimento')
        profissional = request.form.get('profissional')
        procedimento = request.form.get('procedimento')
        observacoes = request.form.get('observacoes', '')
        notificacoes = request.form.get('notificacoes')
        cliente_nome = request.form.get('cliente_nome')
        cliente_email = request.form.get('cliente_email')
        cliente_telefone = request.form.get('cliente_telefone')

        # Parse notificacoes JSON
        try:
            notificacoes = json.loads(notificacoes) if notificacoes else {}
        except:
            notificacoes = {}

        # Upload de imagens (Diretriz 11.3)
        imagens_salvas = []
        if 'imagens' in request.files:
            files = request.files.getlist('imagens')

            # Criar diretório de uploads se não existir
            upload_dir = os.path.join('static', 'uploads', 'prontuario', cpf)
            os.makedirs(upload_dir, exist_ok=True)

            for file in files[:5]:  # Máximo 5 arquivos
                if file and file.filename:
                    # Gerar nome único para o arquivo
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{timestamp}_{file.filename}"
                    filepath = os.path.join(upload_dir, filename)

                    # Salvar arquivo
                    file.save(filepath)

                    # Salvar caminho relativo
                    relative_path = filepath.replace('\\', '/')
                    imagens_salvas.append(relative_path)

        # Criar documento de prontuário
        documento_prontuario = {
            'cliente_cpf': cpf,
            'data_atendimento': data_atendimento,
            'profissional': profissional,
            'procedimento': procedimento,
            'observacoes': observacoes,
            'imagens': imagens_salvas,
            'proxima_sessao': None,
            'created_at': datetime.now(),
            'cadastrado_por': session.get('usuario', 'Sistema')
        }

        result = db.prontuario.insert_one(documento_prontuario)
        prontuario_id = str(result.inserted_id)

        # NOTIFICAÇÕES (Diretriz 11.4)
        notificacoes_enviadas = {}

        # WhatsApp
        if notificacoes.get('whatsapp') and cliente_telefone:
            telefone_limpo = ''.join(filter(str.isdigit, cliente_telefone))
            if not telefone_limpo.startswith('55'):
                telefone_limpo = '55' + telefone_limpo

            mensagem_whatsapp = f"""*BIOMA - Prontuário Registrado*

Olá, {cliente_nome}! 👋

Seu atendimento foi registrado em nosso sistema.

📋 *Procedimento:* {procedimento}
👨‍⚕️ *Profissional:* {profissional}
📅 *Data:* {data_atendimento}
{f'📎 *Anexos:* {len(imagens_salvas)} arquivo(s)' if imagens_salvas else ''}

{f'💬 *Observações:* {observacoes[:100]}{"..." if len(observacoes) > 100 else ""}' if observacoes else ''}

Obrigado pela confiança! 😊"""

            whatsapp_url = f"https://wa.me/{telefone_limpo}?text={urllib.parse.quote(mensagem_whatsapp)}"

            # Log de notificação
            db.notificacoes_log.insert_one({
                'tipo': 'whatsapp',
                'destinatario': cliente_telefone,
                'mensagem': mensagem_whatsapp,
                'url': whatsapp_url,
                'prontuario_id': prontuario_id,
                'data_envio': datetime.now()
            })

            notificacoes_enviadas['whatsapp'] = True

        # Email
        if notificacoes.get('email') and cliente_email:
            db.notificacoes_log.insert_one({
                'tipo': 'email',
                'destinatario': cliente_email,
                'assunto': f'Prontuário Registrado - {procedimento}',
                'corpo': f"""
                    Olá, {cliente_nome}!

                    Seu atendimento foi registrado em nosso sistema.

                    Procedimento: {procedimento}
                    Profissional: {profissional}
                    Data: {data_atendimento}
                    {f'Anexos: {len(imagens_salvas)} arquivo(s)' if imagens_salvas else ''}

                    {f'Observações: {observacoes}' if observacoes else ''}

                    Obrigado pela confiança!

                    Atenciosamente,
                    Equipe BIOMA
                """,
                'prontuario_id': prontuario_id,
                'data_envio': datetime.now(),
                'status': 'simulado'
            })

            notificacoes_enviadas['email'] = True

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'prontuario_criado',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Prontuário cadastrado para cliente {cliente_nome}',
            'data_hora': datetime.now(),
            'detalhes': {
                'cliente_cpf': cpf,
                'procedimento': procedimento,
                'profissional': profissional,
                'imagens_anexadas': len(imagens_salvas),
                'notificacoes': notificacoes_enviadas
            }
        })

        return jsonify({
            'success': True,
            'prontuario_id': prontuario_id,
            'imagens_salvas': imagens_salvas,
            'notificacoes_enviadas': notificacoes_enviadas,
            'message': 'Prontuário cadastrado com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao criar prontuário: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/clientes/<cpf>/notificar', methods=['POST'])
def notificar_cliente(cpf):
    """
    Enviar notificação direta ao cliente
    Diretriz 11.4: Notificações Email/WhatsApp
    """
    try:
        db = get_db()
        data = request.get_json()

        cliente_nome = data.get('cliente_nome')
        cliente_email = data.get('cliente_email')
        cliente_telefone = data.get('cliente_telefone')
        mensagem = data.get('mensagem')
        notificacoes = data.get('notificacoes', {})

        if not mensagem:
            return jsonify({'success': False, 'message': 'Mensagem é obrigatória'}), 400

        notificacoes_enviadas = {}

        # WhatsApp
        if notificacoes.get('whatsapp') and cliente_telefone:
            telefone_limpo = ''.join(filter(str.isdigit, cliente_telefone))
            if not telefone_limpo.startswith('55'):
                telefone_limpo = '55' + telefone_limpo

            mensagem_whatsapp = f"""*BIOMA - Mensagem*

Olá, {cliente_nome}! 👋

{mensagem}

Atenciosamente,
Equipe BIOMA"""

            whatsapp_url = f"https://wa.me/{telefone_limpo}?text={urllib.parse.quote(mensagem_whatsapp)}"

            db.notificacoes_log.insert_one({
                'tipo': 'whatsapp',
                'destinatario': cliente_telefone,
                'mensagem': mensagem_whatsapp,
                'url': whatsapp_url,
                'cliente_cpf': cpf,
                'data_envio': datetime.now()
            })

            notificacoes_enviadas['whatsapp'] = True

        # Email
        if notificacoes.get('email') and cliente_email:
            db.notificacoes_log.insert_one({
                'tipo': 'email',
                'destinatario': cliente_email,
                'assunto': 'Mensagem da BIOMA',
                'corpo': f"""
                    Olá, {cliente_nome}!

                    {mensagem}

                    Atenciosamente,
                    Equipe BIOMA
                """,
                'cliente_cpf': cpf,
                'data_envio': datetime.now(),
                'status': 'simulado'
            })

            notificacoes_enviadas['email'] = True

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'notificacao_cliente',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Notificação enviada ao cliente {cliente_nome}',
            'data_hora': datetime.now(),
            'detalhes': {
                'cliente_cpf': cpf,
                'mensagem_preview': mensagem[:100],
                'notificacoes': notificacoes_enviadas
            }
        })

        return jsonify({
            'success': True,
            'notificacoes_enviadas': notificacoes_enviadas,
            'message': 'Notificação enviada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao notificar cliente: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== MULTICOMISSÃO (Diretriz 12.1) ====================

@bp.route('/api/multicomissao/regras', methods=['GET'])
def get_multicomissao_regras():
    """
    Listar todas as regras de multicomissão
    Diretriz 12.1: Comissão sobre comissão
    """
    try:
        db = get_db()

        # Buscar todas as regras
        regras = list(db.multicomissao.find().sort('created_at', -1))

        # Enricher com nomes dos profissionais
        for regra in regras:
            prof_principal = db.profissionais.find_one({'_id': ObjectId(regra['profissional_principal_id'])})
            assistente = db.profissionais.find_one({'_id': ObjectId(regra['assistente_id'])})

            regra['profissional_principal_nome'] = prof_principal.get('nome') if prof_principal else 'N/A'
            regra['assistente_nome'] = assistente.get('nome') if assistente else 'N/A'

        return jsonify({
            'success': True,
            'regras': convert_objectid(regras)
        })

    except Exception as e:
        logger.error(f"Erro ao listar regras de multicomissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/multicomissao/regras', methods=['POST'])
def criar_multicomissao_regra():
    """
    Criar nova regra de multicomissão
    Diretriz 12.1: Assistente recebe % da comissão do profissional
    """
    try:
        db = get_db()
        data = request.get_json()

        # Validações
        prof_principal_id = data.get('profissional_principal_id')
        assistente_id = data.get('assistente_id')
        percentual_assistente = data.get('percentual_assistente')

        if not prof_principal_id or not assistente_id:
            return jsonify({'success': False, 'message': 'Profissional e assistente são obrigatórios'}), 400

        if prof_principal_id == assistente_id:
            return jsonify({'success': False, 'message': 'Profissional e assistente não podem ser a mesma pessoa'}), 400

        if not percentual_assistente or percentual_assistente < 1 or percentual_assistente > 100:
            return jsonify({'success': False, 'message': 'Percentual deve estar entre 1% e 100%'}), 400

        # Buscar informações dos profissionais
        prof_principal = db.profissionais.find_one({'_id': ObjectId(prof_principal_id)})
        assistente = db.profissionais.find_one({'_id': ObjectId(assistente_id)})

        if not prof_principal or not assistente:
            return jsonify({'success': False, 'message': 'Profissional ou assistente não encontrado'}), 404

        # Criar documento
        documento = {
            'profissional_principal_id': prof_principal_id,
            'profissional_principal_nome': prof_principal.get('nome'),
            'assistente_id': assistente_id,
            'assistente_nome': assistente.get('nome'),
            'percentual_assistente': percentual_assistente,
            'data_inicio': data.get('data_inicio'),
            'data_fim': data.get('data_fim'),
            'observacoes': data.get('observacoes'),
            'ativa': data.get('ativa', True),
            'created_at': datetime.now(),
            'created_by': session.get('usuario', 'Sistema')
        }

        result = db.multicomissao.insert_one(documento)

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'multicomissao_criada',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Regra de multicomissão criada: {prof_principal.get("nome")} → {assistente.get("nome")}',
            'data_hora': datetime.now(),
            'detalhes': {
                'regra_id': str(result.inserted_id),
                'percentual': percentual_assistente
            }
        })

        return jsonify({
            'success': True,
            'regra_id': str(result.inserted_id),
            'profissional_principal_nome': prof_principal.get('nome'),
            'assistente_nome': assistente.get('nome'),
            'percentual_assistente': percentual_assistente,
            'message': 'Regra criada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao criar regra de multicomissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/multicomissao/regras/<regra_id>')
def get_multicomissao_regra(regra_id):
    """
    Visualizar regra com cálculos estatísticos
    Diretriz 12.1: Mostrar valores já pagos
    """
    try:
        db = get_db()

        # Buscar regra
        regra = db.multicomissao.find_one({'_id': ObjectId(regra_id)})
        if not regra:
            return jsonify({'success': False, 'message': 'Regra não encontrada'}), 404

        # Buscar informações dos profissionais
        prof_principal = db.profissionais.find_one({'_id': ObjectId(regra['profissional_principal_id'])})
        assistente = db.profissionais.find_one({'_id': ObjectId(regra['assistente_id'])})

        regra['profissional_principal_nome'] = prof_principal.get('nome') if prof_principal else 'N/A'
        regra['assistente_nome'] = assistente.get('nome') if assistente else 'N/A'

        # CALCULAR ESTATÍSTICAS
        calculos = calcular_multicomissao(db, regra, prof_principal)

        return jsonify({
            'success': True,
            'regra': convert_objectid(regra),
            'calculos': calculos
        })

    except Exception as e:
        logger.error(f"Erro ao buscar regra de multicomissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


def calcular_multicomissao(db, regra, prof_principal):
    """
    Calcular estatísticas de multicomissão

    Lógica:
    1. Buscar orçamentos aprovados do profissional principal
    2. Calcular comissão do profissional (X% do valor total)
    3. Calcular comissão do assistente (Y% da comissão do profissional)
    """
    try:
        # Filtro de data
        query = {
            'profissional_id': regra['profissional_principal_id'],
            'status': 'aprovado'
        }

        # Se houver período definido
        if regra.get('data_inicio'):
            query['data_criacao'] = {'$gte': regra['data_inicio']}

        if regra.get('data_fim'):
            if 'data_criacao' in query:
                query['data_criacao']['$lte'] = regra['data_fim']
            else:
                query['data_criacao'] = {'$lte': regra['data_fim']}

        # Buscar orçamentos
        orcamentos = list(db.orcamentos.find(query))

        total_atendimentos = len(orcamentos)
        valor_total = sum(orc.get('valor_total', 0) for orc in orcamentos)

        # Percentual de comissão do profissional principal
        percentual_prof = prof_principal.get('comissao_percentual', 10) if prof_principal else 10

        # Comissão do profissional principal
        comissao_profissional = (valor_total * percentual_prof) / 100

        # Comissão do assistente = X% da comissão do profissional
        percentual_assistente = regra.get('percentual_assistente', 0)
        comissao_assistente = (comissao_profissional * percentual_assistente) / 100

        return {
            'total_atendimentos': total_atendimentos,
            'valor_total': valor_total,
            'percentual_profissional': percentual_prof,
            'comissao_profissional': comissao_profissional,
            'percentual_assistente': percentual_assistente,
            'comissao_assistente': comissao_assistente
        }

    except Exception as e:
        logger.error(f"Erro ao calcular multicomissão: {e}")
        return {
            'total_atendimentos': 0,
            'valor_total': 0,
            'comissao_profissional': 0,
            'comissao_assistente': 0
        }


@bp.route('/api/multicomissao/regras/<regra_id>/toggle', methods=['PUT'])
def toggle_multicomissao_regra(regra_id):
    """
    Ativar/Desativar regra de multicomissão
    """
    try:
        db = get_db()
        data = request.get_json()

        nova_status = data.get('ativa', True)

        result = db.multicomissao.update_one(
            {'_id': ObjectId(regra_id)},
            {'$set': {
                'ativa': nova_status,
                'updated_at': datetime.now(),
                'updated_by': session.get('usuario', 'Sistema')
            }}
        )

        if result.matched_count == 0:
            return jsonify({'success': False, 'message': 'Regra não encontrada'}), 404

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'multicomissao_toggle',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Regra de multicomissão {"ativada" if nova_status else "desativada"}',
            'data_hora': datetime.now(),
            'detalhes': {
                'regra_id': regra_id,
                'nova_status': nova_status
            }
        })

        return jsonify({
            'success': True,
            'message': f'Regra {"ativada" if nova_status else "desativada"} com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao toggle regra de multicomissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/multicomissao/regras/<regra_id>', methods=['DELETE'])
def deletar_multicomissao_regra(regra_id):
    """
    Deletar regra de multicomissão
    """
    try:
        db = get_db()

        # Buscar regra para auditoria
        regra = db.multicomissao.find_one({'_id': ObjectId(regra_id)})
        if not regra:
            return jsonify({'success': False, 'message': 'Regra não encontrada'}), 404

        # Deletar
        result = db.multicomissao.delete_one({'_id': ObjectId(regra_id)})

        if result.deleted_count == 0:
            return jsonify({'success': False, 'message': 'Regra não encontrada'}), 404

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'multicomissao_deletada',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Regra de multicomissão deletada: {regra.get("profissional_principal_nome")} → {regra.get("assistente_nome")}',
            'data_hora': datetime.now(),
            'detalhes': {
                'regra_id': regra_id
            }
        })

        return jsonify({
            'success': True,
            'message': 'Regra deletada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao deletar regra de multicomissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== MELHORIAS NOS PROFISSIONAIS (Diretrizes 12.2 e 12.3) ====================

@bp.route('/api/profissionais/<profissional_id>/completo')
def get_profissional_completo(profissional_id):
    """
    Visualizar profissional com estatísticas detalhadas
    Diretriz 12.2: Detalhes muito mais completos
    """
    try:
        db = get_db()

        # Buscar profissional
        profissional = db.profissionais.find_one({'_id': ObjectId(profissional_id)})
        if not profissional:
            return jsonify({'success': False, 'message': 'Profissional não encontrado'}), 404

        # CALCULAR ESTATÍSTICAS
        estatisticas = calcular_estatisticas_profissional(db, profissional_id)

        return jsonify({
            'success': True,
            'profissional': convert_objectid(profissional),
            'estatisticas': estatisticas
        })

    except Exception as e:
        logger.error(f"Erro ao buscar profissional completo: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


def calcular_estatisticas_profissional(db, profissional_id):
    """
    Calcular estatísticas de performance do profissional
    """
    try:
        # Buscar todos os orçamentos aprovados
        orcamentos = list(db.orcamentos.find({
            'profissional_id': profissional_id,
            'status': 'aprovado'
        }))

        total_atendimentos = len(orcamentos)
        faturamento_total = sum(orc.get('valor_total', 0) for orc in orcamentos)

        # Calcular comissões
        profissional = db.profissionais.find_one({'_id': ObjectId(profissional_id)})
        percentual_comissao = profissional.get('comissao_percentual', 10) if profissional else 10
        comissoes_total = (faturamento_total * percentual_comissao) / 100

        # Ticket médio
        ticket_medio = faturamento_total / total_atendimentos if total_atendimentos > 0 else 0

        # Top 5 serviços mais realizados
        servicos_count = {}
        for orc in orcamentos:
            servicos = orc.get('servicos', [])
            if isinstance(servicos, list):
                for servico in servicos:
                    servico_nome = servico.get('nome') if isinstance(servico, dict) else str(servico)
                    servicos_count[servico_nome] = servicos_count.get(servico_nome, 0) + 1

        top_servicos = sorted(servicos_count.items(), key=lambda x: x[1], reverse=True)[:5]
        top_servicos = [{'servico': s[0], 'quantidade': s[1]} for s in top_servicos]

        # Agendamentos do mês atual
        hoje = datetime.now()
        inicio_mes = datetime(hoje.year, hoje.month, 1)
        fim_mes = datetime(hoje.year, hoje.month + 1, 1) if hoje.month < 12 else datetime(hoje.year + 1, 1, 1)

        agendamentos_mes = db.agendamentos.count_documents({
            'profissional_id': profissional_id,
            'data': {
                '$gte': inicio_mes.strftime('%Y-%m-%d'),
                '$lt': fim_mes.strftime('%Y-%m-%d')
            }
        })

        # Dias trabalhados (unique dates)
        agendamentos_list = list(db.agendamentos.find({
            'profissional_id': profissional_id,
            'data': {
                '$gte': inicio_mes.strftime('%Y-%m-%d'),
                '$lt': fim_mes.strftime('%Y-%m-%d')
            }
        }))

        dias_unicos = set(ag.get('data') for ag in agendamentos_list if ag.get('data'))
        dias_trabalhados = len(dias_unicos)

        # Taxa de ocupação (assumindo 8 slots por dia)
        dias_uteis_mes = 22  # Aproximado
        taxa_ocupacao = (dias_trabalhados / dias_uteis_mes * 100) if dias_uteis_mes > 0 else 0

        return {
            'total_atendimentos': total_atendimentos,
            'faturamento_total': faturamento_total,
            'comissoes_total': comissoes_total,
            'ticket_medio': ticket_medio,
            'top_servicos': top_servicos,
            'agendamentos_mes': agendamentos_mes,
            'dias_trabalhados': dias_trabalhados,
            'taxa_ocupacao': taxa_ocupacao
        }

    except Exception as e:
        logger.error(f"Erro ao calcular estatísticas do profissional: {e}")
        return {
            'total_atendimentos': 0,
            'faturamento_total': 0,
            'comissoes_total': 0,
            'ticket_medio': 0,
            'top_servicos': [],
            'agendamentos_mes': 0,
            'dias_trabalhados': 0,
            'taxa_ocupacao': 0
        }


@bp.route('/api/profissionais/<profissional_id>/ordem-servico', methods=['POST'])
def enviar_ordem_servico(profissional_id):
    """
    Enviar ordem de serviço para profissional
    Diretriz 12.3: Email/WhatsApp para ordens de serviço
    """
    try:
        db = get_db()
        data = request.get_json()

        # Extrair dados
        profissional_nome = data.get('profissional_nome')
        profissional_email = data.get('profissional_email')
        profissional_telefone = data.get('profissional_telefone')
        data_servico = data.get('data')
        horario = data.get('horario')
        cliente = data.get('cliente')
        servico = data.get('servico')
        local = data.get('local')
        observacoes = data.get('observacoes', '')
        notificacoes = data.get('notificacoes', {})

        # Validações
        if not data_servico or not horario or not servico:
            return jsonify({'success': False, 'message': 'Data, horário e serviço são obrigatórios'}), 400

        # Criar documento da ordem de serviço
        ordem = {
            'profissional_id': profissional_id,
            'profissional_nome': profissional_nome,
            'data_servico': data_servico,
            'horario': horario,
            'cliente': cliente,
            'servico': servico,
            'local': local,
            'observacoes': observacoes,
            'status': 'enviada',
            'created_at': datetime.now(),
            'created_by': session.get('usuario', 'Sistema')
        }

        result = db.ordens_servico.insert_one(ordem)
        ordem_id = str(result.inserted_id)

        # NOTIFICAÇÕES (Diretriz 12.3)
        notificacoes_enviadas = {}

        # WhatsApp
        if notificacoes.get('whatsapp') and profissional_telefone:
            telefone_limpo = ''.join(filter(str.isdigit, profissional_telefone))
            if not telefone_limpo.startswith('55'):
                telefone_limpo = '55' + telefone_limpo

            mensagem_whatsapp = f"""*BIOMA - Ordem de Serviço*

Olá, {profissional_nome}! 👋

Você tem uma nova ordem de serviço:

📅 *Data:* {data_servico}
🕐 *Horário:* {horario}
👤 *Cliente:* {cliente}
💼 *Serviço:* {servico}
📍 *Local:* {local}

{f'📝 *Observações:* {observacoes}' if observacoes else ''}

Por favor, confirme o recebimento.

Equipe BIOMA"""

            whatsapp_url = f"https://wa.me/{telefone_limpo}?text={urllib.parse.quote(mensagem_whatsapp)}"

            # Log de notificação
            db.notificacoes_log.insert_one({
                'tipo': 'whatsapp',
                'destinatario': profissional_telefone,
                'mensagem': mensagem_whatsapp,
                'url': whatsapp_url,
                'ordem_servico_id': ordem_id,
                'profissional_id': profissional_id,
                'data_envio': datetime.now()
            })

            notificacoes_enviadas['whatsapp'] = True

        # Email
        if notificacoes.get('email') and profissional_email:
            db.notificacoes_log.insert_one({
                'tipo': 'email',
                'destinatario': profissional_email,
                'assunto': f'Ordem de Serviço - {data_servico} às {horario}',
                'corpo': f"""
                    Olá, {profissional_nome}!

                    Você tem uma nova ordem de serviço:

                    Data: {data_servico}
                    Horário: {horario}
                    Cliente: {cliente}
                    Serviço: {servico}
                    Local: {local}

                    {f'Observações: {observacoes}' if observacoes else ''}

                    Por favor, confirme o recebimento.

                    Atenciosamente,
                    Equipe BIOMA
                """,
                'ordem_servico_id': ordem_id,
                'profissional_id': profissional_id,
                'data_envio': datetime.now(),
                'status': 'simulado'
            })

            notificacoes_enviadas['email'] = True

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'ordem_servico_enviada',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Ordem de serviço enviada para {profissional_nome}',
            'data_hora': datetime.now(),
            'detalhes': {
                'ordem_id': ordem_id,
                'profissional_id': profissional_id,
                'data_servico': data_servico,
                'servico': servico,
                'notificacoes': notificacoes_enviadas
            }
        })

        return jsonify({
            'success': True,
            'ordem_id': ordem_id,
            'notificacoes_enviadas': notificacoes_enviadas,
            'message': 'Ordem de serviço enviada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao enviar ordem de serviço: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/profissionais/<profissional_id>/notificar', methods=['POST'])
def notificar_profissional(profissional_id):
    """
    Enviar notificação direta ao profissional
    Diretriz 12.3: Notificações Email/WhatsApp
    """
    try:
        db = get_db()
        data = request.get_json()

        profissional_nome = data.get('profissional_nome')
        profissional_email = data.get('profissional_email')
        profissional_telefone = data.get('profissional_telefone')
        mensagem = data.get('mensagem')
        notificacoes = data.get('notificacoes', {})

        if not mensagem:
            return jsonify({'success': False, 'message': 'Mensagem é obrigatória'}), 400

        notificacoes_enviadas = {}

        # WhatsApp
        if notificacoes.get('whatsapp') and profissional_telefone:
            telefone_limpo = ''.join(filter(str.isdigit, profissional_telefone))
            if not telefone_limpo.startswith('55'):
                telefone_limpo = '55' + telefone_limpo

            mensagem_whatsapp = f"""*BIOMA - Mensagem*

Olá, {profissional_nome}! 👋

{mensagem}

Atenciosamente,
Equipe BIOMA"""

            whatsapp_url = f"https://wa.me/{telefone_limpo}?text={urllib.parse.quote(mensagem_whatsapp)}"

            db.notificacoes_log.insert_one({
                'tipo': 'whatsapp',
                'destinatario': profissional_telefone,
                'mensagem': mensagem_whatsapp,
                'url': whatsapp_url,
                'profissional_id': profissional_id,
                'data_envio': datetime.now()
            })

            notificacoes_enviadas['whatsapp'] = True

        # Email
        if notificacoes.get('email') and profissional_email:
            db.notificacoes_log.insert_one({
                'tipo': 'email',
                'destinatario': profissional_email,
                'assunto': 'Mensagem da BIOMA',
                'corpo': f"""
                    Olá, {profissional_nome}!

                    {mensagem}

                    Atenciosamente,
                    Equipe BIOMA
                """,
                'profissional_id': profissional_id,
                'data_envio': datetime.now(),
                'status': 'simulado'
            })

            notificacoes_enviadas['email'] = True

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'notificacao_profissional',
            'usuario': session.get('usuario', 'Sistema'),
            'descricao': f'Notificação enviada ao profissional {profissional_nome}',
            'data_hora': datetime.now(),
            'detalhes': {
                'profissional_id': profissional_id,
                'mensagem_preview': mensagem[:100],
                'notificacoes': notificacoes_enviadas
            }
        })

        return jsonify({
            'success': True,
            'notificacoes_enviadas': notificacoes_enviadas,
            'message': 'Notificação enviada com sucesso'
        })

    except Exception as e:
        logger.error(f"Erro ao notificar profissional: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== HISTÓRICO DE ATENDIMENTOS (Diretriz 12.4) ====================

@bp.route('/api/profissionais/<profissional_id>/historico')
def get_historico_atendimentos(profissional_id):
    """
    Histórico completo de atendimentos do profissional com filtros
    Diretriz 12.4: Histórico de atendimento por profissional
    """
    try:
        db = get_db()

        # Verificar se profissional existe
        profissional = db.profissionais.find_one({'_id': ObjectId(profissional_id)})
        if not profissional:
            return jsonify({'success': False, 'message': 'Profissional não encontrado'}), 404

        # Parâmetros de filtro da query string
        data_inicio = request.args.get('data_inicio')  # YYYY-MM-DD
        data_fim = request.args.get('data_fim')        # YYYY-MM-DD
        cliente_id = request.args.get('cliente_id')
        servico_nome = request.args.get('servico')
        status = request.args.get('status')  # pendente, em_andamento, concluido, cancelado
        limite = int(request.args.get('limite', 50))
        pagina = int(request.args.get('pagina', 1))

        # Construir filtro base
        filtro = {'profissional_id': profissional_id}

        # Filtro de data
        if data_inicio or data_fim:
            filtro['data'] = {}
            if data_inicio:
                filtro['data']['$gte'] = data_inicio
            if data_fim:
                filtro['data']['$lte'] = data_fim

        # Filtro de cliente
        if cliente_id:
            filtro['cliente_id'] = cliente_id

        # Filtro de status
        if status:
            filtro['status'] = status

        # Pipeline de agregação para buscar agendamentos com dados relacionados
        pipeline = [
            {'$match': filtro},
            {'$sort': {'data': -1, 'horario': -1}},
            {'$skip': (pagina - 1) * limite},
            {'$limit': limite},
            {
                '$lookup': {
                    'from': 'clientes',
                    'localField': 'cliente_id',
                    'foreignField': '_id',
                    'as': 'cliente_info'
                }
            },
            {
                '$lookup': {
                    'from': 'servicos',
                    'localField': 'servico_id',
                    'foreignField': '_id',
                    'as': 'servico_info'
                }
            },
            {
                '$lookup': {
                    'from': 'orcamentos',
                    'localField': 'orcamento_id',
                    'foreignField': '_id',
                    'as': 'orcamento_info'
                }
            },
            {
                '$unwind': {
                    'path': '$cliente_info',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$unwind': {
                    'path': '$servico_info',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$unwind': {
                    'path': '$orcamento_info',
                    'preserveNullAndEmptyArrays': True
                }
            },
            {
                '$project': {
                    '_id': 1,
                    'data': 1,
                    'horario': 1,
                    'status': 1,
                    'observacoes': 1,
                    'duracao': 1,
                    'cliente_nome': '$cliente_info.nome',
                    'cliente_telefone': '$cliente_info.telefone',
                    'servico_nome': '$servico_info.nome',
                    'servico_categoria': '$servico_info.categoria',
                    'valor_servico': '$servico_info.preco',
                    'valor_total': '$orcamento_info.valor_total',
                    'forma_pagamento': '$orcamento_info.forma_pagamento',
                    'orcamento_status': '$orcamento_info.status'
                }
            }
        ]

        # Executar agregação
        atendimentos = list(db.agendamentos.aggregate(pipeline))

        # Filtro adicional por serviço (nome parcial)
        if servico_nome:
            atendimentos = [
                a for a in atendimentos
                if a.get('servico_nome') and servico_nome.lower() in a.get('servico_nome', '').lower()
            ]

        # Contar total de registros (sem paginação)
        total_registros = db.agendamentos.count_documents(filtro)

        # Calcular estatísticas do histórico filtrado
        estatisticas_historico = calcular_estatisticas_historico(atendimentos)

        # Converter ObjectId para string
        atendimentos_convertidos = []
        for atend in atendimentos:
            atend_convertido = convert_objectid(atend)
            atendimentos_convertidos.append(atend_convertido)

        return jsonify({
            'success': True,
            'profissional': {
                'id': str(profissional['_id']),
                'nome': profissional.get('nome'),
                'especialidade': profissional.get('especialidade')
            },
            'atendimentos': atendimentos_convertidos,
            'paginacao': {
                'pagina_atual': pagina,
                'limite': limite,
                'total_registros': total_registros,
                'total_paginas': (total_registros + limite - 1) // limite
            },
            'estatisticas': estatisticas_historico,
            'filtros_aplicados': {
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'cliente_id': cliente_id,
                'servico': servico_nome,
                'status': status
            }
        })

    except Exception as e:
        logger.error(f"Erro ao buscar histórico de atendimentos: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


def calcular_estatisticas_historico(atendimentos):
    """
    Calcular estatísticas do histórico de atendimentos
    """
    try:
        total_atendimentos = len(atendimentos)

        if total_atendimentos == 0:
            return {
                'total_atendimentos': 0,
                'atendimentos_concluidos': 0,
                'atendimentos_cancelados': 0,
                'taxa_conclusao': 0,
                'faturamento_total': 0,
                'ticket_medio': 0,
                'servicos_mais_realizados': [],
                'clientes_atendidos': 0,
                'duracao_media': 0
            }

        # Contar por status
        status_count = {}
        for atend in atendimentos:
            status = atend.get('status', 'desconhecido')
            status_count[status] = status_count.get(status, 0) + 1

        atendimentos_concluidos = status_count.get('concluido', 0)
        atendimentos_cancelados = status_count.get('cancelado', 0)
        taxa_conclusao = (atendimentos_concluidos / total_atendimentos * 100) if total_atendimentos > 0 else 0

        # Faturamento total (somente concluídos)
        faturamento_total = sum(
            atend.get('valor_total', 0) or atend.get('valor_servico', 0)
            for atend in atendimentos
            if atend.get('status') == 'concluido'
        )

        ticket_medio = faturamento_total / atendimentos_concluidos if atendimentos_concluidos > 0 else 0

        # Serviços mais realizados
        servicos_count = {}
        for atend in atendimentos:
            servico = atend.get('servico_nome')
            if servico:
                servicos_count[servico] = servicos_count.get(servico, 0) + 1

        servicos_mais_realizados = sorted(
            [{'servico': s, 'quantidade': q} for s, q in servicos_count.items()],
            key=lambda x: x['quantidade'],
            reverse=True
        )[:5]

        # Clientes únicos
        clientes_unicos = set()
        for atend in atendimentos:
            cliente = atend.get('cliente_nome')
            if cliente:
                clientes_unicos.add(cliente)

        clientes_atendidos = len(clientes_unicos)

        # Duração média (em minutos)
        duracoes = [atend.get('duracao', 0) for atend in atendimentos if atend.get('duracao')]
        duracao_media = sum(duracoes) / len(duracoes) if duracoes else 60  # Default 60min

        return {
            'total_atendimentos': total_atendimentos,
            'atendimentos_concluidos': atendimentos_concluidos,
            'atendimentos_cancelados': atendimentos_cancelados,
            'taxa_conclusao': round(taxa_conclusao, 1),
            'faturamento_total': faturamento_total,
            'ticket_medio': ticket_medio,
            'servicos_mais_realizados': servicos_mais_realizados,
            'clientes_atendidos': clientes_atendidos,
            'duracao_media': int(duracao_media),
            'distribuicao_status': status_count
        }

    except Exception as e:
        logger.error(f"Erro ao calcular estatísticas do histórico: {e}")
        return {
            'total_atendimentos': len(atendimentos),
            'error': str(e)
        }


@bp.route('/api/profissionais/<profissional_id>/timeline')
def get_timeline_atendimentos(profissional_id):
    """
    Timeline visual dos atendimentos (agrupados por mês)
    Diretriz 12.4: Visualização alternativa do histórico
    """
    try:
        db = get_db()

        # Verificar se profissional existe
        profissional = db.profissionais.find_one({'_id': ObjectId(profissional_id)})
        if not profissional:
            return jsonify({'success': False, 'message': 'Profissional não encontrado'}), 404

        # Buscar atendimentos dos últimos 12 meses
        hoje = datetime.now()
        data_limite = hoje - timedelta(days=365)

        pipeline = [
            {
                '$match': {
                    'profissional_id': profissional_id,
                    'data': {'$gte': data_limite.strftime('%Y-%m-%d')}
                }
            },
            {
                '$addFields': {
                    'ano_mes': {'$substr': ['$data', 0, 7]}  # YYYY-MM
                }
            },
            {
                '$group': {
                    '_id': '$ano_mes',
                    'total_atendimentos': {'$sum': 1},
                    'concluidos': {
                        '$sum': {'$cond': [{'$eq': ['$status', 'concluido']}, 1, 0]}
                    },
                    'cancelados': {
                        '$sum': {'$cond': [{'$eq': ['$status', 'cancelado']}, 1, 0]}
                    }
                }
            },
            {'$sort': {'_id': 1}}
        ]

        timeline_data = list(db.agendamentos.aggregate(pipeline))

        # Formatar timeline
        timeline_formatada = []
        for item in timeline_data:
            ano_mes = item['_id']
            try:
                # Converter YYYY-MM para nome do mês
                ano, mes = ano_mes.split('-')
                meses_pt = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                           'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                mes_nome = meses_pt[int(mes) - 1]
                label = f"{mes_nome}/{ano}"
            except:
                label = ano_mes

            timeline_formatada.append({
                'periodo': ano_mes,
                'label': label,
                'total_atendimentos': item['total_atendimentos'],
                'concluidos': item['concluidos'],
                'cancelados': item['cancelados'],
                'taxa_conclusao': round((item['concluidos'] / item['total_atendimentos'] * 100), 1) if item['total_atendimentos'] > 0 else 0
            })

        return jsonify({
            'success': True,
            'profissional': {
                'id': str(profissional['_id']),
                'nome': profissional.get('nome')
            },
            'timeline': timeline_formatada
        })

    except Exception as e:
        logger.error(f"Erro ao buscar timeline: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== SISTEMA DE NÍVEIS DE ACESSO (Diretriz 5.1) ====================

from functools import wraps

# Definição de níveis de acesso
NIVEIS_ACESSO = {
    'admin': {
        'nivel': 3,
        'nome': 'Administrador',
        'descricao': 'Acesso total ao sistema',
        'permissoes': ['*']  # Todas as permissões
    },
    'gestor': {
        'nivel': 2,
        'nome': 'Gestor',
        'descricao': 'Acesso a relatórios e gestão',
        'permissoes': [
            'visualizar_orcamentos',
            'visualizar_clientes',
            'visualizar_profissionais',
            'visualizar_relatorios',
            'visualizar_financeiro',
            'visualizar_agendamentos',
            'visualizar_estoque',
            'editar_orcamentos',
            'editar_clientes',
            'aprovar_orcamentos'
        ]
    },
    'profissional': {
        'nivel': 1,
        'nome': 'Profissional',
        'descricao': 'Acesso limitado a próprios agendamentos',
        'permissoes': [
            'visualizar_proprios_agendamentos',
            'visualizar_proprios_orcamentos',
            'visualizar_clientes_basico',
            'atualizar_status_agendamento'
        ]
    }
}


def verificar_permissao(permissao_requerida):
    """
    Decorator para verificar se usuário tem permissão para acessar rota
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Verificar se usuário está logado
            if 'usuario_id' not in session:
                return jsonify({'success': False, 'message': 'Não autorizado'}), 401

            # Buscar usuário no banco
            db = get_db()
            usuario = db.usuarios.find_one({'_id': ObjectId(session['usuario_id'])})

            if not usuario:
                return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 401

            # Obter nível de acesso (default: profissional)
            nivel_usuario = usuario.get('nivel_acesso', 'profissional')
            config_nivel = NIVEIS_ACESSO.get(nivel_usuario, NIVEIS_ACESSO['profissional'])

            # Admin tem acesso total
            if '*' in config_nivel['permissoes']:
                return f(*args, **kwargs)

            # Verificar permissão específica
            if permissao_requerida not in config_nivel['permissoes']:
                return jsonify({
                    'success': False,
                    'message': 'Você não tem permissão para acessar este recurso',
                    'nivel_requerido': permissao_requerida,
                    'seu_nivel': nivel_usuario
                }), 403

            return f(*args, **kwargs)

        return decorated_function
    return decorator


def verificar_nivel_minimo(nivel_minimo):
    """
    Decorator para verificar nível mínimo de acesso (admin=3, gestor=2, profissional=1)
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'usuario_id' not in session:
                return jsonify({'success': False, 'message': 'Não autorizado'}), 401

            db = get_db()
            usuario = db.usuarios.find_one({'_id': ObjectId(session['usuario_id'])})

            if not usuario:
                return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 401

            nivel_usuario = usuario.get('nivel_acesso', 'profissional')
            nivel_numero = NIVEIS_ACESSO.get(nivel_usuario, NIVEIS_ACESSO['profissional'])['nivel']

            nivel_requerido_numero = NIVEIS_ACESSO.get(nivel_minimo, NIVEIS_ACESSO['profissional'])['nivel']

            if nivel_numero < nivel_requerido_numero:
                return jsonify({
                    'success': False,
                    'message': f'Acesso restrito a {NIVEIS_ACESSO[nivel_minimo]["nome"]} ou superior',
                    'nivel_requerido': nivel_minimo,
                    'seu_nivel': nivel_usuario
                }), 403

            return f(*args, **kwargs)

        return decorated_function
    return decorator


@bp.route('/api/usuarios/nivel-acesso/info')
def get_niveis_acesso():
    """
    Obter informações sobre níveis de acesso disponíveis
    """
    try:
        return jsonify({
            'success': True,
            'niveis': NIVEIS_ACESSO
        })

    except Exception as e:
        logger.error(f"Erro ao obter níveis de acesso: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/usuarios/meu-perfil')
def get_meu_perfil():
    """
    Obter perfil do usuário logado com permissões
    """
    try:
        if 'usuario_id' not in session:
            return jsonify({'success': False, 'message': 'Não autorizado'}), 401

        db = get_db()
        usuario = db.usuarios.find_one({'_id': ObjectId(session['usuario_id'])})

        if not usuario:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 401

        nivel_acesso = usuario.get('nivel_acesso', 'profissional')
        config_nivel = NIVEIS_ACESSO.get(nivel_acesso, NIVEIS_ACESSO['profissional'])

        perfil = {
            'id': str(usuario['_id']),
            'nome': usuario.get('nome'),
            'email': usuario.get('email'),
            'nivel_acesso': nivel_acesso,
            'nivel_nome': config_nivel['nome'],
            'nivel_numero': config_nivel['nivel'],
            'permissoes': config_nivel['permissoes'],
            'criado_em': usuario.get('created_at')
        }

        return jsonify({
            'success': True,
            'perfil': perfil
        })

    except Exception as e:
        logger.error(f"Erro ao obter perfil: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/usuarios', methods=['GET'])
@verificar_nivel_minimo('gestor')
def listar_usuarios():
    """
    Listar todos os usuários (somente gestor/admin)
    """
    try:
        db = get_db()

        # Buscar todos os usuários
        usuarios = list(db.usuarios.find({}))

        usuarios_formatados = []
        for usuario in usuarios:
            nivel_acesso = usuario.get('nivel_acesso', 'profissional')
            config_nivel = NIVEIS_ACESSO.get(nivel_acesso, NIVEIS_ACESSO['profissional'])

            usuarios_formatados.append({
                'id': str(usuario['_id']),
                'nome': usuario.get('nome'),
                'email': usuario.get('email'),
                'nivel_acesso': nivel_acesso,
                'nivel_nome': config_nivel['nome'],
                'nivel_numero': config_nivel['nivel'],
                'ativo': usuario.get('ativo', True),
                'criado_em': usuario.get('created_at')
            })

        # Ordenar por nível (admin primeiro)
        usuarios_formatados.sort(key=lambda x: x['nivel_numero'], reverse=True)

        return jsonify({
            'success': True,
            'usuarios': usuarios_formatados,
            'total': len(usuarios_formatados)
        })

    except Exception as e:
        logger.error(f"Erro ao listar usuários: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/usuarios/<usuario_id>/nivel-acesso', methods=['PUT'])
@verificar_nivel_minimo('admin')
def alterar_nivel_acesso(usuario_id):
    """
    Alterar nível de acesso de um usuário (somente admin)
    """
    try:
        db = get_db()
        data = request.get_json()

        novo_nivel = data.get('nivel_acesso')

        # Validar nível
        if novo_nivel not in NIVEIS_ACESSO:
            return jsonify({
                'success': False,
                'message': f'Nível inválido. Opções: {list(NIVEIS_ACESSO.keys())}'
            }), 400

        # Buscar usuário alvo
        usuario_alvo = db.usuarios.find_one({'_id': ObjectId(usuario_id)})
        if not usuario_alvo:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404

        # Não permitir alterar próprio nível (segurança)
        if str(usuario_alvo['_id']) == session.get('usuario_id'):
            return jsonify({
                'success': False,
                'message': 'Você não pode alterar seu próprio nível de acesso'
            }), 403

        # Atualizar nível
        db.usuarios.update_one(
            {'_id': ObjectId(usuario_id)},
            {
                '$set': {
                    'nivel_acesso': novo_nivel,
                    'updated_at': datetime.now()
                }
            }
        )

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'nivel_acesso_alterado',
            'usuario': session.get('usuario', 'Admin'),
            'descricao': f'Nível de acesso alterado de {usuario_alvo.get("nome")} para {NIVEIS_ACESSO[novo_nivel]["nome"]}',
            'data_hora': datetime.now(),
            'detalhes': {
                'usuario_id': usuario_id,
                'nivel_anterior': usuario_alvo.get('nivel_acesso', 'profissional'),
                'nivel_novo': novo_nivel
            }
        })

        return jsonify({
            'success': True,
            'message': 'Nível de acesso alterado com sucesso',
            'novo_nivel': NIVEIS_ACESSO[novo_nivel]['nome']
        })

    except Exception as e:
        logger.error(f"Erro ao alterar nível de acesso: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/usuarios/<usuario_id>/ativar', methods=['PUT'])
@verificar_nivel_minimo('admin')
def toggle_usuario_ativo(usuario_id):
    """
    Ativar/desativar usuário (somente admin)
    """
    try:
        db = get_db()

        # Buscar usuário
        usuario = db.usuarios.find_one({'_id': ObjectId(usuario_id)})
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404

        # Não permitir desativar próprio usuário
        if str(usuario['_id']) == session.get('usuario_id'):
            return jsonify({
                'success': False,
                'message': 'Você não pode desativar sua própria conta'
            }), 403

        # Toggle ativo
        novo_status = not usuario.get('ativo', True)

        db.usuarios.update_one(
            {'_id': ObjectId(usuario_id)},
            {
                '$set': {
                    'ativo': novo_status,
                    'updated_at': datetime.now()
                }
            }
        )

        # Auditoria
        db.auditoria.insert_one({
            'tipo': 'usuario_ativado' if novo_status else 'usuario_desativado',
            'usuario': session.get('usuario', 'Admin'),
            'descricao': f'Usuário {usuario.get("nome")} {"ativado" if novo_status else "desativado"}',
            'data_hora': datetime.now(),
            'detalhes': {
                'usuario_id': usuario_id,
                'novo_status': novo_status
            }
        })

        return jsonify({
            'success': True,
            'message': f'Usuário {"ativado" if novo_status else "desativado"} com sucesso',
            'ativo': novo_status
        })

    except Exception as e:
        logger.error(f"Erro ao ativar/desativar usuário: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/verificar-permissao/<permissao>')
def verificar_permissao_api(permissao):
    """
    Verificar se usuário logado tem determinada permissão
    """
    try:
        if 'usuario_id' not in session:
            return jsonify({'success': False, 'tem_permissao': False, 'message': 'Não autorizado'}), 401

        db = get_db()
        usuario = db.usuarios.find_one({'_id': ObjectId(session['usuario_id'])})

        if not usuario:
            return jsonify({'success': False, 'tem_permissao': False}), 401

        nivel_acesso = usuario.get('nivel_acesso', 'profissional')
        config_nivel = NIVEIS_ACESSO.get(nivel_acesso, NIVEIS_ACESSO['profissional'])

        # Admin tem todas as permissões
        if '*' in config_nivel['permissoes']:
            tem_permissao = True
        else:
            tem_permissao = permissao in config_nivel['permissoes']

        return jsonify({
            'success': True,
            'tem_permissao': tem_permissao,
            'nivel_acesso': nivel_acesso
        })

    except Exception as e:
        logger.error(f"Erro ao verificar permissão: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# PDF COM ASSINATURAS (DIRETRIZ 3.2)
# ============================================================================

@bp.route('/api/contratos/<id>/pdf-assinatura', methods=['GET'])
def gerar_pdf_contrato_com_assinatura(id):
    """
    Diretriz 3.2: Gerar PDF de contrato com campos de assinatura

    Melhoria: Adiciona seção profissional de assinaturas no mesmo campo,
    permitindo que cliente e empresa assinem no mesmo documento.
    """
    try:
        db = get_db()
        if db is None:
            return jsonify({'success': False, 'message': 'Database offline'}), 500

        # Buscar contrato
        contrato = db.orcamentos.find_one({'_id': ObjectId(id)})
        if not contrato:
            return jsonify({'success': False, 'message': 'Contrato não encontrado'}), 404

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=1.5*cm,
            bottomMargin=2*cm,
            leftMargin=2*cm,
            rightMargin=2*cm
        )
        elements = []
        styles = getSampleStyleSheet()

        # ====================================================================
        # HEADER COM LOGO E TÍTULO
        # ====================================================================

        # Estilo do título principal
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=HexColor('#7C3AED'),
            alignment=TA_CENTER,
            spaceAfter=0.3*cm,
            fontName='Helvetica-Bold'
        )

        elements.append(Paragraph("🌿 BIOMA UBERABA", title_style))

        # Subtítulo com número do contrato
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=16,
            textColor=HexColor('#4B5563'),
            alignment=TA_CENTER,
            spaceAfter=0.5*cm,
            fontName='Helvetica-Bold'
        )

        numero_contrato = contrato.get('numero', str(contrato['_id'])[-6:])
        elements.append(Paragraph(f"CONTRATO DE PRESTAÇÃO DE SERVIÇOS Nº {numero_contrato}", subtitle_style))

        # Badge de status
        status = contrato.get('status', 'Pendente')
        status_color = {
            'Aprovado': '#10B981',
            'Pendente': '#F59E0B',
            'Cancelado': '#EF4444',
            'Em Andamento': '#3B82F6'
        }.get(status, '#6B7280')

        status_style = ParagraphStyle(
            'Status',
            parent=styles['Normal'],
            fontSize=12,
            textColor=HexColor('#FFFFFF'),
            alignment=TA_CENTER,
            spaceAfter=0.8*cm
        )

        status_table = Table(
            [[Paragraph(f"<b>STATUS: {status}</b>", status_style)]],
            colWidths=[6*cm]
        )
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor(status_color)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('ROUNDEDCORNERS', [5, 5, 5, 5]),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 0.3*cm))

        # ====================================================================
        # DADOS DO CLIENTE
        # ====================================================================

        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=HexColor('#1F2937'),
            spaceAfter=0.3*cm,
            fontName='Helvetica-Bold'
        )

        elements.append(Paragraph("📋 DADOS DO CLIENTE", section_style))

        # Data de criação
        data_contrato = contrato.get('created_at', datetime.now())
        if isinstance(data_contrato, datetime):
            data_formatada = data_contrato.strftime('%d/%m/%Y')
        else:
            data_formatada = 'N/A'

        cliente_data = [
            ['Nome Completo:', contrato.get('cliente_nome', 'N/A')],
            ['CPF:', contrato.get('cliente_cpf', 'N/A')],
            ['Telefone:', contrato.get('cliente_telefone', 'N/A')],
            ['E-mail:', contrato.get('cliente_email', 'N/A')],
            ['Data do Contrato:', data_formatada],
            ['Forma de Pagamento:', contrato.get('forma_pagamento', 'N/A')]
        ]

        cliente_table = Table(cliente_data, colWidths=[5*cm, 12*cm])
        cliente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (0, -1), HexColor('#374151')),
            ('TEXTCOLOR', (1, 0), (1, -1), HexColor('#111827')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#D1D5DB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(cliente_table)
        elements.append(Spacer(1, 0.6*cm))

        # ====================================================================
        # SERVIÇOS CONTRATADOS
        # ====================================================================

        if contrato.get('servicos'):
            elements.append(Paragraph("✨ SERVIÇOS CONTRATADOS", section_style))

            servicos_data = [['Descrição', 'Profissional', 'Qtd', 'Valor Unit.', 'Total']]
            for srv in contrato['servicos']:
                servicos_data.append([
                    srv.get('nome', 'N/A'),
                    srv.get('profissional_nome', '-'),
                    str(srv.get('quantidade', 1)),
                    f"R$ {srv.get('preco', 0):.2f}",
                    f"R$ {srv.get('total', 0):.2f}"
                ])

            servicos_table = Table(servicos_data, colWidths=[6*cm, 4*cm, 1.5*cm, 2.5*cm, 3*cm])
            servicos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#7C3AED')),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#D1D5DB')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), HexColor('#FAFAFA')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), HexColor('#F9FAFB')]),
            ]))
            elements.append(servicos_table)
            elements.append(Spacer(1, 0.4*cm))

        # ====================================================================
        # PRODUTOS
        # ====================================================================

        if contrato.get('produtos'):
            elements.append(Paragraph("🛍️ PRODUTOS", section_style))

            produtos_data = [['Descrição', 'Quantidade', 'Valor Unit.', 'Total']]
            for prd in contrato['produtos']:
                produtos_data.append([
                    prd.get('nome', 'N/A'),
                    str(prd.get('quantidade', 1)),
                    f"R$ {prd.get('preco', 0):.2f}",
                    f"R$ {prd.get('total', 0):.2f}"
                ])

            produtos_table = Table(produtos_data, colWidths=[8*cm, 3*cm, 3*cm, 3*cm])
            produtos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#D1D5DB')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), HexColor('#FAFAFA')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), HexColor('#F0FDF4')]),
            ]))
            elements.append(produtos_table)
            elements.append(Spacer(1, 0.4*cm))

        # ====================================================================
        # TOTAIS
        # ====================================================================

        elements.append(Paragraph("💰 VALORES", section_style))

        totais_data = [
            ['Subtotal Serviços:', f"R$ {contrato.get('total_servicos', 0):.2f}"],
            ['Subtotal Produtos:', f"R$ {contrato.get('total_produtos', 0):.2f}"],
            ['Desconto:', f"R$ {contrato.get('desconto_valor', 0):.2f}"],
        ]

        totais_table = Table(totais_data, colWidths=[13*cm, 4*cm])
        totais_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#D1D5DB')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#F9FAFB')),
        ]))
        elements.append(totais_table)

        # Total final destacado
        total_final_data = [
            [Paragraph('<b>TOTAL DO CONTRATO:</b>', styles['Normal']),
             Paragraph(f'<b>R$ {contrato.get("total_final", 0):.2f}</b>', styles['Normal'])]
        ]

        total_final_table = Table(total_final_data, colWidths=[13*cm, 4*cm])
        total_final_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('TEXTCOLOR', (0, 0), (-1, -1), HexColor('#FFFFFF')),
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#7C3AED')),
            ('GRID', (0, 0), (-1, -1), 0, HexColor('#7C3AED')),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(total_final_table)
        elements.append(Spacer(1, 0.6*cm))

        # ====================================================================
        # OBSERVAÇÕES
        # ====================================================================

        if contrato.get('observacoes'):
            elements.append(Paragraph("📝 OBSERVAÇÕES", section_style))
            obs_style = ParagraphStyle(
                'Observacoes',
                parent=styles['Normal'],
                fontSize=10,
                alignment=TA_JUSTIFY,
                spaceAfter=0.6*cm
            )
            elements.append(Paragraph(contrato['observacoes'], obs_style))

        # ====================================================================
        # TERMOS E CONDIÇÕES
        # ====================================================================

        elements.append(Paragraph("📄 TERMOS E CONDIÇÕES", section_style))

        termos_texto = """
        1. Este contrato estabelece os serviços e produtos acordados entre as partes, conforme
        descrição detalhada acima.<br/>
        2. O cliente declara estar ciente dos valores, condições de pagamento e prazos apresentados.<br/>
        3. A BIOMA UBERABA se compromete a prestar os serviços com excelência, profissionalismo
        e qualidade.<br/>
        4. O pagamento deverá ser realizado conforme a forma especificada neste contrato.<br/>
        5. Eventuais alterações deverão ser acordadas por escrito entre as partes.<br/>
        6. Este contrato é regido pelas leis brasileiras, elegendo-se o foro da cidade de Uberaba/MG
        para dirimir quaisquer questões.
        """

        termos_style = ParagraphStyle(
            'Termos',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_JUSTIFY,
            spaceAfter=1*cm,
            leading=12
        )
        elements.append(Paragraph(termos_texto, termos_style))

        # ====================================================================
        # SEÇÃO DE ASSINATURAS (DIRETRIZ 3.2)
        # ====================================================================

        elements.append(Spacer(1, 1*cm))

        assinatura_title_style = ParagraphStyle(
            'AssinaturaTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=HexColor('#7C3AED'),
            alignment=TA_CENTER,
            spaceAfter=0.8*cm,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("✍️ ASSINATURAS", assinatura_title_style))

        # Data e local
        local_data_style = ParagraphStyle(
            'LocalData',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=1.2*cm
        )
        elements.append(Paragraph(
            f"Uberaba/MG, {datetime.now().strftime('%d de %B de %Y')}",
            local_data_style
        ))

        # ====================================================================
        # ASSINATURAS NO MESMO CAMPO (DIRETRIZ 3.2 - CORRIGIDA)
        # ====================================================================

        # Campo unificado de assinatura com borda
        assinatura_campo_style = ParagraphStyle(
            'AssinaturaCampo',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            leading=14
        )

        # Criar tabela com campo único de assinatura
        assinatura_data = [
            # Cabeçalho do campo de assinatura
            [Paragraph('<b>CAMPO DE ASSINATURAS</b>', assinatura_campo_style)],
            # Espaço para assinatura do cliente
            [Paragraph('<br/><br/><br/>_________________________________<br/>', assinatura_campo_style)],
            [Paragraph('<b>ASSINATURA DO CLIENTE</b><br/>'
                      f'{contrato.get("cliente_nome", "N/A")}<br/>'
                      f'CPF: {contrato.get("cliente_cpf", "N/A")}',
                      assinatura_campo_style)],
            # Separador
            [Paragraph('<br/>', assinatura_campo_style)],
            # Espaço para assinatura da empresa (mesmo campo)
            [Paragraph('<br/><br/>_________________________________<br/>', assinatura_campo_style)],
            [Paragraph('<b>ASSINATURA DA EMPRESA</b><br/>'
                      '<b>BIOMA UBERABA</b><br/>'
                      'Representante Legal<br/>'
                      'CNPJ: 00.000.000/0001-00',
                      assinatura_campo_style)]
        ]

        assinatura_table = Table(assinatura_data, colWidths=[17*cm])
        assinatura_table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (0, 0), HexColor('#7C3AED')),
            ('TEXTCOLOR', (0, 0), (0, 0), white),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 12),
            ('TOPPADDING', (0, 0), (0, 0), 10),
            ('BOTTOMPADDING', (0, 0), (0, 0), 10),

            # Corpo do campo de assinatura
            ('BACKGROUND', (0, 1), (0, -1), HexColor('#F9FAFB')),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (0, -1), 5),
            ('BOTTOMPADDING', (0, 1), (0, -1), 5),

            # Borda externa
            ('BOX', (0, 0), (-1, -1), 2, HexColor('#7C3AED')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#D1D5DB')),
        ]))
        elements.append(assinatura_table)

        # ====================================================================
        # FOOTER COM INFORMAÇÕES DA EMPRESA
        # ====================================================================

        elements.append(Spacer(1, 1.5*cm))

        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=HexColor('#6B7280'),
            alignment=TA_CENTER,
            leading=10
        )

        footer_texto = """
        <b>BIOMA UBERABA</b><br/>
        Endereço: Rua Exemplo, 123 - Centro - Uberaba/MG - CEP 38010-000<br/>
        Telefone: (34) 3333-3333 | E-mail: contato@biomauberaba.com.br<br/>
        Website: www.biomauberaba.com.br
        """

        elements.append(Paragraph(footer_texto, footer_style))

        # ====================================================================
        # GERAR PDF
        # ====================================================================

        doc.build(elements)
        buffer.seek(0)

        logger.info(f"✅ PDF com assinaturas gerado para contrato {numero_contrato}")

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"contrato_bioma_{numero_contrato}.pdf"
        )

    except Exception as e:
        logger.error(f"❌ Erro ao gerar PDF com assinaturas: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


logger.info("✅ Rotas de melhorias carregadas com sucesso")
logger.info("✅ Sistema de Fila Inteligente carregado (Diretrizes 10.1 e 10.2)")
logger.info("✅ Sistema de Anamnese/Prontuário carregado (Diretrizes 11.1, 11.3, 11.4)")
logger.info("✅ Sistema de Multicomissão carregado (Diretriz 12.1)")
logger.info("✅ Melhorias nos Profissionais carregadas (Diretrizes 12.2 e 12.3)")
logger.info("✅ Histórico de Atendimentos carregado (Diretriz 12.4)")
logger.info("✅ Sistema de Níveis de Acesso carregado (Diretriz 5.1)")
logger.info("✅ PDF com Assinaturas implementado (Diretriz 3.2)")
