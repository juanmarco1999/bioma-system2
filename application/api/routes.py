#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BIOMA v3.7 - Todas as Rotas (Consolidado)
Migrado automaticamente do app.py monolítico
"""

from flask import request, jsonify, session, current_app, send_file, render_template, Response, stream_with_context
from flask_cors import CORS
from pymongo import MongoClient, ASCENDING, DESCENDING
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from bson import ObjectId
from functools import wraps, lru_cache
from datetime import datetime, timedelta
from dotenv import load_dotenv
import urllib.parse
import os
import io
from io import BytesIO
import csv
import json
import re
import requests
import logging
import random
from time import time

# ReportLab para PDFs
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.colors import HexColor, black, white
from reportlab.lib.units import cm

# OpenPyXL para Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from application.api import bp
from application.decorators import login_required, permission_required, get_user_permissions
from application.utils import convert_objectid, allowed_file, registrar_auditoria, update_cliente_denormalized_fields, get_assistente_details
from application.constants import ANAMNESE_FORM, PRONTUARIO_FORM, default_form_state
from application.extensions import get_from_cache, set_in_cache

logger = logging.getLogger(__name__)

# Importar DB diretamente de extensions
from application.extensions import db as database_connection

# Helper para obter DB
def get_db():
    """Retorna instância do MongoDB"""
    # Tentar primeiro de current_app (produção)
    try:
        db_from_app = current_app.config.get('DB_CONNECTION')
        if db_from_app is not None:
            return db_from_app
    except:
        pass

    # Fallback: usar db de extensions
    return database_connection




@bp.route('/')
def index():
    """Renderiza index.html com cache busting via timestamp"""
    from flask import make_response
    import time

    db = get_db()

    # Renderizar template
    html = render_template('index.html', cache_buster=str(int(time.time())))

    # Criar response com headers anti-cache
    response = make_response(html)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'

    return response

@bp.route('/health')
def health():
    """Health check endpoint com diagnóstico detalhado"""
    import sys
    import os

    db = get_db()
    db_status = 'disconnected'
    db_error = None

    if db is not None:
        try:
            db.command('ping')
            db_status = 'connected'
        except Exception as e:
            db_status = 'error'
            db_error = str(e)

    # Verificar variáveis de ambiente (sem expor valores)
    env_check = {
        'MONGO_USERNAME': bool(os.getenv('MONGO_USERNAME')),
        'MONGO_PASSWORD': bool(os.getenv('MONGO_PASSWORD')),
        'MONGO_CLUSTER': bool(os.getenv('MONGO_CLUSTER')),
        'FLASK_ENV': os.getenv('FLASK_ENV', 'not_set'),
        'SECRET_KEY': bool(os.getenv('SECRET_KEY'))
    }

    response = {
        'status': 'healthy' if db_status == 'connected' else 'degraded',
        'version': '6.0',
        'time': datetime.now().isoformat(),
        'database': db_status,
        'python_version': f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        'environment': env_check
    }

    if db_error:
        response['database_error'] = db_error

    # Retornar 200 mesmo se degraded (para Render não marcar como down)
    return jsonify(response), 200

@bp.route('/api/ping', methods=['GET'])
def ping():
    """
    Endpoint de ping para verificação de conectividade
    Sistema offline v4.0
    """
    return jsonify({'success': True, 'status': 'online', 'timestamp': datetime.now().isoformat()}), 200

@bp.route('/api/login', methods=['POST'])
def login():
    db = get_db()
    data = request.json
    logger.info(f"🔐 Login attempt: {data.get('username')}")
    
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        user = db.users.find_one({'$or': [{'username': data['username']}, {'email': data['username']}]})
        
        if user and check_password_hash(user['password'], data['password']):
            session.permanent = True
            session['user_id'] = str(user['_id'])
            session['username'] = user['username']
            session['role'] = user.get('role', 'admin')
            session['tipo_acesso'] = user.get('tipo_acesso', 'Admin')  # Admin, Gestão, Profissional

            logger.info(f"✅ Login SUCCESS: {user['username']} (role: {session['role']}, tipo: {session['tipo_acesso']})")

            return jsonify({
                'success': True,
                'user': {
                    'id': str(user['_id']),
                    'name': user['name'],
                    'username': user['username'],
                    'email': user['email'],
                    'role': user.get('role', 'admin'),
                    'tipo_acesso': user.get('tipo_acesso', 'Admin'),
                    'theme': user.get('theme', 'light')
                }
            })
        
        logger.warning(f"❌ Login FAILED: {data.get('username')}")
        return jsonify({'success': False, 'message': 'Usuário ou senha inválidos'})
        
    except Exception as e:
        logger.error(f"❌ Login ERROR: {e}")
        return jsonify({'success': False, 'message': 'Erro no servidor'}), 500

@bp.route('/api/register', methods=['POST'])
def register():
    db = get_db()
    data = request.json
    logger.info(f"👤 Register attempt: {data.get('username')}")

    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        if db.users.find_one({'$or': [{'username': data['username']}, {'email': data['email']}]}):
            return jsonify({'success': False, 'message': 'Usuário ou email já existe'})
        
        user_data = {
            'name': data['name'],
            'username': data['username'],
            'email': data['email'],
            'telefone': data.get('telefone', ''),
            'password': generate_password_hash(data['password']),
            'role': 'admin',
            'tipo_acesso': data.get('tipo_acesso', 'Admin'),  # Admin, Gestão, Profissional (Diretriz #6)
            'theme': 'light',
            'created_at': datetime.now()
        }

        db.users.insert_one(user_data)
        logger.info(f"✅ User registered: {data['username']} with ADMIN role and {user_data['tipo_acesso']} access")
        
        return jsonify({'success': True, 'message': 'Conta criada com sucesso! Faça login.'})
        
    except Exception as e:
        logger.error(f"❌ Register ERROR: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/logout', methods=['POST'])
def logout():
    db = get_db()
    username = session.get('username', 'Unknown')
    session.clear()
    logger.info(f"🚪 Logout: {username}")
    return jsonify({'success': True})

@bp.route('/api/current-user')
def current_user():
    db = get_db()
    if 'user_id' in session and db is not None:
        try:
            user = db.users.find_one({'_id': ObjectId(session['user_id'])})
            if user:
                # Obter permissões do usuário
                permissions_data = get_user_permissions()

                return jsonify({
                    'success': True,
                    'user': {
                        'id': str(user['_id']),
                        'name': user['name'],
                        'username': user['username'],
                        'email': user['email'],
                        'role': user.get('role', 'admin'),
                        'tipo_acesso': user.get('tipo_acesso', 'Admin'),
                        'theme': user.get('theme', 'light'),
                        'permissions': permissions_data.get('permissoes', []) if permissions_data else []
                    }
                })
        except Exception as e:
            logger.error(f"❌ Current user error: {e}")
    return jsonify({'success': False})

@bp.route('/api/permissions')
@login_required
def get_permissions():
    db = get_db()
    """Retorna as permissões do usuário atual"""
    permissions_data = get_user_permissions()

    if permissions_data:
        return jsonify({
            'success': True,
            'tipo_acesso': permissions_data.get('tipo_acesso', 'Profissional'),
            'permissions': permissions_data.get('permissoes', [])
        })
    else:
        return jsonify({
            'success': False,
            'message': 'Não foi possível obter permissões'
        }), 500

@bp.route('/api/update-theme', methods=['POST'])
@login_required
def update_theme():
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    try:
        db.users.update_one({'_id': ObjectId(session['user_id'])}, {'$set': {'theme': request.json['theme']}})
        return jsonify({'success': True})
    except:
        return jsonify({'success': False}), 500


# ==================== GESTÃO DE USUÁRIOS E TIPOS DE ACESSO (Diretriz #6) ====================

@bp.route('/api/users', methods=['GET'])
@login_required
@permission_required('Admin')
def list_users():
    db = get_db()
    """Listar todos os usuários do sistema (Admin only)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        users = list(db.users.find({}, {'password': 0}).sort('name', ASCENDING))
        return jsonify({'success': True, 'users': convert_objectid(users)})

    except Exception as e:
        logger.error(f"Erro ao listar usuários: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/users/<id>/tipo-acesso', methods=['PUT'])
@login_required
@permission_required('Admin')
def update_user_tipo_acesso(id):
    db = get_db()
    """Atualizar tipo de acesso de um usuário (Admin only)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        data = request.json
        novo_tipo = data.get('tipo_acesso')

        # Validar tipo de acesso
        tipos_validos = ['Admin', 'Gestão', 'Profissional']
        if novo_tipo not in tipos_validos:
            return jsonify({'success': False, 'message': f'Tipo de acesso inválido. Use: {", ".join(tipos_validos)}'}), 400

        # Atualizar usuário
        result = db.users.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'tipo_acesso': novo_tipo, 'updated_at': datetime.now()}}
        )

        if result.modified_count > 0:
            logger.info(f"✅ Tipo de acesso atualizado para usuário {id}: {novo_tipo}")
            return jsonify({'success': True, 'message': f'Tipo de acesso alterado para {novo_tipo}'})
        else:
            return jsonify({'success': False, 'message': 'Usuário não encontrado ou tipo já definido'}), 404

    except Exception as e:
        logger.error(f"Erro ao atualizar tipo de acesso: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/users/<id>', methods=['GET'])
@login_required
def get_user(id):
    db = get_db()
    """Obter detalhes de um usuário específico"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        user = db.users.find_one({'_id': ObjectId(id)}, {'password': 0})
        if not user:
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404

        return jsonify({'success': True, 'user': convert_objectid(user)})

    except Exception as e:
        logger.error(f"Erro ao buscar usuário: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/system/status')
@login_required
def system_status():
    db = get_db()
    mongo_ok = False
    mongo_msg = 'Desconectado'
    try:
        if db is not None:
            db.command('ping')
            mongo_ok = True
            mongo_msg = 'Conectado'
    except Exception as e:
        mongo_msg = f'Erro: {str(e)[:100]}'
    
    return jsonify({
        'success': True,
        'status': {
            'mongodb': {'operational': mongo_ok, 'message': mongo_msg, 'last_check': datetime.now().isoformat()},
            'mailersend': {'operational': bool(os.getenv('MAILERSEND_API_KEY')), 'message': 'Configurado' if bool(os.getenv('MAILERSEND_API_KEY')) else 'Não configurado'},
            'server': {'time': datetime.now().isoformat(), 'version': '3.7.0'}
        }
    })

@bp.route('/api/dashboard/stats')
@login_required
def dashboard_stats():
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    
    try:
        hoje_inicio = datetime.now().replace(hour=0, minute=0, second=0)
        hoje_fim = datetime.now().replace(hour=23, minute=59, second=59)
        
        agendamentos_hoje = 0
        if 'agendamentos' in db.list_collection_names():
            agendamentos_hoje = db.agendamentos.count_documents({'data': {'$gte': hoje_inicio, '$lte': hoje_fim}})
        
        stats = {
            'total_orcamentos': db.orcamentos.count_documents({}),
            'total_clientes': db.clientes.count_documents({}),
            'total_servicos': db.servicos.count_documents({}),
            'faturamento': sum(o.get('total_final', 0) for o in db.orcamentos.find({'status': 'Aprovado'})),
            'agendamentos_hoje': agendamentos_hoje
        }
        return jsonify({'success': True, 'stats': stats})
    except Exception as e:
        logger.error(f"Erro ao buscar stats: {e}")
        return jsonify({'success': False}), 500

@bp.route('/api/dashboard/stats/realtime')
@login_required
def dashboard_stats_realtime():
    db = get_db()
    """Estatísticas em tempo real com métricas avançadas"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        agora = datetime.now()
        hoje_inicio = agora.replace(hour=0, minute=0, second=0, microsecond=0)
        semana_inicio = agora - timedelta(days=7)
        mes_inicio = agora - timedelta(days=30)
        
        # Estatísticas básicas
        total_clientes = db.clientes.count_documents({})
        total_profissionais = db.profissionais.count_documents({})
        total_produtos = db.produtos.count_documents({})
        total_servicos = db.servicos.count_documents({})
        
        # Orçamentos
        total_orcamentos = db.orcamentos.count_documents({})
        orcamentos_pendentes = db.orcamentos.count_documents({'status': 'Pendente'})
        orcamentos_aprovados = db.orcamentos.count_documents({'status': 'Aprovado'})
        
        # Faturamento OTIMIZADO com Aggregation Pipeline (Roadmap - Query Optimization)
        # ANTES: 3 queries separadas (ineficiente)
        # DEPOIS: 1 aggregation pipeline (usa índices, processa no servidor)
        faturamento_pipeline = [
            {'$match': {'status': 'Aprovado'}},
            {'$facet': {
                'total': [
                    {'$group': {'_id': None, 'valor': {'$sum': '$total_final'}}}
                ],
                'mes': [
                    {'$match': {'data_criacao': {'$gte': mes_inicio}}},
                    {'$group': {'_id': None, 'valor': {'$sum': '$total_final'}}}
                ],
                'hoje': [
                    {'$match': {'data_criacao': {'$gte': hoje_inicio}}},
                    {'$group': {'_id': None, 'valor': {'$sum': '$total_final'}}}
                ]
            }}
        ]

        faturamento_result = list(db.orcamentos.aggregate(faturamento_pipeline))
        if faturamento_result:
            faturamento_total = faturamento_result[0]['total'][0]['valor'] if faturamento_result[0]['total'] else 0
            faturamento_mes = faturamento_result[0]['mes'][0]['valor'] if faturamento_result[0]['mes'] else 0
            faturamento_hoje = faturamento_result[0]['hoje'][0]['valor'] if faturamento_result[0]['hoje'] else 0
        else:
            faturamento_total = faturamento_mes = faturamento_hoje = 0
        
        # Agendamentos
        agendamentos_hoje = 0
        agendamentos_semana = 0
        if 'agendamentos' in db.list_collection_names():
            agendamentos_hoje = db.agendamentos.count_documents({
                'data': {'$gte': hoje_inicio, '$lte': agora}
            })
            agendamentos_semana = db.agendamentos.count_documents({
                'data': {'$gte': semana_inicio}
            })
        
        # Estoque OTIMIZADO com Aggregation (Roadmap - Query Optimization)
        # ANTES: Carregar todos produtos em memória Python para calcular valor
        # DEPOIS: Calcular valor no servidor MongoDB com $multiply
        estoque_pipeline = [
            {'$facet': {
                'valor_total': [
                    {'$project': {
                        'valor_item': {'$multiply': ['$preco', '$estoque']}
                    }},
                    {'$group': {'_id': None, 'total': {'$sum': '$valor_item'}}}
                ],
                'baixo_estoque': [
                    {'$match': {'$expr': {'$lt': ['$estoque', '$estoque_minimo']}}},
                    {'$count': 'total'}
                ],
                'sem_estoque': [
                    {'$match': {'estoque': 0}},
                    {'$count': 'total'}
                ]
            }}
        ]

        estoque_result = list(db.produtos.aggregate(estoque_pipeline))
        if estoque_result:
            valor_estoque = estoque_result[0]['valor_total'][0]['total'] if estoque_result[0]['valor_total'] else 0
            produtos_baixo_estoque = estoque_result[0]['baixo_estoque'][0]['total'] if estoque_result[0]['baixo_estoque'] else 0
            produtos_sem_estoque = estoque_result[0]['sem_estoque'][0]['total'] if estoque_result[0]['sem_estoque'] else 0
        else:
            valor_estoque = produtos_baixo_estoque = produtos_sem_estoque = 0
        
        # Pendências
        entradas_pendentes = 0
        if 'estoque_pendencias' in db.list_collection_names():
            entradas_pendentes = db.estoque_pendencias.count_documents({'status': 'pendente'})
        
        # Comissões do mês
        comissoes_mes = 0
        if 'comissoes_historico' in db.list_collection_names():
            comissoes_mes = sum(c.get('valor_total', 0) for c in db.comissoes_historico.find({
                'data': {'$gte': mes_inicio}
            }))
        
        # Clientes novos (últimos 30 dias)
        clientes_novos = db.clientes.count_documents({
            'data_cadastro': {'$gte': mes_inicio}
        })
        
        return jsonify({
            'success': True,
            'timestamp': agora.isoformat(),
            'stats': {
                'totais': {
                    'clientes': total_clientes,
                    'profissionais': total_profissionais,
                    'produtos': total_produtos,
                    'servicos': total_servicos,
                    'orcamentos': total_orcamentos
                },
                'orcamentos': {
                    'pendentes': orcamentos_pendentes,
                    'aprovados': orcamentos_aprovados,
                    'taxa_aprovacao': round((orcamentos_aprovados / total_orcamentos * 100) if total_orcamentos > 0 else 0, 2)
                },
                'faturamento': {
                    'total': round(faturamento_total, 2),
                    'mes': round(faturamento_mes, 2),
                    'hoje': round(faturamento_hoje, 2),
                    'ticket_medio': round(faturamento_total / orcamentos_aprovados if orcamentos_aprovados > 0 else 0, 2)
                },
                'agendamentos': {
                    'hoje': agendamentos_hoje,
                    'semana': agendamentos_semana
                },
                'estoque': {
                    'valor_total': round(valor_estoque, 2),
                    'baixo_estoque': produtos_baixo_estoque,
                    'sem_estoque': produtos_sem_estoque,
                    'alerta': produtos_baixo_estoque > 0 or produtos_sem_estoque > 0
                },
                'pendencias': {
                    'entradas_estoque': entradas_pendentes,
                    'orcamentos': orcamentos_pendentes,
                    'total': entradas_pendentes + orcamentos_pendentes
                },
                'crescimento': {
                    'clientes_novos_mes': clientes_novos,
                    'comissoes_mes': round(comissoes_mes, 2)
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao gerar estatísticas em tempo real: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    except Exception as e:
        logger.error(f"❌ Dashboard stats error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes', methods=['GET', 'POST'])
@login_required
def clientes():
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500

    if request.method == 'GET':
        try:
            # ==================== PERFORMANCE OPTIMIZATION ====================
            # Implement pagination to prevent loading thousands of records at once
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 50, type=int)
            skip = (page - 1) * per_page

            # Get total count for pagination metadata
            total_count = db.clientes.count_documents({})
            total_pages = (total_count + per_page - 1) // per_page

            # Use projection to only return needed fields (reduces data transfer)
            projection = {
                'nome': 1,
                'cpf': 1,
                'email': 1,
                'telefone': 1,
                'total_faturado': 1,  # Denormalized field
                'ultima_visita': 1,    # Denormalized field
                'total_visitas': 1,    # Denormalized field
                'created_at': 1
            }

            clientes_list = list(
                db.clientes.find({}, projection)
                .sort('nome', ASCENDING)
                .skip(skip)
                .limit(per_page)
            )

            # If denormalized fields don't exist, calculate and store them
            # This is a one-time migration for existing records
            for cliente in clientes_list:
                cliente_cpf = cliente.get('cpf')

                if 'total_faturado' not in cliente:
                    # Calculate and denormalize
                    total_faturado = sum(
                        o.get('total_final', 0)
                        for o in db.orcamentos.find(
                            {'cliente_cpf': cliente_cpf, 'status': 'Aprovado'},
                            {'total_final': 1}
                        )
                    )
                    ultimo_orc = db.orcamentos.find_one(
                        {'cliente_cpf': cliente_cpf},
                        sort=[('created_at', DESCENDING)],
                        projection={'created_at': 1}
                    )
                    ultima_visita = ultimo_orc['created_at'] if ultimo_orc else None
                    total_visitas = db.orcamentos.count_documents({'cliente_cpf': cliente_cpf})

                    # Store denormalized values for future queries
                    db.clientes.update_one(
                        {'_id': cliente['_id']},
                        {'$set': {
                            'total_faturado': total_faturado,
                            'ultima_visita': ultima_visita,
                            'total_visitas': total_visitas
                        }}
                    )

                    cliente['total_faturado'] = total_faturado
                    cliente['ultima_visita'] = ultima_visita
                    cliente['total_visitas'] = total_visitas

                # Maintain backwards compatibility
                cliente['total_gasto'] = cliente.get('total_faturado', 0)

            return jsonify({
                'success': True,
                'clientes': convert_objectid(clientes_list),
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total_count': total_count,
                    'total_pages': total_pages,
                    'has_next': page < total_pages,
                    'has_prev': page > 1
                }
            })
        except Exception as e:
            logger.error(f"❌ Error loading clientes: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500
    
    # Aceitar tanto JSON quanto FormData (multipart para upload de foto)
    if request.is_json:
        data = request.json
    else:
        # FormData (multipart/form-data)
        data = request.form.to_dict()

    try:
        # Validar campos obrigatórios
        if not data.get('nome') or not data.get('cpf'):
            return jsonify({'success': False, 'message': 'Nome e CPF são obrigatórios'}), 400

        existing = db.clientes.find_one({'cpf': data['cpf']})

        # Dados completos do cliente
        cliente_data = {
            'nome': data['nome'],
            'cpf': data['cpf'],
            'email': data.get('email', ''),
            'telefone': data.get('telefone', ''),
            'data_nascimento': data.get('data_nascimento', ''),
            'cep': data.get('cep', ''),
            'estado': data.get('estado', ''),
            'cidade': data.get('cidade', ''),
            'bairro': data.get('bairro', ''),
            'rua': data.get('rua', ''),
            'numero': data.get('numero', ''),
            'observacoes': data.get('observacoes', ''),
            'updated_at': datetime.now()
        }

        # Upload de foto (se enviado)
        if 'foto' in request.files:
            foto = request.files['foto']
            if foto and foto.filename:
                # Processar foto (salvar como base64 ou URL)
                import base64
                foto_b64 = base64.b64encode(foto.read()).decode('utf-8')
                cliente_data['foto_url'] = f"data:image/jpeg;base64,{foto_b64}"

        if existing:
            db.clientes.update_one({'cpf': data['cpf']}, {'$set': cliente_data})
            logger.info(f"✅ Cliente atualizado: {data['nome']} (CPF: {data['cpf']})")
        else:
            cliente_data['created_at'] = datetime.now()
            cliente_data['total_faturado'] = 0
            cliente_data['total_visitas'] = 0
            db.clientes.insert_one(cliente_data)
            logger.info(f"✅ Cliente criado: {data['nome']} (CPF: {data['cpf']})")

        return jsonify({'success': True, 'message': 'Cliente salvo com sucesso'})
    except Exception as e:
        logger.error(f"❌ Erro ao salvar cliente: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes/<id>', methods=['DELETE'])
@login_required
def delete_cliente(id):
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    try:
        result = db.clientes.delete_one({'_id': ObjectId(id)})
        return jsonify({'success': result.deleted_count > 0})
    except:
        return jsonify({'success': False}), 500

@bp.route('/api/clientes/<id>', methods=['GET'])
@login_required
def get_cliente(id):
    db = get_db()
    """Visualizar um cliente específico"""
    if db is None:
        return jsonify({'success': False}), 500
    try:
        cliente = db.clientes.find_one({'_id': ObjectId(id)})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
        
        # Adicionar estatísticas
        cliente_cpf = cliente.get('cpf')
        # ALTERADO: total_gasto -> total_faturado (Diretriz #11)
        cliente['total_faturado'] = sum(o.get('total_final', 0) for o in db.orcamentos.find({'cliente_cpf': cliente_cpf, 'status': 'Aprovado'}))
        cliente['total_gasto'] = cliente['total_faturado']  # Mantém compatibilidade
        ultimo_orc = db.orcamentos.find_one({'cliente_cpf': cliente_cpf}, sort=[('created_at', DESCENDING)])
        cliente['ultima_visita'] = ultimo_orc['created_at'] if ultimo_orc else None
        cliente['total_visitas'] = db.orcamentos.count_documents({'cliente_cpf': cliente_cpf})
        
        return jsonify({'success': True, 'cliente': convert_objectid(cliente)})
    except Exception as e:
        logger.error(f"Erro ao buscar cliente: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes/<id>', methods=['PUT'])
@login_required
def update_cliente(id):
    db = get_db()
    """Editar um cliente existente"""
    if db is None:
        return jsonify({'success': False}), 500
    
    data = request.json
    try:
        cliente_existente = db.clientes.find_one({'_id': ObjectId(id)})
        if not cliente_existente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
        
        # Verificar se o CPF já existe em outro cliente
        if data.get('cpf') != cliente_existente.get('cpf'):
            cpf_duplicado = db.clientes.find_one({'cpf': data['cpf'], '_id': {'$ne': ObjectId(id)}})
            if cpf_duplicado:
                return jsonify({'success': False, 'message': 'CPF já cadastrado em outro cliente'}), 400
        
        anamnese_respostas = data.get('anamnese')
        prontuario_respostas = data.get('prontuario')
        if anamnese_respostas is None:
            anamnese_respostas = cliente_existente.get('anamnese', default_form_state(ANAMNESE_FORM))
        if prontuario_respostas is None:
            prontuario_respostas = cliente_existente.get('prontuario', default_form_state(PRONTUARIO_FORM))

        update_data = {
            'nome': data.get('nome', cliente_existente.get('nome')),
            'cpf': data.get('cpf', cliente_existente.get('cpf')),
            'email': data.get('email', cliente_existente.get('email', '')),
            'telefone': data.get('telefone', cliente_existente.get('telefone', '')),
            'endereco': data.get('endereco', cliente_existente.get('endereco', '')),
            'data_nascimento': data.get('data_nascimento', cliente_existente.get('data_nascimento', '')),
            'genero': data.get('genero', cliente_existente.get('genero', '')),
            'estado_civil': data.get('estado_civil', cliente_existente.get('estado_civil', '')),
            'profissao': data.get('profissao', cliente_existente.get('profissao', '')),
            'instagram': data.get('instagram', cliente_existente.get('instagram', '')),
            'indicacao': data.get('indicacao', cliente_existente.get('indicacao', '')),
            'preferencias': data.get('preferencias', cliente_existente.get('preferencias', '')),
            'restricoes': data.get('restricoes', cliente_existente.get('restricoes', '')),
            'tipo_cabelo': data.get('tipo_cabelo', cliente_existente.get('tipo_cabelo', '')),
            'historico_tratamentos': data.get('historico_tratamentos', cliente_existente.get('historico_tratamentos', '')),
            'observacoes': data.get('observacoes', cliente_existente.get('observacoes', '')),
            'anamnese': anamnese_respostas if isinstance(anamnese_respostas, dict) else cliente_existente.get('anamnese', default_form_state(ANAMNESE_FORM)),
            'prontuario': prontuario_respostas if isinstance(prontuario_respostas, dict) else cliente_existente.get('prontuario', default_form_state(PRONTUARIO_FORM)),
            'updated_at': datetime.now()
        }
        
        db.clientes.update_one({'_id': ObjectId(id)}, {'$set': update_data})
        logger.info(f"✅ Cliente atualizado: {update_data['nome']}")
        
        return jsonify({'success': True, 'message': 'Cliente atualizado com sucesso'})
    except Exception as e:
        logger.error(f"Erro ao atualizar cliente: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes/buscar')
@login_required
def buscar_clientes():
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    termo = request.args.get('termo', '').strip()
    cache_key = f"clientes_busca_{termo}"
    
    # Tentar buscar do cache primeiro
    cached = get_from_cache(cache_key)
    if cached:
        return jsonify(cached)
    
    try:
        regex = {'$regex': termo, '$options': 'i'}
        clientes = list(db.clientes.find({
            '$or': [
                {'nome': regex},
                {'cpf': regex},
                {'email': regex},
                {'telefone': regex}
            ]
        }).sort('nome', ASCENDING).limit(50))
        
        # Adicionar informação completa formatada
        for c in clientes:
            c['display_name'] = f"{c.get('nome', '')} - CPF: {c.get('cpf', '')}"
        
        result = {'success': True, 'clientes': convert_objectid(clientes)}
        set_in_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        logger.error(f"Erro ao buscar clientes: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/search/suggest', methods=['GET'])
@login_required
def search_suggest():
    db = get_db()
    """Sugestões de busca para autocomplete"""
    if db is None:
        return jsonify({'success': False, 'suggestions': []}), 500

    query = request.args.get('q', '').strip()

    if not query or len(query) < 2:
        return jsonify({'success': True, 'suggestions': []})

    try:
        regex = {'$regex': query, '$options': 'i'}
        suggestions = []

        # Buscar em clientes
        clientes = list(db.clientes.find({'nome': regex}, {'nome': 1}).limit(5))
        for c in clientes:
            suggestions.append({'text': c['nome'], 'type': 'cliente', 'id': str(c['_id'])})

        # Buscar em produtos
        produtos = list(db.produtos.find({'nome': regex}, {'nome': 1}).limit(5))
        for p in produtos:
            suggestions.append({'text': p['nome'], 'type': 'produto', 'id': str(p['_id'])})

        # Buscar em profissionais
        profissionais = list(db.profissionais.find({'nome': regex}, {'nome': 1}).limit(3))
        for prof in profissionais:
            suggestions.append({'text': prof['nome'], 'type': 'profissional', 'id': str(prof['_id'])})

        # Buscar em serviços
        servicos = list(db.servicos.find({'nome': regex}, {'nome': 1}).limit(5))
        for s in servicos:
            suggestions.append({'text': s['nome'], 'type': 'servico', 'id': str(s['_id'])})

        return jsonify({'success': True, 'suggestions': suggestions[:15]})

    except Exception as e:
        logger.error(f"❌ Erro ao buscar sugestões: {e}")
        return jsonify({'success': False, 'suggestions': []}), 500


@bp.route('/api/busca/global', methods=['GET'])
@login_required
def busca_global():
    db = get_db()
    """Busca global em múltiplas collections"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    termo = request.args.get('termo', '').strip()
    
    if not termo or len(termo) < 2:
        return jsonify({'success': True, 'resultados': {'clientes': [], 'profissionais': [], 'produtos': [], 'servicos': []}})
    
    cache_key = f"busca_global_{termo}"
    cached = get_from_cache(cache_key)
    if cached:
        return jsonify(cached)
    
    try:
        regex = {'$regex': termo, '$options': 'i'}

        # OTIMIZAÇÃO: Usar projection para selecionar apenas campos necessários (Roadmap - Query Optimization)

        # Buscar em clientes (projection: apenas campos essenciais para busca)
        clientes = list(db.clientes.find({
            '$or': [
                {'nome': regex},
                {'cpf': regex},
                {'email': regex},
                {'telefone': regex}
            ]
        }, {
            '_id': 1,
            'nome': 1,
            'cpf': 1,
            'email': 1,
            'telefone': 1,
            'foto_url': 1
        }).limit(10))

        # Buscar em profissionais (projection: apenas campos essenciais)
        profissionais = list(db.profissionais.find({
            '$or': [
                {'nome': regex},
                {'cpf': regex},
                {'email': regex},
                {'especialidade': regex}
            ]
        }, {
            '_id': 1,
            'nome': 1,
            'cpf': 1,
            'email': 1,
            'especialidade': 1,
            'foto_url': 1,
            'ativo': 1
        }).limit(10))

        # Buscar em produtos (projection: apenas campos essenciais)
        produtos = list(db.produtos.find({
            '$or': [
                {'nome': regex},
                {'marca': regex},
                {'sku': regex}
            ]
        }, {
            '_id': 1,
            'nome': 1,
            'marca': 1,
            'sku': 1,
            'preco': 1,
            'estoque': 1,
            'ativo': 1
        }).limit(10))

        # Buscar em serviços (projection: apenas campos essenciais)
        servicos = list(db.servicos.find({
            '$or': [
                {'nome': regex},
                {'categoria': regex}
            ]
        }, {
            '_id': 1,
            'nome': 1,
            'categoria': 1,
            'preco': 1,
            'duracao': 1,
            'ativo': 1
        }).limit(10))
        
        result = {
            'success': True,
            'resultados': {
                'clientes': convert_objectid(clientes),
                'profissionais': convert_objectid(profissionais),
                'produtos': convert_objectid(produtos),
                'servicos': convert_objectid(servicos)
            },
            'total': len(clientes) + len(profissionais) + len(produtos) + len(servicos)
        }
        
        set_in_cache(cache_key, result)
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Erro na busca global: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/profissionais', methods=['GET', 'POST'])
@login_required
def profissionais():
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    if request.method == 'GET':
        cache_key = 'profissionais_list'
        cached = get_from_cache(cache_key)
        if cached:
            return jsonify(cached)
        
        try:
            profs = list(db.profissionais.find({}).sort('nome', ASCENDING).limit(500))

            # Agregar métricas de avaliação para exibição rápida na lista
            avaliacoes_map = {}
            try:
                stats_pipeline = [
                    {'$group': {
                        '_id': '$profissional_id',
                        'media': {'$avg': '$nota'},
                        'total': {'$sum': 1}
                    }}
                ]
                for stat in db.profissionais_avaliacoes.aggregate(stats_pipeline):
                    avaliacoes_map[stat['_id']] = {
                        'media': round(stat.get('media', 0), 2),
                        'total': stat.get('total', 0)
                    }
            except Exception as agg_error:
                logger.debug(f"Falha ao agregar avaliações de profissionais: {agg_error}")

            for prof in profs:
                assistente_info = get_assistente_details(
                    prof.get('assistente_id'),
                    prof.get('assistente_tipo')
                )
                if assistente_info:
                    prof['assistente'] = {
                        'id': assistente_info.get('_id'),
                        'nome': assistente_info.get('nome'),
                        'tipo': assistente_info.get('tipo_origem'),
                        'foto_url': assistente_info.get('foto_url', '')
                    }
                # Avaliações agregadas
                stat = avaliacoes_map.get(str(prof.get('_id')))
                if stat:
                    prof['avaliacao_media'] = stat['media']
                    prof['avaliacoes_total'] = stat['total']
                else:
                    prof['avaliacao_media'] = 0
                    prof['avaliacoes_total'] = 0

            result = {'success': True, 'profissionais': convert_objectid(profs)}
            set_in_cache(cache_key, result)
            return jsonify(result)
        except Exception as e:
            logger.error(f"Erro ao listar profissionais: {e}")
            return jsonify({'success': False, 'message': 'Erro ao carregar profissionais'}), 500
    
    # Aceitar tanto JSON quanto FormData
    if request.is_json:
        data = request.json
    else:
        data = request.form.to_dict()

    try:
        #Validar nome obrigatório
        if not data.get('nome'):
            return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400

        assistente_id = data.get('assistente_id') or None
        assistente_tipo = data.get('assistente_tipo')
        if assistente_id:
            assistente_id = str(assistente_id)
        if assistente_tipo not in {'profissional', 'assistente', None}:
            assistente_tipo = None

        profissional_data = {
            'nome': data['nome'],
            'cpf': data.get('cpf', ''),  # CPF OPCIONAL
            'email': data.get('email', ''),
            'telefone': data.get('telefone', ''),
            'especialidade': data.get('especialidade', ''),
            'comissao_perc': float(data.get('comissao_perc', 0)),
            'foto_url': data.get('foto_url', ''),
            'observacoes': data.get('observacoes', ''),
            'assistente_id': assistente_id,
            'assistente_tipo': assistente_tipo if assistente_id else None,
            'comissao_assistente_perc': float(data.get('comissao_assistente_perc', 0)),
            'ativo': True,
            'created_at': datetime.now()
        }

        # Upload de foto (se enviado via FormData)
        if 'foto' in request.files:
            foto = request.files['foto']
            if foto and foto.filename:
                import base64
                foto_b64 = base64.b64encode(foto.read()).decode('utf-8')
                profissional_data['foto_url'] = f"data:image/jpeg;base64,{foto_b64}"

        result = db.profissionais.insert_one(profissional_data)
        inserted_id = str(result.inserted_id)
        logger.info(f"✅ Profissional cadastrado: {profissional_data['nome']} (ID: {inserted_id})")
        return jsonify({'success': True, 'message': 'Profissional cadastrado com sucesso', 'id': inserted_id})
    except Exception as e:
        logger.error(f"❌ Erro ao cadastrar profissional: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/profissionais/<id>', methods=['DELETE'])
@login_required
def delete_profissional(id):
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Erro ao conectar ao banco'}), 500
    try:
        result = db.profissionais.delete_one({'_id': ObjectId(id)})

        if result.deleted_count > 0:
            logger.info(f"✅ Profissional deletado: {id}")
            return jsonify({'success': True, 'message': 'Profissional deletado com sucesso'})
        else:
            logger.warning(f"⚠️ Profissional não encontrado: {id}")
            return jsonify({'success': False, 'message': 'Profissional não encontrado'}), 404

    except Exception as e:
        logger.error(f"❌ Erro ao deletar profissional {id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao deletar: {str(e)}'}), 500

@bp.route('/api/profissionais/<id>', methods=['GET', 'PUT'])
@login_required
def profissional_detalhes(id):
    db = get_db()

    # PUT - Atualizar profissional
    if request.method == 'PUT':
        """Atualizar dados de um profissional"""
        try:
            data = request.get_json()

            update_data = {}
            if 'nome' in data:
                update_data['nome'] = data['nome']
            if 'cpf' in data:
                update_data['cpf'] = data['cpf']
            if 'especialidade' in data:
                update_data['especialidade'] = data['especialidade']
            if 'telefone' in data:
                update_data['telefone'] = data.get('telefone', '')
            if 'email' in data:
                update_data['email'] = data.get('email', '')
            if 'comissao_perc' in data:
                update_data['comissao_perc'] = float(data['comissao_perc'])
            if 'ativo' in data:
                update_data['ativo'] = bool(data['ativo'])

            if not update_data:
                return jsonify({'success': False, 'message': 'Nenhum dado para atualizar'}), 400

            result = db.profissionais.update_one(
                {'_id': ObjectId(id)},
                {'$set': update_data}
            )

            if result.modified_count > 0:
                logger.info(f"✅ Profissional {id} atualizado: {update_data}")
                return jsonify({'success': True, 'message': 'Profissional atualizado com sucesso'})

            return jsonify({'success': False, 'message': 'Nenhuma alteração realizada'}), 400

        except Exception as e:
            logger.error(f"❌ Erro ao atualizar profissional: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # GET - Visualizar profissional com estatísticas
    """Visualizar um profissional especifico com estatisticas completas"""
    if db is None:
        return jsonify({'success': False}), 500
    try:
        profissional = db.profissionais.find_one({'_id': ObjectId(id)})
        if not profissional:
            return jsonify({'success': False, 'message': 'Profissional nao encontrado'}), 404

        profissional_id_str = str(profissional['_id'])

        orcamentos_prof = list(db.orcamentos.find({
            'servicos.profissional_id': profissional_id_str
        }))

        total_comissao = 0.0
        total_comissao_assistente = 0.0
        servicos_realizados = 0
        orcamentos_aprovados = 0
        desempenho_mensal = {}
        multicomissao_detalhes = []

        # v7.1: Tratamento robusto de erros para evitar 500
        try:
            assistente_info = get_assistente_details(
                profissional.get('assistente_id'),
                profissional.get('assistente_tipo')
            )
        except Exception as e:
            logger.warning(f"⚠️ Erro ao buscar assistente: {e}")
            assistente_info = None

        try:
            comissao_prof_perc = float(profissional.get('comissao_perc', 0))
        except (ValueError, TypeError):
            comissao_prof_perc = 0.0

        try:
            comissao_assistente_padrao = float(profissional.get('comissao_assistente_perc', 0))
        except (ValueError, TypeError):
            comissao_assistente_padrao = 0.0

        for orc in orcamentos_prof:
            data_orc = orc.get('created_at', datetime.now())
            mes_key = data_orc.strftime('%Y-%m') if isinstance(data_orc, datetime) else 'desconhecido'
            if mes_key not in desempenho_mensal:
                desempenho_mensal[mes_key] = {
                    'mes': mes_key,
                    'comissao_profissional': 0.0,
                    'comissao_assistente': 0.0,
                    'servicos': 0,
                    'faturamento': 0.0
                }

            if orc.get('status') == 'Aprovado':
                orcamentos_aprovados += 1
                desempenho_mensal[mes_key]['faturamento'] += orc.get('total_final', 0)

            for servico in orc.get('servicos', []):
                if servico.get('profissional_id') != profissional_id_str:
                    continue

                quantidade = servico.get('qtd', 1) or 1
                valor_servico = servico.get('total', 0) or 0
                servicos_realizados += quantidade

                comissao_valor = valor_servico * (comissao_prof_perc / 100)
                total_comissao += comissao_valor
                desempenho_mensal[mes_key]['comissao_profissional'] += comissao_valor
                desempenho_mensal[mes_key]['servicos'] += quantidade

                assistente_item = None
                assistente_perc = servico.get('assistente_comissao_perc')
                assistente_servico = servico.get('assistente_servico') or servico.get('nome')

                # v7.1: Tratamento robusto para assistentes
                try:
                    if servico.get('assistente_id'):
                        assistente_item = get_assistente_details(
                            servico.get('assistente_id'),
                            servico.get('assistente_tipo')
                        )
                        if assistente_perc is None:
                            assistente_perc = servico.get('assistente_comissao_perc', comissao_assistente_padrao)
                    elif assistente_info:
                        assistente_item = assistente_info
                        if assistente_perc is None:
                            assistente_perc = comissao_assistente_padrao
                except Exception as e:
                    logger.warning(f"⚠️ Erro ao processar assistente do serviço: {e}")
                    assistente_item = None

                comissao_assistente_valor = 0.0
                try:
                    if assistente_item and assistente_perc:
                        assistente_perc = float(assistente_perc)
                        comissao_assistente_valor = comissao_valor * (assistente_perc / 100)
                        total_comissao_assistente += comissao_assistente_valor
                        desempenho_mensal[mes_key]['comissao_assistente'] += comissao_assistente_valor
                except (ValueError, TypeError, KeyError) as e:
                    logger.warning(f"⚠️ Erro ao calcular comissão assistente: {e}")

                multicomissao_detalhes.append({
                    'orcamento_id': str(orc['_id']),
                    'orcamento_numero': orc.get('numero'),
                    'cliente': orc.get('cliente_nome'),
                    'status': orc.get('status'),
                    'data': data_orc.isoformat() if isinstance(data_orc, datetime) else data_orc,
                    'servico': servico.get('nome'),
                    'valor_servico': valor_servico,
                    'comissao_profissional': comissao_valor,
                    'comissao_assistente': comissao_assistente_valor,
                    'assistente_nome': assistente_item.get('nome') if assistente_item else None,
                    'assistente_tipo': assistente_item.get('tipo_origem') if assistente_item else None,
                    'descricao': f"Profissional {profissional.get('nome')} - {servico.get('nome')}" + (
                        f" | Assistente {assistente_item.get('nome')} - {assistente_servico}" if assistente_item else ''
                    )
                })

        desempenho_ordenado = sorted(desempenho_mensal.items())
        grafico_labels = [item[0] for item in desempenho_ordenado]
        grafico_dados_prof = [round(item[1]['comissao_profissional'], 2) for item in desempenho_ordenado]
        grafico_dados_assist = [round(item[1]['comissao_assistente'], 2) for item in desempenho_ordenado]
        grafico_servicos = [item[1]['servicos'] for item in desempenho_ordenado]

        avaliacoes = []
        try:
            avaliacoes = list(db.profissionais_avaliacoes.find(
                {'profissional_id': profissional_id_str}
            ).sort('created_at', DESCENDING).limit(20))
        except Exception as avaliacao_error:
            logger.debug(f"Falha ao carregar avaliacoes do profissional {id}: {avaliacao_error}")

        if assistente_info:
            profissional['assistente'] = assistente_info

        profissional['estatisticas'] = {
            'total_comissao': round(total_comissao, 2),
            'total_comissao_assistente': round(total_comissao_assistente, 2),
            'servicos_realizados': servicos_realizados,
            'total_orcamentos': len(orcamentos_prof),
            'orcamentos_aprovados': orcamentos_aprovados,
            'comissao_media': round(total_comissao / servicos_realizados, 2) if servicos_realizados else 0
        }

        profissional['desempenho'] = {
            'labels': grafico_labels,
            'comissao_profissional': grafico_dados_prof,
            'comissao_assistente': grafico_dados_assist,
            'servicos': grafico_servicos
        }

        profissional['multicomissao'] = multicomissao_detalhes
        profissional['avaliacoes'] = convert_objectid(avaliacoes)

        return jsonify({'success': True, 'profissional': convert_objectid(profissional)})
    except Exception as e:
        logger.error(f"Erro ao buscar profissional: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/comissao/calcular', methods=['POST'])
@login_required
def calcular_comissao():
    db = get_db()
    """Calcular comissoes de profissional e assistente para um orcamento"""
    if db is None:
        return jsonify({'success': False}), 500

    data = request.json or {}
    try:
        orcamento_id = data.get('orcamento_id')
        if not orcamento_id:
            return jsonify({'success': False, 'message': 'ID do orcamento nao fornecido'}), 400

        orcamento = db.orcamentos.find_one({'_id': ObjectId(orcamento_id)})
        if not orcamento:
            return jsonify({'success': False, 'message': 'Orcamento nao encontrado'}), 404

        comissoes = []
        total_comissoes = 0.0

        for servico in orcamento.get('servicos', []):
            profissional_id = servico.get('profissional_id')
            if not profissional_id:
                continue

            profissional = db.profissionais.find_one({'_id': ObjectId(profissional_id)})
            if not profissional:
                continue

            valor_servico = servico.get('total', 0) or 0
            comissao_perc = float(profissional.get('comissao_perc', 0))
            comissao_valor = valor_servico * (comissao_perc / 100)

            comissao_info = {
                'profissional_id': str(profissional['_id']),
                'profissional_nome': profissional.get('nome', ''),
                'servico': servico.get('nome', ''),
                'valor_servico': valor_servico,
                'comissao_perc': comissao_perc,
                'comissao_valor': comissao_valor
            }

            assistente_info = None
            assistente_perc = servico.get('assistente_comissao_perc')
            assistente_servico = servico.get('assistente_servico') or servico.get('nome')

            if servico.get('assistente_id'):
                assistente_info = get_assistente_details(
                    servico.get('assistente_id'),
                    servico.get('assistente_tipo')
                )
                if assistente_perc is None:
                    assistente_perc = servico.get('assistente_comissao_perc', profissional.get('comissao_assistente_perc', 0))
            else:
                assistente_info = get_assistente_details(
                    profissional.get('assistente_id'),
                    profissional.get('assistente_tipo')
                )
                if assistente_perc is None:
                    assistente_perc = profissional.get('comissao_assistente_perc', 0)

            if assistente_info and assistente_perc:
                assistente_perc = float(assistente_perc)
                comissao_assistente_valor = comissao_valor * (assistente_perc / 100)
                comissao_info['assistente'] = {
                    'assistente_id': assistente_info.get('_id'),
                    'assistente_nome': assistente_info.get('nome', ''),
                    'assistente_tipo': assistente_info.get('tipo_origem'),
                    'comissao_perc': assistente_perc,
                    'comissao_valor': comissao_assistente_valor,
                    'servico': assistente_servico
                }
                total_comissoes += comissao_assistente_valor

            comissao_info['descricao'] = f"Profissional {profissional.get('nome')} - {servico.get('nome')}"
            if assistente_info:
                comissao_info['descricao'] += f" | Assistente {assistente_info.get('nome', '')} - {assistente_servico}"

            comissoes.append(comissao_info)
            total_comissoes += comissao_valor

        return jsonify({
            'success': True,
            'orcamento_numero': orcamento.get('numero'),
            'comissoes': comissoes,
            'total_comissoes': total_comissoes
        })

    except Exception as e:
        logger.error(f"Erro ao calcular comissoes: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/profissionais/<id>/avaliacoes', methods=['GET', 'POST'])
@login_required
def profissional_avaliacoes(id):
    db = get_db()
    """Registrar e listar avaliacoes de profissionais."""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        ObjectId(id)
    except Exception:
        return jsonify({'success': False, 'message': 'ID de profissional invalido'}), 400

    if request.method == 'GET':
        try:
            avaliacoes = list(db.profissionais_avaliacoes.find(
                {'profissional_id': id}
            ).sort('created_at', DESCENDING).limit(50))
            return jsonify({'success': True, 'avaliacoes': convert_objectid(avaliacoes)})
        except Exception as e:
            logger.error(f"Erro ao listar avaliacoes: {e}")
            return jsonify({'success': False, 'message': 'Erro ao listar avaliacoes'}), 500

    data = request.json or {}
    try:
        nota = float(data.get('nota', 0))
        nota = max(0, min(5, nota))
        avaliacao = {
            'profissional_id': id,
            'nota': nota,
            'comentario': data.get('comentario', '').strip(),
            'autor': data.get('autor', session.get('username', 'anonimo')),
            'created_at': datetime.now()
        }
        db.profissionais_avaliacoes.insert_one(avaliacao)
        # clear_cache('profissionais_list')  # Cache invalidation - opcional
        return jsonify({'success': True, 'avaliacao': convert_objectid(avaliacao)})
    except Exception as e:
        logger.error(f"Erro ao registrar avaliacao: {e}")
        return jsonify({'success': False, 'message': 'Erro ao registrar avaliacao'}), 500


@bp.route('/api/profissionais/<id>/upload-foto', methods=['POST'])
@login_required
def upload_foto_profissional(id):
    """Upload de foto de perfil para profissionais (Diretriz #12)"""
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Verificar se o profissional existe
        profissional = db.profissionais.find_one({'_id': ObjectId(id)})
        if not profissional:
            return jsonify({'success': False, 'message': 'Profissional não encontrado'}), 404

        # Verificar se há arquivo na requisição
        if 'foto' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['foto']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Arquivo sem nome'}), 400

        if file and allowed_file(file.filename):
            # Ler arquivo DIRETO EM MEMÓRIA (SEM salvar em /tmp)
            import base64
            file_content = file.read()
            foto_base64 = base64.b64encode(file_content).decode('utf-8')

            # Determinar tipo MIME
            filename = secure_filename(file.filename)
            ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'png'
            mime_types = {
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'webp': 'image/webp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            # Montar data URI
            foto_data_uri = f"data:{mime_type};base64,{foto_base64}"

            # Atualizar profissional com a foto
            db.profissionais.update_one(
                {'_id': ObjectId(id)},
                {'$set': {
                    'foto': foto_data_uri,
                    'foto_url': foto_data_uri,
                    'foto_atualizada_em': datetime.now()
                }}
            )

            # Limpar cache
            # clear_cache('profissionais_list')  # Cache invalidation - opcional

            logger.info(f"✅ Foto de perfil atualizada para profissional {id}")
            return jsonify({
                'success': True,
                'message': 'Foto atualizada com sucesso',
                'foto_url': foto_data_uri[:100] + '...'  # Retorna preview truncado
            })
        else:
            return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido'}), 400

    except Exception as e:
        logger.error(f"Erro ao fazer upload de foto: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/assistentes', methods=['GET', 'POST'])
@login_required
def assistentes():
    db = get_db()
    """Gerenciar assistentes (que podem não ser profissionais ativos)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    if request.method == 'GET':
        try:
            # Listar todos os assistentes
            assistentes_list = list(db.assistentes.find({}).sort('nome', ASCENDING))
            return jsonify({'success': True, 'assistentes': convert_objectid(assistentes_list)})
        except Exception as e:
            logger.error(f"Erro ao listar assistentes: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # POST - Cadastrar novo assistente
    try:
        data = request.get_json()

        if not data:
            return jsonify({'success': False, 'message': 'Dados inválidos'}), 400

        if 'nome' not in data or not data['nome']:
            return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400

        assistente_data = {
            'nome': data['nome'].strip(),
            'cpf': data.get('cpf', '').strip(),
            'email': data.get('email', '').strip(),
            'telefone': data.get('telefone', '').strip(),
            'comissao_perc': float(data.get('comissao_perc', 0)),
            'tipo': data.get('tipo', 'assistente'),
            'ativo': True,
            'created_at': datetime.now()
        }

        result = db.assistentes.insert_one(assistente_data)

        if result and result.inserted_id:
            inserted_id = str(result.inserted_id)
            logger.info(f"✅ Assistente cadastrado: {assistente_data['nome']} (ID: {inserted_id})")
            return jsonify({
                'success': True,
                'message': 'Assistente cadastrado com sucesso',
                'id': inserted_id,
                'assistente': {
                    **assistente_data,
                    '_id': inserted_id
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Erro ao inserir assistente no banco'}), 500

    except KeyError as e:
        logger.error(f"Campo obrigatório faltando: {e}")
        return jsonify({'success': False, 'message': f'Campo obrigatório faltando: {e}'}), 400
    except Exception as e:
        logger.error(f"Erro ao cadastrar assistente: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/assistentes/<id>', methods=['DELETE'])
@login_required
def delete_assistente(id):
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    try:
        result = db.assistentes.delete_one({'_id': ObjectId(id)})
        return jsonify({'success': result.deleted_count > 0})
    except:
        return jsonify({'success': False}), 500

# --- INÍCIO DA SEÇÃO MODIFICADA ---
class GradientHeader(Flowable):
    """Um cabeçalho com fundo em gradiente."""
    def __init__(self, width, text):
        Flowable.__init__(self)
        self.width = width
        self.text = text
        self.height = 1.5*cm

    def draw(self):
        c = self.canv
        c.saveState()
        start_color = HexColor('#7C3AED')
        end_color = HexColor('#EC4899')
        steps = 100
        for i in range(steps):
            ratio = i / float(steps)
            r = start_color.red * (1 - ratio) + end_color.red * ratio
            g = start_color.green * (1 - ratio) + end_color.green * ratio
            b = start_color.blue * (1 - ratio) + end_color.blue * ratio
            c.setFillColorRGB(r, g, b)
            c.rect((self.width / steps) * i, 0, self.width / steps, self.height, stroke=0, fill=1)
        c.setFont('Helvetica-Bold', 36)
        c.setFillColor(white)
        c.drawCentredString(self.width / 2, self.height / 2 - (0.5*cm), self.text)
        c.restoreState()

class HRFlowable(Flowable):
    """Linha horizontal com cor customizada."""
    def __init__(self, width, thickness=1, color=black):
        Flowable.__init__(self)
        self.width = width
        self.thickness = thickness
        self.color = color

    def draw(self):
        db = get_db()
        self.canv.saveState()
        self.canv.setStrokeColor(self.color)
        self.canv.setLineWidth(self.thickness)
        self.canv.line(0, 0, self.width, 0)
        self.canv.restoreState()

def format_date_pt_br(dt):
    db = get_db()
    """Formata a data para Português-BR manualmente para evitar problemas de locale."""
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    return f"{dt.day} de {meses[dt.month - 1]} de {dt.year}"

@bp.route('/api/orcamento/<id>/pdf')
@login_required
def gerar_pdf_orcamento_singular(id):
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    try:
        orcamento = db.orcamentos.find_one({'_id': ObjectId(id)})
        if not orcamento:
            return jsonify({'success': False, 'message': 'Orçamento não encontrado'}), 404

        buffer = io.BytesIO()
        doc_width, doc_height = A4
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=1.5*cm, bottomMargin=2.5*cm)

        COLOR_PRIMARY = HexColor('#7C3AED')
        COLOR_SECONDARY = HexColor('#6B7280')
        COLOR_LIGHT_BG = HexColor('#F9FAFB')
        COLOR_WHITE = HexColor('#FFFFFF')
        
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='MainTitle', fontName='Helvetica-Bold', fontSize=20, textColor=COLOR_PRIMARY, spaceAfter=18))
        styles.add(ParagraphStyle(name='SubTitle', fontName='Helvetica', fontSize=10, textColor=COLOR_SECONDARY, spaceAfter=24))
        styles.add(ParagraphStyle(name='SectionTitle', fontName='Helvetica-Bold', fontSize=10, textColor=white))
        styles.add(ParagraphStyle(name='Body', fontName='Helvetica', fontSize=9, leading=14))
        styles.add(ParagraphStyle(name='BodyRight', fontName='Helvetica', fontSize=9, leading=14, alignment=TA_RIGHT))
        styles.add(ParagraphStyle(name='Clause', fontName='Helvetica', fontSize=8, leading=14, alignment=TA_JUSTIFY, leftIndent=12))
        styles.add(ParagraphStyle(name='Signature', fontName='Helvetica', fontSize=9, alignment=TA_CENTER))

        story = []
        
        story.append(GradientHeader(doc_width - 4*cm, "BIOMA"))
        story.append(Spacer(1, 1.5*cm))
        story.append(Paragraph("Contrato de Prestação de Serviços", styles['MainTitle']))
        story.append(Paragraph(
            "Pelo presente instrumento particular, as 'Partes' resolvem celebrar o presente 'Contrato', de acordo com as cláusulas e condições a seguir.", 
            styles['SubTitle']
        ))

        data_contrato = orcamento.get('created_at', datetime.now())
        info_data = [
            [Paragraph('<b>NÚMERO DO CONTRATO</b>', styles['Body']), Paragraph(f"#{orcamento.get('numero', 'N/A')}", styles['Body'])],
            [Paragraph('<b>DATA DE EMISSÃO</b>', styles['Body']), Paragraph(format_date_pt_br(data_contrato), styles['Body'])]
        ]
        story.append(Table(info_data, colWidths=[5*cm, '*'], style=[('VALIGN', (0,0), (-1,-1), 'TOP')]))
        story.append(Spacer(1, 1*cm))

        story.append(HRFlowable(doc_width - 4*cm, color=HexColor('#E5E7EB'), thickness=1))
        story.append(Spacer(1, 1*cm))

        # Fix: Garantir que nenhum valor seja None
        contratante_details = f"""
            <b>Nome:</b> {orcamento.get('cliente_nome') or 'N/A'}<br/>
            <b>CPF:</b> {orcamento.get('cliente_cpf') or 'N/A'}<br/>
            <b>Telefone:</b> {orcamento.get('cliente_telefone') or 'N/A'}<br/>
            <b>E-mail:</b> {orcamento.get('cliente_email') or 'N/A'}
        """
        contratada_details = f"""
            <b>Razão Social:</b> BIOMA UBERABA<br/>
            <b>CNPJ:</b> 49.470.937/0001-10<br/>
            <b>Endereço:</b> Av. Santos Dumont 3110, Santa Maria, Uberaba/MG<br/>
            <b>Contato:</b> (34) 99235-5890
        """
        partes_data = [
            [Paragraph('<b>CONTRATANTE</b>', styles['Body']), Paragraph('<b>CONTRATADA</b>', styles['Body'])],
            [Paragraph(contratante_details, styles['Body']), Paragraph(contratada_details, styles['Body'])]
        ]
        partes_table = Table(partes_data, colWidths=['*', '*'], hAlign='LEFT')
        partes_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'), ('LINEBELOW', (0,0), (-1,0), 1, COLOR_PRIMARY),
            ('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 12),
            ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ]))
        story.append(partes_table)
        story.append(Spacer(1, 1.5*cm))

        story.append(Paragraph("SERVIÇOS E PRODUTOS CONTRATADOS", styles['MainTitle']))
        table_header = [Paragraph(c, styles['SectionTitle']) for c in ['Item', 'Descrição', 'Qtd', 'Vl. Unit.', 'Total']]
        items_data = [table_header]
        all_items = orcamento.get('servicos', []) + orcamento.get('produtos', [])
        
        table_style_commands = [
            ('BACKGROUND', (0,0), (-1,0), COLOR_PRIMARY), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 10), ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('GRID', (0,0), (-1,-1), 1, COLOR_PRIMARY),
            ('ALIGN', (2,1), (-1,-1), 'RIGHT'), # Alinha colunas numéricas à direita
            ('LEFTPADDING', (0,0), (-1,-1), 8), ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]

        for i, item in enumerate(all_items):
            desc = f"{item.get('nome', '')} {item.get('tamanho', '')}".strip() if 'servico' in item.get('id', '') else f"{item.get('nome', '')} {item.get('marca', '')}".strip()
            items_data.append([
                Paragraph(str(i+1), styles['Body']), Paragraph(desc, styles['Body']),
                Paragraph(str(item.get('qtd', 1)), styles['BodyRight']), 
                Paragraph(f"R$ {item.get('preco_unit', 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), styles['BodyRight']),
                Paragraph(f"R$ {item.get('total', 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), styles['BodyRight']),
            ])
            # Adiciona estilo de zebra (cores alternadas)
            if i % 2 == 0:
                table_style_commands.append(('BACKGROUND', (0, i + 1), (-1, i + 1), COLOR_LIGHT_BG))
            else:
                table_style_commands.append(('BACKGROUND', (0, i + 1), (-1, i + 1), COLOR_WHITE))

        items_table = Table(items_data, colWidths=[1.5*cm, '*', 1.5*cm, 3*cm, 3*cm], repeatRows=1)
        items_table.setStyle(TableStyle(table_style_commands))
        story.append(items_table)
        story.append(Spacer(1, 1*cm))

        subtotal = orcamento.get('subtotal', 0)
        desconto = orcamento.get('desconto_valor', 0)
        total = orcamento.get('total_final', 0)
        pag_tipo = orcamento.get('pagamento', {}).get('tipo', 'Não especificado')
        
        valores_data = [
            [Paragraph('Subtotal:', styles['Body']), Paragraph(f'R$ {subtotal:,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."), styles['BodyRight'])],
            [Paragraph('Desconto Global:', styles['Body']), Paragraph(f'R$ {desconto:,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."), styles['BodyRight'])],
            ['', HRFlowable(8*cm, color=COLOR_PRIMARY, thickness=1.5)],
            [Paragraph('<b>Valor Total a Pagar:</b>', styles['Body']), Paragraph(f'<b>R$ {total:,.2f}</b>'.replace(",", "X").replace(".", ",").replace("X", "."), styles['BodyRight'])],
            [Paragraph('Forma de Pagamento:', styles['Body']), Paragraph(pag_tipo, styles['BodyRight'])],
        ]
        
        # Tabela para alinhar o bloco de valores à direita
        container_valores = Table([[Table(valores_data, colWidths=[4*cm, 4*cm])]], colWidths=[doc_width-4*cm])
        container_valores.setStyle(TableStyle([('ALIGN', (0,0), (0,0), 'RIGHT')]))
        story.append(container_valores)
        story.append(Spacer(1, 1.5*cm))

        story.append(Paragraph("DISPOSIÇÕES GERAIS E CLÁUSULAS", styles['MainTitle']))
        clausulas = [
            "O Contrato tem por objeto a prestação de serviços acima descritos, pela Contratada à Contratante, mediante agendamento prévio. A Contratada utilizará produtos com ingredientes naturais para a saúde dos cabelos, de alta qualidade, que serão manipulados dentro das normas de higiene e limpeza exigidas pela Vigilância Sanitária.",
            "A Contratante declara e está ciente que (i) os serviços têm caráter pessoal e são intransferíveis; (ii) só poderá alterar os Serviços contratados com a anuência da Contratada e desde que a utilização seja no prazo originalmente contratado; (iii) não tem nenhum impedimento médico e/ou alergias que impeçam de realizar os serviços contratados; (iv) escolheu os tratamentos de acordo com o seu tipo de cabelo; (v) concorda em realizar os tratamentos com a frequência indicada pela Contratada; e (vi) o resultado pretendido depende do respeito à frequência indicada pela Contratada.",
            "Os serviços deverão ser utilizados em conformidade com o prazo de 18 (dezoito) meses e a Contratante está ciente de que não haverá prorrogação do prazo previsto para a utilização dos serviços, ou seja, ao final de 18 (dezoito) meses, o Contrato será extinto e a Contratante não terá direito ao reembolso de tratamentos não realizados no prazo contratual.",
            "A Contratante poderá desistir dos serviços no prazo de até 90 (noventa) dias a contar da assinatura deste Contrato e, neste caso, está de acordo com a restituição do valor equivalente a 80% (oitenta por cento) dos tratamentos não realizados, no prazo de até 5 (cinco) dias úteis da desistência. Eventuais descontos ou promoções nos valores dos serviços e/ou tratamentos não serão reembolsáveis.",
            "No caso de devolução de valor pago por cartão de crédito, o cancelamento será efetuado junto à administradora do seu cartão e o estorno poderá ocorrer em até 2 (duas) faturas posteriores de acordo com procedimentos internos da operadora do cartão de crédito, ou outro prazo definido pela administradora do cartão de crédito, ou, a exclusivo arbítrio da Contratada, mediante transferência direta do valor equivalente ao reembolso.",
            "Na hipótese de responsabilidade civil da Contratada, independentemente da natureza do dano, fica desde já limitada a responsabilidade da Contratada ao valor máximo equivalente a 2 (duas) sessões de tratamento dos serviços.",
            "No caso de alergias decorrentes dos produtos utilizados pela Contratada, a Contratante poderá optar pela suspensão dos serviços com a retomada após o reestabelecimento de sua saúde, ou pela concessão de crédito do valor remanescente em outros serviços junto à Contratada. A Contratada não é responsável por qualquer perda, independentemente do valor, incluindo danos diretos, indiretos, à imagem, lucros cessantes e/ou morais que se tornem exigíveis em decorrência de eventual alergia.",
            "As Partes se comprometem a tratar apenas os dados pessoais estritamente necessários para atingir as finalidades específicas do objeto do Contrato, em cumprimento ao disposto na Lei nº 13.709/2018 (\"LGPD\") e na regulamentação aplicável.",
            "Fica eleito o Foro da Comarca de UBERABA, Estado de MINAS GERAIS, como o competente para dirimir as dúvidas e controvérsias decorrentes do presente Contrato, com renúncia a qualquer outro, por mais privilegiado que seja.",
            "Este Contrato poderá ser assinado e entregue eletronicamente e terá a mesma validade e efeitos de um documento original com assinaturas físicas."
        ]
        for i, clausula in enumerate(clausulas):
            story.append(Paragraph(f"<b>{i+1}.</b> {clausula}", styles['Clause']))
            story.append(Spacer(1, 0.4*cm))
        
        story.append(PageBreak())
        story.append(Paragraph("ASSINATURAS", styles['MainTitle']))
        story.append(Spacer(1, 4*cm))

        # Fix: Garantir que cliente_nome nunca seja None
        cliente_nome = orcamento.get('cliente_nome') or 'N/A'
        assinatura_contratante = Paragraph(f"________________________________________<br/><b>CONTRATANTE</b><br/>{cliente_nome}", styles['Signature'])
        assinatura_contratada = Paragraph("________________________________________<br/><b>CONTRATADA</b><br/>BIOMA UBERABA", styles['Signature'])
        
        assinaturas_table = Table([[assinatura_contratante, assinatura_contratada]], colWidths=['*', '*'])
        story.append(assinaturas_table)
        
        def on_each_page(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(COLOR_SECONDARY)
            page_num = canvas.getPageNumber()
            canvas.drawCentredString(doc_width/2, 1.5*cm, f"Página {page_num} | Contrato BIOMA Uberaba")
            canvas.restoreState()

        doc.build(story, onFirstPage=on_each_page, onLaterPages=on_each_page)
        
        buffer.seek(0)
        # Fix: Garantir que numero nunca seja None no nome do arquivo
        numero_orcamento = orcamento.get("numero") or str(orcamento.get('_id', 'sem_numero'))[-6:]
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'contrato_bioma_{numero_orcamento}.pdf')
        
    except Exception as e:
        logger.error(f"❌ PDF error: {e}")
        return jsonify({'success': False, 'message': f'Erro interno ao gerar PDF: {e}'}), 500

# --- FIM DA SEÇÃO MODIFICADA ---


@bp.route('/api/contratos')
@login_required
def contratos():
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    try:
        contratos_list = list(db.orcamentos.find({'status': 'Aprovado'}).sort('created_at', DESCENDING))
        return jsonify({'success': True, 'contratos': convert_objectid(contratos_list)})
    except:
        return jsonify({'success': False}), 500

@bp.route('/api/agendamentos', methods=['GET', 'POST'])
@login_required
def agendamentos():
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    if request.method == 'GET':
        try:
            agora = datetime.now()
            agends = list(db.agendamentos.find({'data': {'$gte': agora}}).sort('data', ASCENDING).limit(10))
            return jsonify({'success': True, 'agendamentos': convert_objectid(agends)})
        except:
            return jsonify({'success': False}), 500
    # POST - Criar novo agendamento com VALIDAÇÃO FLEXÍVEL
    data = request.json

    # Validação de campos obrigatórios mínimos (data e horario)
    if not data.get('data') or not data.get('horario'):
        return jsonify({
            'success': False,
            'message': 'Data e horário são obrigatórios'
        }), 400

    # Validar que pelo menos nome do cliente foi fornecido
    if not data.get('cliente_nome') and not data.get('cliente_id'):
        return jsonify({
            'success': False,
            'message': 'Nome ou ID do cliente é obrigatório'
        }), 400

    try:
        # Validar e converter data
        data_agendamento = datetime.fromisoformat(data['data'].replace('Z', '+00:00'))

        # Validar que a data não está no passado
        if data_agendamento < datetime.now():
            return jsonify({
                'success': False,
                'message': 'Não é possível agendar para data/hora no passado'
            }), 400

        # Validar formato do horário (HH:MM)
        horario = data['horario']
        if not isinstance(horario, str) or not len(horario.split(':')) == 2:
            return jsonify({
                'success': False,
                'message': 'Formato de horário inválido. Use HH:MM'
            }), 400

        # Verificar se já existe agendamento conflitante
        conflito = db.agendamentos.find_one({
            'profissional_id': data['profissional_id'],
            'data': data_agendamento,
            'horario': horario,
            'status': {'$ne': 'cancelado'}
        })

        if conflito:
            return jsonify({
                'success': False,
                'message': 'Já existe um agendamento para este profissional neste horário'
            }), 409  # HTTP 409 Conflict

        # Inserir agendamento
        agend_id = db.agendamentos.insert_one({
            'cliente_id': data.get('cliente_id', 'temp'),
            'cliente_nome': data.get('cliente_nome', 'N/A'),
            'cliente_telefone': data.get('cliente_telefone'),
            'profissional_id': data.get('profissional_id', 'temp'),
            'profissional_nome': data.get('profissional_nome', 'N/A'),
            'servico_id': data.get('servico_id', 'temp'),
            'servico_nome': data.get('servico_nome', 'N/A'),
            'tamanho': data.get('tamanho', ''),
            'data': data_agendamento,
            'horario': horario,
            'status': data.get('status', 'confirmado'),
            'observacoes': data.get('observacoes', ''),
            'created_at': datetime.now(),
            'created_by': session.get('user_email', 'sistema')
        }).inserted_id

        logger.info(f"✅ Agendamento criado: {agend_id} para {data.get('cliente_nome')} em {data_agendamento}")
        return jsonify({'success': True, 'id': str(agend_id)})

    except ValueError as e:
        return jsonify({
            'success': False,
            'message': f'Erro de validação: {str(e)}'
        }), 400
    except Exception as e:
        logger.error(f"Erro ao criar agendamento: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno ao criar agendamento'
        }), 500

@bp.route('/api/agendamentos/horarios-disponiveis', methods=['GET'])
@login_required
def horarios_disponiveis():
    db = get_db()
    """Verificar horários disponíveis para uma data específica"""
    if db is None:
        return jsonify({'success': False}), 500
    
    try:
        data_str = request.args.get('data')
        profissional_id = request.args.get('profissional_id')
        
        if not data_str:
            return jsonify({'success': False, 'message': 'Data não fornecida'}), 400
        
        # Converter a data
        data = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
        data_inicio = data.replace(hour=0, minute=0, second=0)
        data_fim = data.replace(hour=23, minute=59, second=59)
        
        # Buscar agendamentos do dia
        query = {'data': {'$gte': data_inicio, '$lte': data_fim}}
        if profissional_id:
            query['profissional_id'] = profissional_id
        
        agendamentos = list(db.agendamentos.find(query))
        horarios_ocupados = [a.get('horario') for a in agendamentos if a.get('horario')]
        
        # Gerar horários disponíveis (08:00 - 18:00)
        horarios_todos = []
        for hora in range(8, 18):
            for minuto in [0, 30]:
                horarios_todos.append(f"{hora:02d}:{minuto:02d}")
        
        horarios_disponiveis = [h for h in horarios_todos if h not in horarios_ocupados]
        
        return jsonify({
            'success': True,
            'horarios_disponiveis': horarios_disponiveis,
            'horarios_ocupados': horarios_ocupados,
            'total_slots': len(horarios_todos),
            'slots_disponiveis': len(horarios_disponiveis)
        })
    except Exception as e:
        logger.error(f"Erro ao buscar horários disponíveis: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/agendamentos/mapa-calor', methods=['GET'])
@login_required
def mapa_calor_agendamentos():
    db = get_db()
    """Gerar mapa de calor de agendamentos e orçamentos"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        # Parâmetro de dias (padrão: 30)
        dias = request.args.get('dias', 30, type=int)
        data_inicio = datetime.now() - timedelta(days=dias)
        data_fim = datetime.now()
        
        # Buscar agendamentos
        agendamentos = list(db.agendamentos.find({
            'created_at': {'$gte': data_inicio, '$lte': data_fim}
        }))
        
        # Buscar orçamentos
        orcamentos = list(db.orcamentos.find({
            'created_at': {'$gte': data_inicio, '$lte': data_fim}
        }))
        
        # Agrupar por dia
        mapa_calor = {}
        
        for agend in agendamentos:
            data_agend = agend.get('created_at', datetime.now())
            dia = data_agend.strftime('%Y-%m-%d')
            if dia not in mapa_calor:
                mapa_calor[dia] = {'agendamentos': 0, 'orcamentos': 0, 'total': 0}
            mapa_calor[dia]['agendamentos'] += 1
            mapa_calor[dia]['total'] += 1
        
        for orc in orcamentos:
            data_orc = orc.get('created_at', datetime.now())
            dia = data_orc.strftime('%Y-%m-%d')
            if dia not in mapa_calor:
                mapa_calor[dia] = {'agendamentos': 0, 'orcamentos': 0, 'total': 0}
            mapa_calor[dia]['orcamentos'] += 1
            mapa_calor[dia]['total'] += 1
        
        # Converter para lista ordenada
        mapa_lista = [
            {
                'data': dia,
                'agendamentos': dados['agendamentos'],
                'orcamentos': dados['orcamentos'],
                'total': dados['total']
            }
            for dia, dados in sorted(mapa_calor.items())
        ]
        
        return jsonify({'success': True, 'mapa_calor': mapa_lista})
    except Exception as e:
        logger.error(f"Erro ao gerar mapa de calor: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/agendamentos/<id>', methods=['DELETE'])
@login_required
@permission_required('Admin', 'Gestão')
def delete_agendamento(id):
    db = get_db()
    """Deletar um agendamento (Apenas Admin e Gestão - Roadmap RBAC)"""
    if db is None:
        return jsonify({'success': False}), 500
    try:
        result = db.agendamentos.delete_one({'_id': ObjectId(id)})
        if result.deleted_count > 0:
            logger.info(f"✅ Agendamento {id} deletado por {session.get('user_email')}")
        return jsonify({'success': result.deleted_count > 0})
    except Exception as e:
        logger.error(f"Erro ao deletar agendamento {id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== NOVOS ENDPOINTS AGENDAMENTOS - SUB-TABS ==========
@bp.route('/api/agendamentos/hoje', methods=['GET'])
@login_required
def agendamentos_hoje():
    db = get_db()
    """Buscar agendamentos de hoje com estatísticas"""
    if db is None:
        return jsonify({'success': False, 'message': 'Banco de dados indisponível'}), 500
    
    try:
        hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        amanha = hoje + timedelta(days=1)
        
        # Buscar agendamentos de hoje
        agendamentos = list(db.agendamentos.find({
            'data': {'$gte': hoje, '$lt': amanha}
        }).sort('horario', ASCENDING))
        
        resultado = []
        confirmados = 0
        pendentes = 0
        concluidos = 0
        cancelados = 0
        
        for a in agendamentos:
            # Buscar dados relacionados
            cliente = db.clientes.find_one({'_id': ObjectId(a['cliente_id'])}) if a.get('cliente_id') else None
            servico = db.servicos.find_one({'_id': ObjectId(a['servico_id'])}) if a.get('servico_id') else None
            prof = db.profissionais.find_one({'_id': ObjectId(a['profissional_id'])}) if a.get('profissional_id') else None
            
            status = a.get('status', 'Pendente')
            
            # Contar por status
            if status == 'Confirmado':
                confirmados += 1
            elif status == 'Pendente':
                pendentes += 1
            elif status == 'Concluído':
                concluidos += 1
            elif status == 'Cancelado':
                cancelados += 1
            
            resultado.append({
                '_id': str(a['_id']),
                'horario': a.get('horario', ''),
                'cliente_nome': cliente.get('nome') if cliente else 'Desconhecido',
                'servico': servico.get('nome') if servico else 'Desconhecido',
                'profissional': prof.get('nome') if prof else 'Desconhecido',
                'status': status,
                'observacoes': a.get('observacoes', '')
            })
        
        logger.info(f"Agendamentos hoje: {len(resultado)} encontrados")
        
        return jsonify({
            'success': True,
            'agendamentos': resultado,
            'total': len(resultado),
            'confirmados': confirmados,
            'pendentes': pendentes,
            'concluidos': concluidos,
            'cancelados': cancelados
        })
        
    except Exception as e:
        logger.error(f"Erro em agendamentos_hoje: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/agendamentos/semana', methods=['GET'])
@login_required
def agendamentos_semana():
    db = get_db()
    """Buscar agendamentos da semana atual"""
    if db is None:
        return jsonify({'success': False, 'message': 'Banco de dados indisponível'}), 500
    
    try:
        hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        inicio_semana = hoje - timedelta(days=hoje.weekday())  # Segunda-feira
        fim_semana = inicio_semana + timedelta(days=7)  # Próxima segunda
        
        agendamentos = list(db.agendamentos.find({
            'data': {'$gte': inicio_semana, '$lt': fim_semana}
        }).sort('data', ASCENDING))
        
        resultado = []
        for a in agendamentos:
            cliente = db.clientes.find_one({'_id': ObjectId(a['cliente_id'])}) if a.get('cliente_id') else None
            servico = db.servicos.find_one({'_id': ObjectId(a['servico_id'])}) if a.get('servico_id') else None
            prof = db.profissionais.find_one({'_id': ObjectId(a['profissional_id'])}) if a.get('profissional_id') else None
            
            resultado.append({
                '_id': str(a['_id']),
                'data': a.get('data').isoformat() if a.get('data') else '',
                'horario': a.get('horario', ''),
                'cliente_nome': cliente.get('nome') if cliente else 'Desconhecido',
                'servico': servico.get('nome') if servico else 'Desconhecido',
                'profissional': prof.get('nome') if prof else 'Desconhecido',
                'status': a.get('status', 'Pendente')
            })
        
        logger.info(f"Agendamentos semana: {len(resultado)} encontrados")
        
        return jsonify({
            'success': True,
            'agendamentos': resultado,
            'periodo': 'semana',
            'inicio': inicio_semana.isoformat(),
            'fim': fim_semana.isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erro em agendamentos_semana: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/agendamentos/mes', methods=['GET'])
@login_required
def agendamentos_mes():
    db = get_db()
    """Buscar agendamentos do mês atual"""
    if db is None:
        return jsonify({'success': False, 'message': 'Banco de dados indisponível'}), 500
    
    try:
        hoje = datetime.now()
        inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Calcular último dia do mês
        if hoje.month == 12:
            fim_mes = hoje.replace(year=hoje.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            fim_mes = hoje.replace(month=hoje.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
        
        agendamentos = list(db.agendamentos.find({
            'data': {'$gte': inicio_mes, '$lt': fim_mes}
        }).sort('data', ASCENDING))
        
        resultado = []
        for a in agendamentos:
            cliente = db.clientes.find_one({'_id': ObjectId(a['cliente_id'])}) if a.get('cliente_id') else None
            servico = db.servicos.find_one({'_id': ObjectId(a['servico_id'])}) if a.get('servico_id') else None
            prof = db.profissionais.find_one({'_id': ObjectId(a['profissional_id'])}) if a.get('profissional_id') else None
            
            resultado.append({
                '_id': str(a['_id']),
                'data': a.get('data').isoformat() if a.get('data') else '',
                'horario': a.get('horario', ''),
                'cliente_nome': cliente.get('nome') if cliente else 'Desconhecido',
                'servico': servico.get('nome') if servico else 'Desconhecido',
                'profissional': prof.get('nome') if prof else 'Desconhecido',
                'status': a.get('status', 'Pendente')
            })
        
        logger.info(f"Agendamentos mês: {len(resultado)} encontrados")
        
        return jsonify({
            'success': True,
            'agendamentos': resultado,
            'periodo': 'mes',
            'mes': hoje.strftime('%B %Y'),
            'inicio': inicio_mes.isoformat(),
            'fim': fim_mes.isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erro em agendamentos_mes: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/fila', methods=['GET', 'POST'])
@login_required
def fila():
    db = get_db()
    if db is None:
        logger.error("❌ Erro: db is None em /api/fila")
        return jsonify({'success': False, 'message': 'Database not available'}), 500

    if request.method == 'GET':
        try:
            # Ensure collection exists and has proper structure
            if 'fila_atendimento' not in db.list_collection_names():
                logger.info("📋 Criando collection fila_atendimento")
                db.create_collection('fila_atendimento')

            fila_list = list(db.fila_atendimento.find(
                {'status': {'$in': ['aguardando', 'atendendo']}}
            ).sort('created_at', ASCENDING))

            logger.info(f"✅ Fila carregada: {len(fila_list)} pessoas aguardando")
            return jsonify({'success': True, 'fila': convert_objectid(fila_list)})
        except Exception as e:
            logger.error(f"❌ Erro ao buscar fila de atendimento: {e}")
            import traceback
            logger.error(f"   Traceback: {traceback.format_exc()}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # POST
    data = request.json
    try:
        total = db.fila_atendimento.count_documents({'status': {'$in': ['aguardando', 'atendendo']}})
        db.fila_atendimento.insert_one({
            'cliente_nome': data['cliente_nome'],
            'cliente_telefone': data['cliente_telefone'],
            'servico': data.get('servico', ''),
            'profissional': data.get('profissional', ''),
            'posicao': total + 1,
            'status': 'aguardando',
            'created_at': datetime.now()
        })
        logger.info(f"✅ Cliente adicionado à fila: {data['cliente_nome']}")
        return jsonify({'success': True, 'posicao': total + 1})
    except Exception as e:
        logger.error(f"❌ Erro ao adicionar à fila: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/fila/<id>', methods=['DELETE'])
@login_required
def delete_fila(id):
    db = get_db()
    if db is None:
        logger.error("❌ Erro: db is None em DELETE /api/fila/<id>")
        return jsonify({'success': False, 'message': 'Database not available'}), 500
    try:
        result = db.fila_atendimento.delete_one({'_id': ObjectId(id)})
        if result.deleted_count > 0:
            logger.info(f"✅ Removido da fila: {id}")
        return jsonify({'success': result.deleted_count > 0})
    except Exception as e:
        logger.error(f"❌ Erro ao remover da fila {id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== NOTIFICAÇÕES INTELIGENTES PARA FILA (Diretriz #10) ====================

@bp.route('/api/fila/<id>/notificar', methods=['POST'])
@login_required
def notificar_cliente_fila(id):
    db = get_db()
    """Enviar notificação SMS/Email para cliente na fila"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        # Buscar cliente na fila
        cliente_fila = db.fila_atendimento.find_one({'_id': ObjectId(id)})
        if not cliente_fila:
            return jsonify({'success': False, 'message': 'Cliente não encontrado na fila'}), 404

        data = request.json
        tipo_notificacao = data.get('tipo', 'whatsapp')  # whatsapp, sms, email
        mensagem_custom = data.get('mensagem', '')

        # Gerar mensagem padrão se não fornecida
        if not mensagem_custom:
            posicao = cliente_fila.get('posicao', 0)
            cliente_nome = cliente_fila.get('cliente_nome', 'Cliente')
            servico = cliente_fila.get('servico', 'atendimento')

            if tipo_notificacao == 'whatsapp':
                mensagem = f"""🌿 *BIOMA - Sua vez está próxima!*

Olá, {cliente_nome}! 👋

📍 *Posição na fila:* {posicao}
✨ *Serviço:* {servico}

Aguarde mais alguns instantes. Em breve você será atendido!

Obrigado pela preferência! 🌿"""
            elif tipo_notificacao == 'email':
                mensagem = f"""Olá {cliente_nome},

Você está na posição {posicao} da fila de atendimento.
Serviço: {servico}

Aguarde mais alguns instantes, você será atendido em breve!

Atenciosamente,
Equipe BIOMA"""
            else:  # sms
                mensagem = f"BIOMA: Ola {cliente_nome}! Voce esta na posicao {posicao} da fila para {servico}. Aguarde!"
        else:
            mensagem = mensagem_custom

        telefone = cliente_fila.get('cliente_telefone', '').replace('(', '').replace(')', '').replace('-', '').replace(' ', '')

        # Registrar notificação no banco
        notificacao = {
            'fila_id': ObjectId(id),
            'cliente_nome': cliente_fila.get('cliente_nome'),
            'cliente_telefone': telefone,
            'tipo': tipo_notificacao,
            'mensagem': mensagem,
            'status': 'preparada',  # Em produção: 'enviada', 'erro'
            'created_at': datetime.now(),
            'sent_by': session.get('username', 'sistema')
        }

        # Em produção, aqui você integraria com:
        # - Twilio para SMS
        # - SendGrid para Email
        # - WhatsApp Business API

        # Por enquanto, apenas salvamos a notificação preparada
        result = db.notificacoes.insert_one(notificacao)

        # Se for WhatsApp, gerar URL
        whatsapp_url = None
        if tipo_notificacao == 'whatsapp':
            mensagem_encoded = urllib.parse.quote(mensagem)
            whatsapp_url = f"https://wa.me/55{telefone}?text={mensagem_encoded}"

        logger.info(f"✅ Notificação {tipo_notificacao} preparada para {cliente_fila.get('cliente_nome')}")

        return jsonify({
            'success': True,
            'message': f'Notificação {tipo_notificacao} preparada com sucesso',
            'notificacao_id': str(result.inserted_id),
            'whatsapp_url': whatsapp_url,
            'preview': mensagem[:100] + '...' if len(mensagem) > 100 else mensagem
        })

    except Exception as e:
        logger.error(f"Erro ao preparar notificação: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/fila/notificar-todos', methods=['POST'])
@login_required
def notificar_todos_fila():
    db = get_db()
    """Notificar todos os clientes na fila sobre sua posição"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        data = request.json
        tipo_notificacao = data.get('tipo', 'whatsapp')

        # Buscar todos na fila aguardando
        clientes_fila = list(db.fila_atendimento.find({'status': 'aguardando'}).sort('posicao', ASCENDING))

        notificacoes_criadas = []
        for cliente in clientes_fila:
            posicao = cliente.get('posicao', 0)
            cliente_nome = cliente.get('cliente_nome', 'Cliente')
            servico = cliente.get('servico', 'atendimento')
            telefone = cliente.get('cliente_telefone', '').replace('(', '').replace(')', '').replace('-', '').replace(' ', '')

            # Mensagem personalizada pela posição
            if posicao == 1:
                mensagem = f"🌿 BIOMA: {cliente_nome}, é sua vez! Dirija-se ao atendimento para {servico}. Obrigado!"
            elif posicao <= 3:
                mensagem = f"🌿 BIOMA: {cliente_nome}, você está próximo! Posição {posicao} na fila. Prepare-se!"
            else:
                mensagem = f"🌿 BIOMA: {cliente_nome}, você está na posição {posicao}. Aguarde, em breve será sua vez!"

            notificacao = {
                'fila_id': cliente['_id'],
                'cliente_nome': cliente_nome,
                'cliente_telefone': telefone,
                'tipo': tipo_notificacao,
                'mensagem': mensagem,
                'posicao': posicao,
                'status': 'preparada',
                'created_at': datetime.now(),
                'sent_by': session.get('username', 'sistema')
            }

            result = db.notificacoes.insert_one(notificacao)
            notificacoes_criadas.append(str(result.inserted_id))

        logger.info(f"✅ {len(notificacoes_criadas)} notificações preparadas para a fila")

        return jsonify({
            'success': True,
            'message': f'{len(notificacoes_criadas)} notificações preparadas',
            'total': len(notificacoes_criadas)
        })

    except Exception as e:
        logger.error(f"Erro ao notificar fila: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/notificacoes', methods=['GET'])
@login_required
def listar_notificacoes():
    db = get_db()
    """Listar histórico de notificações"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        limite = int(request.args.get('limit', 50))
        tipo = request.args.get('tipo')

        query = {}
        if tipo:
            query['tipo'] = tipo

        notificacoes = list(db.notificacoes.find(query).sort('created_at', DESCENDING).limit(limite))
        return jsonify({'success': True, 'notificacoes': convert_objectid(notificacoes)})

    except Exception as e:
        logger.error(f"Erro ao listar notificações: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/estoque/entrada', methods=['POST'])
@login_required
def registrar_entrada_estoque():
    db = get_db()
    '''Registrar uma nova entrada de estoque que deve ser aprovada.'''
    if db is None:
        return jsonify({'success': False}), 500

    data = request.json or {}
    try:
        produto_id = data.get('produto_id')
        if not produto_id:
            return jsonify({'success': False, 'message': 'Produto obrigatorio'}), 400

        produto = db.produtos.find_one({'_id': ObjectId(produto_id)})
        if not produto:
            return jsonify({'success': False, 'message': 'Produto nao encontrado'}), 404

        quantidade = int(data.get('quantidade', 0))
        if quantidade <= 0:
            return jsonify({'success': False, 'message': 'Quantidade invalida'}), 400

        entrada_data = {
            'produto_id': ObjectId(produto_id),
            'produto_nome': produto.get('nome'),
            'quantidade': quantidade,
            'fornecedor': data.get('fornecedor', ''),
            'motivo': data.get('motivo', 'Entrada manual'),
            'nota_fiscal': data.get('nota_fiscal', ''),
            'previsao_chegada': data.get('previsao_chegada'),
            'status': 'Pendente',
            'usuario': session.get('username', 'sistema'),
            'data': datetime.now(),
            'updated_at': datetime.now()
        }

        db.estoque_entradas_pendentes.insert_one(entrada_data)
        logger.info(f"Entrada de estoque registrada para produto {produto_id}")
        return jsonify({'success': True, 'message': 'Entrada registrada e aguardando aprovacao'})
    except ValueError:
        return jsonify({'success': False, 'message': 'Quantidade deve ser numerica'}), 400
    except Exception as e:
        logger.error(f"Erro ao registrar entrada: {e}")
        return jsonify({'success': False, 'message': 'Erro interno ao registrar entrada'}), 500


@bp.route('/api/estoque/entrada/pendentes', methods=['GET'])
@login_required
def estoque_entradas_pendentes():
    db = get_db()
    '''Listar entradas de estoque pendentes de aprovacao.'''
    if db is None:
        return jsonify({'success': False}), 500

    try:
        entradas = list(db.estoque_entradas_pendentes.find({'status': 'Pendente'}).sort('data', DESCENDING))
        for entrada in entradas:
            produto = db.produtos.find_one({'_id': entrada['produto_id']})
            if produto:
                entrada['estoque_atual'] = produto.get('estoque', 0)
        return jsonify({'success': True, 'entradas': convert_objectid(entradas)})
    except Exception as e:
        logger.error(f"Erro ao listar pendencias de estoque: {e}")
        return jsonify({'success': False, 'message': 'Erro ao listar entradas pendentes'}), 500


@bp.route('/api/estoque/entrada/<id>', methods=['PUT'])
@login_required
def atualizar_entrada_estoque(id):
    db = get_db()
    '''Atualizar uma entrada pendente antes da aprovacao.'''
    if db is None:
        return jsonify({'success': False}), 500

    data = request.json or {}
    try:
        entrada = db.estoque_entradas_pendentes.find_one({'_id': ObjectId(id)})
        if not entrada:
            return jsonify({'success': False, 'message': 'Entrada nao encontrada'}), 404
        if entrada.get('status') != 'Pendente':
            return jsonify({'success': False, 'message': 'Entrada ja processada'}), 400

        campos_atualizar = {}
        for campo in ['quantidade', 'fornecedor', 'motivo', 'nota_fiscal', 'previsao_chegada']:
            if campo in data:
                campos_atualizar[campo] = data[campo]

        if 'quantidade' in campos_atualizar:
            try:
                campos_atualizar['quantidade'] = int(campos_atualizar['quantidade'])
                if campos_atualizar['quantidade'] <= 0:
                    return jsonify({'success': False, 'message': 'Quantidade invalida'}), 400
            except ValueError:
                return jsonify({'success': False, 'message': 'Quantidade deve ser numerica'}), 400

        if not campos_atualizar:
            return jsonify({'success': False, 'message': 'Nenhum dado para atualizar'}), 400

        campos_atualizar['updated_at'] = datetime.now()
        db.estoque_entradas_pendentes.update_one({'_id': ObjectId(id)}, {'$set': campos_atualizar})
        logger.info(f"Entrada de estoque {id} atualizada")
        return jsonify({'success': True, 'message': 'Entrada atualizada com sucesso'})
    except Exception as e:
        logger.error(f"Erro ao atualizar entrada: {e}")
        return jsonify({'success': False, 'message': 'Erro ao atualizar entrada'}), 500


@bp.route('/api/estoque/entrada/<id>/aprovar', methods=['POST'])
@login_required
def aprovar_entrada_estoque(id):
    db = get_db()
    '''Aprovar uma entrada de estoque pendente.'''
    if db is None:
        return jsonify({'success': False}), 500

    try:
        entrada = db.estoque_entradas_pendentes.find_one({'_id': ObjectId(id)})
        if not entrada:
            return jsonify({'success': False, 'message': 'Entrada nao encontrada'}), 404
        if entrada.get('status') != 'Pendente':
            return jsonify({'success': False, 'message': 'Entrada ja processada'}), 400

        produto = db.produtos.find_one({'_id': entrada['produto_id']})
        if not produto:
            return jsonify({'success': False, 'message': 'Produto nao encontrado para entrada'}), 404

        novo_estoque = produto.get('estoque', 0) + int(entrada.get('quantidade', 0))
        db.produtos.update_one({'_id': produto['_id']}, {'$set': {'estoque': novo_estoque, 'updated_at': datetime.now()}})

        db.estoque_movimentacoes.insert_one({
            'produto_id': produto['_id'],
            'tipo': 'entrada',
            'quantidade': int(entrada.get('quantidade', 0)),
            'motivo': entrada.get('motivo', 'Entrada aprovada'),
            'usuario': session.get('username', 'sistema'),
            'data': datetime.now()
        })

        db.estoque_entradas_pendentes.update_one({
            '_id': ObjectId(id)
        }, {
            '$set': {'status': 'Aprovado', 'aprovado_em': datetime.now(), 'aprovado_por': session.get('username')}
        })

        return jsonify({'success': True, 'message': 'Entrada aprovada e estoque atualizado'})
    except Exception as e:
        logger.error(f"Erro ao aprovar entrada: {e}")
        return jsonify({'success': False, 'message': 'Erro ao aprovar entrada'}), 500


@bp.route('/api/estoque/entrada/<id>/rejeitar', methods=['POST'])
@login_required
def rejeitar_entrada_estoque(id):
    db = get_db()
    '''Rejeitar uma entrada pendente de estoque.'''
    if db is None:
        return jsonify({'success': False}), 500

    data = request.json or {}
    try:
        entrada = db.estoque_entradas_pendentes.find_one({'_id': ObjectId(id)})
        if not entrada:
            return jsonify({'success': False, 'message': 'Entrada nao encontrada'}), 404
        if entrada.get('status') != 'Pendente':
            return jsonify({'success': False, 'message': 'Entrada ja processada'}), 400

        motivo = data.get('motivo', '').strip() or 'Rejeitada sem motivo informado'
        db.estoque_entradas_pendentes.update_one({
            '_id': ObjectId(id)
        }, {
            '$set': {'status': 'Rejeitado', 'motivo_rejeicao': motivo, 'rejeitado_em': datetime.now(), 'rejeitado_por': session.get('username')}})

        logger.info(f"Entrada de estoque {id} rejeitada")
        return jsonify({'success': True, 'message': 'Entrada rejeitada'})
    except Exception as e:
        logger.error(f"Erro ao rejeitar entrada: {e}")
        return jsonify({'success': False, 'message': 'Erro ao rejeitar entrada'}), 500


@bp.route('/api/estoque/alerta')
@login_required
def estoque_alerta():
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Erro ao conectar ao banco'}), 500
    try:
        # Buscar TODOS os produtos primeiro
        todos_produtos = list(db.produtos.find({}))

        # Filtrar produtos onde estoque <= estoque_minimo (default 5)
        produtos_baixos = []
        for p in todos_produtos:
            estoque_atual = p.get('estoque', 0) or 0
            estoque_minimo = p.get('estoque_minimo', 5) or 5

            if estoque_atual <= estoque_minimo:
                produtos_baixos.append(p)

        logger.info(f"📦 Alerta de estoque: {len(produtos_baixos)}/{len(todos_produtos)} produtos abaixo do mínimo")
        return jsonify({'success': True, 'produtos': convert_objectid(produtos_baixos)})

    except Exception as e:
        logger.error(f"❌ Erro ao buscar alertas de estoque: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/estoque/movimentacao', methods=['POST'])
@login_required
def estoque_movimentacao():
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    data = request.json
    try:
        produto = db.produtos.find_one({'_id': ObjectId(data['produto_id'])})
        if not produto:
            return jsonify({'success': False})
        qtd = int(data['quantidade'])
        tipo = data['tipo']
        novo_estoque = produto.get('estoque', 0)
        if tipo == 'entrada':
            novo_estoque += qtd
        else:
            novo_estoque -= qtd
        db.produtos.update_one({'_id': ObjectId(data['produto_id'])}, {'$set': {'estoque': novo_estoque}})
        db.estoque_movimentacoes.insert_one({'produto_id': ObjectId(data['produto_id']), 'tipo': tipo, 'quantidade': qtd, 'motivo': data.get('motivo', ''), 'usuario': session.get('username'), 'data': datetime.now()})
        return jsonify({'success': True})
    except:
        return jsonify({'success': False}), 500

@bp.route('/api/estoque/movimentacoes', methods=['GET'])
@login_required
def listar_movimentacoes():
    db = get_db()
    """Listar movimentações de estoque com paginação e filtros"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        # Paginação
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        skip = (page - 1) * per_page
        
        # Filtros opcionais
        filtro = {}
        
        # Filtro por tipo de movimentação
        tipo = request.args.get('tipo')
        if tipo:
            filtro['tipo'] = tipo
        
        # Filtro por data (início)
        data_inicio = request.args.get('data_inicio')
        if data_inicio:
            try:
                filtro['data'] = {'$gte': datetime.fromisoformat(data_inicio)}
            except:
                pass
        
        # Filtro por data (fim)
        data_fim = request.args.get('data_fim')
        if data_fim:
            try:
                if 'data' in filtro:
                    filtro['data']['$lte'] = datetime.fromisoformat(data_fim)
                else:
                    filtro['data'] = {'$lte': datetime.fromisoformat(data_fim)}
            except:
                pass
        
        # Filtro por produto
        produto_id = request.args.get('produto_id')
        if produto_id:
            try:
                filtro['produto_id'] = ObjectId(produto_id)
            except:
                pass
        
        # Total de registros
        total = db.estoque_movimentacoes.count_documents(filtro)
        
        # Buscar movimentações com paginação
        movimentacoes = list(db.estoque_movimentacoes.find(filtro)
                            .sort('data', DESCENDING)
                            .skip(skip)
                            .limit(per_page))

        # Enriquecer com nomes dos produtos
        for mov in movimentacoes:
            produto_id = mov.get('produto_id')
            if produto_id:
                try:
                    produto = db.produtos.find_one({'_id': ObjectId(produto_id) if isinstance(produto_id, str) else produto_id})
                    if produto:
                        mov['produto_nome'] = produto.get('nome', '[Produto sem nome]')
                    else:
                        mov['produto_nome'] = f'[Produto removido: {produto_id}]'
                except Exception as e:
                    logger.warning(f"Erro ao buscar produto {produto_id}: {e}")
                    mov['produto_nome'] = f'[Erro: {produto_id}]'
            else:
                mov['produto_nome'] = '[Sem produto]'

        return jsonify({
            'success': True,
            'movimentacoes': convert_objectid(movimentacoes),
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': (total + per_page - 1) // per_page
            }
        })
    except Exception as e:
        logger.error(f"Erro ao listar movimentações: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/estoque/saida', methods=['POST'])
@login_required
def registrar_saida_estoque():
    db = get_db()
    """Registrar saída de estoque"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    data = request.json or {}
    try:
        produto_id = data.get('produto_id')
        quantidade = int(data.get('quantidade', 0))
        motivo = data.get('motivo', '')
        
        if not produto_id or quantidade <= 0:
            return jsonify({'success': False, 'message': 'Dados inválidos'}), 400
        
        # Verificar se produto existe
        produto = db.produtos.find_one({'_id': ObjectId(produto_id)})
        if not produto:
            return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404
        
        # Verificar estoque disponível
        estoque_atual = produto.get('estoque', 0)
        if estoque_atual < quantidade:
            return jsonify({
                'success': False, 
                'message': f'Estoque insuficiente. Disponível: {estoque_atual}'
            }), 400
        
        # Atualizar estoque
        novo_estoque = estoque_atual - quantidade
        db.produtos.update_one(
            {'_id': ObjectId(produto_id)},
            {'$set': {'estoque': novo_estoque, 'updated_at': datetime.now()}}
        )
        
        # Registrar movimentação
        movimentacao = {
            'produto_id': ObjectId(produto_id),
            'produto_nome': produto.get('nome'),
            'tipo': 'saida',
            'quantidade': quantidade,
            'motivo': motivo,
            'usuario': session.get('username', 'Desconhecido'),
            'data': datetime.now()
        }
        db.estoque_movimentacoes.insert_one(movimentacao)
        
        return jsonify({
            'success': True,
            'message': 'Saída registrada com sucesso',
            'estoque_atual': novo_estoque
        })
        
    except Exception as e:
        logger.error(f"Erro ao registrar saída: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/estoque/relatorio-simples', methods=['GET'])
@login_required
def relatorio_estoque():
    db = get_db()
    """Gerar relatório completo de estoque com filtros e opção de export"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        # Filtros opcionais
        data_inicio_str = request.args.get('data_inicio')
        data_fim_str = request.args.get('data_fim')
        formato = request.args.get('formato', 'json')  # json ou excel
        
        # Total de produtos
        total_produtos = db.produtos.count_documents({})
        
        # Valor total do estoque
        produtos = list(db.produtos.find({}))
        valor_total = sum(p.get('preco', 0) * p.get('estoque', 0) for p in produtos)
        
        # Produtos com baixo estoque
        baixo_estoque = db.produtos.count_documents({
            '$expr': {'$lt': ['$estoque', '$estoque_minimo']}
        })
        
        # Produtos sem estoque
        sem_estoque = db.produtos.count_documents({'estoque': 0})
        
        # Filtro de data para movimentações
        filtro_data = {}
        if data_inicio_str:
            try:
                filtro_data['data'] = {'$gte': datetime.fromisoformat(data_inicio_str)}
            except:
                pass
        
        if data_fim_str:
            try:
                if 'data' in filtro_data:
                    filtro_data['data']['$lte'] = datetime.fromisoformat(data_fim_str)
                else:
                    filtro_data['data'] = {'$lte': datetime.fromisoformat(data_fim_str)}
            except:
                pass
        
        # Últimas movimentações (com filtro de data se aplicável)
        ultimas_movimentacoes = list(db.estoque_movimentacoes.find(filtro_data)
                                     .sort('data', DESCENDING)
                                     .limit(10))
        
        # Produtos mais movimentados (com filtro de data)
        if not filtro_data:
            filtro_data['data'] = {'$gte': datetime.now() - timedelta(days=30)}
        
        pipeline = [
            {'$match': filtro_data},
            {'$group': {
                '_id': '$produto_id',
                'total_movimentacoes': {'$sum': 1},
                'total_quantidade': {'$sum': '$quantidade'}
            }},
            {'$sort': {'total_movimentacoes': -1}},
            {'$limit': 5}
        ]
        mais_movimentados = list(db.estoque_movimentacoes.aggregate(pipeline))
        
        # Enriquecer com nomes dos produtos
        for item in mais_movimentados:
            produto = db.produtos.find_one({'_id': item['_id']})
            if produto:
                item['produto_nome'] = produto.get('nome')
        
        relatorio_data = {
            'total_produtos': total_produtos,
            'valor_total_estoque': round(valor_total, 2),
            'produtos_baixo_estoque': baixo_estoque,
            'produtos_sem_estoque': sem_estoque,
            'ultimas_movimentacoes': convert_objectid(ultimas_movimentacoes),
            'mais_movimentados': convert_objectid(mais_movimentados),
            'periodo': {
                'inicio': data_inicio_str,
                'fim': data_fim_str
            }
        }
        
        # Export para Excel
        if formato == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Relatório de Estoque'
            
            # Cabeçalho
            ws.append(['RELATÓRIO DE ESTOQUE'])
            ws.append([''])
            ws.append(['Total de Produtos', total_produtos])
            ws.append(['Valor Total Estoque', f'R$ {valor_total:.2f}'])
            ws.append(['Produtos Baixo Estoque', baixo_estoque])
            ws.append(['Produtos Sem Estoque', sem_estoque])
            ws.append([''])
            
            # Produtos mais movimentados
            ws.append(['PRODUTOS MAIS MOVIMENTADOS'])
            ws.append(['Produto', 'Total Movimentações', 'Quantidade Total'])
            for item in mais_movimentados:
                ws.append([
                    item.get('produto_nome', 'N/A'),
                    item.get('total_movimentacoes', 0),
                    item.get('total_quantidade', 0)
                ])
            
            # Salvar em buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return send_file(
                excel_buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        
        # Retorno JSON padrão
        return jsonify({
            'success': True,
            'relatorio': relatorio_data
        })
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def normalizar_coluna(texto):
    """Normalização extrema de nomes de colunas para máxima compatibilidade"""
    import unicodedata
    import re

    if not texto:
        return ''

    # Converter para string
    texto = str(texto)

    # Remover acentos (NFD = Canonical Decomposition)
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(char for char in texto if unicodedata.category(char) != 'Mn')

    # Converter para lowercase
    texto = texto.lower()

    # Remover caracteres especiais, manter apenas alfanuméricos e underscores
    texto = re.sub(r'[^a-z0-9_]', '', texto)

    # Remover espaços/underscores duplicados
    texto = re.sub(r'_+', '_', texto)

    # Remover underscores no início/fim
    texto = texto.strip('_')

    return texto

@bp.route('/api/importar', methods=['POST'])
@login_required
def importar():
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Erro ao conectar ao banco de dados'}), 500
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400
    file = request.files['file']
    tipo = request.form.get('tipo', 'produtos')
    if not file.filename:
        return jsonify({'success': False, 'message': 'Nome do arquivo inválido'}), 400
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if ext not in ['csv', 'xlsx', 'xls']:
        return jsonify({'success': False, 'message': 'Formato de arquivo inválido'}), 400

    # Criar diretório de upload se não existir
    upload_folder = current_app.config.get('UPLOAD_FOLDER', '/tmp')
    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder, exist_ok=True)

    filepath = None
    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join(upload_folder, f"{int(time())}_{filename}")
        file.save(filepath)
        count_success = 0
        count_error = 0
        rows = []
        if ext == 'csv':
            encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
            for encoding in encodings:
                try:
                    with open(filepath, 'r', encoding=encoding) as csvfile:
                        rows = list(csv.DictReader(csvfile))
                        break
                except:
                    continue
        else:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i in range(len(headers)):
                    if i < len(row) and row[i] is not None:
                        val = row[i]
                        if isinstance(val, str):
                            val = val.strip()
                        row_dict[headers[i]] = val
                if row_dict:
                    rows.append(row_dict)
            wb.close()
        if tipo == 'produtos':
            for idx, row in enumerate(rows, 1):
                try:
                    r = {k.lower().strip(): v for k, v in row.items() if k and v is not None}
                    if not r or all(not v for v in r.values()):
                        continue
                    nome = None
                    for col in ['nome', 'produto', 'name']:
                        if col in r and r[col]:
                            nome = str(r[col]).strip()
                            break
                    if not nome or len(nome) < 2:
                        count_error += 1
                        continue
                    marca = ''
                    for col in ['marca', 'brand']:
                        if col in r and r[col]:
                            marca = str(r[col]).strip()
                            break
                    sku = f"PROD-{count_success+1}"
                    for col in ['sku', 'codigo']:
                        if col in r and r[col]:
                            sku = str(r[col]).strip()
                            break
                    preco = 0.0
                    for col in ['preco', 'preço', 'price', 'valor']:
                        if col in r and r[col]:
                            try:
                                val = str(r[col]).replace('R$', '').strip()
                                if ',' in val:
                                    val = val.replace(',', '.')
                                preco = float(val)
                                break
                            except:
                                continue
                    if preco <= 0:
                        count_error += 1
                        continue
                    custo = 0.0
                    for col in ['custo', 'cost']:
                        if col in r and r[col]:
                            try:
                                val = str(r[col]).replace('R$', '').strip()
                                if ',' in val:
                                    val = val.replace(',', '.')
                                custo = float(val)
                                break
                            except:
                                continue
                    estoque = 0
                    for col in ['estoque', 'quantidade', 'qtd']:
                        if col in r and r[col]:
                            try:
                                estoque = int(float(r[col]))
                                break
                            except:
                                continue
                    categoria = 'Produto'
                    for col in ['categoria', 'category']:
                        if col in r and r[col]:
                            categoria = str(r[col]).strip().title()
                            break
                    # Código de barras
                    codigo_barras = 'Sem código de barras no momento'
                    for col in ['codigo_barras', 'código_barras', 'codigo barras', 'barcode', 'codigo de barras', 'código de barras']:
                        if col in r and r[col]:
                            codigo_barras = str(r[col]).strip()
                            break
                    db.produtos.insert_one({'nome': nome, 'marca': marca, 'sku': sku, 'preco': preco, 'custo': custo, 'estoque': estoque, 'estoque_minimo': 5, 'categoria': categoria, 'codigo_barras': codigo_barras, 'ativo': True, 'created_at': datetime.now()})
                    count_success += 1
                except:
                    count_error += 1
        elif tipo == 'servicos':
            # Importação de serviços com NORMALIZAÇÃO EXTREMA
            logger.info(f"🔄 Importando {len(rows)} linhas de serviços com normalização extrema...")

            for idx, row in enumerate(rows, 1):
                try:
                    # NORMALIZAÇÃO EXTREMA: remover acentos, caracteres especiais, etc
                    r = {normalizar_coluna(k): v for k, v in row.items() if k and v is not None}

                    if not r or all(not v for v in r.values()):
                        logger.debug(f"Linha {idx}: linha vazia, pulando")
                        continue

                    logger.debug(f"📋 Linha {idx}: Colunas detectadas: {list(r.keys())}")

                    # Nome do serviço - NORMALIZADO
                    nome = None
                    nome_cols = [normalizar_coluna(x) for x in ['nome', 'servico', 'serviço', 'name', 'service']]
                    for col in nome_cols:
                        if col in r and r[col]:
                            nome = str(r[col]).strip()
                            logger.debug(f"  ✓ Nome encontrado na coluna '{col}': {nome}")
                            break

                    if not nome or len(nome) < 2:
                        logger.warning(f"❌ Linha {idx}: nome do serviço inválido ou ausente - Colunas: {list(r.keys())}")
                        count_error += 1
                        continue

                    # Categoria - NORMALIZADO
                    categoria = 'Serviço'
                    cat_cols = [normalizar_coluna(x) for x in ['categoria', 'category', 'tipo', 'type']]
                    for col in cat_cols:
                        if col in r and r[col]:
                            categoria = str(r[col]).strip().title()
                            logger.debug(f"  ✓ Categoria encontrada: {categoria}")
                            break

                    # Duração (em minutos) - NORMALIZADO
                    duracao = 60  # Padrão: 60 minutos
                    dur_cols = [normalizar_coluna(x) for x in ['duracao', 'duração', 'tempo', 'duration', 'minutos', 'minutes', 'min']]
                    for col in dur_cols:
                        if col in r and r[col]:
                            try:
                                duracao = int(float(r[col]))
                                logger.debug(f"  ✓ Duração encontrada: {duracao} min")
                                break
                            except:
                                continue

                    # Preços por tamanho - DETECÇÃO MELHORADA
                    # Busca por padrões DENTRO do nome da coluna (ex: "Preço P" contém "p")
                    tamanhos_patterns = {
                        'kids': ['kids', 'crianca', 'infantil', 'child', 'kid', 'bebe'],
                        'masculino': ['masculino', 'male', 'homem', 'masc', 'barba', 'beard'],
                        'curto': ['curto', 'short', 'pequeno', 'mini', 'small'],
                        'medio': ['medio', 'medium', 'media', 'normal'],
                        'longo': ['longo', 'long', 'grande', 'large', 'big'],
                        'extra_longo': ['extralongo', 'extralong', 'extralarge', 'muitolongo']
                    }

                    # Letras únicas para tamanhos (P, M, G, etc) - busca EXATA
                    tamanhos_letras = {
                        'curto': ['p', 's'],
                        'medio': ['m'],
                        'longo': ['g', 'l'],
                        'extra_longo': ['gg', 'xl', 'xxl']
                    }

                    tamanhos_precos = {}

                    # Buscar por colunas normalizadas - DETECÇÃO MELHORADA v7.0
                    logger.debug(f"  🔍 Iniciando detecção de tamanhos...")
                    for coluna_normalizada, valor in r.items():
                        if not valor:
                            continue

                        tamanho_detectado = None
                        metodo_deteccao = None

                        # 1. Verificar se a coluna é EXATAMENTE uma letra (p, m, g, etc)
                        if len(coluna_normalizada) <= 3:  # Colunas curtas como 'p', 'm', 'g', 'gg', 'xl'
                            for tam, letras in tamanhos_letras.items():
                                if coluna_normalizada in letras:
                                    tamanho_detectado = tam
                                    metodo_deteccao = f"letra exata '{coluna_normalizada}'"
                                    break

                        # 2. Se não achou, verificar se a coluna TERMINA com letra de tamanho
                        if not tamanho_detectado and len(coluna_normalizada) > 1:
                            ultima_letra = coluna_normalizada[-1]
                            if ultima_letra in ['p', 's']:
                                tamanho_detectado = 'curto'
                                metodo_deteccao = f"termina com '{ultima_letra}'"
                            elif ultima_letra == 'm':
                                tamanho_detectado = 'medio'
                                metodo_deteccao = f"termina com '{ultima_letra}'"
                            elif ultima_letra in ['g', 'l']:
                                tamanho_detectado = 'longo'
                                metodo_deteccao = f"termina com '{ultima_letra}'"

                            # Verificar se termina com GG/XL
                            if not tamanho_detectado and len(coluna_normalizada) >= 2:
                                ultimas_duas = coluna_normalizada[-2:]
                                if ultimas_duas in ['gg', 'xl']:
                                    tamanho_detectado = 'extra_longo'
                                    metodo_deteccao = f"termina com '{ultimas_duas}'"

                        # 3. Se não achou, verificar se algum padrão está CONTIDO na coluna
                        if not tamanho_detectado:
                            for tam, patterns in tamanhos_patterns.items():
                                for pattern in patterns:
                                    pattern_norm = normalizar_coluna(pattern)
                                    # Buscar o padrão DENTRO da coluna
                                    if pattern_norm in coluna_normalizada or coluna_normalizada in pattern_norm:
                                        tamanho_detectado = tam
                                        metodo_deteccao = f"padrão '{pattern}'"
                                        break
                                if tamanho_detectado:
                                    break

                        # Se detectou um tamanho, extrair o preço (REMOVIDO o check "not in tamanhos_precos" para pegar TODOS)
                        if tamanho_detectado:
                            try:
                                val = str(valor).replace('R$', '').replace('$', '').strip()
                                if ',' in val:
                                    val = val.replace(',', '.')
                                preco = float(val)
                                if preco > 0:
                                    # Permitir sobrescrever se encontrar preço maior
                                    if tamanho_detectado not in tamanhos_precos or preco > tamanhos_precos[tamanho_detectado]:
                                        tamanhos_precos[tamanho_detectado] = preco
                                        logger.debug(f"  ✓ Preço {tamanho_detectado} encontrado: R$ {preco:.2f} (coluna: '{coluna_normalizada}', método: {metodo_deteccao})")
                            except Exception as e:
                                logger.debug(f"  ⚠ Erro ao converter preço da coluna '{coluna_normalizada}': {e}")

                    # Se não há nenhum preço válido, tentar preço único
                    if not tamanhos_precos:
                        # Tentar importar com preço único - NORMALIZADO
                        preco_unico = 0.0
                        preco_cols = [normalizar_coluna(x) for x in ['preco', 'preço', 'price', 'valor', 'value', 'cost']]
                        for col in preco_cols:
                            if col in r and r[col]:
                                try:
                                    val = str(r[col]).replace('R$', '').replace('$', '').strip()
                                    if ',' in val:
                                        val = val.replace(',', '.')
                                    preco_unico = float(val)
                                    logger.debug(f"  ✓ Preço único encontrado: R$ {preco_unico:.2f}")
                                    break
                                except:
                                    continue

                        if preco_unico > 0:
                            # Criar serviço com preço único
                            sku = f"{nome.upper().replace(' ', '-')}"
                            db.servicos.insert_one({
                                'nome': nome,
                                'sku': sku,
                                'preco': preco_unico,
                                'preco_kids': preco_unico,
                                'preco_masculino': preco_unico,
                                'preco_curto': preco_unico,
                                'preco_medio': preco_unico,
                                'preco_longo': preco_unico,
                                'preco_extra_longo': preco_unico,
                                'categoria': categoria,
                                'duracao': duracao,
                                'ativo': True,
                                'created_at': datetime.now()
                            })
                            logger.info(f"✅ Linha {idx}: '{nome}' importado com preço único R$ {preco_unico:.2f}")
                            count_success += 1
                        else:
                            logger.warning(f"❌ Linha {idx}: '{nome}' sem preços válidos - Colunas: {list(r.keys())}")
                            count_error += 1
                        continue
                    
                    # Criar um serviço para cada tamanho com preço definido
                    tamanhos_labels = {
                        'kids': 'Kids',
                        'masculino': 'Masculino',
                        'curto': 'Curto',
                        'medio': 'Médio',
                        'longo': 'Longo',
                        'extra_longo': 'Extra Longo'
                    }

                    logger.info(f"✅ Linha {idx}: '{nome}' com {len(tamanhos_precos)} tamanhos detectados")
                    for tamanho_key, preco in tamanhos_precos.items():
                        tamanho_label = tamanhos_labels.get(tamanho_key, tamanho_key.title())
                        sku = f"{nome.upper().replace(' ', '-')}-{tamanho_label.upper().replace(' ', '-')}"

                        db.servicos.insert_one({
                            'nome': nome,
                            'sku': sku,
                            'tamanho': tamanho_label,
                            'preco': preco,
                            'categoria': categoria,
                            'duracao': duracao,
                            'ativo': True,
                            'created_at': datetime.now()
                        })
                        logger.debug(f"  ➕ Criado: {nome} - {tamanho_label} - R$ {preco:.2f}")

                    count_success += 1
                except Exception as e:
                    logger.error(f"❌ Erro na linha {idx}: {str(e)}")
                    count_error += 1
        if filepath and os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': True, 'message': f'{count_success} importados!', 'count_success': count_success, 'count_error': count_error})
    except Exception as e:
        logger.error(f"Erro na importação: {e}")
        if filepath and os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': False, 'message': 'Erro ao processar arquivo'}), 500

@bp.route('/api/importar/desfazer', methods=['POST'])
@login_required
def desfazer_importacao():
    """Desfaz a última importação removendo os registros mais recentes."""
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Erro ao conectar ao banco de dados'}), 500

    try:
        data = request.get_json()
        tipo = data.get('tipo')
        timestamp = data.get('timestamp')
        count = data.get('count', 0)

        if not tipo or not timestamp:
            return jsonify({'success': False, 'message': 'Dados inválidos'}), 400

        # Converter timestamp para datetime
        import_date = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))

        # Determinar a coleção correta
        collection_map = {
            'produtos': db.produtos,
            'servicos': db.servicos,
            'clientes': db.clientes,
            'profissionais': db.profissionais
        }

        collection = collection_map.get(tipo)
        if not collection:
            return jsonify({'success': False, 'message': 'Tipo inválido'}), 400

        # Deletar registros criados após (ou próximo a) o timestamp da importação
        # Consideramos uma margem de 5 minutos para garantir que pegamos apenas os registros da importação
        time_margin = timedelta(minutes=5)
        query = {
            'created_at': {
                '$gte': import_date - time_margin,
                '$lte': import_date + time_margin
            }
        }

        # Deletar os registros
        result = collection.delete_many(query)
        deleted_count = result.deleted_count

        logger.info(f"Desfazer importação: {deleted_count} registros de {tipo} removidos")

        return jsonify({
            'success': True,
            'message': f'{deleted_count} registros de {tipo} foram removidos',
            'deleted': deleted_count
        })

    except Exception as e:
        logger.error(f"Erro ao desfazer importação: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/estoque/importar', methods=['POST'])
@login_required
def estoque_importar_alias():
    """Alias para importação de planilhas do estoque; reutiliza a rota /api/importar existente."""
    return importar()

@bp.route('/api/template/download/<tipo>')
@login_required
def download_template(tipo):
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    if tipo == 'produtos':
        ws.title = 'Produtos BIOMA'
        headers = ['nome','marca','sku','preco','custo','estoque','estoque_minimo','categoria']
    elif tipo == 'clientes':
        ws.title = 'Clientes BIOMA'
        headers = ['cpf','nome','email','telefone','genero','data_nascimento','endereco']
    elif tipo == 'profissionais':
        ws.title = 'Profissionais BIOMA'
        headers = ['cpf','nome','especialidade','email','telefone','comissao_perc','assistente_id','comissao_assistente_perc']
    else:
        # serviços (padrão)
        ws.title = 'Servicos BIOMA'
        headers = ['nome','tamanho','sku','preco','categoria','duracao_min']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    # largura de colunas básica (A..I)
    for i, col in enumerate(['A','B','C','D','E','F','G','H','I'], start=1):
        try:
            ws.column_dimensions[col].width = 18
        except Exception:
            pass
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'template_{tipo}_bioma.xlsx')



@bp.route('/api/clientes/formularios', methods=['GET'])
@login_required
def clientes_formularios():
    db = get_db()
    """Retorna a configuração dos formulários de Anamnese e Prontuário para renderização dinâmica no front."""
    try:
        return jsonify({
            'success': True,
            'anamnese': ANAMNESE_FORM,
            'prontuario': PRONTUARIO_FORM
        })
    except Exception as e:
        logger.exception("Erro ao carregar formulários")
        return jsonify({'success': False, 'message': str(e)}), 500
@bp.route('/api/config', methods=['GET', 'POST'])
@login_required
def config():
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500
    if request.method == 'GET':
        try:
            cfg = db.config.find_one({'key': 'unidade'}) or {}
            return jsonify({'success': True, 'config': convert_objectid(cfg)})
        except:
            return jsonify({'success': False}), 500
    data = request.json
    try:
        # Atualizar configurações incluindo logo_url
        config_data = {
            'key': 'unidade',
            'nome_empresa': data.get('nome_empresa', 'BIOMA UBERABA'),
            'logo_url': data.get('logo_url', ''),
            'logo_login_url': data.get('logo_login_url', ''),
            'endereco': data.get('endereco', ''),
            'telefone': data.get('telefone', ''),
            'email': data.get('email', ''),
            'cnpj': data.get('cnpj', ''),
            'cor_primaria': data.get('cor_primaria', '#7C3AED'),
            'cor_secundaria': data.get('cor_secundaria', '#EC4899'),
            # (Ponto 2) CAMPOS OPERACIONAIS ADICIONADOS
            'horario_abertura': data.get('horario_abertura', '08:00'),
            'horario_fechamento': data.get('horario_fechamento', '19:00'),
            'duracao_padrao': data.get('duracao_padrao', 60),
            'cashback_padrao': data.get('cashback_padrao', 5),
            'comissao_padrao': data.get('comissao_padrao', 10),
            'updated_at': datetime.now()
        }
        
        db.config.update_one({'key': 'unidade'}, {'$set': config_data}, upsert=True)
        logger.info("✅ Configurações atualizadas")
        return jsonify({'success': True, 'message': 'Configurações salvas com sucesso!'})
    except Exception as e:
        logger.error(f"Erro ao salvar configurações: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
        return jsonify({'success': False}), 500

@bp.route('/api/relatorios/completo', methods=['GET'])
@login_required
def relatorio_completo():
    db = get_db()
    """Relatório completo com todas as estatísticas do sistema"""
    if db is None:
        return jsonify({'success': False}), 500
    
    try:
        periodo = request.args.get('periodo', '30')  # Padrão 30 dias
        dias = int(periodo)
        data_inicio = datetime.now() - timedelta(days=dias)
        
        # === ESTATÍSTICAS GERAIS ===
        total_clientes = db.clientes.count_documents({})
        total_produtos = db.produtos.count_documents({'ativo': True})
        total_servicos = db.servicos.count_documents({'ativo': True})
        total_profissionais = db.profissionais.count_documents({'ativo': True})
        
        # === FATURAMENTO ===
        orcamentos_aprovados = list(db.orcamentos.find({
            'status': 'Aprovado',
            'created_at': {'$gte': data_inicio}
        }))
        
        faturamento_total = sum(o.get('total_final', 0) for o in orcamentos_aprovados)
        faturamento_servicos = sum(
            sum(s.get('total', 0) for s in o.get('servicos', []))
            for o in orcamentos_aprovados
        )
        faturamento_produtos = sum(
            sum(p.get('total', 0) for p in o.get('produtos', []))
            for o in orcamentos_aprovados
        )
        
        # === VENDAS ===
        total_orcamentos = db.orcamentos.count_documents({
            'created_at': {'$gte': data_inicio}
        })
        orcamentos_aprovados_count = len(orcamentos_aprovados)
        orcamentos_pendentes = db.orcamentos.count_documents({
            'status': 'Pendente',
            'created_at': {'$gte': data_inicio}
        })
        taxa_conversao = (orcamentos_aprovados_count / total_orcamentos * 100) if total_orcamentos > 0 else 0
        ticket_medio = faturamento_total / orcamentos_aprovados_count if orcamentos_aprovados_count > 0 else 0
        
        # === ESTOQUE ===
        produtos = list(db.produtos.find({'ativo': True}))
        estoque_total_valor_custo = sum(p.get('estoque', 0) * p.get('custo', 0) for p in produtos)
        estoque_total_valor_venda = sum(p.get('estoque', 0) * p.get('preco', 0) for p in produtos)
        produtos_zerados = len([p for p in produtos if p.get('estoque', 0) == 0])
        produtos_baixo_estoque = len([p for p in produtos if p.get('estoque', 0) <= p.get('estoque_minimo', 5) and p.get('estoque', 0) > 0])
        produtos_criticos = len([p for p in produtos if p.get('estoque', 0) == 0])
        
        # === CLIENTES ===
        novos_clientes = db.clientes.count_documents({
            'created_at': {'$gte': data_inicio}
        })
        
        # Top 10 clientes por gasto
        pipeline_clientes = [
            {'$match': {'status': 'Aprovado', 'created_at': {'$gte': data_inicio}}},
            {'$group': {
                '_id': '$cliente_cpf',
                'total_gasto': {'$sum': '$total_final'},
                'total_compras': {'$sum': 1},
                'cliente_nome': {'$first': '$cliente_nome'}
            }},
            {'$sort': {'total_gasto': -1}},
            {'$limit': 10}
        ]
        top_clientes = list(db.orcamentos.aggregate(pipeline_clientes))
        
        # === PRODUTOS MAIS VENDIDOS ===
        produtos_vendidos = {}
        for orc in orcamentos_aprovados:
            for prod in orc.get('produtos', []):
                prod_id = prod.get('id')
                if prod_id:
                    if prod_id not in produtos_vendidos:
                        produtos_vendidos[prod_id] = {
                            'nome': prod.get('nome', 'N/A'),
                            'quantidade': 0,
                            'faturamento': 0
                        }
                    produtos_vendidos[prod_id]['quantidade'] += prod.get('qtd', 1)
                    produtos_vendidos[prod_id]['faturamento'] += prod.get('total', 0)
        
        top_produtos = sorted(
            [{'id': k, **v} for k, v in produtos_vendidos.items()],
            key=lambda x: x['faturamento'],
            reverse=True
        )[:10]
        
        # === SERVIÇOS MAIS REALIZADOS ===
        servicos_realizados = {}
        for orc in orcamentos_aprovados:
            for serv in orc.get('servicos', []):
                serv_id = serv.get('id')
                if serv_id:
                    if serv_id not in servicos_realizados:
                        servicos_realizados[serv_id] = {
                            'nome': serv.get('nome', 'N/A'),
                            'tamanho': serv.get('tamanho', ''),
                            'quantidade': 0,
                            'faturamento': 0
                        }
                    servicos_realizados[serv_id]['quantidade'] += 1
                    servicos_realizados[serv_id]['faturamento'] += serv.get('total', 0)
        
        top_servicos = sorted(
            [{'id': k, **v} for k, v in servicos_realizados.items()],
            key=lambda x: x['faturamento'],
            reverse=True
        )[:10]
        
        # === FATURAMENTO POR DIA (para gráficos) ===
        faturamento_por_dia = {}
        for orc in orcamentos_aprovados:
            data = orc.get('created_at', datetime.now())
            dia = data.strftime('%Y-%m-%d')
            if dia not in faturamento_por_dia:
                faturamento_por_dia[dia] = 0
            faturamento_por_dia[dia] += orc.get('total_final', 0)
        
        faturamento_timeline = [
            {'data': dia, 'valor': valor}
            for dia, valor in sorted(faturamento_por_dia.items())
        ]
        
        # === PROFISSIONAIS COM MELHOR DESEMPENHO ===
        desempenho_profissionais = {}
        for orc in orcamentos_aprovados:
            for serv in orc.get('servicos', []):
                prof_id = serv.get('profissional_id')
                if prof_id:
                    if prof_id not in desempenho_profissionais:
                        prof = db.profissionais.find_one({'_id': ObjectId(prof_id)})
                        nome = prof.get('nome', 'N/A') if prof else 'N/A'
                        desempenho_profissionais[prof_id] = {
                            'nome': nome,
                            'servicos_realizados': 0,
                            'faturamento': 0
                        }
                    desempenho_profissionais[prof_id]['servicos_realizados'] += 1
                    desempenho_profissionais[prof_id]['faturamento'] += serv.get('total', 0)
        
        top_profissionais = sorted(
            [{'id': k, **v} for k, v in desempenho_profissionais.items()],
            key=lambda x: x['faturamento'],
            reverse=True
        )[:10]
        
        relatorio = {
            'periodo': f'Últimos {dias} dias',
            'data_inicio': data_inicio.isoformat(),
            'data_fim': datetime.now().isoformat(),
            
            'geral': {
                'total_clientes': total_clientes,
                'total_produtos': total_produtos,
                'total_servicos': total_servicos,
                'total_profissionais': total_profissionais,
                'novos_clientes': novos_clientes
            },
            
            'faturamento': {
                'total': faturamento_total,
                'servicos': faturamento_servicos,
                'produtos': faturamento_produtos,
                'timeline': faturamento_timeline
            },
            
            'vendas': {
                'total_orcamentos': total_orcamentos,
                'aprovados': orcamentos_aprovados_count,
                'pendentes': orcamentos_pendentes,
                'taxa_conversao': taxa_conversao,
                'ticket_medio': ticket_medio
            },
            
            'estoque': {
                'valor_total_custo': estoque_total_valor_custo,
                'valor_total_venda': estoque_total_valor_venda,
                'margem_potencial': estoque_total_valor_venda - estoque_total_valor_custo,
                'produtos_zerados': produtos_zerados,
                'produtos_baixo_estoque': produtos_baixo_estoque,
                'produtos_criticos': produtos_criticos
            },
            
            'rankings': {
                'top_clientes': convert_objectid(top_clientes),
                'top_produtos': top_produtos,
                'top_servicos': top_servicos,
                'top_profissionais': top_profissionais
            }
        }
        
        return jsonify({'success': True, 'relatorio': relatorio})
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/auditoria', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def consultar_auditoria():
    db = get_db()
    """Consultar logs de auditoria do sistema (Admin/Gestão only)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        # Paginação
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        skip = (page - 1) * per_page
        
        # Filtros
        filtro = {}
        
        # Filtro por usuário
        username = request.args.get('username')
        if username:
            filtro['username'] = {'$regex': username, '$options': 'i'}
        
        # Filtro por ação
        acao = request.args.get('acao')
        if acao:
            filtro['acao'] = acao
        
        # Filtro por entidade
        entidade = request.args.get('entidade')
        if entidade:
            filtro['entidade'] = entidade
        
        # Filtro por data
        data_inicio = request.args.get('data_inicio')
        if data_inicio:
            try:
                filtro['timestamp'] = {'$gte': datetime.fromisoformat(data_inicio)}
            except:
                pass
        
        data_fim = request.args.get('data_fim')
        if data_fim:
            try:
                if 'timestamp' in filtro:
                    filtro['timestamp']['$lte'] = datetime.fromisoformat(data_fim)
                else:
                    filtro['timestamp'] = {'$lte': datetime.fromisoformat(data_fim)}
            except:
                pass
        
        # Total de registros
        total = db.auditoria.count_documents(filtro)
        
        # Buscar registros
        registros = list(db.auditoria.find(filtro)
                        .sort('timestamp', DESCENDING)
                        .skip(skip)
                        .limit(per_page))
        
        # Estatísticas rápidas
        stats = {
            'total_acoes': total,
            'acoes_por_tipo': list(db.auditoria.aggregate([
                {'$match': filtro},
                {'$group': {'_id': '$acao', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}}
            ])),
            'usuarios_ativos': list(db.auditoria.aggregate([
                {'$match': filtro},
                {'$group': {'_id': '$username', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}},
                {'$limit': 10}
            ]))
        }
        
        return jsonify({
            'success': True,
            'registros': convert_objectid(registros),
            'stats': stats,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': (total + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao consultar auditoria: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/sistema/limpar-banco', methods=['POST'])
@login_required
@permission_required('Admin')
def limpar_banco_dados():
    """
    ATENÇÃO: Esta função limpa TODOS os dados do banco de dados!
    Requer permissão de Admin e código de confirmação.
    """
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        data = request.get_json()
        codigo_confirmacao = data.get('codigo_confirmacao', '')

        # Código de segurança: LIMPAR2025
        if codigo_confirmacao != 'LIMPAR2025':
            return jsonify({
                'success': False,
                'message': 'Código de confirmação inválido'
            }), 403

        # Registrar ação de auditoria ANTES de limpar
        usuario = session.get('user', {})
        db.auditoria.insert_one({
            'username': usuario.get('username', 'Desconhecido'),
            'acao': 'LIMPAR_BANCO_DADOS',
            'entidade': 'SISTEMA',
            'timestamp': datetime.now(),
            'detalhes': 'Limpeza completa do banco de dados'
        })

        # Coleções a serem limpas (exceto users e auditoria para manter histórico)
        colecoes_para_limpar = [
            'clientes',
            'profissionais',
            'servicos',
            'produtos',
            'orcamentos',
            'contratos',
            'agendamentos',
            'estoque_movimentacoes',
            'comissoes',
            'despesas',
            'assistentes',
            'fila'
        ]

        stats = {}
        for colecao_nome in colecoes_para_limpar:
            colecao = getattr(db, colecao_nome)
            count_antes = colecao.count_documents({})
            resultado = colecao.delete_many({})
            stats[colecao_nome] = {
                'antes': count_antes,
                'deletados': resultado.deleted_count
            }
            logger.warning(f"🗑️ Coleção '{colecao_nome}': {resultado.deleted_count} documentos deletados")

        logger.warning(f"⚠️ BANCO DE DADOS LIMPO por {usuario.get('username', 'Desconhecido')}")

        return jsonify({
            'success': True,
            'message': 'Banco de dados limpo com sucesso!',
            'stats': stats
        })

    except Exception as e:
        logger.error(f"Erro ao limpar banco de dados: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def init_db():
    db = get_db()
    if db is None:
        logger.warning("⚠️ DB não disponível para inicialização")
        return
    logger.info("🔧 Initializing DB...")
    if db.users.count_documents({}) == 0:
        db.users.insert_one({'name': 'Administrador', 'username': 'admin', 'email': 'admin@bioma.com', 'telefone': '', 'password': generate_password_hash('admin123'), 'role': 'admin', 'theme': 'light', 'created_at': datetime.now()})
        logger.info("✅ Admin user created: admin/admin123 (role: admin)")
    if db.servicos.count_documents({}) == 0:
        services = [('Hidratação', 'Tratamento', [50, 60, 80, 100, 120, 150]), ('Corte', 'Cabelo', [40, 50, 60, 80, 100, 120])]
        tamanhos = ['Kids', 'Masculino', 'Curto', 'Médio', 'Longo', 'Extra Longo']
        for nome, cat, precos in services:
            for tam, preco in zip(tamanhos, precos):
                db.servicos.insert_one({'nome': nome, 'sku': f"{nome.upper()}-{tam.upper()}", 'tamanho': tam, 'preco': preco, 'categoria': cat, 'duracao': 60, 'ativo': True, 'created_at': datetime.now()})
        logger.info(f"✅ {len(services) * 6} service SKUs created")
    try:
        # ==================== STRATEGIC INDEXING FOR PERFORMANCE ====================
        # User indexes (existing)
        db.users.create_index([('username', ASCENDING)], unique=True)
        db.users.create_index([('email', ASCENDING)], unique=True)

        # Clientes - For lookups and filtering
        db.clientes.create_index([('cpf', ASCENDING)])
        db.clientes.create_index([('email', ASCENDING)])
        db.clientes.create_index([('nome', ASCENDING)])

        # Produtos - For catalog search and filtering
        db.produtos.create_index([('nome', ASCENDING), ('categoria', ASCENDING)])
        db.produtos.create_index([('sku', ASCENDING)], unique=True, sparse=True)
        db.produtos.create_index([('ativo', ASCENDING)])

        # Servicos - For service catalog
        db.servicos.create_index([('nome', ASCENDING), ('categoria', ASCENDING)])
        db.servicos.create_index([('sku', ASCENDING)], unique=True, sparse=True)
        db.servicos.create_index([('ativo', ASCENDING)])

        # Agendamentos - CRITICAL for calendar performance
        db.agendamentos.create_index([('profissional_id', ASCENDING), ('data', ASCENDING)])
        db.agendamentos.create_index([('data', ASCENDING), ('hora', ASCENDING)])
        db.agendamentos.create_index([('cliente_id', ASCENDING), ('data', DESCENDING)])
        db.agendamentos.create_index([('status', ASCENDING)])

        # Orçamentos
        db.orcamentos.create_index([('numero', ASCENDING)], unique=True)
        db.orcamentos.create_index([('cliente_id', ASCENDING), ('created_at', DESCENDING)])
        db.orcamentos.create_index([('status', ASCENDING)])

        # Profissionais
        db.profissionais.create_index([('nome', ASCENDING)])
        db.profissionais.create_index([('email', ASCENDING)])
        db.profissionais.create_index([('ativo', ASCENDING)])

        # Profissionais Avaliacoes (existing)
        db.profissionais_avaliacoes.create_index([('profissional_id', ASCENDING), ('created_at', DESCENDING)])

        # Auditoria - For filtering audit logs by user and time
        db.auditoria.create_index([('user_id', ASCENDING), ('timestamp', DESCENDING)])
        db.auditoria.create_index([('timestamp', DESCENDING)])
        db.auditoria.create_index([('action', ASCENDING)])

        # Estoque/Movimentacoes - For stock tracking
        db.movimentacoes.create_index([('produto_id', ASCENDING), ('data', DESCENDING)])
        db.movimentacoes.create_index([('tipo', ASCENDING), ('data', DESCENDING)])

        # Fila - For queue management
        db.fila.create_index([('status', ASCENDING), ('created_at', ASCENDING)])
        db.fila.create_index([('cliente_id', ASCENDING)])

        logger.info('✅ Strategic database indexes created for optimal performance')
    except Exception as e:
        logger.warning(f"⚠️ Index creation warning: {e}")
    logger.info("🎉 DB initialization complete!")

@bp.errorhandler(404)
def not_found(e):
    return jsonify({'success': False, 'message': 'Not found'}), 404

@bp.errorhandler(500)
def internal_error(e):
    logger.error(f"❌ 500 Internal Error: {e}")
    return jsonify({'success': False, 'message': 'Internal server error'}), 500

# ==================== NOVAS FUNCIONALIDADES v3.7 ====================

# 1. Upload de Logo da Empresa
@bp.route('/api/upload/logo', methods=['POST'])
@login_required
def upload_logo():
    """Upload de logo da empresa (armazenado como base64, SEM arquivos externos)"""
    try:
        if 'logo' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['logo']
        tipo = request.form.get('tipo', 'principal')  # principal ou login

        if file.filename == '':
            return jsonify({'success': False, 'message': 'Arquivo vazio'}), 400

        if file and allowed_file(file.filename):
            import base64

            # Ler arquivo em memória (SEM salvar em /tmp)
            file_content = file.read()
            foto_base64 = base64.b64encode(file_content).decode('utf-8')

            # Determinar tipo MIME
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
            mime_types = {
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'webp': 'image/webp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            # Montar data URI (armazenado diretamente no MongoDB)
            data_uri = f"data:{mime_type};base64,{foto_base64}"

            # Salvar referência no banco COM data URI
            db.uploads.insert_one({
                'tipo': f'logo_{tipo}',
                'filename': secure_filename(file.filename),
                'data_uri': data_uri,  # Armazena base64 diretamente
                'mime_type': mime_type,
                'data_upload': datetime.now()
            })

            logger.info(f"✅ Logo {tipo} salvo como base64 no MongoDB (SEM arquivos externos)")

            return jsonify({
                'success': True,
                'message': 'Logo enviado com sucesso',
                'url': data_uri  # Retorna Data URI diretamente
            })

        return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido'}), 400
    except Exception as e:
        logger.error(f"Erro no upload de logo: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 1.1. Upload de Imagem (Alias para compatibilidade)
@bp.route('/api/upload/imagem', methods=['POST'])
@login_required
def upload_imagem():
    db = get_db()
    """Alias para compatibilidade com frontend"""
    return upload_logo()

# 2. Obter Logo Configurado
@bp.route('/api/config/logo', methods=['GET'])
def get_logo():
    db = get_db()
    """Obter Data URI do logo configurado (sem arquivos externos)"""
    try:
        tipo = request.args.get('tipo', 'principal')
        logo = db.uploads.find_one({'tipo': f'logo_{tipo}'}, sort=[('data_upload', DESCENDING)])

        if logo:
            # Retorna data_uri em vez de url de arquivo
            return jsonify({'success': True, 'url': logo.get('data_uri')})
        return jsonify({'success': True, 'url': None})
    except Exception as e:
        logger.error(f"Erro ao obter logo: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 3. Servir Arquivos Uploaded (DEPRECATED - Sistema usa Data URI agora)
@bp.route('/uploads/<filename>')
def uploaded_file(filename):
    db = get_db()
    """
    DEPRECATED: Sistema agora usa Data URI (base64) armazenado diretamente no MongoDB.
    Arquivos não são mais salvos no filesystem.
    Esta rota existe apenas para compatibilidade retroativa.
    """
    try:
        # Tentar buscar do MongoDB caso seja um registro antigo
        upload = db.uploads.find_one({'filename': filename})
        if upload and 'data_uri' in upload:
            # Retorna o data_uri como JSON (frontend deve usar esse valor direto)
            return jsonify({'success': True, 'data_uri': upload['data_uri']})

        # Se não encontrou no MongoDB, retorna 404
        logger.warning(f"⚠️ Tentativa de acessar arquivo que não existe: {filename}")
        return jsonify({
            'success': False,
            'message': 'Sistema não usa mais arquivos externos. Todos os uploads são armazenados como Data URI no MongoDB.'
        }), 404
    except Exception as e:
        logger.error(f"Erro ao buscar upload: {e}")
        return jsonify({'success': False, 'message': 'Arquivo não encontrado'}), 404

# 4. Upload de Foto de Profissional (via form data)
@bp.route('/api/upload/foto-profissional', methods=['POST'])
@login_required
def upload_foto_profissional_form():
    """Upload de foto de perfil de profissional via form data (armazenado como base64, SEM arquivos externos)"""
    db = get_db()
    if db is None:
        logger.error("❌ Database offline em upload foto")
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        logger.info(f"📤 Upload foto - Files: {list(request.files.keys())}, Form: {list(request.form.keys())}")

        if 'foto' not in request.files:
            logger.warning("⚠️ Nenhum arquivo 'foto' encontrado")
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['foto']
        profissional_id = request.form.get('profissional_id')

        if not profissional_id:
            logger.warning("⚠️ profissional_id não fornecido")
            return jsonify({'success': False, 'message': 'ID do profissional não fornecido'}), 400

        if file.filename == '':
            logger.warning("⚠️ Arquivo com nome vazio")
            return jsonify({'success': False, 'message': 'Arquivo vazio'}), 400

        # Função allowed_file inline (se não existir)
        def allowed_file_check(filename):
            ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

        if file and allowed_file_check(file.filename):
            import base64

            # Ler arquivo em memória (SEM salvar em /tmp)
            file_content = file.read()
            foto_base64 = base64.b64encode(file_content).decode('utf-8')

            # Determinar tipo MIME
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
            mime_types = {
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'webp': 'image/webp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            # Montar data URI (armazenado diretamente no MongoDB)
            data_uri = f"data:{mime_type};base64,{foto_base64}"

            # Atualizar profissional com a foto (Data URI em vez de URL de arquivo)
            db.profissionais.update_one(
                {'_id': ObjectId(profissional_id)},
                {'$set': {
                    'foto': data_uri,
                    'foto_url': data_uri,
                    'foto_atualizada_em': datetime.now()
                }}
            )

            # Salvar referência no banco COM data URI
            db.uploads.insert_one({
                'tipo': 'foto_profissional',
                'profissional_id': ObjectId(profissional_id),
                'filename': secure_filename(file.filename),
                'data_uri': data_uri,  # Armazena base64 diretamente
                'mime_type': mime_type,
                'data_upload': datetime.now()
            })

            logger.info(f"✅ Foto de profissional {profissional_id} salva como base64 no MongoDB (SEM arquivos externos)")

            return jsonify({
                'success': True,
                'message': 'Foto enviada com sucesso',
                'url': data_uri  # Retorna Data URI diretamente
            })

        return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido'}), 400
    except Exception as e:
        logger.error(f"Erro no upload de foto: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 5. Listar Todas as Comissões
@bp.route('/api/comissoes', methods=['GET'])
@login_required
def get_all_comissoes():
    db = get_db()
    """Lista todas as comissões registradas no sistema"""
    try:
        # Buscar todas as comissões do histórico
        comissoes_lista = []
        comissoes = list(db.comissoes_historico.find().sort('data_registro', DESCENDING).limit(100))

        for comissao in comissoes:
            # Buscar informações do profissional
            profissional = db.profissionais.find_one({'_id': comissao.get('profissional_id')})
            profissional_nome = profissional.get('nome', '[Profissional removido]') if profissional else '[Profissional não encontrado]'

            # Buscar informações do orçamento
            orcamento = db.orcamentos.find_one({'_id': comissao.get('orcamento_id')})
            cliente_nome = '[Cliente não encontrado]'
            orcamento_numero = '[Removido]'

            if orcamento:
                cliente_nome = orcamento.get('cliente_nome') or orcamento.get('nome_cliente') or '[Nome não informado]'
                orcamento_numero = orcamento.get('numero', orcamento.get('orcamento_numero', '[S/N]'))

            comissoes_lista.append({
                'id': str(comissao['_id']),
                'orcamento_id': str(comissao.get('orcamento_id', '')),
                'orcamento_numero': orcamento_numero,
                'profissional_id': str(comissao.get('profissional_id', '')),
                'profissional_nome': profissional_nome,
                'cliente_nome': cliente_nome,
                'comissao_valor': comissao.get('comissao_valor', 0),
                'data_registro': comissao.get('data_registro', datetime.now()).isoformat() if comissao.get('data_registro') else datetime.now().isoformat()
            })

        total = sum(c.get('comissao_valor', 0) for c in comissoes_lista)

        logger.info(f"✅ Listadas {len(comissoes_lista)} comissões - Total: R$ {total:.2f}")

        return jsonify({
            'success': True,
            'comissoes': comissoes_lista,
            'total': round(total, 2),
            'quantidade': len(comissoes_lista)
        })
    except Exception as e:
        logger.error(f"❌ Erro ao listar comissões: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 6. Calcular Comissões Multiníveis
@bp.route('/api/comissoes/calcular-orcamento', methods=['POST'])
@login_required
def calcular_comissoes_orcamento():
    db = get_db()
    """Calcula comissões em cascata: profissional + assistente"""
    try:
        data = request.get_json()
        orcamento_id = data.get('orcamento_id')
        
        if not orcamento_id:
            return jsonify({'success': False, 'message': 'ID do orçamento não fornecido'}), 400
        
        orcamento = db.orcamentos.find_one({'_id': ObjectId(orcamento_id)})
        if not orcamento:
            return jsonify({'success': False, 'message': 'Orçamento não encontrado'}), 404
        
        # Buscar dados do profissional
        profissional = db.profissionais.find_one({'_id': ObjectId(orcamento.get('profissional_id'))})
        if not profissional:
            return jsonify({'success': False, 'message': 'Profissional não encontrado'}), 404
        
        # Calcular comissão do profissional
        valor_total = orcamento.get('total', 0)
        comissao_perc = profissional.get('comissao_perc', 10)
        comissao_profissional = valor_total * (comissao_perc / 100)
        
        resultado = {
            'orcamento_id': str(orcamento_id),
            'valor_total': valor_total,
            'profissional': {
                'id': str(profissional['_id']),
                'nome': profissional.get('nome'),
                'foto': profissional.get('foto'),
                'comissao_percentual': comissao_perc,
                'comissao_valor': round(comissao_profissional, 2)
            }
        }
        
        # Calcular comissão do assistente (10% da comissão do profissional)
        assistentes = profissional.get('assistentes', [])
        if assistentes:
            assistente_principal = assistentes[0]
            assistente_id = assistente_principal.get('_id') if isinstance(assistente_principal, dict) else assistente_principal
            
            assistente_doc = db.assistentes.find_one({'_id': ObjectId(assistente_id)})
            if assistente_doc:
                comissao_assistente = comissao_profissional * 0.1  # 10% da comissão do profissional
                resultado['assistente'] = {
                    'id': str(assistente_doc['_id']),
                    'nome': assistente_doc.get('nome'),
                    'foto': assistente_doc.get('foto'),
                    'comissao_percentual': 10,
                    'comissao_valor': round(comissao_assistente, 2)
                }
                resultado['total_comissoes'] = round(comissao_profissional + comissao_assistente, 2)
            else:
                resultado['total_comissoes'] = round(comissao_profissional, 2)
        else:
            resultado['total_comissoes'] = round(comissao_profissional, 2)
        
        # Salvar no histórico
        db.comissoes_historico.insert_one({
            'orcamento_id': ObjectId(orcamento_id),
            'profissional_id': ObjectId(profissional['_id']),
            'profissional_comissao': resultado['profissional']['comissao_valor'],
            'assistente_id': ObjectId(resultado['assistente']['id']) if 'assistente' in resultado else None,
            'assistente_comissao': resultado.get('assistente', {}).get('comissao_valor', 0),
            'total': resultado['total_comissoes'],
            'data_calculo': datetime.now()
        })
        
        return jsonify({'success': True, 'comissoes': resultado})
    except Exception as e:
        logger.error(f"Erro ao calcular comissões: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 6. Histórico de Comissões
@bp.route('/api/comissoes/historico', methods=['GET'])
@login_required
def historico_comissoes():
    db = get_db()
    """Lista histórico de comissões calculadas"""
    try:
        historico = list(db.comissoes_historico.find().sort('data_calculo', DESCENDING).limit(50))
        
        for item in historico:
            item['_id'] = str(item['_id'])
            item['orcamento_id'] = str(item.get('orcamento_id'))
            item['profissional_id'] = str(item.get('profissional_id'))
            if item.get('assistente_id'):
                item['assistente_id'] = str(item['assistente_id'])
            
            # Buscar nomes
            prof = db.profissionais.find_one({'_id': ObjectId(item['profissional_id'])})
            if prof:
                item['profissional_nome'] = prof.get('nome')
            
            if item.get('assistente_id'):
                asst = db.assistentes.find_one({'_id': ObjectId(item['assistente_id'])})
                if asst:
                    item['assistente_nome'] = asst.get('nome')
        
        return jsonify({'success': True, 'historico': historico})
    except Exception as e:
        logger.error(f"Erro ao buscar histórico: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 7. Cadastrar Assistente Independente
@bp.route('/api/assistentes/cadastrar-independente', methods=['POST'])
@login_required
def cadastrar_assistente_independente():
    db = get_db()
    """Cadastra assistente sem vínculo obrigatório com profissional"""
    try:
        data = request.get_json()
        
        assistente = {
            'nome': data.get('nome'),
            'cpf': data.get('cpf'),
            'telefone': data.get('telefone'),
            'email': data.get('email'),
            'profissional_id': None,  # Independente
            'ativo': True,
            'created_at': datetime.now()
        }
        
        result = db.assistentes.insert_one(assistente)
        assistente['_id'] = str(result.inserted_id)
        
        return jsonify({'success': True, 'assistente': assistente})
    except Exception as e:
        logger.error(f"Erro ao cadastrar assistente: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 8. Registrar Entrada de Produto (com aprovação)
@bp.route('/api/estoque/produtos/entrada', methods=['POST'])
@login_required
def registrar_entrada_produto():
    db = get_db()
    """Registra entrada de produto que precisa ser aprovada"""
    try:
        data = request.get_json()
        
        produto = db.produtos.find_one({'_id': ObjectId(data.get('produto_id'))})
        if not produto:
            return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404
        
        entrada = {
            'produto_id': ObjectId(data.get('produto_id')),
            'produto_nome': produto.get('nome'),
            'quantidade': data.get('quantidade'),
            'fornecedor': data.get('fornecedor'),
            'motivo': data.get('motivo'),
            'usuario': session.get('user', {}).get('name', 'Desconhecido'),
            'status': 'pendente',
            'data_solicitacao': datetime.now(),
            'data_processamento': None
        }
        
        result = db.estoque_pendencias.insert_one(entrada)
        entrada['_id'] = str(result.inserted_id)
        
        return jsonify({'success': True, 'entrada': entrada, 'message': 'Entrada registrada. Aguardando aprovação.'})
    except Exception as e:
        logger.error(f"Erro ao registrar entrada: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 9. Aprovar Entrada de Produto
@bp.route('/api/estoque/produtos/aprovar/<id>', methods=['POST'])
@login_required
def aprovar_entrada_produto(id):
    db = get_db()
    """Aprova entrada de produto e atualiza estoque"""
    try:
        entrada = db.estoque_pendencias.find_one({'_id': ObjectId(id)})
        if not entrada:
            return jsonify({'success': False, 'message': 'Entrada não encontrada'}), 404
        
        # Atualizar status da entrada
        db.estoque_pendencias.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'status': 'aprovado', 'data_processamento': datetime.now()}}
        )
        
        # Atualizar estoque do produto
        db.produtos.update_one(
            {'_id': entrada['produto_id']},
            {'$inc': {'estoque': entrada['quantidade']}}
        )
        
        # Registrar movimentação
        db.estoque_movimentacoes.insert_one({
            'produto_id': entrada['produto_id'],
            'produto_nome': entrada['produto_nome'],
            'tipo': 'entrada',
            'quantidade': entrada['quantidade'],
            'motivo': f"Aprovação: {entrada.get('motivo', '')}",
            'usuario': session.get('user', {}).get('name'),
            'data': datetime.now()
        })
        
        return jsonify({'success': True, 'message': 'Entrada aprovada e estoque atualizado'})
    except Exception as e:
        logger.error(f"Erro ao aprovar entrada: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 10. Rejeitar Entrada de Produto
@bp.route('/api/estoque/produtos/rejeitar/<id>', methods=['POST'])
@login_required
def rejeitar_entrada_produto(id):
    db = get_db()
    """Rejeita entrada de produto"""
    try:
        db.estoque_pendencias.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'status': 'rejeitado', 'data_processamento': datetime.now()}}
        )
        
        return jsonify({'success': True, 'message': 'Entrada rejeitada'})
    except Exception as e:
        logger.error(f"Erro ao rejeitar entrada: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 11. Listar Entradas Pendentes
@bp.route('/api/estoque/produtos/pendentes', methods=['GET'])
@login_required
def listar_entradas_pendentes():
    db = get_db()
    """Lista entradas de produtos pendentes de aprovação"""
    try:
        pendentes = list(db.estoque_pendencias.find({'status': 'pendente'}).sort('data_solicitacao', DESCENDING))
        
        for item in pendentes:
            item['_id'] = str(item['_id'])
            item['produto_id'] = str(item.get('produto_id'))
        
        return jsonify({'success': True, 'entradas': pendentes})
    except Exception as e:
        logger.error(f"Erro ao listar pendentes: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 12. Mapa de Calor
@bp.route('/api/relatorios/mapa-calor', methods=['GET'])
@login_required
def mapa_calor():
    db = get_db()
    """Retorna dados melhorados para mapa de calor de movimentação (Diretriz #5)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Parâmetros configuráveis
        dias = int(request.args.get('dias', 90))
        incluir_faturamento = request.args.get('faturamento', 'true').lower() == 'true'
        incluir_clientes = request.args.get('clientes', 'true').lower() == 'true'

        # Período
        data_inicio = datetime.now() - timedelta(days=dias)
        data_fim = datetime.now()

        # Pipeline de agregação otimizado
        pipeline_base = [
            {'$match': {'created_at': {'$gte': data_inicio, '$lte': data_fim}}},
            {'$group': {
                '_id': {
                    '$dateToString': {'format': '%Y-%m-%d', 'date': '$created_at'}
                },
                'count': {'$sum': 1}
            }},
            {'$sort': {'_id': 1}}
        ]

        # Buscar dados de diferentes collections
        agendamentos_data = list(db.agendamentos.aggregate(pipeline_base))
        orcamentos_data = list(db.orcamentos.aggregate(pipeline_base))

        # Pipeline com faturamento para orçamentos aprovados
        pipeline_faturamento = [
            {'$match': {
                'created_at': {'$gte': data_inicio, '$lte': data_fim},
                'status': 'Aprovado'
            }},
            {'$group': {
                '_id': {
                    '$dateToString': {'format': '%Y-%m-%d', 'date': '$created_at'}
                },
                'faturamento': {'$sum': '$total_final'},
                'count': {'$sum': 1}
            }},
            {'$sort': {'_id': 1}}
        ]

        faturamento_data = list(db.orcamentos.aggregate(pipeline_faturamento)) if incluir_faturamento else []

        # Pipeline para novos clientes
        pipeline_clientes = [
            {'$match': {'created_at': {'$gte': data_inicio, '$lte': data_fim}}},
            {'$group': {
                '_id': {
                    '$dateToString': {'format': '%Y-%m-%d', 'date': '$created_at'}
                },
                'count': {'$sum': 1}
            }},
            {'$sort': {'_id': 1}}
        ]

        clientes_data = list(db.clientes.aggregate(pipeline_clientes)) if incluir_clientes else []

        # Combinar todos os dados em um mapa por dia
        dias_map = {}

        # Inicializar todos os dias do período
        current_date = data_inicio
        while current_date <= data_fim:
            dia_str = current_date.strftime('%Y-%m-%d')
            dias_map[dia_str] = {
                'data': dia_str,
                'agendamentos': 0,
                'orcamentos': 0,
                'faturamento': 0,
                'novos_clientes': 0,
                'intensidade_total': 0
            }
            current_date += timedelta(days=1)

        # Preencher agendamentos
        for item in agendamentos_data:
            dia = item['_id']
            if dia in dias_map:
                dias_map[dia]['agendamentos'] = item['count']
                dias_map[dia]['intensidade_total'] += item['count']

        # Preencher orçamentos
        for item in orcamentos_data:
            dia = item['_id']
            if dia in dias_map:
                dias_map[dia]['orcamentos'] = item['count']
                dias_map[dia]['intensidade_total'] += item['count']

        # Preencher faturamento
        for item in faturamento_data:
            dia = item['_id']
            if dia in dias_map:
                dias_map[dia]['faturamento'] = round(item['faturamento'], 2)
                # Adicionar peso ao faturamento para intensidade
                dias_map[dia]['intensidade_total'] += item['count'] * 2  # Peso maior para vendas

        # Preencher novos clientes
        for item in clientes_data:
            dia = item['_id']
            if dia in dias_map:
                dias_map[dia]['novos_clientes'] = item['count']
                dias_map[dia]['intensidade_total'] += item['count']

        # Converter para lista ordenada
        dados = sorted(dias_map.values(), key=lambda x: x['data'])

        # Calcular estatísticas gerais
        total_agendamentos = sum(d['agendamentos'] for d in dados)
        total_orcamentos = sum(d['orcamentos'] for d in dados)
        total_faturamento = sum(d['faturamento'] for d in dados)
        total_clientes = sum(d['novos_clientes'] for d in dados)

        # Dia mais movimentado
        dia_mais_movimentado = max(dados, key=lambda x: x['intensidade_total']) if dados else None
        dia_maior_faturamento = max(dados, key=lambda x: x['faturamento']) if dados and incluir_faturamento else None

        return jsonify({
            'success': True,
            'data': dados,  # Frontend espera 'data', não 'dados'
            'dados': dados,  # Manter 'dados' por compatibilidade
            'periodo': {
                'inicio': data_inicio.strftime('%Y-%m-%d'),
                'fim': data_fim.strftime('%Y-%m-%d'),
                'total_dias': dias
            },
            'estatisticas': {
                'total_agendamentos': total_agendamentos,
                'total_orcamentos': total_orcamentos,
                'total_faturamento': total_faturamento,
                'total_novos_clientes': total_clientes,
                'media_diaria_agendamentos': round(total_agendamentos / dias, 2) if dias > 0 else 0,
                'dia_mais_movimentado': dia_mais_movimentado,
                'dia_maior_faturamento': dia_maior_faturamento
            }
        })

    except Exception as e:
        logger.error(f"Erro ao gerar mapa de calor: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# NOVOS ENDPOINTS PARA GRÁFICOS AVANÇADOS (Roadmap Section V - Relatórios)
# ============================================================================

@bp.route('/api/relatorios/vendas-por-mes', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def relatorio_vendas_por_mes():
    db = get_db()
    """Vendas e faturamento agregados por mês para Chart.js (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # (Ponto 1) Funções de Relatório atualizadas para aceitar filtros de data
        data_inicio_str = request.args.get('data_inicio')
        data_fim_str = request.args.get('data_fim')

        match_query = {'status': 'Aprovado'}
        if data_inicio_str:
            match_query['created_at'] = {'$gte': datetime.fromisoformat(data_inicio_str)}
        if data_fim_str:
            match_query.setdefault('created_at', {})['$lte'] = datetime.fromisoformat(data_fim_str)

        # Pipeline de agregação - vendas por mês
        pipeline = [
            {'$match': match_query},
            {'$group': {
                '_id': {
                    'ano': {'$year': '$created_at'},
                    'mes': {'$month': '$created_at'}
                },
                'total_vendas': {'$sum': 1},
                'faturamento': {'$sum': '$total_final'},
                'ticket_medio': {'$avg': '$total_final'}
            }},
            {'$sort': {'_id.ano': 1, '_id.mes': 1}},
            {'$project': {
                'mes_ano': {
                    '$concat': [
                        {'$toString': '$_id.mes'},
                        '/',
                        {'$toString': '$_id.ano'}
                    ]
                },
                'total_vendas': 1,
                'faturamento': {'$round': ['$faturamento', 2]},
                'ticket_medio': {'$round': ['$ticket_medio', 2]}
            }}
        ]

        resultados = list(db.orcamentos.aggregate(pipeline))

        # Formatar para Chart.js
        labels = [r['mes_ano'] for r in resultados]
        vendas = [r['total_vendas'] for r in resultados]
        faturamento = [r['faturamento'] for r in resultados]
        ticket_medio = [r['ticket_medio'] for r in resultados]

        return jsonify({
            'success': True,
            'chart_data': {
                'labels': labels,
                'datasets': [
                    {
                        'label': 'Total de Vendas',
                        'data': vendas,
                        'backgroundColor': 'rgba(54, 162, 235, 0.5)',
                        'borderColor': 'rgba(54, 162, 235, 1)',
                        'borderWidth': 2
                    },
                    {
                        'label': 'Faturamento (R$)',
                        'data': faturamento,
                        'backgroundColor': 'rgba(75, 192, 192, 0.5)',
                        'borderColor': 'rgba(75, 192, 192, 1)',
                        'borderWidth': 2,
                        'yAxisID': 'y-faturamento'
                    }
                ]
            },
            'ticket_medio': ticket_medio,
            'total_periodo': {
                'vendas': sum(vendas),
                'faturamento': round(sum(faturamento), 2),
                'ticket_medio_geral': round(sum(faturamento) / sum(vendas), 2) if sum(vendas) > 0 else 0
            }
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório de vendas por mês: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/relatorios/servicos-top', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def relatorio_servicos_top():
    db = get_db()
    """Top serviços por faturamento e quantidade (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Parâmetros atualizados para filtros de data
        limite = int(request.args.get('limite', 10))
        data_inicio_str = request.args.get('data_inicio')
        data_fim_str = request.args.get('data_fim')

        match_query = {'status': 'Aprovado'}
        if data_inicio_str:
            match_query['created_at'] = {'$gte': datetime.fromisoformat(data_inicio_str)}
        if data_fim_str:
            match_query.setdefault('created_at', {})['$lte'] = datetime.fromisoformat(data_fim_str)

        # Buscar orçamentos aprovados
        orcamentos = list(db.orcamentos.find(match_query))

        # Agregar serviços
        servicos_map = {}
        for orc in orcamentos:
            for serv in orc.get('servicos', []):
                serv_id = serv.get('id')
                if serv_id:
                    if serv_id not in servicos_map:
                        servicos_map[serv_id] = {
                            'nome': serv.get('nome', 'N/A'),
                            'quantidade': 0,
                            'faturamento': 0
                        }
                    servicos_map[serv_id]['quantidade'] += serv.get('qtd', 1)
                    servicos_map[serv_id]['faturamento'] += serv.get('total', 0)

        # Ordenar por faturamento
        servicos_ranking = sorted(
            servicos_map.values(),
            key=lambda x: x['faturamento'],
            reverse=True
        )[:limite]

        # Formatar para Chart.js (horizontal bar chart)
        labels = [s['nome'] for s in servicos_ranking]
        faturamento = [round(s['faturamento'], 2) for s in servicos_ranking]
        quantidade = [s['quantidade'] for s in servicos_ranking]

        return jsonify({
            'success': True,
            'chart_data': {
                'labels': labels,
                'datasets': [
                    {
                        'label': 'Faturamento (R$)',
                        'data': faturamento,
                        'backgroundColor': 'rgba(255, 159, 64, 0.7)',
                        'borderColor': 'rgba(255, 159, 64, 1)',
                        'borderWidth': 1
                    }
                ]
            },
            'quantidade_execucoes': quantidade,
            'total_servicos': len(servicos_map),
            'faturamento_total': round(sum(s['faturamento'] for s in servicos_map.values()), 2)
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório de top serviços: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/relatorios/profissionais-desempenho', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def relatorio_profissionais_desempenho():
    db = get_db()
    """Desempenho de profissionais por comissões e serviços (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Parâmetros
        dias = int(request.args.get('dias', 90))
        data_inicio = datetime.now() - timedelta(days=dias)

        # Pipeline de agregação - comissões por profissional
        pipeline_comissoes = [
            {'$match': {
                'data_registro': {'$gte': data_inicio},
                'status': {'$in': ['Pago', 'Pendente']}
            }},
            {'$group': {
                '_id': '$profissional_id',
                'total_comissoes': {'$sum': '$valor'},
                'total_servicos': {'$sum': 1}
            }},
            {'$sort': {'total_comissoes': -1}}
        ]

        comissoes_data = list(db.comissoes.aggregate(pipeline_comissoes))

        # Buscar nomes dos profissionais
        profissionais_ranking = []
        for item in comissoes_data:
            prof = db.profissionais.find_one({'_id': item['_id']})
            if prof:
                profissionais_ranking.append({
                    'id': str(prof['_id']),
                    'nome': prof.get('nome', 'N/A'),
                    'total_comissoes': round(item['total_comissoes'], 2),
                    'total_servicos': item['total_servicos'],
                    'media_por_servico': round(item['total_comissoes'] / item['total_servicos'], 2) if item['total_servicos'] > 0 else 0
                })

        # Formatar para Chart.js
        labels = [p['nome'] for p in profissionais_ranking[:10]]  # Top 10
        comissoes = [p['total_comissoes'] for p in profissionais_ranking[:10]]
        servicos = [p['total_servicos'] for p in profissionais_ranking[:10]]

        return jsonify({
            'success': True,
            'chart_data': {
                'labels': labels,
                'datasets': [
                    {
                        'label': 'Total Comissões (R$)',
                        'data': comissoes,
                        'backgroundColor': 'rgba(153, 102, 255, 0.7)',
                        'borderColor': 'rgba(153, 102, 255, 1)',
                        'borderWidth': 1
                    }
                ]
            },
            'servicos_realizados': servicos,
            'profissionais_completo': profissionais_ranking,
            'total_profissionais_ativos': len(profissionais_ranking)
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório de desempenho de profissionais: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/relatorios/produtos-top', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def relatorio_produtos_top():
    db = get_db()
    """Top produtos por faturamento e quantidade vendida (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Parâmetros
        dias = int(request.args.get('dias', 90))
        limite = int(request.args.get('limite', 10))
        data_inicio = datetime.now() - timedelta(days=dias)

        # Buscar orçamentos aprovados
        orcamentos = list(db.orcamentos.find({
            'status': 'Aprovado',
            'created_at': {'$gte': data_inicio}
        }))

        # Agregar produtos
        produtos_map = {}
        for orc in orcamentos:
            for prod in orc.get('produtos', []):
                prod_id = prod.get('id')
                if prod_id:
                    if prod_id not in produtos_map:
                        produtos_map[prod_id] = {
                            'nome': prod.get('nome', 'N/A'),
                            'quantidade': 0,
                            'faturamento': 0
                        }
                    produtos_map[prod_id]['quantidade'] += prod.get('qtd', 1)
                    produtos_map[prod_id]['faturamento'] += prod.get('total', 0)

        # Ordenar por faturamento
        produtos_ranking = sorted(
            produtos_map.values(),
            key=lambda x: x['faturamento'],
            reverse=True
        )[:limite]

        # Formatar para Chart.js
        labels = [p['nome'] for p in produtos_ranking]
        faturamento = [round(p['faturamento'], 2) for p in produtos_ranking]
        quantidade = [p['quantidade'] for p in produtos_ranking]

        return jsonify({
            'success': True,
            'chart_data': {
                'labels': labels,
                'datasets': [
                    {
                        'label': 'Faturamento (R$)',
                        'data': faturamento,
                        'backgroundColor': 'rgba(255, 99, 132, 0.7)',
                        'borderColor': 'rgba(255, 99, 132, 1)',
                        'borderWidth': 1
                    }
                ]
            },
            'quantidade_vendida': quantidade,
            'total_produtos': len(produtos_map),
            'faturamento_total': round(sum(p['faturamento'] for p in produtos_map.values()), 2)
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório de top produtos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/relatorios/taxa-conversao', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def relatorio_taxa_conversao():
    db = get_db()
    """Taxa de conversão de orçamentos (Pendente → Aprovado) por período (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Parâmetros
        dias = int(request.args.get('dias', 90))
        data_inicio = datetime.now() - timedelta(days=dias)

        # Orçamentos por status
        total = db.orcamentos.count_documents({'created_at': {'$gte': data_inicio}})
        aprovados = db.orcamentos.count_documents({
            'created_at': {'$gte': data_inicio},
            'status': 'Aprovado'
        })
        pendentes = db.orcamentos.count_documents({
            'created_at': {'$gte': data_inicio},
            'status': 'Pendente'
        })
        cancelados = db.orcamentos.count_documents({
            'created_at': {'$gte': data_inicio},
            'status': {'$in': ['Cancelado', 'Recusado']}
        })

        # Calcular taxas
        taxa_conversao = round((aprovados / total * 100), 2) if total > 0 else 0
        taxa_cancelamento = round((cancelados / total * 100), 2) if total > 0 else 0

        # Formatar para Chart.js (pizza/doughnut chart)
        return jsonify({
            'success': True,
            'chart_data': {
                'labels': ['Aprovados', 'Pendentes', 'Cancelados/Recusados'],
                'datasets': [{
                    'label': 'Orçamentos',
                    'data': [aprovados, pendentes, cancelados],
                    'backgroundColor': [
                        'rgba(75, 192, 192, 0.7)',
                        'rgba(255, 206, 86, 0.7)',
                        'rgba(255, 99, 132, 0.7)'
                    ],
                    'borderColor': [
                        'rgba(75, 192, 192, 1)',
                        'rgba(255, 206, 86, 1)',
                        'rgba(255, 99, 132, 1)'
                    ],
                    'borderWidth': 1
                }]
            },
            'metricas': {
                'total_orcamentos': total,
                'taxa_conversao': taxa_conversao,
                'taxa_cancelamento': taxa_cancelamento,
                'pendentes_percentual': round((pendentes / total * 100), 2) if total > 0 else 0
            }
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório de taxa de conversão: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/relatorios/exportar-pdf', methods=['POST'])
@login_required
@permission_required('Admin', 'Gestão')
def exportar_relatorio_pdf_com_grafico():
    """
    (Ponto 3) NOVO ENDPOINT: Recebe imagens base64 dos gráficos e as insere em um PDF.
    """
    db = get_db()
    if db is None:
        return jsonify({'success': False}), 500

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.utils import ImageReader
        from reportlab.lib.colors import HexColor
        import base64
        import io

        data = request.json
        data_inicio = data.get('data_inicio', (datetime.now() - timedelta(days=30)).strftime('%d/%m/%Y'))
        data_fim = data.get('data_fim', datetime.now().strftime('%d/%m/%Y'))

        # Imagens dos gráficos (base64)
        img_vendas_mes = data.get('img_vendas_mes')  # Gráfico 1
        img_top_servicos = data.get('img_top_servicos')  # Gráfico 2

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                              topMargin=1.5*cm, bottomMargin=2.5*cm)
        elements = []

        # Configurar estilos
        styles = getSampleStyleSheet()
        COLOR_PRIMARY = HexColor('#7C3AED')
        styles.add(ParagraphStyle(name='MainTitle', fontName='Helvetica-Bold',
                                 fontSize=20, textColor=COLOR_PRIMARY, spaceAfter=18, alignment=1))
        styles.add(ParagraphStyle(name='SectionTitle', fontName='Helvetica-Bold',
                                 fontSize=14, textColor=COLOR_PRIMARY, spaceAfter=10))

        # Título
        elements.append(Paragraph("RELATÓRIO FINANCEIRO", styles['MainTitle']))
        elements.append(Paragraph(f"Período: {data_inicio} a {data_fim}", styles['Normal']))
        elements.append(Spacer(1, 1*cm))

        # Função para adicionar imagem base64
        def add_base64_image(img_data_uri):
            if not img_data_uri or ',' not in img_data_uri:
                return Paragraph("Erro: Imagem do gráfico indisponível.", styles['Normal'])

            img_data = base64.b64decode(img_data_uri.split(',')[1])
            img_file = io.BytesIO(img_data)
            img = ImageReader(img_file)

            # Redimensionar para caber na página
            width, height = img.getSize()
            max_width = A4[0] - 4*cm  # Largura da página menos margens
            if width > max_width:
                ratio = max_width / width
                height = height * ratio
                width = max_width

            return Image(img_file, width=width, height=height)

        # Adicionar Gráfico 1 (Vendas por Mês)
        if img_vendas_mes:
            elements.append(Paragraph("Vendas por Mês", styles['SectionTitle']))
            elements.append(add_base64_image(img_vendas_mes))
            elements.append(Spacer(1, 1*cm))

        # Adicionar Gráfico 2 (Top Serviços)
        if img_top_servicos:
            elements.append(Paragraph("Top Serviços (Faturamento)", styles['SectionTitle']))
            elements.append(add_base64_image(img_top_servicos))
            elements.append(Spacer(1, 1*cm))

        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"relatorio_financeiro_{datetime.now().strftime('%Y%m%d')}.pdf"
        )

    except Exception as e:
        logger.error(f"Erro ao gerar PDF com gráficos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 13. Editar Serviço
@bp.route('/api/servicos/<id>/editar', methods=['PUT'])
@login_required
def editar_servico(id):
    db = get_db()
    """Edita informações de um serviço"""
    try:
        data = request.get_json()
        
        update_data = {}
        if 'nome' in data:
            update_data['nome'] = data['nome']
        if 'preco' in data:
            update_data['preco'] = float(data['preco'])
        if 'duracao' in data:
            update_data['duracao'] = int(data['duracao'])
        if 'categoria' in data:
            update_data['categoria'] = data['categoria']
        if 'tamanho' in data:
            update_data['tamanho'] = data['tamanho']
        
        if not update_data:
            return jsonify({'success': False, 'message': 'Nenhum dado para atualizar'}), 400
        
        result = db.servicos.update_one(
            {'_id': ObjectId(id)},
            {'$set': update_data}
        )
        
        if result.modified_count > 0:
            return jsonify({'success': True, 'message': 'Serviço atualizado com sucesso'})
        return jsonify({'success': False, 'message': 'Nenhuma alteração realizada'}), 400
    except Exception as e:
        logger.error(f"Erro ao editar serviço: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 14. Editar Produto
@bp.route('/api/produtos/<id>/editar', methods=['PUT'])
@login_required
def editar_produto(id):
    db = get_db()
    """Edita informações de um produto"""
    try:
        data = request.get_json()

        update_data = {}
        if 'nome' in data:
            update_data['nome'] = data['nome']
        if 'marca' in data:
            update_data['marca'] = data['marca']
        if 'preco' in data:
            update_data['preco'] = float(data['preco'])
        if 'estoque' in data:
            update_data['estoque'] = int(data.get('estoque') or 0)
        if 'estoque_minimo' in data:
            update_data['estoque_minimo'] = int(data['estoque_minimo'])
        if 'sku' in data:
            update_data['sku'] = data['sku']
        if 'codigo_barras' in data:
            update_data['codigo_barras'] = data['codigo_barras']
        if 'categoria' in data:
            update_data['categoria'] = data['categoria']
        if 'status' in data:
            update_data['status'] = data['status']
            update_data['ativo'] = (data['status'] == 'Ativo')
        if 'ativo' in data:
            update_data['ativo'] = bool(data['ativo'])
            update_data['status'] = 'Ativo' if data['ativo'] else 'Inativo'

        if not update_data:
            return jsonify({'success': False, 'message': 'Nenhum dado para atualizar'}), 400

        result = db.produtos.update_one(
            {'_id': ObjectId(id)},
            {'$set': update_data}
        )

        # Aceitar tanto modificações quanto quando não há mudanças
        if result.matched_count > 0:
            return jsonify({'success': True, 'message': 'Produto atualizado com sucesso'})
        return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404
    except Exception as e:
        logger.error(f"Erro ao editar produto: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 15. Faturamento Total do Cliente
@bp.route('/api/clientes/<id>/faturamento', methods=['GET'])
@login_required
def faturamento_cliente(id):
    db = get_db()
    """Retorna faturamento total de um cliente"""
    try:
        pipeline = [
            {'$match': {'cliente_id': ObjectId(id), 'status': 'aprovado'}},
            {'$group': {'_id': None, 'total': {'$sum': '$total'}}}
        ]
        
        resultado = list(db.orcamentos.aggregate(pipeline))
        total = resultado[0]['total'] if resultado else 0
        
        return jsonify({'success': True, 'faturamento_total': round(total, 2)})
    except Exception as e:
        logger.error(f"Erro ao calcular faturamento: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 16. Anamnese do Cliente (GET/PUT)
@bp.route('/api/clientes/<id>/anamnese', methods=['GET', 'PUT', 'POST'])
@login_required
def anamnese_cliente(id):
    db = get_db()
    """Gerencia anamnese do cliente"""
    try:
        # Tentar buscar por ObjectId primeiro, se falhar buscar por CPF
        try:
            cliente = db.clientes.find_one({'_id': ObjectId(id)})
        except:
            # Se não for um ObjectId válido, buscar por CPF
            cliente = db.clientes.find_one({'cpf': id})

        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        if request.method == 'GET':
            # Buscar todas as anamneses do cliente (histórico)
            anamneses = cliente.get('anamneses', [])
            anamnese_atual = cliente.get('anamnese', {})

            return jsonify({
                'success': True,
                'anamnese': anamnese_atual,  # Última anamnese
                'anamneses': anamneses,  # Histórico de anamneses
                'cliente': {
                    'nome': cliente.get('nome'),
                    'cpf': cliente.get('cpf')
                }
            })

        elif request.method in ['PUT', 'POST']:
            data = request.get_json()

            # Adicionar timestamp à anamnese
            nova_anamnese = {
                **data,
                'data': datetime.now().isoformat(),
                'profissional': session.get('user', {}).get('nome', 'Sistema')
            }

            # Atualizar anamnese atual e adicionar ao histórico
            db.clientes.update_one(
                {'_id': cliente['_id']},
                {
                    '$set': {'anamnese': nova_anamnese},
                    '$push': {'anamneses': nova_anamnese}
                }
            )

            return jsonify({'success': True, 'message': 'Anamnese salva com sucesso', 'anamnese': nova_anamnese})
    except Exception as e:
        logger.error(f"Erro na anamnese: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 17. Prontuário do Cliente (GET/PUT/POST)
@bp.route('/api/clientes/<id>/prontuario', methods=['GET', 'PUT', 'POST'])
@login_required
def prontuario_cliente(id):
    db = get_db()
    """Gerencia prontuário do cliente"""
    try:
        # Tentar buscar por ObjectId primeiro, se falhar buscar por CPF
        try:
            cliente = db.clientes.find_one({'_id': ObjectId(id)})
        except:
            # Se não for um ObjectId válido, buscar por CPF
            cliente = db.clientes.find_one({'cpf': id})

        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        if request.method == 'GET':
            prontuario = cliente.get('prontuario', [])
            # Ordenar por data (mais recente primeiro)
            prontuario_ordenado = sorted(prontuario, key=lambda x: x.get('data', ''), reverse=True)

            return jsonify({
                'success': True,
                'prontuario': prontuario_ordenado,
                'cliente': {
                    'nome': cliente.get('nome'),
                    'cpf': cliente.get('cpf')
                }
            })

        elif request.method in ['PUT', 'POST']:
            data = request.get_json()

            # Criar novo registro de prontuário
            novo_registro = {
                'data': datetime.now().isoformat(),
                'procedimento': data.get('procedimento', ''),
                'observacoes': data.get('observacoes', ''),
                'profissional': data.get('profissional', session.get('user', {}).get('nome', 'Sistema')),
                'documento_url': data.get('documento_url', None)  # URL do documento escaneado se houver
            }

            # Adicionar novo registro ao prontuário
            db.clientes.update_one(
                {'_id': cliente['_id']},
                {'$push': {'prontuario': novo_registro}}
            )

            return jsonify({
                'success': True,
                'message': 'Registro adicionado ao prontuário',
                'registro': novo_registro
            })
    except Exception as e:
        logger.error(f"Erro no prontuário: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 17.1 Upload de Documento Físico - Anamnese
@bp.route('/api/clientes/<id>/anamnese/upload', methods=['POST'])
@login_required
def upload_anamnese_documento(id):
    """Upload de foto/documento físico de anamnese"""
    db = get_db()
    try:
        # Buscar cliente
        try:
            cliente = db.clientes.find_one({'_id': ObjectId(id)})
        except:
            cliente = db.clientes.find_one({'cpf': id})

        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # Verificar se há arquivo
        if 'documento' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['documento']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Arquivo vazio'}), 400

        if file and allowed_file(file.filename):
            import base64

            # Ler arquivo em memória
            file_content = file.read()
            doc_base64 = base64.b64encode(file_content).decode('utf-8')

            # Determinar tipo MIME
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
            mime_types = {
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'webp': 'image/webp',
                'pdf': 'application/pdf'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            # Montar data URI
            data_uri = f"data:{mime_type};base64,{doc_base64}"

            # Criar registro de documento
            documento = {
                'tipo': 'anamnese_fisica',
                'cliente_id': str(cliente['_id']),
                'cliente_cpf': cliente.get('cpf'),
                'cliente_nome': cliente.get('nome'),
                'data_upload': datetime.now().isoformat(),
                'data_uri': data_uri,
                'mime_type': mime_type,
                'filename': secure_filename(file.filename),
                'uploaded_by': session.get('user', {}).get('nome', 'Sistema')
            }

            # Salvar no array de documentos do cliente
            db.clientes.update_one(
                {'_id': cliente['_id']},
                {
                    '$push': {'documentos_anamnese': documento},
                    '$set': {'anamnese_fisica_disponivel': True}
                }
            )

            logger.info(f"✅ Documento de anamnese física uploaded para cliente {cliente.get('nome')}")

            return jsonify({
                'success': True,
                'message': 'Documento enviado com sucesso',
                'documento': documento
            })

        return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido. Use PNG, JPG ou PDF'}), 400

    except Exception as e:
        logger.error(f"❌ Erro ao fazer upload de anamnese: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 17.2 Upload de Documento Físico - Prontuário
@bp.route('/api/clientes/<id>/prontuario/upload', methods=['POST'])
@login_required
def upload_prontuario_documento(id):
    """Upload de foto/documento físico de prontuário"""
    db = get_db()
    try:
        # Buscar cliente
        try:
            cliente = db.clientes.find_one({'_id': ObjectId(id)})
        except:
            cliente = db.clientes.find_one({'cpf': id})

        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # Verificar se há arquivo
        if 'documento' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['documento']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Arquivo vazio'}), 400

        if file and allowed_file(file.filename):
            import base64

            # Ler arquivo em memória
            file_content = file.read()
            doc_base64 = base64.b64encode(file_content).decode('utf-8')

            # Determinar tipo MIME
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
            mime_types = {
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'webp': 'image/webp',
                'pdf': 'application/pdf'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            # Montar data URI
            data_uri = f"data:{mime_type};base64,{doc_base64}"

            # Obter observações do form (opcional)
            observacoes = request.form.get('observacoes', '')

            # Criar registro de documento
            documento = {
                'tipo': 'prontuario_fisico',
                'cliente_id': str(cliente['_id']),
                'cliente_cpf': cliente.get('cpf'),
                'cliente_nome': cliente.get('nome'),
                'data_upload': datetime.now().isoformat(),
                'data_uri': data_uri,
                'mime_type': mime_type,
                'filename': secure_filename(file.filename),
                'uploaded_by': session.get('user', {}).get('nome', 'Sistema'),
                'observacoes': observacoes
            }

            # Salvar no array de documentos do cliente
            db.clientes.update_one(
                {'_id': cliente['_id']},
                {
                    '$push': {'documentos_prontuario': documento},
                    '$set': {'prontuario_fisico_disponivel': True}
                }
            )

            logger.info(f"✅ Documento de prontuário físico uploaded para cliente {cliente.get('nome')}")

            return jsonify({
                'success': True,
                'message': 'Documento enviado com sucesso',
                'documento': documento
            })

        return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido. Use PNG, JPG ou PDF'}), 400

    except Exception as e:
        logger.error(f"❌ Erro ao fazer upload de prontuário: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 17.3 Listar Documentos - Anamnese
@bp.route('/api/clientes/<id>/anamnese/documentos', methods=['GET'])
@login_required
def listar_documentos_anamnese(id):
    """Lista todos os documentos físicos de anamnese do cliente"""
    db = get_db()
    try:
        try:
            cliente = db.clientes.find_one({'_id': ObjectId(id)})
        except:
            cliente = db.clientes.find_one({'cpf': id})

        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        documentos = cliente.get('documentos_anamnese', [])

        return jsonify({
            'success': True,
            'documentos': documentos,
            'cliente': {
                'nome': cliente.get('nome'),
                'cpf': cliente.get('cpf')
            }
        })

    except Exception as e:
        logger.error(f"❌ Erro ao listar documentos de anamnese: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 17.4 Listar Documentos - Prontuário
@bp.route('/api/clientes/<id>/prontuario/documentos', methods=['GET'])
@login_required
def listar_documentos_prontuario(id):
    """Lista todos os documentos físicos de prontuário do cliente"""
    db = get_db()
    try:
        try:
            cliente = db.clientes.find_one({'_id': ObjectId(id)})
        except:
            cliente = db.clientes.find_one({'cpf': id})

        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        documentos = cliente.get('documentos_prontuario', [])

        return jsonify({
            'success': True,
            'documentos': documentos,
            'cliente': {
                'nome': cliente.get('nome'),
                'cpf': cliente.get('cpf')
            }
        })

    except Exception as e:
        logger.error(f"❌ Erro ao listar documentos de prontuário: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# 18. Gerar PDF Resumo do Cliente
@bp.route('/api/clientes/<id>/resumo-pdf', methods=['GET'])
@login_required
def resumo_pdf_cliente(id):
    db = get_db()
    """Gera PDF completo com todos os dados do cliente"""
    try:
        cliente = db.clientes.find_one({'_id': ObjectId(id)})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
        
        # Criar PDF em memória
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=HexColor('#7C3AED'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph('RESUMO DO CLIENTE', title_style))
        elements.append(Spacer(1, 20))
        
        # Dados pessoais
        elements.append(Paragraph('<b>DADOS PESSOAIS</b>', styles['Heading2']))
        dados = [
            ['Nome:', cliente.get('nome', '-')],
            ['CPF:', cliente.get('cpf', '-')],
            ['Telefone:', cliente.get('telefone', '-')],
            ['Email:', cliente.get('email', '-')],
        ]
        t = Table(dados, colWidths=[4*cm, 12*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, HexColor('#E5E7EB'))
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        # Anamnese
        anamnese = cliente.get('anamnese', {})
        if anamnese:
            elements.append(Paragraph('<b>ANAMNESE</b>', styles['Heading2']))
            for key, value in anamnese.items():
                elements.append(Paragraph(f'<b>{key}:</b> {value}', styles['Normal']))
            elements.append(Spacer(1, 20))
        
        # Prontuário
        prontuario = cliente.get('prontuario', [])
        if prontuario:
            elements.append(Paragraph('<b>PRONTUÁRIO</b>', styles['Heading2']))
            for registro in prontuario:
                elements.append(Paragraph(
                    f"<b>Data:</b> {registro.get('data', '-')} | "
                    f"<b>Procedimento:</b> {registro.get('procedimento', '-')}<br/>"
                    f"<b>Observações:</b> {registro.get('observacoes', '-')}",
                    styles['Normal']
                ))
                elements.append(Spacer(1, 10))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'cliente_{cliente.get("nome", "resumo")}.pdf'
        )
    except Exception as e:
        logger.error(f"Erro ao gerar PDF: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== SISTEMA DE ANAMNESE E PRONTUÁRIO ====================

@bp.route('/api/clientes/<cpf>/anamnese', methods=['GET', 'POST'])
@login_required
def handle_anamnese(cpf):
    db = get_db()
    """Gerenciar histórico de anamneses de um cliente"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        # Buscar cliente
        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
        
        if request.method == 'GET':
            # Listar histórico de anamneses
            anamneses = list(db.anamneses.find({'cliente_cpf': cpf}).sort('data_cadastro', DESCENDING))
            
            for anamnese in anamneses:
                anamnese['_id'] = str(anamnese['_id'])
                if 'data_cadastro' in anamnese and isinstance(anamnese['data_cadastro'], datetime):
                    anamnese['data_cadastro'] = anamnese['data_cadastro'].isoformat()
            
            logger.info(f"📋 Listando {len(anamneses)} anamneses do cliente {cpf}")
            return jsonify({
                'success': True,
                'cliente': {
                    'cpf': cliente.get('cpf'),
                    'nome': cliente.get('nome')
                },
                'anamneses': anamneses
            })
        
        elif request.method == 'POST':
            # Criar nova anamnese
            data = request.json
            logger.info(f"📝 Criando anamnese para cliente {cpf}")
            
            anamnese = {
                'cliente_cpf': cpf,
                'cliente_nome': cliente.get('nome'),
                'respostas': data.get('respostas', {}),
                'observacoes': data.get('observacoes', ''),
                'data_cadastro': datetime.now(),
                'cadastrado_por': session.get('username'),
                'versao': db.anamneses.count_documents({'cliente_cpf': cpf}) + 1
            }
            
            result = db.anamneses.insert_one(anamnese)
            anamnese['_id'] = str(result.inserted_id)
            anamnese['data_cadastro'] = anamnese['data_cadastro'].isoformat()
            
            # Atualizar campo anamnese_atualizada no cliente
            db.clientes.update_one(
                {'cpf': cpf},
                {'$set': {'anamnese_atualizada': datetime.now()}}
            )
            
            logger.info(f"✅ Anamnese v{anamnese['versao']} criada para {cpf}")
            return jsonify({'success': True, 'anamnese': anamnese})
            
    except Exception as e:
        logger.error(f"❌ Erro em handle_anamnese: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes/<cpf>/anamnese/<id>', methods=['GET', 'DELETE'])
@login_required
def handle_anamnese_by_id(cpf, id):
    db = get_db()
    """Gerenciar anamnese específica"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        if request.method == 'GET':
            anamnese = db.anamneses.find_one({'_id': ObjectId(id), 'cliente_cpf': cpf})
            if not anamnese:
                return jsonify({'success': False, 'message': 'Anamnese não encontrada'}), 404
            
            anamnese['_id'] = str(anamnese['_id'])
            if 'data_cadastro' in anamnese and isinstance(anamnese['data_cadastro'], datetime):
                anamnese['data_cadastro'] = anamnese['data_cadastro'].isoformat()
            
            return jsonify({'success': True, 'anamnese': anamnese})
        
        elif request.method == 'DELETE':
            result = db.anamneses.delete_one({'_id': ObjectId(id), 'cliente_cpf': cpf})
            if result.deleted_count == 0:
                return jsonify({'success': False, 'message': 'Anamnese não encontrada'}), 404
            
            logger.info(f"🗑️ Anamnese {id} deletada")
            return jsonify({'success': True, 'message': 'Anamnese deletada com sucesso'})
            
    except Exception as e:
        logger.error(f"❌ Erro em handle_anamnese_by_id: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes/<cpf>/prontuario', methods=['GET', 'POST'])
@login_required
def handle_prontuario(cpf):
    db = get_db()
    """Gerenciar histórico de prontuários de um cliente"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        # Buscar cliente
        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
        
        if request.method == 'GET':
            # Listar histórico de prontuários
            prontuarios = list(db.prontuarios.find({'cliente_cpf': cpf}).sort('data_atendimento', DESCENDING))
            
            for pront in prontuarios:
                pront['_id'] = str(pront['_id'])
                if 'data_atendimento' in pront and isinstance(pront['data_atendimento'], datetime):
                    pront['data_atendimento'] = pront['data_atendimento'].isoformat()
                if 'data_cadastro' in pront and isinstance(pront['data_cadastro'], datetime):
                    pront['data_cadastro'] = pront['data_cadastro'].isoformat()
            
            logger.info(f"📋 Listando {len(prontuarios)} prontuários do cliente {cpf}")
            return jsonify({
                'success': True,
                'cliente': {
                    'cpf': cliente.get('cpf'),
                    'nome': cliente.get('nome')
                },
                'prontuarios': prontuarios
            })
        
        elif request.method == 'POST':
            # Criar novo prontuário
            data = request.json
            logger.info(f"📝 Criando prontuário para cliente {cpf}")
            
            prontuario = {
                'cliente_cpf': cpf,
                'cliente_nome': cliente.get('nome'),
                'data_atendimento': datetime.fromisoformat(data.get('data_atendimento', datetime.now().isoformat())),
                'profissional': data.get('profissional', ''),
                'procedimento': data.get('procedimento', ''),
                'produtos_utilizados': data.get('produtos_utilizados', []),
                'observacoes': data.get('observacoes', ''),
                'fotos_antes': data.get('fotos_antes', []),
                'fotos_depois': data.get('fotos_depois', []),
                'proxima_sessao': data.get('proxima_sessao', ''),
                'data_cadastro': datetime.now(),
                'cadastrado_por': session.get('username')
            }
            
            result = db.prontuarios.insert_one(prontuario)
            prontuario['_id'] = str(result.inserted_id)
            prontuario['data_atendimento'] = prontuario['data_atendimento'].isoformat()
            prontuario['data_cadastro'] = prontuario['data_cadastro'].isoformat()
            
            # Atualizar campo prontuario_atualizado no cliente
            db.clientes.update_one(
                {'cpf': cpf},
                {'$set': {'prontuario_atualizado': datetime.now()}}
            )
            
            logger.info(f"✅ Prontuário criado para {cpf}")
            return jsonify({'success': True, 'prontuario': prontuario})
            
    except Exception as e:
        logger.error(f"❌ Erro em handle_prontuario: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes/<cpf>/prontuario/<id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def handle_prontuario_by_id(cpf, id):
    db = get_db()
    """Gerenciar prontuário específico"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        if request.method == 'GET':
            prontuario = db.prontuarios.find_one({'_id': ObjectId(id), 'cliente_cpf': cpf})
            if not prontuario:
                return jsonify({'success': False, 'message': 'Prontuário não encontrado'}), 404
            
            prontuario['_id'] = str(prontuario['_id'])
            if 'data_atendimento' in prontuario and isinstance(prontuario['data_atendimento'], datetime):
                prontuario['data_atendimento'] = prontuario['data_atendimento'].isoformat()
            if 'data_cadastro' in prontuario and isinstance(prontuario['data_cadastro'], datetime):
                prontuario['data_cadastro'] = prontuario['data_cadastro'].isoformat()
            
            return jsonify({'success': True, 'prontuario': prontuario})
        
        elif request.method == 'PUT':
            data = request.json
            logger.info(f"📝 Atualizando prontuário {id}")
            
            update_data = {
                'data_atendimento': datetime.fromisoformat(data.get('data_atendimento', datetime.now().isoformat())),
                'profissional': data.get('profissional', ''),
                'procedimento': data.get('procedimento', ''),
                'produtos_utilizados': data.get('produtos_utilizados', []),
                'observacoes': data.get('observacoes', ''),
                'fotos_antes': data.get('fotos_antes', []),
                'fotos_depois': data.get('fotos_depois', []),
                'proxima_sessao': data.get('proxima_sessao', ''),
                'atualizado_em': datetime.now(),
                'atualizado_por': session.get('username')
            }
            
            result = db.prontuarios.update_one(
                {'_id': ObjectId(id), 'cliente_cpf': cpf},
                {'$set': update_data}
            )
            
            if result.matched_count == 0:
                return jsonify({'success': False, 'message': 'Prontuário não encontrado'}), 404
            
            logger.info(f"✅ Prontuário {id} atualizado")
            return jsonify({'success': True, 'message': 'Prontuário atualizado com sucesso'})
        
        elif request.method == 'DELETE':
            result = db.prontuarios.delete_one({'_id': ObjectId(id), 'cliente_cpf': cpf})
            if result.deleted_count == 0:
                return jsonify({'success': False, 'message': 'Prontuário não encontrado'}), 404
            
            logger.info(f"🗑️ Prontuário {id} deletado")
            return jsonify({'success': True, 'message': 'Prontuário deletado com sucesso'})
            
    except Exception as e:
        logger.error(f"❌ Erro em handle_prontuario_by_id: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== HISTÓRICO COMPLETO ANAMNESE/PRONTUÁRIO (Diretriz #21) ====================

@bp.route('/api/clientes/<cpf>/historico-completo', methods=['GET'])
@login_required
def historico_completo_cliente(cpf):
    db = get_db()
    """Obter histórico completo de anamnese e prontuários de um cliente (COM PAGINAÇÃO - Roadmap #11)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        # Parâmetros de paginação (Roadmap Section V - Clientes #11)
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))  # 25 itens por página por padrão
        skip = (page - 1) * limit

        # Buscar cliente
        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # PAGINAÇÃO: Buscar anamneses com limit e skip
        anamneses_query = db.anamneses.find({'cliente_cpf': cpf}).sort('data_cadastro', DESCENDING)
        total_anamneses = db.anamneses.count_documents({'cliente_cpf': cpf})
        anamneses_paginadas = list(anamneses_query.skip(skip).limit(limit))

        # PAGINAÇÃO: Buscar prontuários com limit e skip
        prontuarios_query = db.prontuarios.find({'cliente_cpf': cpf}).sort('data_atendimento', DESCENDING)
        total_prontuarios = db.prontuarios.count_documents({'cliente_cpf': cpf})
        prontuarios_paginados = list(prontuarios_query.skip(skip).limit(limit))

        # PAGINAÇÃO: Buscar orçamentos/contratos com limit e skip
        orcamentos_query = db.orcamentos.find({'cliente_cpf': cpf}).sort('created_at', DESCENDING)
        total_orcamentos = db.orcamentos.count_documents({'cliente_cpf': cpf})
        orcamentos_paginados = list(orcamentos_query.skip(skip).limit(limit))

        # Calcular estatísticas (usa totais, não paginados)
        total_atendimentos = total_prontuarios
        total_gasto = sum(
            o.get('total_final', 0)
            for o in db.orcamentos.find({'cliente_cpf': cpf, 'status': 'Aprovado'})
        )

        # Buscar último atendimento (sempre o primeiro resultado ordenado)
        ultimo_prontuario = db.prontuarios.find_one(
            {'cliente_cpf': cpf},
            sort=[('data_atendimento', DESCENDING)]
        )
        ultimo_atendimento = ultimo_prontuario.get('data_atendimento') if ultimo_prontuario else None

        historico = {
            'cliente': {
                'nome': cliente.get('nome'),
                'cpf': cliente.get('cpf'),
                'telefone': cliente.get('telefone'),
                'email': cliente.get('email'),
                'data_nascimento': cliente.get('data_nascimento')
            },
            'estatisticas': {
                'total_atendimentos': total_atendimentos,
                'total_faturado': total_gasto,
                'total_orcamentos': total_orcamentos,
                'ultimo_atendimento': ultimo_atendimento.isoformat() if isinstance(ultimo_atendimento, datetime) else ultimo_atendimento
            },
            'anamneses': convert_objectid(anamneses_paginadas),
            'prontuarios': convert_objectid(prontuarios_paginados),
            'orcamentos': convert_objectid(orcamentos_paginados),
            'paginacao': {
                'page': page,
                'limit': limit,
                'total_anamneses': total_anamneses,
                'total_prontuarios': total_prontuarios,
                'total_orcamentos': total_orcamentos,
                'total_paginas_anamneses': (total_anamneses + limit - 1) // limit,
                'total_paginas_prontuarios': (total_prontuarios + limit - 1) // limit,
                'total_paginas_orcamentos': (total_orcamentos + limit - 1) // limit,
                'tem_proxima': skip + limit < max(total_anamneses, total_prontuarios, total_orcamentos),
                'tem_anterior': page > 1
            }
        }

        return jsonify({'success': True, 'historico': historico})

    except Exception as e:
        logger.error(f"Erro ao buscar histórico completo: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/clientes/<cpf>/prontuario/<id>/pdf', methods=['GET'])
@login_required
def gerar_pdf_prontuario(cpf, id):
    db = get_db()
    """Gerar PDF de prontuário para impressão (Diretriz #21)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        prontuario = db.prontuarios.find_one({'_id': ObjectId(id), 'cliente_cpf': cpf})
        if not prontuario:
            return jsonify({'success': False, 'message': 'Prontuário não encontrado'}), 404

        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER)
        elements.append(Paragraph("PRONTUÁRIO DE ATENDIMENTO", title_style))
        elements.append(Spacer(1, 0.3*cm))

        # Dados do cliente
        elements.append(Paragraph('<b>DADOS DO CLIENTE</b>', styles['Heading2']))
        cliente_data = [
            ['Nome:', cliente.get('nome', 'N/A')],
            ['CPF:', cliente.get('cpf', 'N/A')],
            ['Telefone:', cliente.get('telefone', 'N/A')],
            ['Email:', cliente.get('email', 'N/A')],
            ['Data Nasc.:', cliente.get('data_nascimento', 'N/A')]
        ]
        cliente_table = Table(cliente_data, colWidths=[4*cm, 12*cm])
        cliente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(cliente_table)
        elements.append(Spacer(1, 0.5*cm))

        # Dados do atendimento
        elements.append(Paragraph('<b>DADOS DO ATENDIMENTO</b>', styles['Heading2']))
        data_atendimento = prontuario.get('data_atendimento')
        if isinstance(data_atendimento, datetime):
            data_atendimento = data_atendimento.strftime('%d/%m/%Y %H:%M')

        atendimento_data = [
            ['Data:', str(data_atendimento) if data_atendimento else 'N/A'],
            ['Profissional:', prontuario.get('profissional', 'N/A')],
            ['Procedimento:', prontuario.get('procedimento', 'N/A')],
            ['Próxima Sessão:', prontuario.get('proxima_sessao', 'N/A')]
        ]
        atendimento_table = Table(atendimento_data, colWidths=[4*cm, 12*cm])
        atendimento_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#e3f2fd')),
            ('GRID', (0, 0), (-1, -1), 0.5, black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(atendimento_table)
        elements.append(Spacer(1, 0.5*cm))

        # Produtos utilizados
        if prontuario.get('produtos_utilizados'):
            elements.append(Paragraph('<b>PRODUTOS UTILIZADOS</b>', styles['Heading3']))
            produtos_list = prontuario['produtos_utilizados']
            if isinstance(produtos_list, list):
                for produto in produtos_list:
                    elements.append(Paragraph(f"• {produto}", styles['Normal']))
            elements.append(Spacer(1, 0.3*cm))

        # Observações
        if prontuario.get('observacoes'):
            elements.append(Paragraph('<b>OBSERVAÇÕES</b>', styles['Heading3']))
            elements.append(Paragraph(prontuario['observacoes'], styles['Normal']))
            elements.append(Spacer(1, 0.3*cm))

        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)
        elements.append(Paragraph('Este documento é confidencial e destinado exclusivamente ao cliente.', footer_style))
        elements.append(Paragraph(f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}', footer_style))

        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"prontuario_{cliente.get('nome', 'cliente')}_{id}.pdf"
        )

    except Exception as e:
        logger.error(f"Erro ao gerar PDF de prontuário: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/clientes/<cpf>/historico-completo/pdf', methods=['GET'])
@login_required
def gerar_pdf_historico_completo(cpf):
    db = get_db()
    """Gerar PDF do histórico completo do cliente (Diretriz #21)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        prontuarios = list(db.prontuarios.find({'cliente_cpf': cpf}).sort('data_atendimento', DESCENDING))
        orcamentos = list(db.orcamentos.find({'cliente_cpf': cpf, 'status': 'Aprovado'}).sort('created_at', DESCENDING))

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER)
        elements.append(Paragraph(f"HISTÓRICO COMPLETO - {cliente.get('nome', 'Cliente')}", title_style))
        elements.append(Spacer(1, 0.5*cm))

        # Estatísticas
        total_atendimentos = len(prontuarios)
        total_faturado = sum(o.get('total_final', 0) for o in orcamentos)

        estatisticas_data = [
            ['Total de Atendimentos:', str(total_atendimentos)],
            ['Total Faturado:', f"R$ {total_faturado:.2f}"],
            ['Última Visita:', prontuarios[0].get('data_atendimento').strftime('%d/%m/%Y') if prontuarios and isinstance(prontuarios[0].get('data_atendimento'), datetime) else 'N/A']
        ]
        estatisticas_table = Table(estatisticas_data, colWidths=[8*cm, 8*cm])
        estatisticas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#e8f5e9')),
            ('GRID', (0, 0), (-1, -1), 0.5, black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(estatisticas_table)
        elements.append(Spacer(1, 0.5*cm))

        # Histórico de prontuários
        if prontuarios:
            elements.append(Paragraph('<b>HISTÓRICO DE ATENDIMENTOS</b>', styles['Heading2']))
            for prontuario in prontuarios[:10]:  # Limitar aos 10 mais recentes
                data_atend = prontuario.get('data_atendimento')
                if isinstance(data_atend, datetime):
                    data_atend = data_atend.strftime('%d/%m/%Y')

                elements.append(Paragraph(f"<b>{data_atend}</b> - {prontuario.get('procedimento', 'Procedimento não especificado')}", styles['Normal']))
                if prontuario.get('profissional'):
                    elements.append(Paragraph(f"Profissional: {prontuario.get('profissional')}", styles['Normal']))
                if prontuario.get('observacoes'):
                    elements.append(Paragraph(f"Obs: {prontuario.get('observacoes')[:150]}...", styles['Normal']))
                elements.append(Spacer(1, 0.2*cm))

        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"historico_completo_{cliente.get('nome', 'cliente')}.pdf"
        )

    except Exception as e:
        logger.error(f"Erro ao gerar PDF de histórico completo: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/clientes/<cpf>/historico-completo/whatsapp', methods=['GET'])
@login_required
def enviar_whatsapp_historico(cpf):
    db = get_db()
    """Gerar link WhatsApp com resumo do histórico (Diretriz #21)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        cliente = db.clientes.find_one({'cpf': cpf})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        prontuarios = list(db.prontuarios.find({'cliente_cpf': cpf}).sort('data_atendimento', DESCENDING).limit(5))
        orcamentos = list(db.orcamentos.find({'cliente_cpf': cpf, 'status': 'Aprovado'}).sort('created_at', DESCENDING))

        telefone = cliente.get('telefone', '').replace('(', '').replace(')', '').replace('-', '').replace(' ', '')

        # Montar mensagem
        mensagem = f"""🌿 *BIOMA - Histórico de Atendimentos*

👤 *Cliente:* {cliente.get('nome', 'N/A')}
📊 *Total de Atendimentos:* {len(prontuarios)}
💰 *Total Investido:* R$ {sum(o.get('total_final', 0) for o in orcamentos):.2f}

📋 *ÚLTIMOS ATENDIMENTOS:*
"""

        for i, prontuario in enumerate(prontuarios, 1):
            data = prontuario.get('data_atendimento')
            if isinstance(data, datetime):
                data = data.strftime('%d/%m/%Y')
            procedimento = prontuario.get('procedimento', 'Procedimento')
            mensagem += f"{i}. {data} - {procedimento}\n"

        mensagem += "\nObrigado por confiar na BIOMA! 🌿"

        # URL encode
        mensagem_encoded = urllib.parse.quote(mensagem)
        whatsapp_url = f"https://wa.me/55{telefone}?text={mensagem_encoded}"

        return jsonify({
            'success': True,
            'whatsapp_url': whatsapp_url,
            'telefone': telefone
        })

    except Exception as e:
        logger.error(f"Erro ao gerar link WhatsApp: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== SISTEMA DE MULTICOMISSÕES ====================

@bp.route('/api/orcamentos', methods=['GET', 'POST'])
@login_required
def handle_orcamentos():
    db = get_db()
    """Gerenciar orçamentos com suporte a múltiplos profissionais"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        if request.method == 'GET':
            # Listar orçamentos
            orcamentos = list(db.orcamentos.find().sort('created_at', DESCENDING).limit(100))
            for orc in orcamentos:
                orc['_id'] = str(orc['_id'])
                if 'created_at' in orc and isinstance(orc['created_at'], datetime):
                    orc['created_at'] = orc['created_at'].isoformat()
                # Converter ObjectIds dos profissionais vinculados
                if 'profissionais_vinculados' in orc:
                    for prof in orc['profissionais_vinculados']:
                        if 'profissional_id' in prof and isinstance(prof['profissional_id'], ObjectId):
                            prof['profissional_id'] = str(prof['profissional_id'])
            
            logger.info(f"📊 Listando {len(orcamentos)} orçamentos")
            return jsonify({'success': True, 'orcamentos': orcamentos})
        
        elif request.method == 'POST':
            # Criar novo orçamento
            data = request.json
            logger.info(f"💼 Criando orçamento para cliente: {data.get('cliente_nome')}")
            
            # Processar profissionais vinculados
            profissionais_vinculados = data.get('profissionais_vinculados', [])
            for prof in profissionais_vinculados:
                if 'profissional_id' in prof and isinstance(prof['profissional_id'], str):
                    try:
                        prof['profissional_id'] = ObjectId(prof['profissional_id'])
                    except:
                        pass
            
            # Gerar número do orçamento
            ultimo_orc = db.orcamentos.find_one(sort=[('numero', DESCENDING)])
            proximo_numero = (ultimo_orc.get('numero', 0) + 1) if ultimo_orc else 1
            
            orcamento = {
                'numero': proximo_numero,
                'cliente_cpf': data.get('cliente_cpf'),
                'cliente_nome': data.get('cliente_nome'),
                'cliente_telefone': data.get('cliente_telefone'),
                'cliente_email': data.get('cliente_email'),
                'servicos': data.get('servicos', []),
                'produtos': data.get('produtos', []),
                'profissionais_vinculados': profissionais_vinculados,
                'total_servicos': data.get('total_servicos', 0),
                'total_produtos': data.get('total_produtos', 0),
                'desconto_perc': data.get('desconto_perc', 0),
                'desconto_valor': data.get('desconto_valor', 0),
                'total_final': data.get('total_final', 0),
                'total_comissoes': sum(p.get('comissao_valor', 0) for p in profissionais_vinculados),
                'forma_pagamento': data.get('forma_pagamento'),
                'observacoes': data.get('observacoes', ''),
                'status': data.get('status', 'Pendente'),
                'created_at': datetime.now(),
                'created_by': session.get('username')
            }
            
            result = db.orcamentos.insert_one(orcamento)
            orcamento['_id'] = str(result.inserted_id)

            # Update cliente denormalized fields for performance
            if orcamento.get('cliente_cpf'):
                update_cliente_denormalized_fields(orcamento['cliente_cpf'])

            # Registrar comissões no histórico
            if profissionais_vinculados:
                for prof in profissionais_vinculados:
                    db.comissoes_historico.insert_one({
                        'orcamento_id': result.inserted_id,
                        'profissional_id': prof.get('profissional_id'),
                        'nome': prof.get('nome'),
                        'tipo': prof.get('tipo'),
                        'comissao_perc': prof.get('comissao_perc'),
                        'comissao_valor': prof.get('comissao_valor'),
                        'valor_base': data.get('total_servicos', 0),
                        'status_orcamento': orcamento['status'],
                        'data_registro': datetime.now()
                    })
            
            logger.info(f"✅ Orçamento #{proximo_numero} criado com {len(profissionais_vinculados)} profissionais")
            return jsonify({'success': True, 'orcamento': orcamento, 'numero': proximo_numero})
            
    except Exception as e:
        logger.error(f"❌ Erro em handle_orcamentos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/orcamentos/<id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def handle_orcamento_by_id(id):
    db = get_db()
    """Gerenciar orçamento específico"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        if request.method == 'GET':
            orcamento = db.orcamentos.find_one({'_id': ObjectId(id)})
            if not orcamento:
                return jsonify({'success': False, 'message': 'Orçamento não encontrado'}), 404
            
            orcamento['_id'] = str(orcamento['_id'])
            if 'created_at' in orcamento and isinstance(orcamento['created_at'], datetime):
                orcamento['created_at'] = orcamento['created_at'].isoformat()
            
            return jsonify({'success': True, 'orcamento': orcamento})
        
        elif request.method == 'PUT':
            data = request.json
            logger.info(f"📝 Atualizando orçamento {id}")
            
            # Processar profissionais vinculados
            profissionais_vinculados = data.get('profissionais_vinculados', [])
            for prof in profissionais_vinculados:
                if 'profissional_id' in prof and isinstance(prof['profissional_id'], str):
                    try:
                        prof['profissional_id'] = ObjectId(prof['profissional_id'])
                    except:
                        pass
            
            update_data = {
                'cliente_cpf': data.get('cliente_cpf'),
                'cliente_nome': data.get('cliente_nome'),
                'cliente_telefone': data.get('cliente_telefone'),
                'cliente_email': data.get('cliente_email'),
                'servicos': data.get('servicos', []),
                'produtos': data.get('produtos', []),
                'profissionais_vinculados': profissionais_vinculados,
                'total_servicos': data.get('total_servicos', 0),
                'total_produtos': data.get('total_produtos', 0),
                'desconto_perc': data.get('desconto_perc', 0),
                'desconto_valor': data.get('desconto_valor', 0),
                'total_final': data.get('total_final', 0),
                'total_comissoes': sum(p.get('comissao_valor', 0) for p in profissionais_vinculados),
                'forma_pagamento': data.get('forma_pagamento'),
                'observacoes': data.get('observacoes', ''),
                'status': data.get('status', 'Pendente'),
                'updated_at': datetime.now(),
                'updated_by': session.get('username')
            }
            
            db.orcamentos.update_one({'_id': ObjectId(id)}, {'$set': update_data})

            # Update cliente denormalized fields for performance
            if update_data.get('cliente_cpf'):
                update_cliente_denormalized_fields(update_data['cliente_cpf'])

            # Atualizar histórico de comissões
            db.comissoes_historico.delete_many({'orcamento_id': ObjectId(id)})
            if profissionais_vinculados:
                for prof in profissionais_vinculados:
                    db.comissoes_historico.insert_one({
                        'orcamento_id': ObjectId(id),
                        'profissional_id': prof.get('profissional_id'),
                        'nome': prof.get('nome'),
                        'tipo': prof.get('tipo'),
                        'comissao_perc': prof.get('comissao_perc'),
                        'comissao_valor': prof.get('comissao_valor'),
                        'valor_base': data.get('total_servicos', 0),
                        'status_orcamento': update_data['status'],
                        'data_registro': datetime.now()
                    })
            
            logger.info(f"✅ Orçamento {id} atualizado")
            return jsonify({'success': True, 'message': 'Orçamento atualizado com sucesso'})
        
        elif request.method == 'DELETE':
            # Get cliente_cpf before deleting for denormalized field update
            orcamento = db.orcamentos.find_one({'_id': ObjectId(id)}, {'cliente_cpf': 1})
            cliente_cpf = orcamento.get('cliente_cpf') if orcamento else None

            db.orcamentos.delete_one({'_id': ObjectId(id)})
            db.comissoes_historico.delete_many({'orcamento_id': ObjectId(id)})

            # Update cliente denormalized fields after deletion
            if cliente_cpf:
                update_cliente_denormalized_fields(cliente_cpf)

            logger.info(f"🗑️ Orçamento {id} deletado")
            return jsonify({'success': True, 'message': 'Orçamento deletado com sucesso'})
            
    except Exception as e:
        logger.error(f"❌ Erro em handle_orcamento_by_id: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/orcamentos/<id>/pdf', methods=['GET'])
@login_required
def gerar_pdf_orcamento(id):
    db = get_db()
    """Gerar PDF de orçamento para impressão (Diretriz #3)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        orcamento = db.orcamentos.find_one({'_id': ObjectId(id)})
        if not orcamento:
            return jsonify({'success': False, 'message': 'Orçamento não encontrado'}), 404

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER)
        elements.append(Paragraph(f"ORÇAMENTO #{orcamento.get('numero', str(orcamento['_id'])[-6:])}", title_style))
        elements.append(Spacer(1, 0.3*cm))

        # Informações do cliente
        elements.append(Paragraph('<b>DADOS DO CLIENTE</b>', styles['Heading2']))
        cliente_data = [
            ['Nome:', orcamento.get('cliente_nome', 'N/A')],
            ['CPF:', orcamento.get('cliente_cpf', 'N/A')],
            ['Telefone:', orcamento.get('cliente_telefone', 'N/A')],
            ['Email:', orcamento.get('cliente_email', 'N/A')],
            ['Data:', orcamento.get('created_at', datetime.now()).strftime('%d/%m/%Y') if isinstance(orcamento.get('created_at'), datetime) else 'N/A']
        ]
        cliente_table = Table(cliente_data, colWidths=[4*cm, 12*cm])
        cliente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(cliente_table)
        elements.append(Spacer(1, 0.5*cm))

        # Serviços
        if orcamento.get('servicos'):
            elements.append(Paragraph('<b>SERVIÇOS</b>', styles['Heading2']))
            servicos_data = [['Descrição', 'Quantidade', 'Valor Unit.', 'Total']]
            for srv in orcamento['servicos']:
                servicos_data.append([
                    srv.get('nome', 'N/A'),
                    str(srv.get('quantidade', 1)),
                    f"R$ {srv.get('preco', 0):.2f}",
                    f"R$ {srv.get('total', 0):.2f}"
                ])
            servicos_table = Table(servicos_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
            servicos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(servicos_table)
            elements.append(Spacer(1, 0.3*cm))

        # Produtos
        if orcamento.get('produtos'):
            elements.append(Paragraph('<b>PRODUTOS</b>', styles['Heading2']))
            produtos_data = [['Descrição', 'Quantidade', 'Valor Unit.', 'Total']]
            for prd in orcamento['produtos']:
                produtos_data.append([
                    prd.get('nome', 'N/A'),
                    str(prd.get('quantidade', 1)),
                    f"R$ {prd.get('preco', 0):.2f}",
                    f"R$ {prd.get('total', 0):.2f}"
                ])
            produtos_table = Table(produtos_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
            produtos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2196F3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(produtos_table)
            elements.append(Spacer(1, 0.3*cm))

        # Totais
        totais_data = [
            ['Subtotal Serviços:', f"R$ {orcamento.get('total_servicos', 0):.2f}"],
            ['Subtotal Produtos:', f"R$ {orcamento.get('total_produtos', 0):.2f}"],
            ['Desconto:', f"R$ {orcamento.get('desconto_valor', 0):.2f}"],
            ['<b>TOTAL FINAL:</b>', f"<b>R$ {orcamento.get('total_final', 0):.2f}</b>"]
        ]
        totais_table = Table(totais_data, colWidths=[12*cm, 4*cm])
        totais_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('#FFD700')),
        ]))
        elements.append(totais_table)

        # Observações
        if orcamento.get('observacoes'):
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph('<b>OBSERVAÇÕES:</b>', styles['Heading3']))
            elements.append(Paragraph(orcamento['observacoes'], styles['Normal']))

        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"orcamento_{orcamento.get('numero', id)}.pdf"
        )

    except Exception as e:
        logger.error(f"Erro ao gerar PDF de orçamento: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/orcamentos/<id>/whatsapp', methods=['GET'])
@login_required
def enviar_whatsapp_orcamento(id):
    db = get_db()
    """Gerar link do WhatsApp com orçamento (Diretriz #3)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        orcamento = db.orcamentos.find_one({'_id': ObjectId(id)})
        if not orcamento:
            return jsonify({'success': False, 'message': 'Orçamento não encontrado'}), 404

        # Montar mensagem do WhatsApp
        telefone = orcamento.get('cliente_telefone', '').replace('(', '').replace(')', '').replace('-', '').replace(' ', '')

        mensagem = f"""🌿 *BIOMA - Orçamento #{orcamento.get('numero', str(orcamento['_id'])[-6:])}*

👤 *Cliente:* {orcamento.get('cliente_nome', 'N/A')}
📅 *Data:* {orcamento.get('created_at', datetime.now()).strftime('%d/%m/%Y') if isinstance(orcamento.get('created_at'), datetime) else 'N/A'}

"""

        # Serviços
        if orcamento.get('servicos'):
            mensagem += "✨ *SERVIÇOS:*\n"
            for srv in orcamento['servicos']:
                mensagem += f"• {srv.get('nome', 'N/A')} - R$ {srv.get('total', 0):.2f}\n"
            mensagem += "\n"

        # Produtos
        if orcamento.get('produtos'):
            mensagem += "🛍️ *PRODUTOS:*\n"
            for prd in orcamento['produtos']:
                mensagem += f"• {prd.get('nome', 'N/A')} - R$ {prd.get('total', 0):.2f}\n"
            mensagem += "\n"

        # Totais
        mensagem += f"""💰 *VALORES:*
Serviços: R$ {orcamento.get('total_servicos', 0):.2f}
Produtos: R$ {orcamento.get('total_produtos', 0):.2f}
Desconto: R$ {orcamento.get('desconto_valor', 0):.2f}

✅ *TOTAL: R$ {orcamento.get('total_final', 0):.2f}*
"""

        if orcamento.get('observacoes'):
            mensagem += f"\n📝 *Observações:* {orcamento['observacoes']}"

        # URL encode da mensagem
        mensagem_encoded = urllib.parse.quote(mensagem)
        whatsapp_url = f"https://wa.me/55{telefone}?text={mensagem_encoded}"

        return jsonify({
            'success': True,
            'whatsapp_url': whatsapp_url,
            'telefone': telefone
        })

    except Exception as e:
        logger.error(f"Erro ao gerar link WhatsApp: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/orcamento/<id>/enviar-email', methods=['POST'])
@login_required
def enviar_orcamento_email(id):
    """Enviar orçamento por e-mail ao cliente"""
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Erro de conexão com banco de dados'}), 500

    try:
        # Buscar orçamento
        orcamento = db.orcamentos.find_one({'_id': ObjectId(id)})
        if not orcamento:
            return jsonify({'success': False, 'message': 'Orçamento não encontrado'}), 404

        # Buscar dados do cliente para obter email
        cliente_id = orcamento.get('cliente_id')
        if not cliente_id:
            return jsonify({'success': False, 'message': 'Cliente não identificado no orçamento'}), 400

        cliente = db.clientes.find_one({'_id': ObjectId(cliente_id)})
        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404

        email_cliente = cliente.get('email')
        if not email_cliente:
            return jsonify({'success': False, 'message': 'Cliente não possui e-mail cadastrado'}), 400

        # Montar dados do orçamento para o email
        numero_orcamento = orcamento.get('numero', str(orcamento['_id'])[-6:])
        cliente_nome = orcamento.get('cliente_nome', 'Cliente')
        total_final = orcamento.get('total_final', 0)
        servicos = orcamento.get('servicos', [])
        produtos = orcamento.get('produtos', [])

        # Verificar configuração MailerSend
        import os
        import requests

        mailersend_key = os.getenv('MAILERSEND_API_KEY')
        mailersend_from = os.getenv('MAILERSEND_FROM_EMAIL')
        mailersend_name = os.getenv('MAILERSEND_FROM_NAME', 'BIOMA Uberaba')

        if not mailersend_key or not mailersend_from:
            return jsonify({
                'success': False,
                'message': 'Configurações de e-mail não encontradas no servidor'
            }), 400

        logger.info(f"📧 Enviando orçamento #{numero_orcamento} para: {email_cliente}")

        # Montar HTML do orçamento
        servicos_html = ""
        for srv in servicos:
            servicos_html += f"""
                <tr>
                    <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;">{srv.get('nome', '')}</td>
                    <td style="padding: 10px; border-bottom: 1px solid #e5e7eb; text-align: center;">{srv.get('quantidade', 1)}</td>
                    <td style="padding: 10px; border-bottom: 1px solid #e5e7eb; text-align: right;">R$ {srv.get('preco', 0):.2f}</td>
                </tr>
            """

        produtos_html = ""
        for prd in produtos:
            produtos_html += f"""
                <tr>
                    <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;">{prd.get('nome', '')}</td>
                    <td style="padding: 10px; border-bottom: 1px solid #e5e7eb; text-align: center;">{prd.get('quantidade', 1)}</td>
                    <td style="padding: 10px; border-bottom: 1px solid #e5e7eb; text-align: right;">R$ {prd.get('preco', 0):.2f}</td>
                </tr>
            """

        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 700px; margin: 0 auto; padding: 20px; background-color: #f9fafb;">
            <div style="background: linear-gradient(135deg, #7C3AED, #EC4899); padding: 40px; border-radius: 15px; text-align: center; color: white; margin-bottom: 30px;">
                <h1 style="margin: 0 0 10px 0; font-size: 32px;">💼 Orçamento</h1>
                <p style="margin: 0; font-size: 18px; opacity: 0.95;">#{numero_orcamento}</p>
            </div>

            <div style="background: white; padding: 30px; border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                <h2 style="color: #7C3AED; margin-top: 0;">Olá, {cliente_nome}!</h2>
                <p style="color: #374151; line-height: 1.6;">
                    Segue o orçamento solicitado para os serviços e produtos da BIOMA Uberaba.
                </p>

                {"<h3 style='color: #374151; margin-top: 30px; margin-bottom: 15px;'>🛠️ Serviços</h3>" if servicos else ""}
                {"<table style='width: 100%; border-collapse: collapse; margin-bottom: 20px;'><thead><tr style='background-color: #f3f4f6;'><th style='padding: 12px; text-align: left; color: #6b7280;'>Serviço</th><th style='padding: 12px; text-align: center; color: #6b7280;'>Qtd</th><th style='padding: 12px; text-align: right; color: #6b7280;'>Valor</th></tr></thead><tbody>" + servicos_html + "</tbody></table>" if servicos else ""}

                {"<h3 style='color: #374151; margin-top: 30px; margin-bottom: 15px;'>📦 Produtos</h3>" if produtos else ""}
                {"<table style='width: 100%; border-collapse: collapse; margin-bottom: 20px;'><thead><tr style='background-color: #f3f4f6;'><th style='padding: 12px; text-align: left; color: #6b7280;'>Produto</th><th style='padding: 12px; text-align: center; color: #6b7280;'>Qtd</th><th style='padding: 12px; text-align: right; color: #6b7280;'>Valor</th></tr></thead><tbody>" + produtos_html + "</tbody></table>" if produtos else ""}

                <div style="background: linear-gradient(135deg, #7C3AED, #EC4899); padding: 20px; border-radius: 10px; margin-top: 30px; text-align: center;">
                    <p style="margin: 0 0 5px 0; color: white; font-size: 14px; opacity: 0.9;">VALOR TOTAL</p>
                    <h2 style="margin: 0; color: white; font-size: 36px;">R$ {total_final:.2f}</h2>
                </div>

                <div style="margin-top: 30px; padding: 20px; background: #f9fafb; border-radius: 8px; border-left: 4px solid #7C3AED;">
                    <p style="margin: 0; color: #6b7280; font-size: 14px; line-height: 1.6;">
                        Em caso de dúvidas ou para confirmar o orçamento, entre em contato conosco.
                    </p>
                </div>
            </div>

            <div style="text-align: center; padding: 20px; color: #9ca3af; font-size: 12px; margin-top: 20px;">
                <p style="margin: 0;">© 2024 BIOMA Uberaba - Sistema de Gestão</p>
                <p style="margin: 5px 0 0 0;">Este é um e-mail automático, não é necessário responder</p>
            </div>
        </body>
        </html>
        """

        try:
            # Enviar via MailerSend API
            mailersend_api_url = "https://api.mailersend.com/v1/email"

            headers = {
                "Authorization": f"Bearer {mailersend_key}",
                "Content-Type": "application/json"
            }

            payload = {
                "from": {
                    "email": mailersend_from,
                    "name": mailersend_name
                },
                "to": [
                    {
                        "email": email_cliente,
                        "name": cliente_nome
                    }
                ],
                "subject": f"💼 Orçamento #{numero_orcamento} - BIOMA Uberaba",
                "html": html_content,
                "text": f"Orçamento #{numero_orcamento} - Total: R$ {total_final:.2f}"
            }

            response = requests.post(mailersend_api_url, json=payload, headers=headers, timeout=10)

            if response.status_code == 202:
                logger.info(f"✅ Orçamento enviado via MailerSend para {email_cliente}")
                return jsonify({
                    'success': True,
                    'destinatario': email_cliente,
                    'message': f'✅ Orçamento enviado com sucesso para {email_cliente}',
                    'orcamento_numero': numero_orcamento,
                    'valor_total': total_final
                })
            else:
                logger.error(f"❌ Erro MailerSend: {response.status_code} - {response.text}")
                return jsonify({
                    'success': False,
                    'message': f'Erro ao enviar e-mail: {response.status_code}'
                }), response.status_code

        except requests.exceptions.Timeout:
            logger.error("❌ Timeout ao conectar com MailerSend API")
            return jsonify({
                'success': False,
                'message': 'Timeout ao conectar com servidor de e-mail'
            }), 408

        except requests.exceptions.RequestException as req_err:
            logger.error(f"❌ Erro de conexão MailerSend: {req_err}")
            return jsonify({
                'success': False,
                'message': f'Erro de conexão: {str(req_err)}'
            }), 503

    except Exception as e:
        logger.error(f"Erro ao enviar e-mail: {e}")
        return jsonify({'success': False, 'message': f'Erro ao enviar e-mail: {str(e)}'}), 500


@bp.route('/api/email/test', methods=['POST'])
@login_required
def test_email():
    """Testar configuração de e-mail enviando um e-mail de teste"""
    try:
        data = request.json
        email_destino = data.get('email')

        if not email_destino:
            return jsonify({'success': False, 'message': 'E-mail de destino não fornecido'}), 400

        # Validar formato de e-mail
        import re
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, email_destino):
            return jsonify({'success': False, 'message': 'Formato de e-mail inválido'}), 400

        # Verificar se as configurações de e-mail estão presentes
        import os
        mailersend_key = os.getenv('MAILERSEND_API_KEY')
        mailersend_from = os.getenv('MAILERSEND_FROM_EMAIL')

        if not mailersend_key or not mailersend_from:
            return jsonify({
                'success': False,
                'message': 'Configurações de e-mail não encontradas no servidor. Verifique as variáveis MAILERSEND_API_KEY e MAILERSEND_FROM_EMAIL no arquivo .env'
            }), 400

        # Implementação real do MailerSend
        import requests

        logger.info(f"✉️ Enviando e-mail de teste para: {email_destino}")
        logger.info(f"📧 Remetente configurado: {mailersend_from}")

        try:
            # Chamar API do MailerSend
            mailersend_api_url = "https://api.mailersend.com/v1/email"

            headers = {
                "Authorization": f"Bearer {mailersend_key}",
                "Content-Type": "application/json"
            }

            payload = {
                "from": {
                    "email": mailersend_from,
                    "name": "BIOMA System"
                },
                "to": [
                    {
                        "email": email_destino,
                        "name": "Teste"
                    }
                ],
                "subject": "🧪 E-mail de Teste - BIOMA System",
                "html": """
                    <html>
                    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                        <div style="background: linear-gradient(135deg, #7C3AED, #EC4899); padding: 30px; border-radius: 15px; text-align: center; color: white;">
                            <h1 style="margin: 0 0 10px 0;">✅ E-mail de Teste</h1>
                            <p style="margin: 0; font-size: 16px;">Sistema BIOMA</p>
                        </div>
                        <div style="padding: 30px 20px; background: #f9fafb; border-radius: 10px; margin-top: 20px;">
                            <h2 style="color: #7C3AED; margin-top: 0;">Parabéns!</h2>
                            <p style="color: #374151; line-height: 1.6;">
                                Se você está lendo esta mensagem, significa que o sistema de e-mail do BIOMA está funcionando corretamente!
                            </p>
                            <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #10B981;">
                                <p style="margin: 0; color: #059669;"><strong>✓ Configuração verificada</strong></p>
                                <p style="margin: 5px 0 0 0; color: #6b7280; font-size: 14px;">
                                    MailerSend API está configurada e operacional
                                </p>
                            </div>
                            <p style="color: #6b7280; font-size: 14px; margin-top: 20px;">
                                Este é um e-mail automatizado. Não é necessário responder.
                            </p>
                        </div>
                        <div style="text-align: center; padding: 20px; color: #9ca3af; font-size: 12px;">
                            <p>© 2024 BIOMA System - Sistema de Gestão</p>
                        </div>
                    </body>
                    </html>
                """,
                "text": "Este é um e-mail de teste do sistema BIOMA. Se você está lendo esta mensagem, o sistema de e-mail está funcionando corretamente!"
            }

            response = requests.post(mailersend_api_url, json=payload, headers=headers, timeout=10)

            if response.status_code == 202:
                logger.info(f"✅ E-mail enviado com sucesso via MailerSend para {email_destino}")
                return jsonify({
                    'success': True,
                    'message': f'✅ E-mail de teste enviado com sucesso para {email_destino}!',
                    'destinatario': email_destino,
                    'remetente': mailersend_from,
                    'status_code': response.status_code,
                    'observacao': 'E-mail enviado via MailerSend API'
                })
            else:
                logger.error(f"❌ Erro MailerSend: {response.status_code} - {response.text}")
                return jsonify({
                    'success': False,
                    'message': f'Erro ao enviar e-mail: {response.status_code}',
                    'detalhes': response.text
                }), response.status_code

        except requests.exceptions.Timeout:
            logger.error("❌ Timeout ao conectar com MailerSend API")
            return jsonify({
                'success': False,
                'message': 'Timeout ao conectar com servidor de e-mail. Tente novamente.'
            }), 408

        except requests.exceptions.RequestException as req_err:
            logger.error(f"❌ Erro de conexão MailerSend: {req_err}")
            return jsonify({
                'success': False,
                'message': f'Erro de conexão: {str(req_err)}'
            }), 503

    except Exception as e:
        logger.error(f"❌ Erro ao testar e-mail: {e}")
        return jsonify({'success': False, 'message': f'Erro ao enviar e-mail de teste: {str(e)}'}), 500


@bp.route('/api/contratos/<id>/pdf', methods=['GET'])
@login_required
def gerar_pdf_contrato(id):
    db = get_db()
    """Gerar PDF de contrato para impressão (Diretriz #4)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        contrato = db.orcamentos.find_one({'_id': ObjectId(id)})
        if not contrato:
            return jsonify({'success': False, 'message': 'Contrato não encontrado'}), 404

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER)
        elements.append(Paragraph(f"CONTRATO #{contrato.get('numero', str(contrato['_id'])[-6:])}", title_style))
        elements.append(Spacer(1, 0.3*cm))

        # Status
        status_style = ParagraphStyle('Status', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)
        status_color = HexColor('#4CAF50') if contrato.get('status') == 'Aprovado' else HexColor('#FF9800')
        elements.append(Paragraph(f'<font color="{status_color}"><b>STATUS: {contrato.get("status", "N/A")}</b></font>', status_style))
        elements.append(Spacer(1, 0.3*cm))

        # Informações do cliente
        elements.append(Paragraph('<b>DADOS DO CLIENTE</b>', styles['Heading2']))
        cliente_data = [
            ['Nome:', contrato.get('cliente_nome', 'N/A')],
            ['CPF:', contrato.get('cliente_cpf', 'N/A')],
            ['Telefone:', contrato.get('cliente_telefone', 'N/A')],
            ['Email:', contrato.get('cliente_email', 'N/A')],
            ['Data Contrato:', contrato.get('created_at', datetime.now()).strftime('%d/%m/%Y') if isinstance(contrato.get('created_at'), datetime) else 'N/A'],
            ['Forma Pagamento:', contrato.get('forma_pagamento', 'N/A')]
        ]
        cliente_table = Table(cliente_data, colWidths=[4*cm, 12*cm])
        cliente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(cliente_table)
        elements.append(Spacer(1, 0.5*cm))

        # Serviços contratados
        if contrato.get('servicos'):
            elements.append(Paragraph('<b>SERVIÇOS CONTRATADOS</b>', styles['Heading2']))
            servicos_data = [['Descrição', 'Profissional', 'Qtd', 'Valor Unit.', 'Total']]
            for srv in contrato['servicos']:
                servicos_data.append([
                    srv.get('nome', 'N/A'),
                    srv.get('profissional_nome', '-'),
                    str(srv.get('quantidade', 1)),
                    f"R$ {srv.get('preco', 0):.2f}",
                    f"R$ {srv.get('total', 0):.2f}"
                ])
            servicos_table = Table(servicos_data, colWidths=[6*cm, 4*cm, 1.5*cm, 2.5*cm, 2*cm])
            servicos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(servicos_table)
            elements.append(Spacer(1, 0.3*cm))

        # Produtos
        if contrato.get('produtos'):
            elements.append(Paragraph('<b>PRODUTOS</b>', styles['Heading2']))
            produtos_data = [['Descrição', 'Quantidade', 'Valor Unit.', 'Total']]
            for prd in contrato['produtos']:
                produtos_data.append([
                    prd.get('nome', 'N/A'),
                    str(prd.get('quantidade', 1)),
                    f"R$ {prd.get('preco', 0):.2f}",
                    f"R$ {prd.get('total', 0):.2f}"
                ])
            produtos_table = Table(produtos_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
            produtos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2196F3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(produtos_table)
            elements.append(Spacer(1, 0.3*cm))

        # Totais
        totais_data = [
            ['Subtotal Serviços:', f"R$ {contrato.get('total_servicos', 0):.2f}"],
            ['Subtotal Produtos:', f"R$ {contrato.get('total_produtos', 0):.2f}"],
            ['Desconto:', f"R$ {contrato.get('desconto_valor', 0):.2f}"],
            ['<b>TOTAL DO CONTRATO:</b>', f"<b>R$ {contrato.get('total_final', 0):.2f}</b>"]
        ]
        totais_table = Table(totais_data, colWidths=[12*cm, 4*cm])
        totais_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('#FFD700')),
        ]))
        elements.append(totais_table)

        # Observações
        if contrato.get('observacoes'):
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph('<b>OBSERVAÇÕES:</b>', styles['Heading3']))
            elements.append(Paragraph(contrato['observacoes'], styles['Normal']))

        # Termos e condições
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph('<b>TERMOS E CONDIÇÕES</b>', styles['Heading3']))
        termos = """Este contrato estabelece os serviços e produtos acordados entre as partes.
        O cliente declara estar ciente dos valores e condições apresentados.
        A BIOMA se compromete a prestar os serviços com excelência e profissionalismo."""
        elements.append(Paragraph(termos, styles['Normal']))

        # Assinaturas na MESMA PÁGINA das cláusulas (v7.0)
        elements.append(Spacer(1, 1.5*cm))

        # Criar tabela de assinaturas (duas colunas lado a lado)
        data_assinatura = datetime.now().strftime('%d/%m/%Y')
        cidade = "Uberaba/MG"

        assinaturas_data = [
            [f'{cidade}, {data_assinatura}', ''],
            ['', ''],
            ['_' * 40, '_' * 40],
            ['<b>BIOMA Uberaba</b>', f'<b>{contrato.get("cliente_nome", "CLIENTE")}</b>'],
            ['CNPJ: __.___.___/____-__', f'CPF: {contrato.get("cliente_cpf", "___.___.___-__")}']
        ]

        assinaturas_table = Table(assinaturas_data, colWidths=[8*cm, 8*cm])
        assinaturas_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 2), (-1, 2), 0),
            ('BOTTOMPADDING', (0, 2), (-1, 2), 2),
        ]))
        elements.append(assinaturas_table)

        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"contrato_{contrato.get('numero', id)}.pdf"
        )

    except Exception as e:
        logger.error(f"Erro ao gerar PDF de contrato: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/contratos/<id>/whatsapp', methods=['GET'])
@login_required
def enviar_whatsapp_contrato(id):
    db = get_db()
    """Gerar link do WhatsApp com contrato (Diretriz #4)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        contrato = db.orcamentos.find_one({'_id': ObjectId(id)})
        if not contrato:
            return jsonify({'success': False, 'message': 'Contrato não encontrado'}), 404

        # Montar mensagem do WhatsApp
        telefone = contrato.get('cliente_telefone', '').replace('(', '').replace(')', '').replace('-', '').replace(' ', '')

        mensagem = f"""🌿 *BIOMA - Contrato #{contrato.get('numero', str(contrato['_id'])[-6:])}*

✅ *Status:* {contrato.get('status', 'N/A')}
👤 *Cliente:* {contrato.get('cliente_nome', 'N/A')}
📅 *Data:* {contrato.get('created_at', datetime.now()).strftime('%d/%m/%Y') if isinstance(contrato.get('created_at'), datetime) else 'N/A'}
💳 *Pagamento:* {contrato.get('forma_pagamento', 'N/A')}

"""

        # Serviços
        if contrato.get('servicos'):
            mensagem += "✨ *SERVIÇOS CONTRATADOS:*\n"
            for srv in contrato['servicos']:
                prof_nome = srv.get('profissional_nome', '')
                if prof_nome:
                    mensagem += f"• {srv.get('nome', 'N/A')} ({prof_nome}) - R$ {srv.get('total', 0):.2f}\n"
                else:
                    mensagem += f"• {srv.get('nome', 'N/A')} - R$ {srv.get('total', 0):.2f}\n"
            mensagem += "\n"

        # Produtos
        if contrato.get('produtos'):
            mensagem += "🛍️ *PRODUTOS:*\n"
            for prd in contrato['produtos']:
                mensagem += f"• {prd.get('nome', 'N/A')} - R$ {prd.get('total', 0):.2f}\n"
            mensagem += "\n"

        # Totais
        mensagem += f"""💰 *VALORES:*
Serviços: R$ {contrato.get('total_servicos', 0):.2f}
Produtos: R$ {contrato.get('total_produtos', 0):.2f}
Desconto: R$ {contrato.get('desconto_valor', 0):.2f}

✅ *TOTAL DO CONTRATO: R$ {contrato.get('total_final', 0):.2f}*

Obrigado por escolher a BIOMA! 🌿
"""

        if contrato.get('observacoes'):
            mensagem += f"\n📝 *Observações:* {contrato['observacoes']}"

        # URL encode da mensagem
        mensagem_encoded = urllib.parse.quote(mensagem)
        whatsapp_url = f"https://wa.me/55{telefone}?text={mensagem_encoded}"

        return jsonify({
            'success': True,
            'whatsapp_url': whatsapp_url,
            'telefone': telefone
        })

    except Exception as e:
        logger.error(f"Erro ao gerar link WhatsApp para contrato: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== MÓDULO FINANCEIRO (Diretriz #7) ====================

@bp.route('/api/financeiro/dashboard', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def financeiro_dashboard():
    db = get_db()
    """Dashboard financeiro com comissões, despesas e lucro (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        # Filtros opcionais
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        # Query base
        query = {}
        if data_inicio and data_fim:
            query['created_at'] = {
                '$gte': datetime.fromisoformat(data_inicio),
                '$lte': datetime.fromisoformat(data_fim)
            }

        # Total de receitas (orçamentos aprovados)
        orcamentos_aprovados = list(db.orcamentos.find({**query, 'status': 'Aprovado'}))
        receita_total = sum(o.get('total_final', 0) for o in orcamentos_aprovados)

        # Total de comissões
        comissoes_total = sum(o.get('total_comissoes', 0) for o in orcamentos_aprovados)

        # Total de despesas
        despesas = list(db.despesas.find(query))
        despesas_total = sum(d.get('valor', 0) for d in despesas)

        # Lucro líquido
        lucro_liquido = receita_total - comissoes_total - despesas_total

        # Comissões por profissional
        comissoes_por_profissional = {}
        for orcamento in orcamentos_aprovados:
            for prof in orcamento.get('profissionais_vinculados', []):
                prof_id = str(prof.get('profissional_id', ''))
                prof_nome = prof.get('nome', 'N/A')
                comissao_valor = prof.get('comissao_valor', 0)

                if prof_id not in comissoes_por_profissional:
                    comissoes_por_profissional[prof_id] = {
                        'nome': prof_nome,
                        'total': 0,
                        'quantidade': 0
                    }

                comissoes_por_profissional[prof_id]['total'] += comissao_valor
                comissoes_por_profissional[prof_id]['quantidade'] += 1

        # Despesas por categoria
        despesas_por_categoria = {}
        for despesa in despesas:
            categoria = despesa.get('categoria', 'Outros')
            if categoria not in despesas_por_categoria:
                despesas_por_categoria[categoria] = 0
            despesas_por_categoria[categoria] += despesa.get('valor', 0)

        return jsonify({
            'success': True,
            'financeiro': {
                'receita_total': receita_total,
                'comissoes_total': comissoes_total,
                'despesas_total': despesas_total,
                'lucro_liquido': lucro_liquido,
                'margem_lucro_perc': (lucro_liquido / receita_total * 100) if receita_total > 0 else 0,
                'comissoes_por_profissional': list(comissoes_por_profissional.values()),
                'despesas_por_categoria': despesas_por_categoria,
                'total_orcamentos': len(orcamentos_aprovados),
                'ticket_medio': receita_total / len(orcamentos_aprovados) if orcamentos_aprovados else 0
            }
        })

    except Exception as e:
        logger.error(f"Erro no dashboard financeiro: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/financeiro/resumo', methods=['GET'])
@login_required
def financeiro_resumo():
    db = get_db()
    """Resumo financeiro simplificado para exibição rápida"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database not available'}), 500

    try:
        # Obter mês atual
        hoje = datetime.now()
        inicio_mes = datetime(hoje.year, hoje.month, 1)

        # Receitas do mês (orçamentos aprovados)
        orcamentos_aprovados = list(db.orcamentos.find({
            'status': 'Aprovado',
            'created_at': {'$gte': inicio_mes}
        }))
        receita_mes = sum(o.get('total_final', 0) for o in orcamentos_aprovados)

        # Despesas do mês
        despesas = list(db.despesas.find({
            'data': {'$gte': inicio_mes}
        }))
        despesas_mes = sum(d.get('valor', 0) for d in despesas)

        # Comissões do mês
        comissoes_mes = sum(o.get('total_comissoes', 0) for o in orcamentos_aprovados)

        # Lucro do mês
        lucro_mes = receita_mes - comissoes_mes - despesas_mes

        # Totais gerais (todos os tempos)
        todos_orcamentos = list(db.orcamentos.find({'status': 'Aprovado'}))
        receita_total = sum(o.get('total_final', 0) for o in todos_orcamentos)

        todas_despesas = list(db.despesas.find())
        despesas_total = sum(d.get('valor', 0) for d in todas_despesas)

        comissoes_total = sum(o.get('total_comissoes', 0) for o in todos_orcamentos)
        lucro_total = receita_total - comissoes_total - despesas_total

        return jsonify({
            'success': True,
            'resumo': {
                'mes_atual': {
                    'receita': round(receita_mes, 2),
                    'despesas': round(despesas_mes, 2),
                    'comissoes': round(comissoes_mes, 2),
                    'lucro': round(lucro_mes, 2),
                    'quantidade_orcamentos': len(orcamentos_aprovados)
                },
                'total_geral': {
                    'receita': round(receita_total, 2),
                    'despesas': round(despesas_total, 2),
                    'comissoes': round(comissoes_total, 2),
                    'lucro': round(lucro_total, 2),
                    'quantidade_orcamentos': len(todos_orcamentos)
                }
            }
        })

    except Exception as e:
        logger.error(f"Erro ao carregar resumo financeiro: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/financeiro/despesas', methods=['GET', 'POST'])
@login_required
def financeiro_despesas():
    db = get_db()
    """Gerenciar despesas (GET: Admin/Gestão, POST: Admin only)"""
    if db is None:
        return jsonify({'success': False}), 500

    # Verificar permissões baseado no método
    user_tipo = session.get('tipo_acesso', 'Profissional')

    if request.method == 'GET':
        # Admin e Gestão podem ler
        if user_tipo not in ['Admin', 'Gestão']:
            return jsonify({'success': False, 'message': 'Acesso negado'}), 403
        try:
            # Filtros opcionais
            categoria = request.args.get('categoria')
            data_inicio = request.args.get('data_inicio')
            data_fim = request.args.get('data_fim')

            query = {}
            if categoria:
                query['categoria'] = categoria
            if data_inicio and data_fim:
                query['data'] = {
                    '$gte': datetime.fromisoformat(data_inicio),
                    '$lte': datetime.fromisoformat(data_fim)
                }

            despesas = list(db.despesas.find(query).sort('data', DESCENDING))
            return jsonify({'success': True, 'despesas': convert_objectid(despesas)})

        except Exception as e:
            logger.error(f"Erro ao listar despesas: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # POST - criar despesa (apenas Admin)
    if user_tipo != 'Admin':
        return jsonify({'success': False, 'message': 'Apenas Admin pode criar despesas'}), 403

    data = request.json
    try:
        despesa = {
            'descricao': data['descricao'],
            'categoria': data.get('categoria', 'Outros'),
            'valor': float(data['valor']),
            'data': datetime.fromisoformat(data.get('data', datetime.now().isoformat())),
            'forma_pagamento': data.get('forma_pagamento', 'Dinheiro'),
            'observacoes': data.get('observacoes', ''),
            'created_by': session.get('username', 'sistema'),
            'created_at': datetime.now()
        }
        result = db.despesas.insert_one(despesa)

        logger.info(f"✅ Despesa registrada: {despesa['descricao']} - R$ {despesa['valor']}")
        return jsonify({'success': True, 'id': str(result.inserted_id), 'message': 'Despesa registrada com sucesso'})

    except Exception as e:
        logger.error(f"Erro ao registrar despesa: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/financeiro/despesas/<id>', methods=['PUT', 'DELETE'])
@login_required
@permission_required('Admin')
def handle_despesa(id):
    db = get_db()
    """Atualizar ou deletar despesa (Admin only)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        if request.method == 'PUT':
            data = request.json
            update_data = {
                'descricao': data.get('descricao'),
                'categoria': data.get('categoria'),
                'valor': float(data.get('valor', 0)),
                'data': datetime.fromisoformat(data.get('data')) if data.get('data') else None,
                'forma_pagamento': data.get('forma_pagamento'),
                'observacoes': data.get('observacoes', ''),
                'updated_at': datetime.now(),
                'updated_by': session.get('username', 'sistema')
            }

            # Remover campos None
            update_data = {k: v for k, v in update_data.items() if v is not None}

            db.despesas.update_one({'_id': ObjectId(id)}, {'$set': update_data})
            logger.info(f"✅ Despesa {id} atualizada")
            return jsonify({'success': True, 'message': 'Despesa atualizada com sucesso'})

        elif request.method == 'DELETE':
            db.despesas.delete_one({'_id': ObjectId(id)})
            logger.info(f"🗑️ Despesa {id} deletada")
            return jsonify({'success': True, 'message': 'Despesa deletada com sucesso'})

    except Exception as e:
        logger.error(f"Erro ao manipular despesa: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/financeiro/relatorio', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def financeiro_relatorio():
    db = get_db()
    """Gerar relatório financeiro detalhado (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False}), 500

    try:
        periodo = request.args.get('periodo', 'mes')  # mes, trimestre, ano
        data_referencia = request.args.get('data', datetime.now().isoformat())
        data_ref = datetime.fromisoformat(data_referencia)

        # Calcular período
        if periodo == 'mes':
            data_inicio = data_ref.replace(day=1)
            if data_ref.month == 12:
                data_fim = data_ref.replace(year=data_ref.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                data_fim = data_ref.replace(month=data_ref.month + 1, day=1) - timedelta(days=1)
        elif periodo == 'trimestre':
            mes_inicio = ((data_ref.month - 1) // 3) * 3 + 1
            data_inicio = data_ref.replace(month=mes_inicio, day=1)
            data_fim = (data_inicio + timedelta(days=90)).replace(day=1) - timedelta(days=1)
        else:  # ano
            data_inicio = data_ref.replace(month=1, day=1)
            data_fim = data_ref.replace(month=12, day=31)

        query_tempo = {'created_at': {'$gte': data_inicio, '$lte': data_fim}}

        # Coletar dados
        orcamentos = list(db.orcamentos.find({**query_tempo, 'status': 'Aprovado'}))
        despesas = list(db.despesas.find({'data': {'$gte': data_inicio, '$lte': data_fim}}))

        receita = sum(o.get('total_final', 0) for o in orcamentos)
        comissoes = sum(o.get('total_comissoes', 0) for o in orcamentos)
        despesas_total = sum(d.get('valor', 0) for d in despesas)
        lucro = receita - comissoes - despesas_total

        return jsonify({
            'success': True,
            'relatorio': {
                'periodo': periodo,
                'data_inicio': data_inicio.isoformat(),
                'data_fim': data_fim.isoformat(),
                'receita_bruta': receita,
                'comissoes': comissoes,
                'despesas': despesas_total,
                'lucro_liquido': lucro,
                'margem_lucro': (lucro / receita * 100) if receita > 0 else 0,
                'total_orcamentos': len(orcamentos),
                'ticket_medio': receita / len(orcamentos) if orcamentos else 0
            }
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório financeiro: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/financeiro/receitas', methods=['GET'])
@login_required
@permission_required('Admin', 'Gestão')
def financeiro_receitas():
    db = get_db()
    """Listar receitas - orçamentos aprovados (Admin/Gestão)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Filtros opcionais
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        cliente_id = request.args.get('cliente_id')

        # Query base para orçamentos aprovados
        query = {'status': 'Aprovado'}

        # Aplicar filtros
        if data_inicio and data_fim:
            query['created_at'] = {
                '$gte': datetime.fromisoformat(data_inicio),
                '$lte': datetime.fromisoformat(data_fim)
            }

        if cliente_id:
            query['cliente_id'] = cliente_id

        # Buscar orçamentos aprovados
        orcamentos = list(db.orcamentos.find(query).sort('created_at', DESCENDING))

        # Formatar receitas
        receitas = []
        total = 0
        for orc in orcamentos:
            valor = orc.get('total_final', 0)
            total += valor

            receitas.append({
                '_id': str(orc['_id']),
                'cliente_nome': orc.get('cliente_nome', 'N/A'),
                'cliente_cpf': orc.get('cliente_cpf', 'N/A'),
                'data': orc.get('created_at', datetime.now()).strftime('%Y-%m-%d') if isinstance(orc.get('created_at'), datetime) else orc.get('data', 'N/A'),
                'valor': valor,
                'forma_pagamento': orc.get('forma_pagamento', 'N/A'),
                'status': orc.get('status', 'N/A'),
                'observacoes': orc.get('observacoes', '')
            })

        logger.info(f"✅ Listadas {len(receitas)} receitas - Total: R$ {total:.2f}")
        return jsonify({
            'success': True,
            'receitas': receitas,
            'total': round(total, 2),
            'quantidade': len(receitas)
        })

    except Exception as e:
        logger.error(f"Erro ao listar receitas: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/financeiro/comissoes', methods=['GET'])
@login_required
def financeiro_comissoes():
    db = get_db()
    """Listar comissões (Admin/Gestão: todas | Profissional: próprias)"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        user_tipo = session.get('tipo_acesso', 'Profissional')
        user_id = session.get('user_id')

        # Filtros opcionais
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        profissional_id = request.args.get('profissional_id')

        # Query base
        query = {}

        # RBAC: Profissional só pode ver suas próprias comissões
        if user_tipo == 'Profissional':
            # Buscar o documento profissional vinculado ao usuário
            prof_doc = db.profissionais.find_one({'user_id': user_id})
            if prof_doc:
                profissional_id = str(prof_doc['_id'])
            else:
                return jsonify({'success': True, 'comissoes': [], 'message': 'Profissional não encontrado'}), 200

        # Aplicar filtros
        if data_inicio and data_fim:
            query['data_registro'] = {
                '$gte': datetime.fromisoformat(data_inicio),
                '$lte': datetime.fromisoformat(data_fim)
            }

        if profissional_id:
            query['profissional_id'] = ObjectId(profissional_id)

        # Buscar comissões do histórico
        comissoes_historico = list(db.comissoes_historico.find(query).sort('data_registro', DESCENDING))

        # Se não há histórico, buscar de orçamentos aprovados
        if not comissoes_historico:
            query_orc = {'status': 'Aprovado'}
            if data_inicio and data_fim:
                query_orc['created_at'] = {
                    '$gte': datetime.fromisoformat(data_inicio),
                    '$lte': datetime.fromisoformat(data_fim)
                }

            orcamentos = list(db.orcamentos.find(query_orc))
            comissoes_lista = []

            for orc in orcamentos:
                for prof in orc.get('profissionais_vinculados', []):
                    if not profissional_id or str(prof.get('profissional_id')) == profissional_id:
                        comissoes_lista.append({
                            'profissional_id': str(prof.get('profissional_id', '')),
                            'profissional_nome': prof.get('nome', 'N/A'),
                            'comissao_percentual': prof.get('comissao_percentual', 0),
                            'comissao_valor': prof.get('comissao_valor', 0),
                            'orcamento_id': str(orc['_id']),
                            'cliente_nome': orc.get('cliente_nome', 'N/A'),
                            'data': orc.get('created_at', datetime.now()).strftime('%Y-%m-%d') if isinstance(orc.get('created_at'), datetime) else 'N/A',
                            'status': 'Aprovado'
                        })

            total = sum(c.get('comissao_valor', 0) for c in comissoes_lista)

            logger.info(f"✅ Listadas {len(comissoes_lista)} comissões de orçamentos - Total: R$ {total:.2f}")
            return jsonify({
                'success': True,
                'comissoes': comissoes_lista,
                'total': round(total, 2),
                'quantidade': len(comissoes_lista)
            })

        # Formatar comissões do histórico
        comissoes_lista = []
        total = 0
        for com in comissoes_historico:
            valor = com.get('comissao_valor', 0)
            total += valor

            comissoes_lista.append({
                '_id': str(com['_id']),
                'profissional_id': str(com.get('profissional_id', '')),
                'profissional_nome': com.get('profissional_nome', 'N/A'),
                'comissao_percentual': com.get('comissao_percentual', 0),
                'comissao_valor': valor,
                'orcamento_id': str(com.get('orcamento_id', '')),
                'cliente_nome': com.get('cliente_nome', 'N/A'),
                'data': com.get('data_registro', datetime.now()).strftime('%Y-%m-%d') if isinstance(com.get('data_registro'), datetime) else 'N/A',
                'status': com.get('status_orcamento', 'N/A')
            })

        logger.info(f"✅ Listadas {len(comissoes_lista)} comissões - Total: R$ {total:.2f}")
        return jsonify({
            'success': True,
            'comissoes': comissoes_lista,
            'total': round(total, 2),
            'quantidade': len(comissoes_lista)
        })

    except Exception as e:
        logger.error(f"Erro ao listar comissões: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/profissionais/<id>/comissoes', methods=['GET'])
@login_required
def get_comissoes_profissional(id):
    db = get_db()
    """Obter estatísticas de comissões de um profissional"""
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500
    
    try:
        # Filtros opcionais
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        status = request.args.get('status')
        
        # Query base
        query = {'profissional_id': ObjectId(id)}
        
        # Aplicar filtros
        if data_inicio or data_fim:
            date_filter = {}
            if data_inicio:
                date_filter['$gte'] = datetime.fromisoformat(data_inicio)
            if data_fim:
                date_filter['$lte'] = datetime.fromisoformat(data_fim)
            query['data_registro'] = date_filter
        
        if status:
            query['status_orcamento'] = status
        
        # Buscar histórico de comissões
        comissoes = list(db.comissoes_historico.find(query).sort('data_registro', DESCENDING))
        
        # Calcular estatísticas
        total_comissoes = sum(c.get('comissao_valor', 0) for c in comissoes)
        total_orcamentos = len(set(str(c.get('orcamento_id')) for c in comissoes))
        media_comissao = total_comissoes / total_orcamentos if total_orcamentos > 0 else 0
        
        # Agrupar por mês
        comissoes_por_mes = {}
        for comissao in comissoes:
            data = comissao.get('data_registro')
            if data and isinstance(data, datetime):
                mes_ano = data.strftime('%Y-%m')
                if mes_ano not in comissoes_por_mes:
                    comissoes_por_mes[mes_ano] = {'valor': 0, 'quantidade': 0}
                comissoes_por_mes[mes_ano]['valor'] += comissao.get('comissao_valor', 0)
                comissoes_por_mes[mes_ano]['quantidade'] += 1
        
        # Converter para formato de resposta
        for c in comissoes:
            c['_id'] = str(c['_id'])
            c['orcamento_id'] = str(c['orcamento_id'])
            c['profissional_id'] = str(c['profissional_id'])
            if 'data_registro' in c and isinstance(c['data_registro'], datetime):
                c['data_registro'] = c['data_registro'].isoformat()
        
        # Buscar dados do profissional
        profissional = db.profissionais.find_one({'_id': ObjectId(id)})
        if not profissional:
            profissional = db.assistentes.find_one({'_id': ObjectId(id)})
        
        resultado = {
            'profissional': {
                'id': id,
                'nome': profissional.get('nome') if profissional else 'Desconhecido',
                'foto': profissional.get('foto') if profissional else None
            },
            'estatisticas': {
                'total_comissoes': round(total_comissoes, 2),
                'total_orcamentos': total_orcamentos,
                'media_comissao': round(media_comissao, 2),
                'comissoes_por_mes': comissoes_por_mes
            },
            'historico': comissoes
        }
        
        logger.info(f"📊 Estatísticas de comissões do profissional {id}: R$ {total_comissoes:.2f}")
        return jsonify({'success': True, 'data': resultado})
        
    except Exception as e:
        logger.error(f"❌ Erro em get_comissoes_profissional: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== FIM DAS NOVAS FUNCIONALIDADES ====================

# ==================== ENDPOINTS PARA SUB-TABS - PRODUTOS ====================

@bp.route('/api/produtos', methods=['GET'])
@login_required
def listar_produtos():
    db = get_db()
    """Lista produtos com filtro opcional por status"""
    try:
        status = request.args.get('status')
        
        query = {}
        if status:
            query['status'] = status
        
        produtos = list(db.produtos.find(query).sort('nome', ASCENDING))
        
        # Formatar resposta
        resultado = []
        for p in produtos:
            # Suportar ambos os campos: 'ativo' (boolean) e 'status' (string)
            ativo_val = p.get('ativo', None)
            if ativo_val is None:
                status_val = p.get('status', 'Ativo')
                ativo_val = (status_val == 'Ativo' or status_val == 'ativo')

            resultado.append({
                '_id': str(p['_id']),
                'id': str(p['_id']),
                'nome': p.get('nome', 'Sem nome'),
                'marca': p.get('marca', 'Sem marca'),
                'preco': float(p.get('preco', 0)),
                'estoque': int(p.get('estoque', 0)),
                'estoque_minimo': int(p.get('estoque_minimo', 0)),
                'status': p.get('status', 'Ativo'),
                'ativo': ativo_val,
                'sku': p.get('sku', ''),
                'categoria': p.get('categoria', 'Geral')
            })
        
        logger.info(f"📦 Produtos listados: {len(resultado)} (status: {status or 'todos'})")
        return jsonify({'success': True, 'produtos': resultado})
        
    except Exception as e:
        logger.error(f"❌ Erro ao listar produtos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/produtos/buscar', methods=['GET'])
@login_required
def buscar_produtos():
    db = get_db()
    """Busca produtos por termo (nome, marca, sku)"""
    try:
        termo = request.args.get('termo', '').strip()

        if not termo or len(termo) < 2:
            return jsonify({'success': True, 'produtos': []})

        regex = {'$regex': termo, '$options': 'i'}

        produtos = list(db.produtos.find({
            '$or': [
                {'nome': regex},
                {'marca': regex},
                {'sku': regex}
            ]
        }).sort('nome', ASCENDING).limit(20))

        resultado = []
        for p in produtos:
            ativo_val = p.get('ativo', None)
            if ativo_val is None:
                status_val = p.get('status', 'Ativo')
                ativo_val = (status_val == 'Ativo' or status_val == 'ativo')

            resultado.append({
                '_id': str(p['_id']),
                'id': str(p['_id']),
                'nome': p.get('nome', 'Sem nome'),
                'marca': p.get('marca', ''),
                'preco': float(p.get('preco', 0)),
                'estoque': int(p.get('estoque', 0)),
                'ativo': ativo_val,
                'sku': p.get('sku', ''),
                'categoria': p.get('categoria', 'Geral')
            })

        logger.info(f"🔍 Busca de produtos: '{termo}' - {len(resultado)} resultados")
        return jsonify({'success': True, 'produtos': resultado})

    except Exception as e:
        logger.error(f"❌ Erro ao buscar produtos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/produto/barcode/<sku>', methods=['GET'])
@login_required
def get_produto_por_barcode(sku):
    """
    Busca um único produto pelo seu SKU (código de barras).
    """
    db = get_db()
    if db is None:
        return jsonify({'success': False, 'message': 'Database offline'}), 500

    try:
        # Tenta buscar por SKU ou um campo futuro 'codigo_barras'
        # Usamos 'sku' como o campo principal para o código de barras
        produto = db.produtos.find_one({
            '$or': [
                {'sku': sku},
                {'codigo_barras': sku}
            ]
        })

        if not produto:
            return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404

        # Se encontrou, retorna o produto completo
        return jsonify({'success': True, 'produto': convert_objectid(produto)})

    except Exception as e:
        logger.error(f"Erro ao buscar produto por barcode {sku}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/produtos/baixo-estoque', methods=['GET'])
@login_required
def produtos_baixo_estoque():
    db = get_db()
    """Retorna produtos com estoque baixo e estatísticas"""
    try:
        # Buscar todos os produtos ativos
        produtos = list(db.produtos.find({'status': 'Ativo'}))
        
        criticos = []
        atencao = []
        normais = []
        
        for p in produtos:
            estoque_atual = int(p.get('estoque', 0))
            estoque_minimo = int(p.get('estoque_minimo', 0))
            
            diferenca = estoque_atual - estoque_minimo
            
            produto_formatado = {
                'id': str(p['_id']),
                'nome': p.get('nome', 'Sem nome'),
                'marca': p.get('marca', 'Sem marca'),
                'estoque_atual': estoque_atual,
                'estoque_minimo': estoque_minimo,
                'diferenca': diferenca,
                'preco': float(p.get('preco', 0))
            }
            
            # Classificar por nível de estoque
            if estoque_atual <= estoque_minimo:
                produto_formatado['nivel'] = 'Crítico'
                criticos.append(produto_formatado)
            elif estoque_atual < estoque_minimo * 1.5:
                produto_formatado['nivel'] = 'Atenção'
                atencao.append(produto_formatado)
            else:
                produto_formatado['nivel'] = 'Normal'
                normais.append(produto_formatado)
        
        resultado = {
            'estatisticas': {
                'criticos': len(criticos),
                'atencao': len(atencao),
                'normais': len(normais)
            },
            'produtos': criticos + atencao  # Retorna apenas os que precisam atenção
        }
        
        logger.info(f"⚠️ Estoque - Críticos: {len(criticos)}, Atenção: {len(atencao)}, Normais: {len(normais)}")
        return jsonify({'success': True, 'data': resultado})
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar produtos com baixo estoque: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/produtos/<id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def produto_detalhes(id):
    db = get_db()

    # DELETE - Deletar produto
    if request.method == 'DELETE':
        """Deleta um produto"""
        try:
            result = db.produtos.delete_one({'_id': ObjectId(id)})

            if result.deleted_count > 0:
                logger.info(f"✅ Produto {id} deletado com sucesso")
                return jsonify({'success': True, 'message': 'Produto deletado com sucesso'})

            return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404

        except Exception as e:
            logger.error(f"❌ Erro ao deletar produto: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # GET - Buscar detalhes do produto
    if request.method == 'GET':
        """Busca detalhes de um produto específico"""
        try:
            produto = db.produtos.find_one({'_id': ObjectId(id)})

            if not produto:
                return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404

            return jsonify({'success': True, 'produto': convert_objectid(produto)})

        except Exception as e:
            logger.error(f"❌ Erro ao buscar produto {id}: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # PUT - Atualizar produto
    """Atualiza produto (incluindo toggle de status)"""
    try:
        data = request.get_json()

        update_data = {}
        if 'nome' in data:
            update_data['nome'] = data['nome']
        if 'marca' in data:
            update_data['marca'] = data['marca']
        if 'preco' in data:
            update_data['preco'] = float(data['preco'])
        if 'estoque' in data:
            # Garante que o valor não seja None nem string vazia antes de converter para int
            update_data['estoque'] = int(data.get('estoque') or 0)
        if 'estoque_minimo' in data:
            update_data['estoque_minimo'] = int(data['estoque_minimo'])
        if 'status' in data:
            update_data['status'] = data['status']
            # Sincronizar campo ativo com status
            update_data['ativo'] = (data['status'] == 'Ativo')
        if 'ativo' in data:
            # Se enviar ativo diretamente (boolean), processar
            update_data['ativo'] = bool(data['ativo'])
            update_data['status'] = 'Ativo' if data['ativo'] else 'Inativo'
        if 'sku' in data:
            update_data['sku'] = data['sku']
        if 'categoria' in data:
            update_data['categoria'] = data['categoria']
        if 'codigo_barras' in data:
            update_data['codigo_barras'] = data['codigo_barras']

        if not update_data:
            return jsonify({'success': False, 'message': 'Nenhum dado para atualizar'}), 400

        result = db.produtos.update_one(
            {'_id': ObjectId(id)},
            {'$set': update_data}
        )

        # Aceitar tanto modificações quanto quando não há mudanças (matched_count > 0)
        if result.matched_count > 0:
            logger.info(f"✅ Produto {id} atualizado: {update_data}")
            return jsonify({'success': True, 'message': 'Produto atualizado com sucesso'})

        return jsonify({'success': False, 'message': 'Produto não encontrado'}), 404

    except Exception as e:
        logger.error(f"❌ Erro ao atualizar produto: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== ENDPOINTS PARA SUB-TABS - SERVIÇOS ====================

@bp.route('/api/servicos/buscar', methods=['GET'])
@login_required
def buscar_servicos():
    db = get_db()
    """Busca serviços por termo (nome, categoria, tipo)"""
    try:
        termo = request.args.get('termo', '').strip()

        if not termo or len(termo) < 2:
            return jsonify({'success': True, 'servicos': []})

        regex = {'$regex': termo, '$options': 'i'}

        servicos = list(db.servicos.find({
            '$or': [
                {'nome': regex},
                {'categoria': regex},
                {'tipo': regex},
                {'descricao': regex}
            ]
        }).sort('nome', ASCENDING).limit(20))

        resultado = []
        for s in servicos:
            ativo_val = s.get('ativo', None)
            if ativo_val is None:
                status_val = s.get('status', 'Ativo')
                ativo_val = (status_val == 'Ativo' or status_val == 'ativo')

            resultado.append({
                '_id': str(s['_id']),
                'id': str(s['_id']),
                'nome': s.get('nome', 'Sem nome'),
                'categoria': s.get('categoria', 'Geral'),
                'preco': float(s.get('preco', 0)),
                'duracao': s.get('duracao', 0),
                'ativo': ativo_val,
                'tipo': s.get('tipo', ''),
                'descricao': s.get('descricao', ''),
                'comissao_profissional': s.get('comissao_profissional', 0),
                'comissao_assistente': s.get('comissao_assistente', 0)
            })

        logger.info(f"🔍 Busca de serviços: '{termo}' - {len(resultado)} resultados")
        return jsonify({'success': True, 'servicos': resultado})

    except Exception as e:
        logger.error(f"❌ Erro ao buscar serviços: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/servicos', methods=['GET'])
@login_required
def listar_servicos():
    db = get_db()
    """Lista serviços com filtro opcional por status"""
    try:
        status = request.args.get('status')
        categoria = request.args.get('categoria')
        
        query = {}
        if status:
            query['status'] = status
        if categoria and categoria != 'Todas':
            query['categoria'] = categoria
        
        servicos = list(db.servicos.find(query).sort('nome', ASCENDING))
        
        # Formatar resposta
        resultado = []
        for s in servicos:
            status_val = s.get('status', 'ativo')
            # Suportar ambos os campos: 'ativo' (boolean) e 'status' (string)
            ativo_bool = s.get('ativo', True) if 'ativo' in s else (status_val.lower() == 'ativo')

            resultado.append({
                '_id': str(s['_id']),
                'id': str(s['_id']),
                'nome': s.get('nome', 'Sem nome'),
                'categoria': s.get('categoria', 'Geral'),
                'tamanho': s.get('tamanho', 'Médio'),
                'preco': float(s.get('preco', 0)),
                'duracao': int(s.get('duracao', 60)),
                'status': status_val,
                'ativo': ativo_bool
            })
        
        logger.info(f"✂️ Serviços listados: {len(resultado)} (status: {status or 'todos'})")
        return jsonify({'success': True, 'servicos': resultado})
        
    except Exception as e:
        logger.error(f"❌ Erro ao listar serviços: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/servicos/<id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def atualizar_servico(id):
    db = get_db()

    # DELETE - Deletar serviço
    if request.method == 'DELETE':
        """Deleta um serviço"""
        try:
            result = db.servicos.delete_one({'_id': ObjectId(id)})

            if result.deleted_count > 0:
                logger.info(f"✅ Serviço {id} deletado com sucesso")
                return jsonify({'success': True, 'message': 'Serviço deletado com sucesso'})

            return jsonify({'success': False, 'message': 'Serviço não encontrado'}), 404

        except Exception as e:
            logger.error(f"❌ Erro ao deletar serviço: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # GET - Buscar detalhes do serviço
    if request.method == 'GET':
        """Busca detalhes de um serviço específico"""
        try:
            servico = db.servicos.find_one({'_id': ObjectId(id)})

            if not servico:
                return jsonify({'success': False, 'message': 'Serviço não encontrado'}), 404

            return jsonify({'success': True, 'servico': convert_objectid(servico)})

        except Exception as e:
            logger.error(f"❌ Erro ao buscar serviço {id}: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # PUT - Atualizar serviço
    """Atualiza serviço (incluindo toggle de status)"""
    try:
        data = request.get_json()
        
        update_data = {}
        if 'nome' in data:
            update_data['nome'] = data['nome']
        if 'categoria' in data:
            update_data['categoria'] = data['categoria']
        if 'tamanho' in data:
            update_data['tamanho'] = data['tamanho']
        if 'preco' in data:
            update_data['preco'] = float(data['preco'])
        if 'duracao' in data:
            update_data['duracao'] = int(data['duracao'])
        if 'status' in data:
            update_data['status'] = data['status']
        
        if not update_data:
            return jsonify({'success': False, 'message': 'Nenhum dado para atualizar'}), 400
        
        result = db.servicos.update_one(
            {'_id': ObjectId(id)},
            {'$set': update_data}
        )
        
        if result.modified_count > 0:
            logger.info(f"✅ Serviço {id} atualizado: {update_data}")
            return jsonify({'success': True, 'message': 'Serviço atualizado com sucesso'})
        
        return jsonify({'success': False, 'message': 'Nenhuma alteração realizada'}), 400
        
    except Exception as e:
        logger.error(f"❌ Erro ao atualizar serviço: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/servicos/toggle-todos', methods=['POST'])
@login_required
@permission_required('Admin', 'Gestão')
def toggle_todos_servicos():
    db = get_db()
    """Ativa ou desativa todos os serviços"""
    try:
        data = request.get_json()
        ativo = data.get('ativo', True)

        # Atualizar AMBOS os campos status e ativo para máxima compatibilidade
        result = db.servicos.update_many(
            {},
            {'$set': {
                'status': 'ativo' if ativo else 'inativo',
                'ativo': ativo
            }}
        )

        logger.info(f"✅ {result.modified_count} serviços {'ativados' if ativo else 'desativados'}")
        return jsonify({
            'success': True,
            'count': result.modified_count,
            'message': f'{result.modified_count} serviços foram {"ativados" if ativo else "desativados"} com sucesso'
        })

    except Exception as e:
        logger.error(f"❌ Erro ao toggle servicos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/servicos/deletar-todos', methods=['DELETE'])
@login_required
@permission_required('Admin')
def deletar_todos_servicos():
    db = get_db()
    """Deleta TODOS os serviços do sistema (apenas Admin)"""
    try:
        count_antes = db.servicos.count_documents({})
        result = db.servicos.delete_many({})

        logger.warning(f"🗑️ TODOS os serviços deletados: {result.deleted_count} registros removidos")
        return jsonify({
            'success': True,
            'count': result.deleted_count,
            'message': f'{result.deleted_count} serviços deletados com sucesso'
        })

    except Exception as e:
        logger.error(f"❌ Erro ao deletar todos servicos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/produtos/deletar-todos', methods=['DELETE'])
@login_required
@permission_required('Admin')
def deletar_todos_produtos():
    db = get_db()
    """Deleta TODOS os produtos do sistema (apenas Admin)"""
    try:
        count_antes = db.produtos.count_documents({})
        result = db.produtos.delete_many({})

        logger.warning(f"🗑️ TODOS os produtos deletados: {result.deleted_count} registros removidos")
        return jsonify({
            'success': True,
            'count': result.deleted_count,
            'message': f'{result.deleted_count} produtos deletados com sucesso'
        })

    except Exception as e:
        logger.error(f"❌ Erro ao deletar todos produtos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/produtos/toggle-todos', methods=['POST'])
@login_required
@permission_required('Admin', 'Gestão')
def toggle_todos_produtos():
    db = get_db()
    """Ativa ou desativa todos os produtos"""
    try:
        data = request.get_json()
        ativo = data.get('ativo', True)

        result = db.produtos.update_many(
            {},
            {'$set': {'ativo': ativo}}
        )

        logger.info(f"✅ {result.modified_count} produtos {'ativados' if ativo else 'desativados'}")
        return jsonify({
            'success': True,
            'count': result.modified_count,
            'message': f'{result.modified_count} produtos foram {"ativados" if ativo else "desativados"} com sucesso'
        })

    except Exception as e:
        logger.error(f"❌ Erro ao toggle produtos: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== ENDPOINTS PARA SUB-TABS - ESTOQUE ====================

@bp.route('/api/estoque/visao-geral', methods=['GET'])
@login_required
def estoque_visao_geral():
    db = get_db()
    """Retorna visão geral do estoque com estatísticas"""
    try:
        # Buscar todos os produtos ativos
        produtos = list(db.produtos.find({'status': 'Ativo'}))
        
        total_produtos = len(produtos)
        valor_total_estoque = 0
        alertas_estoque = 0
        
        produtos_formatados = []
        
        for p in produtos:
            estoque_atual = int(p.get('estoque', 0))
            estoque_minimo = int(p.get('estoque_minimo', 0))
            preco = float(p.get('preco', 0))
            valor_total = estoque_atual * preco
            
            valor_total_estoque += valor_total
            
            # Verificar alertas
            if estoque_atual <= estoque_minimo * 1.5:
                alertas_estoque += 1
            
            # Determinar status
            if estoque_atual <= estoque_minimo:
                nivel = 'Crítico'
            elif estoque_atual < estoque_minimo * 1.5:
                nivel = 'Baixo'
            else:
                nivel = 'Normal'
            
            produtos_formatados.append({
                'id': str(p['_id']),
                'nome': p.get('nome', 'Sem nome'),
                'marca': p.get('marca', 'Sem marca'),
                'estoque_atual': estoque_atual,
                'estoque_minimo': estoque_minimo,
                'preco_unitario': preco,
                'valor_total': round(valor_total, 2),
                'nivel': nivel
            })
        
        # Buscar movimentações do mês atual
        inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        movimentacoes_mes = db.estoque_movimentacoes.count_documents({
            'data': {'$gte': inicio_mes}
        })
        
        resultado = {
            'estatisticas': {
                'total_produtos': total_produtos,
                'valor_estoque': round(valor_total_estoque, 2),
                'alertas': alertas_estoque,
                'movimentacoes_mes': movimentacoes_mes
            },
            'produtos': produtos_formatados
        }
        
        logger.info(f"📊 Visão Geral - {total_produtos} produtos, R$ {valor_total_estoque:.2f}, {alertas_estoque} alertas")
        return jsonify({'success': True, 'data': resultado})
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar visão geral do estoque: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/estoque/alertas', methods=['GET'])
@login_required
def estoque_alertas():
    db = get_db()
    """Retorna alertas de estoque e últimas movimentações"""
    try:
        # Buscar produtos ativos
        produtos = list(db.produtos.find({'status': 'Ativo'}))
        
        criticos = []
        atencao = []
        normais = []
        
        for p in produtos:
            estoque_atual = int(p.get('estoque', 0))
            estoque_minimo = int(p.get('estoque_minimo', 0))
            diferenca = estoque_atual - estoque_minimo
            
            produto_info = {
                'id': str(p['_id']),
                'nome': p.get('nome', 'Sem nome'),
                'marca': p.get('marca', 'Sem marca'),
                'estoque_atual': estoque_atual,
                'estoque_minimo': estoque_minimo,
                'diferenca': diferenca
            }
            
            if estoque_atual <= estoque_minimo:
                produto_info['nivel'] = 'Crítico'
                criticos.append(produto_info)
            elif estoque_atual < estoque_minimo * 1.5:
                produto_info['nivel'] = 'Atenção'
                atencao.append(produto_info)
            else:
                normais.append(produto_info)
        
        # Buscar últimas 10 movimentações
        movimentacoes = list(db.estoque_movimentacoes.find().sort('data', DESCENDING).limit(10))
        
        movimentacoes_formatadas = []
        for m in movimentacoes:
            # Buscar nome do produto
            produto = db.produtos.find_one({'_id': ObjectId(m['produto_id'])})
            produto_nome = produto.get('nome', 'Desconhecido') if produto else 'Desconhecido'
            
            # Buscar nome do responsável
            responsavel_nome = 'Sistema'
            if m.get('responsavel_id'):
                responsavel = db.profissionais.find_one({'_id': ObjectId(m['responsavel_id'])})
                if not responsavel:
                    responsavel = db.assistentes.find_one({'_id': ObjectId(m['responsavel_id'])})
                if responsavel:
                    responsavel_nome = responsavel.get('nome', 'Desconhecido')
            
            movimentacoes_formatadas.append({
                'id': str(m['_id']),
                'data': m['data'].strftime('%d/%m/%Y'),
                'hora': m['data'].strftime('%H:%M'),
                'produto': produto_nome,
                'tipo': m.get('tipo', 'Entrada'),
                'quantidade': int(m.get('quantidade', 0)),
                'responsavel': responsavel_nome
            })
        
        resultado = {
            'estatisticas': {
                'criticos': len(criticos),
                'atencao': len(atencao),
                'normais': len(normais)
            },
            'produtos_baixo': criticos + atencao,
            'ultimas_movimentacoes': movimentacoes_formatadas
        }
        
        logger.info(f"⚠️ Alertas - Críticos: {len(criticos)}, Atenção: {len(atencao)}")
        return jsonify({'success': True, 'data': resultado})
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar alertas de estoque: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/estoque/relatorio', methods=['GET'])
@login_required
def gerar_relatorio_estoque():
    """Gera relatório de estoque personalizado"""
    try:
        tipo = request.args.get('tipo', 'movimentacoes')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        # Converter datas
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
        else:
            data_inicio = datetime.now() - timedelta(days=30)
        
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            data_fim = data_fim.replace(hour=23, minute=59, second=59)
        else:
            data_fim = datetime.now()
        
        resultado = {}
        
        if tipo == 'movimentacoes':
            # Relatório de movimentações
            movimentacoes = list(db.estoque_movimentacoes.find({
                'data': {'$gte': data_inicio, '$lte': data_fim}
            }).sort('data', DESCENDING))
            
            entradas = 0
            saidas = 0
            movs_formatadas = []
            
            for m in movimentacoes:
                produto = db.produtos.find_one({'_id': ObjectId(m['produto_id'])})
                produto_nome = produto.get('nome', 'Desconhecido') if produto else 'Desconhecido'
                
                responsavel_nome = 'Sistema'
                if m.get('responsavel_id'):
                    responsavel = db.profissionais.find_one({'_id': ObjectId(m['responsavel_id'])})
                    if not responsavel:
                        responsavel = db.assistentes.find_one({'_id': ObjectId(m['responsavel_id'])})
                    if responsavel:
                        responsavel_nome = responsavel.get('nome', 'Desconhecido')
                
                tipo_mov = m.get('tipo', 'Entrada')
                quantidade = int(m.get('quantidade', 0))
                
                if tipo_mov == 'Entrada':
                    entradas += quantidade
                else:
                    saidas += quantidade
                
                movs_formatadas.append({
                    'data': m['data'].strftime('%d/%m/%Y %H:%M'),
                    'tipo': tipo_mov,
                    'produto': produto_nome,
                    'quantidade': quantidade,
                    'motivo': m.get('motivo', ''),
                    'responsavel': responsavel_nome
                })
            
            resultado = {
                'tipo': 'movimentacoes',
                'periodo': {
                    'inicio': data_inicio.strftime('%d/%m/%Y'),
                    'fim': data_fim.strftime('%d/%m/%Y')
                },
                'resumo': {
                    'total_movimentacoes': len(movimentacoes),
                    'total_entradas': entradas,
                    'total_saidas': saidas,
                    'saldo': entradas - saidas
                },
                'movimentacoes': movs_formatadas
            }
            
        elif tipo == 'posicao':
            # Relatório de posição de estoque
            produtos = list(db.produtos.find({'status': 'Ativo'}))
            
            produtos_formatados = []
            for p in produtos:
                produtos_formatados.append({
                    'nome': p.get('nome', 'Sem nome'),
                    'marca': p.get('marca', 'Sem marca'),
                    'estoque_atual': int(p.get('estoque', 0)),
                    'estoque_minimo': int(p.get('estoque_minimo', 0)),
                    'status': 'Normal' if int(p.get('estoque', 0)) > int(p.get('estoque_minimo', 0)) * 1.5 else ('Baixo' if int(p.get('estoque', 0)) > int(p.get('estoque_minimo', 0)) else 'Crítico')
                })
            
            resultado = {
                'tipo': 'posicao',
                'data': datetime.now().strftime('%d/%m/%Y %H:%M'),
                'total_produtos': len(produtos),
                'produtos': produtos_formatados
            }
            
        elif tipo == 'valorizado':
            # Relatório de estoque valorizado
            produtos = list(db.produtos.find({'status': 'Ativo'}))
            
            valor_total = 0
            produtos_formatados = []
            
            for p in produtos:
                estoque = int(p.get('estoque', 0))
                preco = float(p.get('preco', 0))
                valor = estoque * preco
                valor_total += valor
                
                produtos_formatados.append({
                    'nome': p.get('nome', 'Sem nome'),
                    'marca': p.get('marca', 'Sem marca'),
                    'estoque': estoque,
                    'preco_unitario': preco,
                    'valor_total': round(valor, 2)
                })
            
            # Ordenar por valor total
            produtos_formatados.sort(key=lambda x: x['valor_total'], reverse=True)
            
            resultado = {
                'tipo': 'valorizado',
                'data': datetime.now().strftime('%d/%m/%Y %H:%M'),
                'valor_total_estoque': round(valor_total, 2),
                'total_produtos': len(produtos),
                'produtos': produtos_formatados
            }
            
        elif tipo == 'criticos':
            # Relatório de produtos críticos
            produtos = list(db.produtos.find({'status': 'Ativo'}))
            
            criticos = []
            atencao = []
            
            for p in produtos:
                estoque_atual = int(p.get('estoque', 0))
                estoque_minimo = int(p.get('estoque_minimo', 0))
                
                if estoque_atual <= estoque_minimo:
                    criticos.append({
                        'nome': p.get('nome', 'Sem nome'),
                        'marca': p.get('marca', 'Sem marca'),
                        'estoque_atual': estoque_atual,
                        'estoque_minimo': estoque_minimo,
                        'diferenca': estoque_atual - estoque_minimo,
                        'nivel': 'Crítico'
                    })
                elif estoque_atual < estoque_minimo * 1.5:
                    atencao.append({
                        'nome': p.get('nome', 'Sem nome'),
                        'marca': p.get('marca', 'Sem marca'),
                        'estoque_atual': estoque_atual,
                        'estoque_minimo': estoque_minimo,
                        'diferenca': estoque_atual - estoque_minimo,
                        'nivel': 'Atenção'
                    })
            
            resultado = {
                'tipo': 'criticos',
                'data': datetime.now().strftime('%d/%m/%Y %H:%M'),
                'total_criticos': len(criticos),
                'total_atencao': len(atencao),
                'produtos_criticos': criticos,
                'produtos_atencao': atencao
            }
        
        logger.info(f"📄 Relatório gerado: {tipo} ({data_inicio.strftime('%d/%m/%Y')} - {data_fim.strftime('%d/%m/%Y')})")
        return jsonify({'success': True, 'relatorio': resultado})
        
    except Exception as e:
        logger.error(f"❌ Erro ao gerar relatório de estoque: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== FIM ENDPOINTS SUB-TABS ====================


# ==================== ROTAS ADMINISTRATIVAS ====================

@bp.route('/api/admin/reset-database', methods=['POST'])
@login_required
def admin_reset_database():
    db = get_db()
    """
    LIMPAR TODO O BANCO DE DADOS (apenas Admin)
    ATENÇÃO: Esta operação é IRREVERSÍVEL!
    """
    try:
        # Verificar se o usuário é Admin
        user_id = session.get('user_id')
        user = db.users.find_one({'_id': ObjectId(user_id)})

        if not user or user.get('tipo_acesso') != 'Admin':
            return jsonify({'success': False, 'message': 'Apenas administradores podem resetar o banco'}), 403

        # Confirmar com senha
        data = request.get_json()
        senha_confirmacao = data.get('senha')

        if not senha_confirmacao:
            return jsonify({'success': False, 'message': 'Senha de confirmação necessária'}), 400

        if not check_password_hash(user.get('password', ''), senha_confirmacao):
            return jsonify({'success': False, 'message': 'Senha incorreta'}), 401

        # RESETAR BANCO DE DADOS
        collections_to_reset = [
            'clientes', 'profissionais', 'servicos', 'produtos',
            'agendamentos', 'fila', 'orcamentos', 'contratos',
            'anamneses', 'prontuarios', 'financeiro_despesas',
            'financeiro_comissoes', 'estoque_movimentos',
            'notificacoes', 'uploads', 'auditoria'
        ]

        reset_count = {}
        for collection_name in collections_to_reset:
            count = db[collection_name].count_documents({})
            result = db[collection_name].delete_many({})
            reset_count[collection_name] = count

        logger.warning(f"🗑️ BANCO DE DADOS RESETADO por {user.get('username')}")
        logger.info(f"Coleções resetadas: {reset_count}")

        # Registrar auditoria
        registrar_auditoria(
            acao='reset_database',
            entidade='sistema',
            entidade_id='global',
            dados_antes={'collections': reset_count},
            dados_depois={'status': 'limpo'}
        )

        return jsonify({
            'success': True,
            'message': 'Banco de dados resetado com sucesso',
            'collections_reset': reset_count,
            'total_documentos_removidos': sum(reset_count.values())
        })

    except Exception as e:
        logger.error(f"❌ Erro ao resetar banco: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/admin/database-stats', methods=['GET'])
@login_required
def admin_database_stats():
    db = get_db()
    """Estatísticas do banco de dados (apenas Admin)"""
    try:
        # Verificar se o usuário é Admin
        user_id = session.get('user_id')
        user = db.users.find_one({'_id': ObjectId(user_id)})

        if not user or user.get('tipo_acesso') != 'Admin':
            return jsonify({'success': False, 'message': 'Apenas administradores'}), 403

        stats = {
            'clientes': db.clientes.count_documents({}),
            'profissionais': db.profissionais.count_documents({}),
            'servicos': db.servicos.count_documents({}),
            'produtos': db.produtos.count_documents({}),
            'agendamentos': db.agendamentos.count_documents({}),
            'fila': db.fila.count_documents({}),
            'orcamentos': db.orcamentos.count_documents({}),
            'contratos': db.contratos.count_documents({}),
            'anamneses': db.anamneses.count_documents({}),
            'prontuarios': db.prontuarios.count_documents({}),
            'usuarios': db.users.count_documents({}),
            'auditoria': db.auditoria.count_documents({}),
        }

        return jsonify({'success': True, 'stats': stats, 'total': sum(stats.values())})

    except Exception as e:
        logger.error(f"❌ Erro ao obter stats: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== FIM ROTAS ADMINISTRATIVAS ====================

# ==================== ROTAS ADICIONAIS ====================

@bp.route('/api/stream')
@login_required
def stream_updates():
    """Route for Server-Sent Events (SSE) for real-time updates - OTIMIZADO v4.0"""
    def generate():
        """Generator function for SSE - apenas heartbeat (sem queries no DB)"""
        try:
            # Enviar heartbeat inicial
            yield f"data: {json.dumps({'type': 'connected', 'message': 'SSE conectado com sucesso'})}\n\n"

            # Loop para manter conexão viva
            import time
            counter = 0
            max_duration = 86400  # 24 horas em segundos (limite de conexão)
            start_time = time.time()

            while True:
                counter += 1

                # Verificar se ultrapassou o tempo máximo de conexão
                if time.time() - start_time > max_duration:
                    logger.info("SSE: Tempo máximo de conexão atingido (24h)")
                    break

                # Enviar apenas heartbeat leve (SEM queries no banco)
                # Os dados devem ser buscados via refresh manual ou polling no frontend
                data = {
                    'type': 'heartbeat',
                    'timestamp': datetime.now().isoformat(),
                    'counter': counter
                }

                yield f"data: {json.dumps(data)}\n\n"

                # Aguardar 60 segundos (vs 30s anterior = 50% menos requisições)
                time.sleep(60)

        except GeneratorExit:
            logger.info("Cliente SSE desconectado")
        except Exception as e:
            logger.error(f"Erro no SSE stream: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    from flask import Response
    # Headers otimizados para SSE
    response = Response(generate(), mimetype='text/event-stream')
    response.headers['Cache-Control'] = 'no-cache, no-transform'
    response.headers['X-Accel-Buffering'] = 'no'  # Desabilita buffering no NGINX
    return response


@bp.route('/api/agendamentos/heatmap', methods=['GET'])
@login_required
def heatmap_agendamentos_alias():
    """Alias route for heatmap - redirects to the actual route"""
    # Redirecionar para a rota existente
    dias = request.args.get('dias', 60, type=int)
    return mapa_calor_agendamentos()


# ==================== FILA DE ATENDIMENTO ====================
# REMOVIDO - Rotas duplicadas movidas para linhas 2122-2142
# As rotas /api/fila já estão definidas anteriormente no arquivo

# Nota: removi get_fila() e add_to_fila() pois conflitavam com a definição anterior
# A rota funcional está na linha 2122 com métodos GET e POST


# REMOVIDO - Todas as rotas de fila duplicadas foram removidas
# As rotas funcionais de fila estão definidas nas linhas 2122-2200
# Isso resolve o erro 500 causado por rotas duplicadas e uso incorreto de 'db'


# ==================== FIM ROTAS ADICIONAIS ====================

